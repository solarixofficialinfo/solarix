from __future__ import annotations
from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import json
import math
import io
import re
import uuid
import logging
import secrets
import requests  # type: ignore
import bcrypt
import jwt
import time
import threading
import asyncio
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any, Tuple, Union
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form, Query
from fastapi.responses import Response as FastAPIResponse, StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from pydantic import BaseModel, EmailStr, field_validator
import pdf_generator
from email_service import send_email, render_otp_email, render_password_changed_email
import httpx
from supabase import create_client as supabase_create_client, Client, ClientOptions

import contextvars
from bson import ObjectId

# ─── In-process auth token cache ─────────────────────────────────────────────
# Caches (user_id → user_profile) keyed by bearer token.
# Each entry expires after TOKEN_CACHE_TTL_S seconds.
# This eliminates the 2 sequential Supabase HTTP round-trips on every request.
TOKEN_CACHE_TTL_S = 300  # 5 minutes
_auth_cache: Dict[str, Dict] = {}          # token -> {"user": dict, "exp": float}
_auth_cache_lock = threading.Lock()

# Test helpers for credentials caching and routing
_test_temp_passwords: Dict[str, str] = {}


def _cache_get_user(token: str) -> Optional[Dict]:
    """Return cached user dict if still fresh, else None."""
    with _auth_cache_lock:
        entry = _auth_cache.get(token)
        if entry and entry["exp"] > time.monotonic():
            return entry["user"]
        if entry:
            del _auth_cache[token]   # evict stale
    return None

def _cache_put_user(token: str, user: Dict) -> None:
    """Store user dict in cache with TTL. Prune if cache grows large."""
    with _auth_cache_lock:
        if len(_auth_cache) > 2000:  # hard cap – evict oldest 500
            oldest = sorted(_auth_cache, key=lambda k: _auth_cache[k]["exp"])[:500]
            for k in oldest:
                del _auth_cache[k]
        _auth_cache[token] = {"user": user, "exp": time.monotonic() + TOKEN_CACHE_TTL_S}

def _cache_invalidate_user(user_id: str) -> None:
    """Remove all cache entries for a given user_id (e.g. after role/permission change)."""
    with _auth_cache_lock:
        stale = [k for k, v in _auth_cache.items() if v["user"].get("id") == user_id]
        for k in stale:
            del _auth_cache[k]

COMPANY_CACHE_TTL_S = 600  # 10 minutes
_company_cache: Dict[str, Dict] = {}        # company_id -> {"company": dict, "exp": float}
_company_cache_lock = threading.Lock()

def _cache_get_company(company_id: str) -> Optional[Dict]:
    if not company_id: return None
    with _company_cache_lock:
        entry = _company_cache.get(company_id)
        if entry and entry["exp"] > time.monotonic():
            return entry["company"]
        if entry:
            del _company_cache[company_id]
    return None

def _cache_put_company(company_id: str, company: Dict) -> None:
    if not company_id or not company: return
    with _company_cache_lock:
        if len(_company_cache) > 1000:
            oldest = sorted(_company_cache, key=lambda k: _company_cache[k]["exp"])[:200]
            for k in oldest:
                del _company_cache[k]
        _company_cache[company_id] = {"company": company, "exp": time.monotonic() + COMPANY_CACHE_TTL_S}

def _cache_invalidate_company(company_id: str) -> None:
    if not company_id: return
    with _company_cache_lock:
        _company_cache.pop(company_id, None)
    _company_logo_cache.clear()

supabase_url = os.environ['SUPABASE_URL']
# Primary key used historically (may be anon or service role)
supabase_key = os.environ.get('SUPABASE_KEY')
# Optional explicit service-role key for privileged RPCs (DO NOT commit this value)
supabase_service_key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')

# Tight timeout — Vercel functions have a 10-second hard limit; fail fast, don't retry.
_shared_timeout = httpx.Timeout(8.0, connect=5.0)
# retries=0 — each retry triples the timeout; on Vercel this guarantees a 504.
_shared_transport = httpx.HTTPTransport(retries=0)

# ── Supabase client cache ────────────────────────────────────────────────────
# Creating a new httpx.Client + supabase_create_client on EVERY request is very
# expensive (new TCP connection pool each time). Cache clients keyed by token.
import functools
_client_cache: Dict[Tuple[Optional[str], bool], Any] = {}
_client_cache_lock = threading.Lock()
_MAX_CLIENT_CACHE = 64  # prevent unbounded growth

def get_supabase_client(token: Optional[str] = None, use_service_key: bool = False):
    cache_key = (token, use_service_key)
    with _client_cache_lock:
        if cache_key in _client_cache:
            return _client_cache[cache_key]

    httpx_client = httpx.Client(timeout=_shared_timeout, transport=_shared_transport)
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    key = (supabase_service_key if (use_service_key and supabase_service_key) else supabase_key) or supabase_key
    opts = ClientOptions(
        httpx_client=httpx_client,
        postgrest_client_timeout=8.0,
        storage_client_timeout=8,
        function_client_timeout=8,
        headers=headers
    )
    client = supabase_create_client(supabase_url, key or "", options=opts)
    with _client_cache_lock:
        if len(_client_cache) >= _MAX_CLIENT_CACHE:
            # Evict oldest entry
            try:
                oldest = next(iter(_client_cache))
                del _client_cache[oldest]
            except StopIteration:
                pass
        _client_cache[cache_key] = client
    return client

# Default client (used for request-scoped and anon operations)
default_supabase = get_supabase_client() if supabase_key else None
# Service client (use for admin RPCs that require elevated privileges)
service_supabase = get_supabase_client(use_service_key=True) if supabase_service_key else None

_supabase_var: contextvars.ContextVar[Client | None] = contextvars.ContextVar("supabase", default=None)

def get_rpc_client() -> Client:
    client = service_supabase if service_supabase is not None else default_supabase
    if client is None:
        raise HTTPException(status_code=500, detail="Database client not initialized")
    return client

async def _record_workflow_details(task: dict, user: dict):
    task_type = task.get("task_type")
    task_id = task.get("id")
    company_id = task.get("company_id")
    client_id = task.get("client_id")
    submission = task.get("submission") or {}

    project_id = client_id
    proj = await db.projects.find_one({"company_id": company_id, "client_id": client_id})
    if proj:
        project_id = proj.get("id") or project_id

    # Form details object
    photos = submission.get("photos") or submission.get("attachments") or {}
    details = {
        "completed_by": user.get("name") or "",
        "completed_by_id": user.get("id") or "",
        "assigned_to_name": task.get("assigned_to_name") or user.get("name") or "",
        "assigned_by": task.get("assigned_by_name") or "",
        "assigned_by_name": task.get("assigned_by_name") or "",
        "completed_date": submission.get("submitted_at") or now_iso(),
        "submitted_at": submission.get("submitted_at") or now_iso(),
        "notes": submission.get("notes") or submission.get("remarks") or task.get("remarks") or "",
        "gps": submission.get("gps") or "",
        "manual_location": submission.get("manual_location") or "",
        "checklist": submission.get("checklist") or [],
        "photos": photos,
        "attachments": photos,
        "task_status": "completed",
    }

    # Custom adjustments for specific task types
    table_name = None
    if task_type == "Survey":
        table_name = "surveys"
    elif task_type in ("Material Delivery", "Material Dispatch"):
        table_name = "material_deliveries"
        # Fetch material request to populate checklist/attachments if not present in submission
        req = await db.material_requests.find_one({"company_id": company_id, "client_id": client_id})
        if req:
            details["checklist"] = [
                {"label": f"{it.get('product')} (Qty: {it.get('quantity')}, Approved: {it.get('approved_quantity')})", "checked": True}
                for it in (req.get("items") or [])
            ]
            del_info = req.get("delivery") or {}
            details["attachments"] = {
                "Delivery Photo": del_info.get("delivery_photo_file_id") or "",
                "Challan Photo": del_info.get("challan_photo_file_id") or "",
            }
            if req.get("remarks"):
                details["notes"] = req.get("remarks")
    elif task_type == "Document Signed":
        table_name = "documents"
    elif task_type in ("Meter Testing Request", "Meter Testing Completed"):
        table_name = "meter_testings"
    elif task_type == "Installation":
        table_name = "installations"
    elif task_type == "Handover":
        table_name = "handovers"
        h_photo = submission.get("handover_photo_id") or submission.get("photo_id") or submission.get("file_id")
        if h_photo:
            photos["Handover Photo"] = h_photo
            details["photos"] = photos
            details["attachments"] = photos
        chk = list(details.get("checklist") or [])
        if submission.get("declaration_confirmed"):
            chk.append({"label": "Owner Declaration Confirmed", "checked": True})
        if submission.get("installer_confirmed"):
            chk.append({"label": "Installer Confirmed", "checked": True})
        if submission.get("owner_name"):
            chk.append({"label": f"Owner: {submission.get('owner_name')}", "checked": True})
        if chk:
            details["checklist"] = chk
    elif task_type == "PM Surya Ghar Upload":
        table_name = "pm_surya_uploads"
    elif task_type == "MSEDCL Upload":
        table_name = "msedcl_uploads"
    elif task_type == "Verification":
        # Verification already has its own table, we can sync or let it be
        pass

    if table_name:
        doc = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "client_id": client_id,
            "project_id": project_id,
            "task_id": task_id,
            "employee_id": task.get("assigned_to"),
            "details": details,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        # Check if record already exists for this task_id
        existing = await db[table_name].find_one({"company_id": company_id, "task_id": task_id})
        if existing:
            await db[table_name].update_one(
                {"id": existing["id"]},
                {"$set": {
                    "details": details,
                    "updated_at": now_iso()
                }}
            )
        else:
            await db[table_name].insert_one(doc)

class SupabaseProxy:
    def __getattr__(self, name):
        client = _supabase_var.get()
        if client is None:
            client = default_supabase
        return getattr(client, name)

supabase = SupabaseProxy()


class InsertOneResult:
    def __init__(self, inserted_id):
        self.inserted_id = inserted_id

class InsertManyResult:
    def __init__(self, inserted_ids):
        self.inserted_ids = inserted_ids

class UpdateResult:
    def __init__(self, matched_count, modified_count):
        self.matched_count = matched_count
        self.modified_count = modified_count

class DeleteResult:
    def __init__(self, deleted_count):
        self.deleted_count = deleted_count

class AsyncIteratorWrapper:
    def __init__(self, coro):
        self.coro = coro
        self.data = None
        self.idx = 0

    async def _load(self):
        if self.data is None:
            self.data = await self.coro

    def __aiter__(self):
        return self

    async def __anext__(self):
        await self._load()
        if self.data is not None and self.idx < len(self.data):
            val = self.data[self.idx]
            self.idx += 1
            return val
        else:
            raise StopAsyncIteration

def get_nested_val(d, path):
    parts = path.split(".")
    curr = d
    for p in parts:
        if isinstance(curr, dict):
            curr = curr.get(p)
        else:
            return None
    return curr

class AggregateCursorAdapter:
    def __init__(self, table_name, pipeline):
        self.table_name = table_name
        self.pipeline = pipeline

    async def to_list(self, length=None):
        match_dict = {}
        group_dict = {}
        for stage in self.pipeline:
            if "$match" in stage:
                match_dict = stage["$match"]
            elif "$group" in stage:
                group_dict = stage["$group"]
        
        select_cols = "*"
        if self.table_name == "clients":
            select_cols = "id,company_id,status,system_kw,created_at,updated_at"
        elif self.table_name == "inverter_monitoring":
            select_cols = "id,company_id,inverter_status"
        elif self.table_name == "service_tickets":
            select_cols = "id,company_id,status"

        builder = supabase.table(self.table_name).select(select_cols)
        if "company_id" in match_dict and not isinstance(match_dict["company_id"], dict):
            builder = builder.eq("company_id", match_dict["company_id"])
        
        builder = builder.limit(100000)
        try:
            res = builder.execute()
            rows = res.data or []
        except Exception as e:
            logger.warning(f"AggregateCursorAdapter query failed for {self.table_name}: {e}")
            rows = []
        
        if not rows:
            local_rows = LocalFileCollection(self.table_name)._read_data()
            if local_rows:
                existing_ids = {r.get("id") for r in rows if isinstance(r, dict) and r.get("id")}
                for lr in local_rows:
                    if lr.get("id") not in existing_ids:
                        rows.append(lr)
        
        filtered_rows = []
        for row in rows:
            match = True
            for k, v in match_dict.items():
                col_val = get_nested_val(row, k)
                if isinstance(v, dict):
                    for op, val in v.items():
                        if op == "$ne" and col_val == val:
                            match = False
                        elif op == "$in" and col_val not in val:
                            match = False
                        elif op == "$nin" and col_val in val:
                            match = False
                        elif op == "$gt" and not (col_val > val):
                            match = False
                        elif op == "$gte" and not (col_val >= val):
                            match = False
                        elif op == "$lt" and not (col_val < val):
                            match = False
                        elif op == "$lte" and not (col_val <= val):
                            match = False
                else:
                    if col_val != v:
                        match = False
            if match:
                filtered_rows.append(row)
                
        group_by_field = group_dict.get("_id")
        
        groups = {}
        for row in filtered_rows:
            if isinstance(group_by_field, dict):
                eval_id = {}
                for alias, expr in group_by_field.items():
                    if isinstance(expr, str) and expr.startswith("$"):
                        f_name = expr[1:]
                        eval_id[alias] = row.get(f_name)
                    else:
                        eval_id[alias] = expr
                g_key = tuple(sorted((k, str(v or "").strip()) for k, v in eval_id.items()))
                if g_key not in groups:
                    groups[g_key] = {"_id": eval_id, "rows": []}
                groups[g_key]["rows"].append(row)
            elif isinstance(group_by_field, str) and group_by_field.startswith("$"):
                f_name = group_by_field[1:]
                g_val = row.get(f_name)
                g_key = str(g_val) if g_val is not None else None
                if g_key not in groups:
                    groups[g_key] = {"_id": g_val, "rows": []}
                groups[g_key]["rows"].append(row)
            else:
                g_key = None
                if g_key not in groups:
                    groups[g_key] = {"_id": None, "rows": []}
                groups[g_key]["rows"].append(row)
            
        result = []
        for g_key, grp_data in groups.items():
            group_rows = grp_data["rows"]
            out = {"_id": grp_data["_id"]}
            for agg_k, agg_v in group_dict.items():
                if agg_k == "_id":
                    continue
                if isinstance(agg_v, dict):
                    for op, val in agg_v.items():
                        if op == "$sum":
                            if isinstance(val, (int, float)):
                                out[agg_k] = sum(val for _ in group_rows)
                            elif isinstance(val, str) and val.startswith("$"):
                                f_name = val[1:]
                                out[agg_k] = sum(float(r.get(f_name) or 0) for r in group_rows)
                        elif op == "$max":
                            if isinstance(val, str) and val.startswith("$"):
                                f_name = val[1:]
                                vals = [r.get(f_name) for r in group_rows if r.get(f_name) is not None]
                                out[agg_k] = max(vals) if vals else None
            result.append(out)
            
        return result

    def __aiter__(self):
        return AsyncIteratorWrapper(self.to_list())

class CursorAdapter:
    def __init__(self, collection, filter, projection):
        self.collection = collection
        self.filter = filter
        self.projection = projection
        self.sort_fields = None
        self.limit_val = None
        self.skip_val = None

    def sort(self, key_or_list, direction=None):
        if isinstance(key_or_list, list):
            self.sort_fields = key_or_list
        else:
            self.sort_fields = [(key_or_list, direction or 1)]
        return self

    def limit(self, val):
        self.limit_val = val
        return self

    def skip(self, val):
        self.skip_val = val
        return self

    async def to_list(self, length=None):
        select_cols = "*"
        if self.projection and isinstance(self.projection, dict):
            inclusions = [k for k, v in self.projection.items() if (v == 1 or v is True) and "->" not in k and "." not in k]
            if self.collection.table_name == "products" and not _PRODUCTS_HAS_OPENING_STOCK and "opening_stock" in inclusions:
                inclusions.remove("opening_stock")
            if inclusions:
                select_cols = ",".join(inclusions)
        builder = supabase.table(self.collection.table_name).select(select_cols)
        
        # Intercept and extract unsupported filters for files table
        filter_to_apply = self.filter
        extracted_filters = {}
        if self.collection.table_name == "files" and self.filter:
            extracted_filters, cleaned_filter = self.collection._extract_and_remove_unsupported_filters(self.filter)
            filter_to_apply = cleaned_filter

        builder = self.collection._apply_filters(builder, filter_to_apply)
        
        if self.sort_fields:
            for k, dir in self.sort_fields:
                desc = (dir == -1)
                builder = builder.order(k, desc=desc)
        
        limit = length if length is not None else self.limit_val
        skip = self.skip_val or 0
        
        if self.collection.table_name == "files" and extracted_filters:
            builder = builder.limit(1000)
        else:
            if limit is not None:
                builder = builder.range(skip, skip + limit - 1)
            elif skip > 0:
                builder = builder.range(skip, 1000000)

        try:
            res = await asyncio.to_thread(builder.execute)
            data = res.data or []
        except Exception as e:
            err_str = str(e).lower()
            if "42501" in err_str or "row-level security" in err_str or "unauthorized" in err_str or "timeout" in err_str or "timed out" in err_str or "connection" in err_str or "400" in err_str or "bad request" in err_str or "pgrst" in err_str:
                logger.warning(f"Supabase query failed ({e}), falling back to local files for {self.collection.table_name}")
                return await LocalFileCollection(self.collection.table_name).find(self.filter, self.projection).sort(self.sort_fields).to_list(length)
            elif "pgrst205" in err_str or "does not exist" in err_str or "schema cache" in err_str:
                data = []
            else:
                raise e
        
        # ── Local-file merge (fallback/offline data) ─────────────────────────────
        # Skip this expensive disk read for the products table when Supabase already
        # returned data — products.json can be 26KB+, reading it on every request
        # was the primary cause of the 30-second product-search delay.
        skip_local_merge = len(data) > 0
        if not skip_local_merge:
            local_records = await LocalFileCollection(self.collection.table_name).find(self.filter, self.projection).sort(self.sort_fields).to_list(length)
            if local_records:
                existing_ids = {d.get("id") for d in data if isinstance(d, dict) and d.get("id")}
                for lr in local_records:
                    if lr.get("id") not in existing_ids:
                        data.append(lr)

        deserialized_data = []
        for doc in data:
            doc = self.collection._deserialize_document(doc)
            if self.collection.table_name == "files" and extracted_filters:
                if not self.collection._matches_filter(doc, extracted_filters):
                    continue
            deserialized_data.append(doc)

        if self.projection:
            for doc in deserialized_data:
                for pk, pv in self.projection.items():
                    if pv == 0:
                        doc.pop(pk, None)
                        
        if self.collection.table_name == "files" and extracted_filters:
            if limit is not None:
                deserialized_data = deserialized_data[skip:skip + limit]
            elif skip > 0:
                deserialized_data = deserialized_data[skip:]
                
        return deserialized_data

    def __aiter__(self):
        return AsyncIteratorWrapper(self.to_list())

_PRODUCTS_HAS_RATE = True
_PRODUCTS_HAS_OPENING_STOCK = True
_PRODUCTS_HAS_HV = True
_PRODUCTS_HAS_SN_REQ = True

def _clean_products_doc(doc: dict) -> dict:
    cleaned = dict(doc)
    if not _PRODUCTS_HAS_RATE:
        cleaned.pop("rate", None)
    if not _PRODUCTS_HAS_OPENING_STOCK:
        cleaned.pop("opening_stock", None)
    if not _PRODUCTS_HAS_HV:
        cleaned.pop("high_value_goods", None)
    if not _PRODUCTS_HAS_SN_REQ:
        cleaned.pop("serial_number_required", None)
    return cleaned

_INWARD_VALID_COLS = {
    "id", "company_id", "product", "size", "quantity", "unit",
    "reference_number", "reference_type", "bill_number", "source_type",
    "source_name", "date", "remarks", "attachment_file_id",
    "attachment_filename", "source", "created_by", "created_by_name",
    "created_at", "import_batch", "updated_at"
}

def _clean_inward_doc(doc: dict) -> dict:
    return {k: v for k, v in doc.items() if k in _INWARD_VALID_COLS}

VALID_CLIENT_COLUMNS = {
    "id", "sol_id", "company_id", "created_by", "full_name", "mobile", "alt_mobile",
    "consumer_number", "address", "city", "state", "pincode", "aadhaar", "system_kw",
    "panel_make", "panel_wattage", "num_panels", "inverter_make", "inverter_capacity",
    "inverter_serial", "phase_type", "subsidy_eligible", "status", "stages", "documents",
    "inverters", "progress", "notes", "created_at", "updated_at"
}

def _prepare_client_supabase_payload(payload: dict) -> dict:
    cleaned = {}
    extra_onboarding = {}

    panel_brand_val = payload.get("panel_brand") or payload.get("panel_make")
    if panel_brand_val is not None:
        cleaned["panel_make"] = panel_brand_val
        extra_onboarding["panel_brand"] = panel_brand_val
        extra_onboarding["panel_make"] = panel_brand_val

    inv_brand_val = payload.get("inverter_make") or payload.get("inverter_brand")
    if inv_brand_val is not None:
        cleaned["inverter_make"] = inv_brand_val
        extra_onboarding["inverter_brand"] = inv_brand_val
        extra_onboarding["inverter_make"] = inv_brand_val

    sec_no_val = payload.get("section_number") if "section_number" in payload else payload.get("section_no")
    if sec_no_val is not None:
        extra_onboarding["section_number"] = sec_no_val
        extra_onboarding["section_no"] = sec_no_val

    cat_val = payload.get("consumer_type") if "consumer_type" in payload else (payload.get("consumer_category") if "consumer_category" in payload else payload.get("category"))
    if cat_val is not None:
        extra_onboarding["consumer_type"] = cat_val
        extra_onboarding["consumer_category"] = cat_val
        extra_onboarding["category"] = cat_val

    sr_val = payload.get("inverter_serial") if "inverter_serial" in payload else payload.get("inverter_sr")
    if sr_val is not None:
        cleaned["inverter_serial"] = sr_val
        extra_onboarding["inverter_serial"] = sr_val
        extra_onboarding["inverter_sr"] = sr_val

    yr_val = payload.get("inverter_year") if "inverter_year" in payload else payload.get("manufacturing_year")
    if yr_val is not None:
        extra_onboarding["inverter_year"] = yr_val
        extra_onboarding["manufacturing_year"] = yr_val

    if "panel_technology" in payload and payload["panel_technology"] is not None:
        extra_onboarding["panel_technology"] = payload["panel_technology"]
    if "inverter_model" in payload and payload["inverter_model"] is not None:
        extra_onboarding["inverter_model"] = payload["inverter_model"]
    if "sanction_number" in payload and payload["sanction_number"] is not None:
        extra_onboarding["sanction_number"] = payload["sanction_number"]
    if "aadhaar_name" in payload and payload["aadhaar_name"] is not None:
        extra_onboarding["aadhaar_name"] = payload["aadhaar_name"]
    if "aadhaar_image" in payload and payload["aadhaar_image"] is not None:
        extra_onboarding["aadhaar_image"] = payload["aadhaar_image"]
    if "bu_number" in payload and payload["bu_number"] is not None:
        extra_onboarding["bu_number"] = payload["bu_number"]
    if "bu_text" in payload and payload["bu_text"] is not None:
        extra_onboarding["bu_text"] = payload["bu_text"]
    if "pan_number" in payload and payload["pan_number"] is not None:
        extra_onboarding["pan_number"] = payload["pan_number"]
    if "pan_card_number" in payload and payload["pan_card_number"] is not None:
        extra_onboarding["pan_number"] = payload["pan_card_number"]
    if "add_no" in payload and payload["add_no"] is not None:
        extra_onboarding["add_no"] = payload["add_no"]
    if "address_no" in payload and payload["address_no"] is not None:
        extra_onboarding["add_no"] = payload["address_no"]

    # Extended Location Fields
    for loc_key in ["district", "landmark", "latitude", "longitude", "state_code", "formatted_address"]:
        if loc_key in payload and payload[loc_key] is not None:
            extra_onboarding[loc_key] = payload[loc_key]

    if "inverters" in payload and payload["inverters"] is not None:
        extra_onboarding["inverters"] = payload["inverters"]
        cleaned["inverters"] = payload["inverters"]

    for k, v in payload.items():
        if k in VALID_CLIENT_COLUMNS:
            cleaned[k] = v

    stages_dict = dict(payload.get("stages") or {})
    existing_ob = dict(stages_dict.get("onboarding_data") or {})
    existing_ob.update(extra_onboarding)
    stages_dict["onboarding_data"] = existing_ob
    cleaned["stages"] = stages_dict
    return cleaned

def _enrich_client_doc(c: dict) -> dict:
    if not isinstance(c, dict):
        return c
    stages = dict(c.get("stages") or {})
    ob = dict(stages.get("onboarding_data") or {})

    p_brand = c.get("panel_brand") or c.get("panel_make") or ob.get("panel_brand") or ob.get("panel_make") or ""
    c["panel_brand"] = p_brand
    c["panel_make"] = p_brand
    inv_brand = c.get("inverter_brand") or c.get("inverter_make") or ob.get("inverter_brand") or ob.get("inverter_make") or ""
    c["inverter_brand"] = inv_brand
    c["inverter_make"] = inv_brand
    c["consumer_type"] = ob.get("consumer_type") or c.get("consumer_type") or ob.get("consumer_category") or ""
    c["consumer_category"] = c["consumer_type"]
    c["category"] = c["consumer_type"]
    c["panel_technology"] = ob.get("panel_technology") or c.get("panel_technology") or ""
    c["inverter_model"] = ob.get("inverter_model") or c.get("inverter_model") or ""
    c["inverter_year"] = ob.get("inverter_year") or c.get("inverter_year") or ob.get("manufacturing_year") or ""
    c["manufacturing_year"] = c["inverter_year"]
    c["section_number"] = ob.get("section_number") or c.get("section_number") or ob.get("section_no") or ""
    c["section_no"] = c["section_number"]
    c["sanction_number"] = ob.get("sanction_number") or c.get("sanction_number") or ""
    c["inverter_serial"] = c.get("inverter_serial") or ob.get("inverter_serial") or ob.get("inverter_sr") or ""
    c["inverter_sr"] = c["inverter_serial"]
    bu_val = ob.get("bu_number") or c.get("bu_number") or ob.get("bu_no") or c.get("bu_no") or ob.get("bu") or c.get("bu") or ""
    c["bu_number"] = bu_val
    c["bu_no"] = bu_val
    c["bu"] = bu_val
    c["bu_text"]   = ob.get("bu_text")   or c.get("bu_text")   or ""
    c["pan_number"] = ob.get("pan_number") or ob.get("pan_card_number") or c.get("pan_number") or c.get("pan_card_number") or ""
    c["pan_card_number"] = c["pan_number"]
    c["add_no"] = ob.get("add_no") or ob.get("address_no") or c.get("add_no") or c.get("address_no") or ""
    c["address_no"] = c["add_no"]
    c["inverters"] = c.get("inverters") if (isinstance(c.get("inverters"), list) and len(c.get("inverters")) > 0) else (ob.get("inverters") if (isinstance(ob.get("inverters"), list) and len(ob.get("inverters")) > 0) else ([{"brand": inv_brand, "capacity": str(c.get("inverter_capacity") or ""), "quantity": 1, "serials": [c["inverter_serial"]] if c.get("inverter_serial") else [], "serial": c.get("inverter_serial") or ""}] if (c.get("inverter_capacity") or inv_brand) else []))
    return c

class CollectionAdapter:
    def __init__(self, table_name: str):
        self.table_name = table_name

    def _extract_and_remove_unsupported_filters(self, filter_dict):
        extracted = {}
        cleaned = {}
        if not filter_dict:
            return extracted, cleaned
        for k, v in filter_dict.items():
            if k in ("doc_type", "document_number", "client_name"):
                extracted[k] = v
            else:
                cleaned[k] = v
        return extracted, cleaned

    def _deserialize_document(self, doc):
        if not doc:
            return doc
        if self.table_name == "files" and "original_filename" in doc:
            orig_filename = doc["original_filename"] or ""
            if orig_filename.startswith("__METADATA__:"):
                try:
                    remaining = orig_filename[len("__METADATA__:"):]
                    parts = remaining.rsplit(":", 1)
                    if len(parts) == 2:
                        metadata_str = parts[0]
                        actual_filename = parts[1]
                        metadata = json.loads(metadata_str)
                        doc["doc_type"] = metadata.get("doc_type")
                        doc["document_number"] = metadata.get("document_number")
                        doc["client_name"] = metadata.get("client_name")
                        doc["prepared_by"] = metadata.get("prepared_by")
                        doc["status"] = metadata.get("status")
                        doc["original_filename"] = actual_filename
                except Exception as e:
                    logger.warning(f"Failed to deserialize files metadata: {e}")
        return doc

    def _matches_filter(self, doc, extracted_filters):
        for k, v in extracted_filters.items():
            doc_val = doc.get(k)
            if isinstance(v, dict):
                for op, val in v.items():
                    if op == "$nin":
                        if doc_val in val:
                            return False
                    elif op == "$in":
                        if doc_val not in val:
                            return False
                    elif op == "$ne":
                        if doc_val == val:
                            return False
                    elif op == "$eq":
                        if doc_val != val:
                            return False
            else:
                if doc_val != v:
                    return False
        return True

    def _apply_filters(self, builder, query):
        if not query:
            return builder
        for k, v in query.items():
            if self.table_name == "products" and k == "brand":
                continue
            if k == "$or":
                parts = []
                for cond in v:
                    for cond_k, cond_v in cond.items():
                        col = cond_k.replace(".", "->")
                        if isinstance(cond_v, dict):
                            for op, val in cond_v.items():
                                if "->" in col and isinstance(val, bool):
                                    val = str(val).lower()
                                op_str = self._get_postgrest_op(op)
                                clean_val = val
                                if op == "$regex" and isinstance(val, str):
                                    clean_val = val.replace("\\", "")
                                    starts_with_caret = clean_val.startswith("^")
                                    ends_with_dollar = clean_val.endswith("$")
                                    if starts_with_caret:
                                        clean_val = clean_val[1:]
                                    if ends_with_dollar:
                                        clean_val = clean_val[:-1]
                                    if starts_with_caret and ends_with_dollar:
                                        pass
                                    elif starts_with_caret:
                                        clean_val = f"{clean_val}%"
                                    elif ends_with_dollar:
                                        clean_val = f"%{clean_val}"
                                    else:
                                        clean_val = f"%{clean_val}%"
                                elif op in ("$in", "$nin") and isinstance(val, (list, tuple)):
                                    clean_val = f"({','.join(str(x) for x in val)})"
                                parts.append(f"{col}.{op_str}.{clean_val}")
                        else:
                            if "->" in col and isinstance(cond_v, bool):
                                cond_v = str(cond_v).lower()
                            parts.append(f"{col}.eq.{cond_v}")
                or_str = ",".join(parts)
                builder = builder.or_(or_str)
            elif k == "$and":
                for cond in v:
                    builder = self._apply_filters(builder, cond)
            else:
                col = k.replace(".", "->")
                if isinstance(v, dict):
                    has_regex = "$regex" in v
                    regex_val = v.get("$regex")
                    for op, val in v.items():
                        if op == "$options":
                            continue
                        if "->" in col and isinstance(val, bool):
                            val = str(val).lower()
                        op_str = self._get_postgrest_op(op)
                        if op == "$in":
                            val_str = f"({','.join(str(x) for x in val)})"
                            builder = builder.filter(col, op_str, val_str)
                        elif op == "$nin":
                            val_str = f"({','.join(str(x) for x in val)})"
                            builder = builder.filter(col, "not.in", val_str)
                        elif op == "$regex":
                            clean_val = val
                            if isinstance(val, str):
                                clean_val = val.replace("\\", "")
                                starts_with_caret = clean_val.startswith("^")
                                ends_with_dollar = clean_val.endswith("$")
                                if starts_with_caret:
                                    clean_val = clean_val[1:]
                                if ends_with_dollar:
                                    clean_val = clean_val[:-1]
                                if starts_with_caret and ends_with_dollar:
                                    pass
                                elif starts_with_caret:
                                    clean_val = f"{clean_val}%"
                                elif ends_with_dollar:
                                    clean_val = f"%{clean_val}"
                                else:
                                    clean_val = f"%{clean_val}%"
                            builder = builder.filter(col, "ilike", clean_val)
                        else:
                            builder = builder.filter(col, op_str, val)
                else:
                    if "->" in col and isinstance(v, bool):
                        v = str(v).lower()
                    builder = builder.eq(col, v)
        return builder

    def _get_postgrest_op(self, op: str) -> str:
        ops = {
            "$eq": "eq",
            "$ne": "neq",
            "$gt": "gt",
            "$gte": "gte",
            "$lt": "lt",
            "$lte": "lte",
            "$in": "in",
            "$nin": "not.in",
            "$regex": "ilike"
        }
        return ops.get(op, "eq")

    def _clean_empty_fks(self, doc):
        if isinstance(doc, dict):
            if self.table_name == "products":
                doc.pop("brand", None)
            for k, v in list(doc.items()):
                if (k.endswith("_id") or k in ["created_by", "assigned_to", "uploader_id", "user_id", "requested_by", "raised_by", "to_user_id", "material_request_id"]) and v == "":
                    doc[k] = None
        return doc

    async def find_one(self, filter=None, projection=None, sort=None, **kwargs):
        if self.table_name == "files":
            extracted_filters, cleaned_filter = self._extract_and_remove_unsupported_filters(filter)
            if extracted_filters:
                cursor = self.find(cleaned_filter, projection=projection)
                if sort:
                    cursor = cursor.sort(sort)
                docs = await cursor.to_list(length=1000)
                for doc in docs:
                    if self._matches_filter(doc, extracted_filters):
                        return doc
                return None

        select_cols = "*"
        if projection and isinstance(projection, dict):
            inclusions = [k for k, v in projection.items() if (v == 1 or v is True) and "->" not in k and "." not in k]
            if inclusions:
                select_cols = ",".join(inclusions)
        builder = supabase.table(self.table_name).select(select_cols)
        builder = self._apply_filters(builder, filter)
        if sort:
            for k, dir in sort:
                desc = (dir == -1)
                builder = builder.order(k, desc=desc)
        builder = builder.limit(1)
        try:
            res = await asyncio.to_thread(builder.execute)
            if res.data:
                doc = res.data[0]
                if self.table_name == "companies":
                    cid = doc.get("id") or (filter.get("id") if isinstance(filter, dict) else None)
                    if cid:
                        local_doc = await LocalFileCollection("companies").find_one({"id": cid})
                        if local_doc:
                            doc = {**doc, **local_doc}
                if projection:
                    for pk, pv in projection.items():
                        if pv == 0:
                            doc.pop(pk, None)
                doc = self._deserialize_document(doc)
                return doc
        except Exception as e:
            err_str = str(e).lower()
            if "42501" in err_str or "row-level security" in err_str or "unauthorized" in err_str or "pgrst205" in err_str or "schema cache" in err_str or "could not find the table" in err_str:
                return await LocalFileCollection(self.table_name).find_one(filter, projection)
            raise e

        if self.table_name != "products":
            local_doc = await LocalFileCollection(self.table_name).find_one(filter, projection)
            if local_doc:
                return local_doc
        return None

    def find(self, filter=None, projection=None):
        return CursorAdapter(self, filter, projection)

    async def insert_one(self, document):
        global _PRODUCTS_HAS_RATE, _PRODUCTS_HAS_OPENING_STOCK, _PRODUCTS_HAS_HV, _PRODUCTS_HAS_SN_REQ
        if "id" not in document and self.table_name not in ["counters", "inventory_defaults", "password_reset_tokens"]:
            document["id"] = str(uuid.uuid4())
        document = self._clean_empty_fks(document)
        
        # Serialize metadata for files
        if self.table_name == "files" and "doc_type" in document:
            doc_type = document.pop("doc_type", None)
            doc_number = document.pop("document_number", None)
            client_name = document.pop("client_name", None)
            prepared_by = document.pop("prepared_by", None)
            status = document.pop("status", None)
            orig_filename = document.get("original_filename") or ""
            metadata = {
                "doc_type": doc_type,
                "document_number": doc_number,
                "client_name": client_name,
                "prepared_by": prepared_by,
                "status": status
            }
            document["original_filename"] = f"__METADATA__:{json.dumps(metadata)}:{orig_filename}"

        if self.table_name == "products":
            document = _clean_products_doc(document)
        elif self.table_name == "inward_entries":
            document = _clean_inward_doc(document)
        
        while True:
            try:
                res = supabase.table(self.table_name).insert(document, returning="minimal").execute()
                await LocalFileCollection(self.table_name).insert_one(document)
                return InsertOneResult(document.get("id"))
            except Exception as e:
                err_str = str(e)
                if "PGRST204" in err_str or "Could not find the" in err_str:
                    logger.warning(f"Supabase table '{self.table_name}' missing schema column. Saving full insert locally: {err_str}")
                    await LocalFileCollection(self.table_name).insert_one(document)
                    unsupported = set()
                    if "Could not find the '" in err_str:
                        col = err_str.split("Could not find the '")[1].split("'")[0]
                        unsupported.add(col)
                    if self.table_name == "products":
                        unsupported.update({"high_value_goods", "serial_number_required", "opening_stock", "rate"})
                    doc_clean = {k: v for k, v in document.items() if k not in unsupported}
                    try:
                        supabase.table(self.table_name).insert(doc_clean, returning="minimal").execute()
                    except Exception as e2:
                        logger.warning(f"Fallback insert_one for {self.table_name}: {e2}")
                    return InsertOneResult(document.get("id"))
                if "42501" in err_str or "409" in err_str or "23503" in err_str or "foreign key" in err_str.lower() or "row-level security" in err_str.lower() or "unauthorized" in err_str.lower() or "401" in err_str:
                    return await LocalFileCollection(self.table_name).insert_one(document)
                raise e

    async def insert_many(self, documents):
        global _PRODUCTS_HAS_RATE
        for doc in documents:
            if "id" not in doc and self.table_name not in ["counters", "inventory_defaults", "password_reset_tokens"]:
                doc["id"] = str(uuid.uuid4())
            doc = self._clean_empty_fks(doc)
        if self.table_name == "products" and not _PRODUCTS_HAS_RATE:
            documents = [{k: v for k, v in doc.items() if k != "rate"} for doc in documents]
        try:
            res = supabase.table(self.table_name).insert(documents, returning="minimal").execute()
        except Exception as e:
            err_str = str(e)
            if "PGRST204" in err_str or "Could not find the" in err_str:
                docs_copy = [{k: v for k, v in doc.items() if k not in ["high_value_asset", "high_value_goods", "serial_number_required", "rate", "opening_stock"]} for doc in documents]
                try:
                    res = supabase.table(self.table_name).insert(docs_copy, returning="minimal").execute()
                except Exception as e2:
                    if "42501" in str(e2) or "row-level security" in str(e2).lower() or "unauthorized" in str(e2).lower() or "401" in str(e2) or "PGRST204" in str(e2):
                        return await LocalFileCollection(self.table_name).insert_many(documents)
                    raise e2
            elif "42501" in err_str or "row-level security" in err_str.lower() or "unauthorized" in err_str.lower() or "401" in err_str:
                return await LocalFileCollection(self.table_name).insert_many(documents)
            else:
                raise e
        await LocalFileCollection(self.table_name).insert_many(documents)
        return InsertManyResult([doc.get("id") for doc in documents])

    async def update_one(self, filter, update, upsert=False):
        global _PRODUCTS_HAS_RATE
        patch = {}
        if "$set" in update:
            patch.update(update["$set"])
            patch = self._clean_empty_fks(patch)
        
        if "$inc" in update:
            builder = supabase.table(self.table_name).select("*")
            builder = self._apply_filters(builder, filter)
            res = builder.execute()
            if res.data:
                existing = res.data[0]
                for inc_k, inc_v in update["$inc"].items():
                    patch[inc_k] = (existing.get(inc_k) or 0) + inc_v
        
        if not patch:
            return UpdateResult(1, 1)

        if self.table_name == "products" and not _PRODUCTS_HAS_RATE:
            patch = {k: v for k, v in patch.items() if k != "rate"}

        if self.table_name == "clients":
            logger.info(f"[CLIENT-SAVE DIAG] ▶ update_one called. filter={filter}")
            logger.info(f"[CLIENT-SAVE DIAG] ▶ raw patch before prepare: {json.dumps({k: v for k, v in patch.items() if k != 'stages'}, default=str)}")
            try:
                builder = supabase.table(self.table_name).select("stages").limit(1)
                builder = self._apply_filters(builder, filter)
                ex_res = builder.execute()
                logger.info(f"[CLIENT-SAVE DIAG] ▶ pre-fetch stages result rows={len(ex_res.data or [])}")
                if ex_res.data and isinstance(ex_res.data[0], dict):
                    ex_doc = ex_res.data[0]
                    ex_stages = dict(ex_doc.get("stages") or {})
                    logger.info(f"[CLIENT-SAVE DIAG] ▶ existing stages keys: {list(ex_stages.keys())}")
                    if "stages" not in patch:
                        patch["stages"] = ex_stages
                    else:
                        merged_stages = dict(ex_stages)
                        incoming_stages = dict(patch.get("stages") or {})
                        merged_stages.update(incoming_stages)
                        ex_ob = dict(ex_stages.get("onboarding_data") or {})
                        inc_ob = dict(incoming_stages.get("onboarding_data") or {})
                        ex_ob.update(inc_ob)
                        merged_stages["onboarding_data"] = ex_ob
                        patch["stages"] = merged_stages
                else:
                    logger.warning(f"[CLIENT-SAVE DIAG] ⚠ pre-fetch returned NO data — WHERE clause may be wrong! filter={filter}")
            except Exception as ex_err:
                logger.warning(f"[CLIENT-SAVE DIAG] ⚠ Could not pre-fetch existing client stages: {ex_err}")
            patch = _prepare_client_supabase_payload(patch)
            ob_data = (patch.get("stages") or {}).get("onboarding_data") or {}
            logger.info(f"[CLIENT-SAVE DIAG] ▶ final Supabase patch keys: {list(patch.keys())}")
            logger.info(f"[CLIENT-SAVE DIAG] ▶ panel_wattage in patch: {patch.get('panel_wattage')}")
            logger.info(f"[CLIENT-SAVE DIAG] ▶ panel_make in patch: {patch.get('panel_make')}")
            logger.info(f"[CLIENT-SAVE DIAG] ▶ onboarding_data.consumer_category: {ob_data.get('consumer_category')}")
            logger.info(f"[CLIENT-SAVE DIAG] ▶ onboarding_data.section_number: {ob_data.get('section_number')}")
            logger.info(f"[CLIENT-SAVE DIAG] ▶ onboarding_data.inverters count: {len(ob_data.get('inverters') or [])}")

        if not patch:
            return UpdateResult(1, 1)

        try:
            builder = supabase.table(self.table_name).update(patch)
            builder = self._apply_filters(builder, filter)
            logger.info(f"[CLIENT-SAVE DIAG] ▶ Executing Supabase UPDATE on table='{self.table_name}' WHERE filter={filter}")
            res = builder.execute()
            logger.info(f"[CLIENT-SAVE DIAG] ▶ Supabase UPDATE raw response: data_count={len(res.data or [])} data={json.dumps(res.data, default=str)[:500]}")
            if self.table_name == "clients":
                if not res.data:
                    logger.error(f"[CLIENT-SAVE DIAG] ✗ Supabase UPDATE returned EMPTY data — row may not exist with filter={filter}, or RLS is blocking the write!")
                else:
                    logger.info(f"[CLIENT-SAVE DIAG] ✓ Supabase UPDATE returned {len(res.data)} row(s)")
        except Exception as e:
            logger.error(f"[CLIENT-SAVE DIAG] ✗ Supabase UPDATE EXCEPTION for table='{self.table_name}': {e}")
            logger.error(f"Supabase update failed for table '{self.table_name}': {e}")
            err_str = str(e)
            if "PGRST204" in err_str or "Could not find the" in err_str:
                logger.warning(f"Supabase table '{self.table_name}' missing schema column on update. Updating locally: {err_str}")
                await LocalFileCollection(self.table_name).update_one(filter, update, upsert=upsert)
                unsupported = set()
                if "Could not find the '" in err_str:
                    col = err_str.split("Could not find the '")[1].split("'")[0]
                    unsupported.add(col)
                if self.table_name == "products":
                    unsupported.update({"high_value_goods", "serial_number_required", "opening_stock", "rate"})
                patch_clean = {k: v for k, v in patch.items() if k not in unsupported}
                if patch_clean:
                    try:
                        builder = supabase.table(self.table_name).update(patch_clean)
                        builder = self._apply_filters(builder, filter)
                        res = builder.execute()
                    except Exception as e2:
                        logger.warning(f"Fallback update_one for {self.table_name}: {e2}")
                return UpdateResult(1, 1)
            if "42501" in err_str or "row-level security" in err_str.lower() or "unauthorized" in err_str.lower() or "401" in err_str:
                await LocalFileCollection(self.table_name).update_one(filter, update, upsert=upsert)
                return UpdateResult(1, 1)
            raise e

        try:
            await LocalFileCollection(self.table_name).update_one(filter, update, upsert=upsert)
        except Exception:
            pass

        if not res.data and upsert:
            insert_doc = {}
            # Flatten filter keys if they are simple equality
            for fk, fv in filter.items():
                if not fk.startswith("$") and not isinstance(fv, dict):
                    insert_doc[fk] = fv
            insert_doc.update(patch)
            if "id" not in insert_doc and self.table_name not in ["counters", "inventory_defaults", "password_reset_tokens"]:
                insert_doc["id"] = str(uuid.uuid4())
            insert_doc = self._clean_empty_fks(insert_doc)
            if self.table_name == "products" and not _PRODUCTS_HAS_RATE:
                insert_doc = {k: v for k, v in insert_doc.items() if k != "rate"}
            try:
                supabase.table(self.table_name).insert(insert_doc, returning="minimal").execute()
            except Exception as e:
                if self.table_name == "products" and "rate" in insert_doc:
                    err_str = str(e)
                    if "PGRST204" in err_str or "rate" in err_str:
                        logger.warning("Supabase table products does not have rate column. Disabling rate writes.")
                        _PRODUCTS_HAS_RATE = False
                        insert_doc_copy = {k: v for k, v in insert_doc.items() if k != "rate"}
                        supabase.table(self.table_name).insert(insert_doc_copy, returning="minimal").execute()
                    else:
                        raise e
                else:
                    raise e
            return UpdateResult(0, 1)

        return UpdateResult(len(res.data) if res.data else 1, len(res.data) if res.data else 1)

    async def update_many(self, filter, update):
        global _PRODUCTS_HAS_RATE
        patch = {}
        if "$set" in update:
            patch.update(update["$set"])
            patch = self._clean_empty_fks(patch)
        if self.table_name == "products" and not _PRODUCTS_HAS_RATE:
            patch = {k: v for k, v in patch.items() if k != "rate"}
        try:
            builder = supabase.table(self.table_name).update(patch)
            builder = self._apply_filters(builder, filter)
            res = builder.execute()
        except Exception as e:
            if self.table_name == "products" and "rate" in patch:
                err_str = str(e)
                if "PGRST204" in err_str or "rate" in err_str:
                    logger.warning("Supabase table products does not have rate column. Disabling rate writes.")
                    _PRODUCTS_HAS_RATE = False
                    patch_copy = {k: v for k, v in patch.items() if k != "rate"}
                    if not patch_copy:
                        return UpdateResult(1, 1)
                    builder = supabase.table(self.table_name).update(patch_copy)
                    builder = self._apply_filters(builder, filter)
                    res = builder.execute()
                else:
                    raise e
            else:
                raise e
        try:
            await LocalFileCollection(self.table_name).update_many(filter, update)
        except Exception:
            pass
        return UpdateResult(len(res.data), len(res.data))

    async def delete_one(self, filter):
        try:
            builder = supabase.table(self.table_name).delete()
            builder = self._apply_filters(builder, filter)
            res = builder.execute()
        except Exception as e:
            err_str = str(e).lower()
            if "pgrst205" in err_str or "does not exist" in err_str or "schema cache" in err_str:
                pass
            elif "42501" not in err_str and "unauthorized" not in err_str:
                raise e
        await LocalFileCollection(self.table_name).delete_one(filter)
        return DeleteResult(1)

    async def delete_many(self, filter):
        try:
            builder = supabase.table(self.table_name).delete()
            builder = self._apply_filters(builder, filter)
            res = builder.execute()
        except Exception as e:
            err_str = str(e).lower()
            if "pgrst205" in err_str or "does not exist" in err_str or "schema cache" in err_str:
                pass
            elif "42501" not in err_str and "unauthorized" not in err_str:
                raise e
        await LocalFileCollection(self.table_name).delete_many(filter)
        return DeleteResult(1)

    async def count_documents(self, filter=None):
        builder = supabase.table(self.table_name).select("id", count="exact")
        builder = self._apply_filters(builder, filter)
        try:
            res = builder.execute()
            return res.count if res.count is not None else len(res.data)
        except Exception as e:
            err_str = str(e).lower()
            if "pgrst205" in err_str or "does not exist" in err_str or "schema cache" in err_str or "could not find the table" in err_str:
                return await LocalFileCollection(self.table_name).count_documents(filter)
            raise e

    async def distinct(self, field, filter=None):
        builder = supabase.table(self.table_name).select(field)
        builder = self._apply_filters(builder, filter)
        res = builder.execute()
        vals = {row[field] for row in res.data if row.get(field) is not None}
        return list(vals)

    async def find_one_and_update(self, filter, update, upsert=False, return_document=True):
        if self.table_name == "counters":
            company_id = filter["company_id"]
            year = filter["year"]
            type_val = filter.get("type", "client")
            try:
                client = get_supabase_client(use_service_key=True)
                res = client.table("counters").select("seq").eq("company_id", company_id).eq("year", year).eq("type", type_val).execute()
                if res.data and isinstance(res.data, list) and len(res.data) > 0:
                    first_row = res.data[0]
                    current_seq = int(first_row.get("seq", 0)) if isinstance(first_row, dict) else 0  # type: ignore
                    next_seq = current_seq + 1
                    client.table("counters").update({"seq": next_seq}).eq("company_id", company_id).eq("year", year).eq("type", type_val).execute()
                else:
                    next_seq = 1
                    client.table("counters").insert({"company_id": company_id, "year": year, "type": type_val, "seq": next_seq}).execute()
                return {"seq": next_seq}
            except Exception as e:
                logger.warning(f"Counters table update failed, using fallback sequence: {e}")
                return {"seq": int(datetime.now().timestamp()) % 100000}
        return await self.update_one(filter, update, upsert)

    async def create_index(self, *args, **kwargs):
        pass

    def aggregate(self, pipeline):
        return AggregateCursorAdapter(self.table_name, pipeline)

class LocalFileCollection:
    def __init__(self, table_name: str):
        self.table_name = table_name
        self.file_path = ROOT_DIR / "local_storage" / f"{table_name}.json"

    def _read_data(self) -> list:
        if not self.file_path.exists():
            return []
        try:
            with open(self.file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []

    def _write_data(self, data: list):
        try:
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Error writing to local storage for {self.table_name}: {e}")

    def _match_val(self, doc_val, filter_val) -> bool:
        if isinstance(filter_val, dict):
            for op, val in filter_val.items():
                if op == "$regex":
                    pattern = val
                    if isinstance(pattern, str):
                        pattern = pattern.replace("\\", "")
                    try:
                        if not re.search(pattern, str(doc_val or ""), re.IGNORECASE):
                            return False
                    except Exception:
                        return False
                elif op == "$nin":
                    if doc_val in val:
                        return False
                elif op == "$in":
                    if doc_val not in val:
                        return False
                elif op == "$eq":
                    if doc_val != val:
                        return False
                elif op == "$ne":
                    if doc_val == val:
                        return False
            return True
        return doc_val == filter_val

    def _match(self, doc: dict, query: Optional[dict]) -> bool:
        if not query:
            return True
        for k, v in query.items():
            if k == "$or":
                if not any(self._match(doc, cond) for cond in v):
                    return False
            elif k == "$and":
                if not all(self._match(doc, cond) for cond in v):
                    return False
            else:
                if not self._match_val(doc.get(k), v):
                    return False
        return True

    def find(self, filter=None, projection=None):
        return LocalCursor(self, filter or {}, projection)

    async def find_one(self, filter=None, projection=None):
        data = self._read_data()
        for doc in data:
            if self._match(doc, filter):
                res = dict(doc)
                if projection:
                    for pk, pv in projection.items():
                        if pv == 0:
                            res.pop(pk, None)
                return res
        return None

    async def insert_one(self, document):
        if "id" not in document:
            document["id"] = str(uuid.uuid4())
        data = self._read_data()
        data.append(document)
        self._write_data(data)
        return InsertOneResult(document["id"])

    async def insert_many(self, documents):
        data = self._read_data()
        ids = []
        for doc in documents:
            if "id" not in doc:
                doc["id"] = str(uuid.uuid4())
            ids.append(doc["id"])
            data.append(doc)
        self._write_data(data)
        return InsertManyResult(ids)

    async def update_one(self, filter, update, upsert=False):
        data = self._read_data()
        matched_idx = -1
        for idx, doc in enumerate(data):
            if self._match(doc, filter):
                matched_idx = idx
                break
        
        if matched_idx == -1:
            if upsert:
                doc = {}
                for k, v in filter.items():
                    if not k.startswith("$") and not isinstance(v, dict):
                        doc[k] = v
                if "$set" in update:
                    doc.update(update["$set"])
                if "id" not in doc:
                    doc["id"] = str(uuid.uuid4())
                data.append(doc)
                self._write_data(data)
                return UpdateResult(0, 1)
            return UpdateResult(0, 0)

        doc = data[matched_idx]
        if "$set" in update:
            doc.update(update["$set"])
        if "$inc" in update:
            for inc_k, inc_v in update["$inc"].items():
                doc[inc_k] = (doc.get(inc_k) or 0) + inc_v
        
        self._write_data(data)
        return UpdateResult(1, 1)

    async def update_many(self, filter, update):
        data = self._read_data()
        modified = 0
        for doc in data:
            if self._match(doc, filter):
                if "$set" in update:
                    doc.update(update["$set"])
                if "$inc" in update:
                    for inc_k, inc_v in update["$inc"].items():
                        doc[inc_k] = (doc.get(inc_k) or 0) + inc_v
                modified += 1
        if modified > 0:
            self._write_data(data)
        return UpdateResult(modified, modified)

    async def delete_one(self, filter):
        data = self._read_data()
        matched_idx = -1
        for idx, doc in enumerate(data):
            if self._match(doc, filter):
                matched_idx = idx
                break
        if matched_idx != -1:
            data.pop(matched_idx)
            self._write_data(data)
            return DeleteResult(1)
        return DeleteResult(0)

    async def delete_many(self, filter):
        data = self._read_data()
        initial_len = len(data)
        data = [doc for doc in data if not self._match(doc, filter)]
        deleted = initial_len - len(data)
        if deleted > 0:
            self._write_data(data)
        return DeleteResult(deleted)

    async def count_documents(self, filter=None):
        data = self._read_data()
        count = 0
        for doc in data:
            if self._match(doc, filter):
                count += 1
        return count

    async def distinct(self, field, filter=None):
        data = self._read_data()
        values = set()
        for doc in data:
            if self._match(doc, filter) and field in doc:
                values.add(doc[field])
        return list(values)

    async def create_index(self, *args, **kwargs):
        pass

class LocalCursor:
    def __init__(self, collection, filter, projection):
        self.collection = collection
        self.filter = filter
        self.projection = projection
        self.sort_fields = None
        self.limit_val = None
        self.skip_val = None

    def sort(self, key_or_list, direction=None):
        if isinstance(key_or_list, list):
            self.sort_fields = key_or_list
        else:
            self.sort_fields = [(key_or_list, direction or 1)]
        return self

    def limit(self, val):
        self.limit_val = val
        return self

    def skip(self, val):
        self.skip_val = val
        return self

    async def to_list(self, length=None):
        data = self.collection._read_data()
        filtered = [doc for doc in data if self.collection._match(doc, self.filter)]

        if self.sort_fields:
            for k, dir in reversed(self.sort_fields):
                desc = (dir == -1)
                def sort_key(x):
                    val = x.get(k)
                    if val is None:
                        return "" if isinstance(val, str) else 0
                    return val
                filtered.sort(key=sort_key, reverse=desc)

        skip = self.skip_val or 0
        limit = length if length is not None else self.limit_val
        
        if limit is not None:
            res_data = filtered[skip:skip + limit]
        else:
            res_data = filtered[skip:]

        final_data = []
        for doc in res_data:
            doc_copy = dict(doc)
            if self.projection:
                for pk, pv in self.projection.items():
                    if pv == 0:
                        doc_copy.pop(pk, None)
            final_data.append(doc_copy)
            
        return final_data

    def __aiter__(self):
        return AsyncIteratorWrapper(self.to_list())

class SupabaseDBAdapter:
    def __getattr__(self, name):
        return CollectionAdapter(name)
        
    def __getitem__(self, name):
        return CollectionAdapter(name)

class SupabaseClientAdapter:
    def __getitem__(self, name):
        return db
    def close(self):
        pass

client = SupabaseClientAdapter()
db = SupabaseDBAdapter()
_company_logo_cache = {}  # Cache company logo bytes to prevent database/storage roundtrips

JWT_SECRET = os.environ['JWT_SECRET']
if (os.environ.get("ENVIRONMENT") == "production" or os.environ.get("NODE_ENV") == "production"):
    if not JWT_SECRET or JWT_SECRET in ("solrix-crm-super-secret-key-1029384756", "secret", "change-me", "123456"):
        raise RuntimeError("CRITICAL SECURITY ERROR: Production deployment requires a cryptographically random JWT_SECRET in environment variables.")

JWT_ALGORITHM = "HS256"
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
APP_NAME = os.environ.get('APP_NAME', 'gvp_solar_energy_app')
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"

from contextlib import asynccontextmanager

async def activity_logs_cleanup_task():
    """Background task that runs a daily cleanup of activity logs older than 30 days."""
    logger.info("Activity logs cleanup task initialized")
    while True:
        # CRITICAL: must await sleep FIRST — without this the while True loop
        # is a tight infinite loop that permanently blocks the event loop.
        await asyncio.sleep(86400)  # Run once every 24 hours
        try:
            from datetime import datetime, timedelta, timezone
            thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
            client_to_use = service_supabase if service_supabase is not None else default_supabase
            if client_to_use:
                res = await asyncio.to_thread(
                    lambda: client_to_use.table("activity_logs").delete().lt("created_at", thirty_days_ago).execute()
                )
                deleted_count = len(res.data) if res.data else 0
                logger.info(f"Scheduled Activity Logs Cleanup: Deleted {deleted_count} logs older than 30 days.")
            else:
                logger.warning("Activity logs cleanup skipped: No Supabase client configured.")
        except asyncio.CancelledError:
            break
        except Exception as e:
            logger.error(f"Scheduled Activity Logs Cleanup failed: {e}", exc_info=True)
async def auto_migrate_product_variants():
    try:
        inward_entries = await db.inward_entries.find({}, {"company_id": 1, "product": 1, "size": 1, "unit": 1}).to_list(100000)
        outward_entries = await db.outward_entries.find({}, {"company_id": 1, "product": 1, "size": 1, "unit": 1}).to_list(100000)
        all_entries = (inward_entries or []) + (outward_entries or [])

        # Group non-empty sizes by (company_id, product_name)
        spec_map: Dict[Tuple[str, str], List[Tuple[str, str]]] = {}
        for entry in all_entries:
            cid = entry.get("company_id")
            pn = (entry.get("product") or "").strip().upper()
            ps = (entry.get("size") or "").strip()
            unit = (entry.get("unit") or "Nos").strip()
            if not cid or not pn:
                continue
            k = (cid, pn)
            if k not in spec_map:
                spec_map[k] = []
            if (ps, unit) not in spec_map[k]:
                spec_map[k].append((ps, unit))

        # Backfill empty-size products in db.products if transaction size specs exist
        empty_size_prods = await db.products.find({"size": {"$in": ["", None]}}).to_list(10000)
        for prod in empty_size_prods:
            cid = prod.get("company_id")
            pn = (prod.get("name") or "").strip().upper()
            k = (cid, pn)
            if k in spec_map and len(spec_map[k]) > 0:
                first_size, first_unit = spec_map[k][0]
                if first_size:
                    await db.products.update_one({"id": prod["id"]}, {"$set": {"size": first_size, "unit": first_unit or prod.get("unit", "Nos")}})

        # Ensure distinct product records exist for all (cid, pn, ps, unit) tuples
        seen = set()
        for entry in all_entries:
            cid = entry.get("company_id")
            pn = (entry.get("product") or "").strip().upper()
            ps = (entry.get("size") or "").strip()
            unit = (entry.get("unit") or "Nos").strip()
            if not cid or not pn:
                continue
            key = (cid, pn, ps, unit)
            if key not in seen:
                seen.add(key)
                await ensure_product(cid, pn, ps, unit=unit)
    except Exception as e:
        logger.warning(f"auto_migrate_product_variants error: {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    deferred_task = asyncio.create_task(_deferred_startup_tasks())
    try:
        logger.info("Solarix backend started")
        yield
    finally:
        if deferred_task:
            deferred_task.cancel()
            try:
                await deferred_task
            except asyncio.CancelledError:
                pass

def _is_migration_completed(mig_key: str) -> bool:
    filepath = ROOT_DIR / "local_storage" / "migrations.json"
    if not filepath.exists():
        return False
    try:
        with open(filepath, "r") as f:
            data = json.load(f)
            return data.get(mig_key, False)
    except Exception:
        return False

def _mark_migration_completed(mig_key: str):
    filepath = ROOT_DIR / "local_storage" / "migrations.json"
    data = {}
    if filepath.exists():
        try:
            with open(filepath, "r") as f:
                data = json.load(f)
        except Exception:
            data = {}
    data[mig_key] = True
    try:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w") as f:
            json.dump(data, f)
    except Exception:
        pass

async def run_one_time_size_standardization_migration():
    """
    ONE-TIME DATA MIGRATION FOR EXISTING INVENTORY DATA:
    1. Scans all existing records in inward_entries, outward_entries, and products.
    2. Converts every size format to standard format using norm_str(size) (e.g. 4Cx95 -> 4C*95).
    3. Merges duplicate product master records created only because of different size formats.
    4. Preserves all transaction history, inward entries, outward entries, and stock balances.
    """
    try:
        mig_key = "migration_size_standardization_v1"
        if _is_migration_completed(mig_key):
            return

        logger.info("Starting one-time size standardization & product merge migration...")
        
        # Step 1: Update inward_entries
        inward_docs = await db.inward_entries.find({}).to_list(100000)
        inward_updates = 0
        for doc in (inward_docs or []):
            raw_size = doc.get("size") or ""
            std_size = norm_str(raw_size)
            if raw_size != std_size:
                await db.inward_entries.update_one({"id": doc["id"]}, {"$set": {"size": std_size}})
                inward_updates += 1

        # Step 2: Update outward_entries
        outward_docs = await db.outward_entries.find({}).to_list(100000)
        outward_updates = 0
        for doc in (outward_docs or []):
            raw_size = doc.get("size") or ""
            std_size = norm_str(raw_size)
            if raw_size != std_size:
                await db.outward_entries.update_one({"id": doc["id"]}, {"$set": {"size": std_size}})
                outward_updates += 1

        # Step 3: Update products size and merge duplicate products
        prod_docs = await db.products.find({}).to_list(100000)
        spec_to_prods = {}
        for p in (prod_docs or []):
            cid = p.get("company_id")
            pn = norm_product_name(p.get("name"))
            raw_size = p.get("size") or ""
            std_size = norm_str(raw_size)
            if not cid or not pn:
                continue
            
            if raw_size != std_size:
                await db.products.update_one({"id": p["id"]}, {"$set": {"size": std_size}})
                p["size"] = std_size

            key = (cid, pn, std_size)
            if key not in spec_to_prods:
                spec_to_prods[key] = []
            spec_to_prods[key].append(p)

        merged_count = 0
        for key, prods in spec_to_prods.items():
            if len(prods) > 1:
                prods.sort(key=lambda x: (0 if (x.get("rate") or x.get("min_stock") or x.get("opening_stock")) else 1, x.get("created_at") or ""))
                primary = prods[0]
                
                merged_patch = {}
                for dup in prods[1:]:
                    if not primary.get("rate") and dup.get("rate"):
                        merged_patch["rate"] = dup["rate"]
                    if not primary.get("min_stock") and dup.get("min_stock"):
                        merged_patch["min_stock"] = dup["min_stock"]
                    if not primary.get("opening_stock") and dup.get("opening_stock"):
                        merged_patch["opening_stock"] = dup["opening_stock"]
                    if not primary.get("category") and dup.get("category"):
                        merged_patch["category"] = dup["category"]
                    
                    await db.products.delete_one({"id": dup["id"]})
                    merged_count += 1

                if merged_patch:
                    await db.products.update_one({"id": primary["id"]}, {"$set": merged_patch})

        _mark_migration_completed(mig_key)
        logger.info(f"Size standardization migration completed: {inward_updates} inward updated, {outward_updates} outward updated, {merged_count} duplicate products merged.")
    except Exception as e:
        logger.warning(f"Error during size standardization migration: {e}")

async def _deferred_startup_tasks():
    """Non-critical startup tasks deferred so they don't block the first request."""
    should_run_migrations = os.environ.get("RUN_STARTUP_MIGRATIONS", "false").lower() in ("true", "1", "yes")
    if should_run_migrations:
        await asyncio.sleep(5)  # Wait 5s for the server to be warm before doing heavy migration work
        try:
            await run_one_time_size_standardization_migration()
            await auto_migrate_product_variants()
            await sync_inventory_master()
            logger.info("Deferred product variant migration & inventory synchronization complete")
        except Exception as e:
            logger.warning(f"Deferred migration error: {e}")
    else:
        logger.info("Startup data migrations skipped (RUN_STARTUP_MIGRATIONS=false)")
    # Schedule daily cleanup — sleep first so the infinite loop starts harmlessly
    await activity_logs_cleanup_task()

app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.middleware("http")
async def supabase_client_middleware(request: Request, call_next):
    path = request.url.path
    is_public_route = any(path.endswith(p) for p in [
        "/auth/login", "/auth/register", "/auth/forgot-password",
        "/auth/verify-otp", "/auth/reset-password", "/auth/refresh", "/auth/google", "/auth/request-access"
    ])
    
    token = None
    if not is_public_route:
        token = request.cookies.get("access_token")
        if not token:
            auth = request.headers.get("Authorization", "")
            if auth.startswith("Bearer "):
                token = auth[7:]
    
    if token:
        try:
            is_custom = False
            try:
                jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                is_custom = True
            except Exception:
                pass

            if is_custom:
                req_client = get_supabase_client(use_service_key=True)
            else:
                jwt.decode(token, options={"verify_signature": False, "verify_exp": True})
                req_client = get_supabase_client(token=token)
            token_token = _supabase_var.set(req_client)
        except Exception as e:
            logger.warning(f"Stale or invalid token detected, falling back to default client: {e}")
            token_token = _supabase_var.set(get_supabase_client())
    else:
        token_token = _supabase_var.set(get_supabase_client())
        
    try:
        response = await call_next(request)
        return response
    finally:
        _supabase_var.reset(token_token)

# ---------- Storage ----------
_storage_init_done = False
def init_storage():
    """Lazily init storage buckets on first use — NOT during cold start."""
    global _storage_init_done
    if _storage_init_done:
        return "supabase"
    buckets = ["customer-documents", "project-images", "vendor-documents", "generated-pdfs", "user-profile-images"]
    for b in buckets:
        try:
            supabase.storage.create_bucket(b, options={"public": b == "user-profile-images"})
        except Exception:
            pass
    _storage_init_done = True
    return "supabase"

def _map_path_to_bucket_and_name(path: str) -> tuple:
    parts = path.split("/")
    company_id = "default"
    category = "general"
    filename = parts[-1]
    
    if len(parts) >= 3:
        company_id = parts[1]
        category = parts[2]
        
    bucket_map = {
        "profile": "user-profile-images",
        "profile_photo": "user-profile-images",
        "user": "user-profile-images",
        "avatar": "user-profile-images",
        "generated": "generated-pdfs",
        "generated_pdf": "generated-pdfs",
        "inward": "vendor-documents",
        "vendor": "vendor-documents",
        "verification": "project-images",
        "project": "project-images",
        "images": "project-images",
        "assets": "project-images",
        "templates": "customer-documents",
        "template": "customer-documents",
        "clients": "customer-documents",
        "client": "customer-documents",
        "general": "customer-documents"
    }
    
    bucket = bucket_map.get(category, "customer-documents")
    file_path = f"{company_id}/{category}/{filename}"
    return bucket, file_path

def put_object(path: str, data: bytes, content_type: str) -> dict:
    bucket, file_path = _map_path_to_bucket_and_name(path)
    try:
        supabase.storage.from_(bucket).upload(
            path=file_path,
            file=data,
            file_options={"content-type": content_type, "upsert": "true"}
        )
    except Exception as e:
        logger.error(f"Error uploading to bucket {bucket} at {file_path}: {e}")
        raise e
    return {"path": path, "size": len(data)}

def get_object(path: str):
    bucket, file_path = _map_path_to_bucket_and_name(path)
    try:
        data = supabase.storage.from_(bucket).download(file_path)
    except Exception as e:
        logger.error(f"Error downloading from bucket {bucket} at {file_path}: {e}")
        raise HTTPException(status_code=404, detail="File not found")
    
    import mimetypes
    content_type, _ = mimetypes.guess_type(file_path)
    if not content_type:
        content_type = "application/octet-stream"
        
    return data, content_type

# ---------- Auth helpers ----------
# We retain these shell helpers for compatibility with other endpoints (like OTP reset tokens)
def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        token = request.query_params.get("auth")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # ── Fast path: return cached user if token is still fresh ──────────────────
    cached = _cache_get_user(token)
    if cached:
        if not cached.get("company_id"):
            cached["company_id"] = "COMP-001"
        if not cached.get("name"):
            cached["name"] = cached.get("full_name") or "User"
        if not cached.get("id"):
            cached["id"] = cached.get("sub") or "user"
        return cached

    # ── Slow path: validate with JWT secret or Supabase and fetch profile ──
    user_id = None
    payload = None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub") or payload.get("user_id") or payload.get("id")
    except Exception:
        try:
            res = supabase.auth.get_user(token)
            if res and res.user:
                user_id = res.user.id
        except Exception as e:
            logger.error(f"Supabase auth validation failed: {e}")

    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")

    user = None
    try:
        rpc_res = get_rpc_client().rpc("get_user_by_id", {"p_user_id": user_id}).execute()
        user = rpc_res.data[0] if isinstance(rpc_res.data, list) and rpc_res.data else None
    except Exception as e:
        logger.warning(f"get_user_by_id RPC failed: {e}")

    if not user or not isinstance(user, dict):
        try:
            user = await db.users.find_one({"id": user_id}, {"_id": 0})
        except Exception as db_err:
            logger.error(f"Direct user lookup failed: {db_err}")
            user = None

    if not user or not isinstance(user, dict):
        user_email = (payload.get("email") if isinstance(payload, dict) else None) or (getattr(locals().get('res', None), 'user', None) and getattr(res.user, 'email', None))
        if user_email:
            try:
                user = await db.users.find_one({"email": user_email.lower().strip()}, {"_id": 0})
            except Exception:
                user = None

    if not user or not isinstance(user, dict):
        if isinstance(payload, dict) and (payload.get("id") or payload.get("sub") or payload.get("user_id")):
            user = {
                "id": payload.get("id") or payload.get("sub") or user_id,
                "company_id": payload.get("company_id") or "COMP-001",
                "role": payload.get("role") or "Staff",
                "name": payload.get("name") or payload.get("full_name") or "User",
                "full_name": payload.get("full_name") or payload.get("name") or "User",
                "email": payload.get("email") or "",
                "permissions": payload.get("permissions") or default_perms_for_role(payload.get("role") or "Staff")
            }
        elif locals().get('res') is not None and hasattr(locals()['res'], 'user') and locals()['res'].user:
            user_meta = getattr(res.user, 'user_metadata', {}) or {}
            u_name = user_meta.get("full_name") or user_meta.get("name") or getattr(res.user, 'email', '') or "User"
            user = {
                "id": res.user.id,
                "company_id": user_meta.get("company_id") or "COMP-001",
                "name": u_name,
                "full_name": u_name,
                "email": getattr(res.user, 'email', '') or "",
                "role": user_meta.get("role") or "Staff",
                "permissions": default_perms_for_role(user_meta.get("role") or "Staff")
            }

    if not user or not isinstance(user, dict):
        raise HTTPException(status_code=401, detail="User not found")

    # Ensure company_id and name are guaranteed on the user dictionary
    if not user.get("company_id"):
        cid = (payload.get("company_id") if isinstance(payload, dict) else None) or "COMP-001"
        user["company_id"] = cid
    if not user.get("name"):
        user["name"] = user.get("full_name") or "User"

    if (user.get("email") or "").strip().lower() in SUPER_ADMIN_EMAILS:
        user["is_super_admin"] = True
        user["is_platform_owner"] = True
        user["user_type"] = "super_admin"
        user["role"] = "Super Admin"

    if user.get("role") == "Installer":
        perms = user.get("permissions")
        if not isinstance(perms, dict):
            user["permissions"] = {}
            perms = user["permissions"]
        if not isinstance(perms.get("reports"), dict):
            perms["reports"] = {}
        perms["reports"]["view"] = False

    # Store in cache so subsequent requests in next 5 min skip Supabase round-trips
    _cache_put_user(token, user)
    return user

# ---------- Models ----------
class RegisterCompanyIn(BaseModel):
    owner_name: str
    company_name: str
    mobile: str
    alt_mobile: Optional[str] = ""
    email: EmailStr
    password: str
    gst_number: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    district: Optional[str] = ""
    state: Optional[str] = ""
    pincode: Optional[str] = ""
    business_type: str = "Solar EPC"

class LoginIn(BaseModel):
    identifier: str
    password: str


class RefreshIn(BaseModel):
    refresh_token: str


class MyProfileUpdate(BaseModel):
    name: Optional[str] = None
    mobile: Optional[str] = None
    profile_photo_file_id: Optional[str] = None


class ChangeEmailIn(BaseModel):
    new_email: EmailStr
    current_password: str


class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str

class CompanyUpdate(BaseModel):
    company_name: Optional[str] = None
    owner_name: Optional[str] = None
    mobile: Optional[str] = None
    alt_mobile: Optional[str] = None
    email: Optional[str] = None
    gst_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    business_type: Optional[str] = None
    website: Optional[str] = None
    support_number: Optional[str] = None
    logo_file_id: Optional[str] = None
    documents: Optional[Dict[str, str]] = None

class ClientIn(BaseModel):
    full_name: str
    mobile: str
    alt_mobile: Optional[str] = ""
    consumer_number: Optional[str] = ""
    section_number: Optional[str] = ""
    section_no: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    pincode: Optional[str] = ""
    district: Optional[str] = ""
    landmark: Optional[str] = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    state_code: Optional[str] = ""
    formatted_address: Optional[str] = ""
    aadhaar: Optional[str] = ""
    aadhaar_name: Optional[str] = ""
    aadhaar_image: Optional[str] = ""
    system_kw: Optional[float] = 0
    panel_make: Optional[str] = ""
    panel_brand: Optional[str] = ""
    panel_technology: Optional[str] = ""
    panel_wattage: Optional[float] = 0
    num_panels: Optional[int] = 0
    inverter_make: Optional[str] = ""
    inverter_capacity: Optional[str] = ""
    inverter_model: Optional[str] = ""
    inverter_serial: Optional[str] = ""
    inverter_year: Optional[str] = ""
    sanction_number: Optional[str] = ""
    consumer_type: Optional[str] = ""
    consumer_category: Optional[str] = ""
    phase_type: Optional[str] = "Single Phase"
    subsidy_eligible: Optional[bool] = False
    status: Optional[str] = "Lead"
    stages: Optional[Dict[str, Any]] = None
    documents: Optional[List[Dict[str, Any]]] = None
    inverters: Optional[List[Dict[str, Any]]] = None
    contract_value: Optional[float] = 0.0
    quotation_value: Optional[float] = 0.0
    initial_payments: Optional[List[Dict[str, Any]]] = None
    loan_setup: Optional[Dict[str, Any]] = None
    payment_plan: Optional[List[Dict[str, Any]]] = None

class StageUpdate(BaseModel):
    stages: Dict[str, Any]

class StatusUpdate(BaseModel):
    status: str

class NoteIn(BaseModel):
    text: str

class EmployeeIn(BaseModel):
    name: str
    mobile: str
    email: EmailStr
    password: str
    role: str
    status: str = "Active"
    permissions: Optional[Dict[str, Dict[str, bool]]] = None
    employee_id: Optional[str] = None

class LeadIn(BaseModel):
    name: str
    mobile: str
    alt_mobile: Optional[str] = ""
    city: Optional[str] = ""
    address: Optional[str] = ""
    estimated_kw: Optional[float] = 0
    consumer_type: Optional[str] = ""
    source: Optional[str] = "Other"
    assigned_to: Optional[str] = ""
    assigned_to_name: Optional[str] = ""
    stage: Optional[str] = "NEW"
    status: Optional[str] = "New Lead"
    next_followup_at: Optional[str] = ""
    remarks: Optional[str] = ""

class LeadCallIn(BaseModel):
    outcome: str
    notes: Optional[str] = ""
    duration_sec: Optional[int] = 0
    next_followup_at: Optional[str] = ""
    assigned_to: Optional[str] = ""
    assigned_to_name: Optional[str] = ""
    stage: Optional[str] = ""

class LeadFollowupUpdate(BaseModel):
    status: Optional[str] = "completed"
    rescheduled_at: Optional[str] = ""
    remarks: Optional[str] = ""

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    status: Optional[str] = None
    permissions: Optional[Dict[str, Dict[str, bool]]] = None
    employee_id: Optional[str] = None

DEFAULT_STAGES = [
    "Onboarding",
    "Survey",
    "Quotation",
    "Material Delivery",
    "Installation",
    "Document Making",
    "Document Signed",
    "Meter Testing Request",
    "Meter Testing Completed",
    "PM Surya Ghar Upload",
    "MSEDCL Upload",
    "Verification",
    "Handover",
]

def now_iso():
    return datetime.now(timezone.utc).isoformat()


def _stages_indicate_onboarding(client: dict) -> bool:
    """Return True if the client's stages indicate onboarding, handling common key casing variations."""
    stages = (client.get("stages") or {})
    # Direct check for canonical key
    if stages.get("Onboarding"):
        return True
    # Case-insensitive fallback (some historical data may have different casing)
    for k, v in stages.items():
        if isinstance(k, str) and k.strip().lower() == "onboarding" and v:
            return True
    return False


def _client_current_stage(client: dict) -> str:
    stages = (client.get("stages") or {})
    for stage in [
        "Onboarding",
        "Survey",
        "Quotation",
        "Material Delivery",
        "Installation",
        "Document Making",
        "Document Signed",
        "Meter Testing Request",
        "Meter Testing Completed",
        "PM Surya Ghar Upload",
        "MSEDCL Upload",
        "Verification",
        "Handover",
    ]:
        if not stages.get(stage):
            return stage
    return "Handover"


async def next_client_id(company_id: str) -> str:
    year = datetime.now(timezone.utc).year
    res = await db.counters.find_one_and_update(
        {"company_id": company_id, "year": year},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True
    )
    seq = res["seq"] if isinstance(res, dict) and "seq" in res else 1
    return f"SOL-{year}-{seq:04d}"

async def next_lead_id(company_id: str) -> str:
    year = datetime.now(timezone.utc).year
    res = await db.counters.find_one_and_update(
        {"company_id": company_id, "type": "lead", "year": year},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True
    )
    seq = res["seq"] if isinstance(res, dict) and "seq" in res else 1
    return f"LEAD-{year}-{seq:04d}"

async def log_activity(company_id: str, user_id: str, user_name: str, action: str, target: str = ""):
    try:
        await db.activity_logs.insert_one({
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "user_id": user_id,
            "user_name": user_name,
            "action": action,
            "target": target,
            "created_at": now_iso(),
        })
    except Exception as e:
        logger.warning(f"Activity logging failed: {e}")

async def push_notification(company_id: str, audience: str, title: str, body: str = "", to_user_id: Optional[str] = None):
    try:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "audience": audience,
            "to_user_id": to_user_id,
            "title": title,
            "body": body,
            "read_by": [],
            "created_at": now_iso(),
        })
    except Exception:
        pass  # Notification failures must never crash the parent operation

def calc_progress(stages: Dict[str, bool]) -> int:
    if not stages:
        return 0
    done = sum(1 for s in DEFAULT_STAGES if stages.get(s))
    return round((done / len(DEFAULT_STAGES)) * 100)

STAGE_CHECKLISTS = {
    "Survey": [
        "Site access is safe",
        "Roof layout has been verified",
        "Panel locations are noted",
        "Meter / grid connection identified",
        "Client requirements confirmed",
    ],
    "Document Signed": [
        "WCR Signed",
        "Annexure Signed",
        "SLDR Signed",
        "Net Meter Agreement Signed",
        "Meter Testing Request Signed",
        "Vendor Agreement Signed",
        "Other Documents Signed",
    ],
    "Meter Testing Completed": [
        "Meter Testing Request Received",
        "Meter Installed",
        "Meter Reading Verified",
        "Meter Testing Completed",
        "MSEDCL Meter Testing Submitted",
        "Meter Approved",
        "Final Notes Added"
    ]
}

def sync_checklist_completed(stages: dict) -> dict:
    raw_checklist = stages.get("checklist_completed")
    checklist_completed = {}
    if isinstance(raw_checklist, dict):
        for k, v in raw_checklist.items():
            if isinstance(v, str):
                checklist_completed[str(k)] = v.strip().lower() in ("true", "1", "yes", "completed")
            else:
                checklist_completed[str(k)] = bool(v)
    for stage, items in STAGE_CHECKLISTS.items():
        if not stages.get(stage):
            for item in items:
                checklist_completed.pop(item, None)
    stages["checklist_completed"] = checklist_completed
    return stages

def serialize_user(u: dict) -> dict:
    u.pop("password_hash", None)
    u.pop("_id", None)
    return u

ROLE_PAGES = ["dashboard", "clients", "documents", "project_execution", "task_portal", "data_management", "client_data", "complaints", "reports", "settings", "team", "sales_documents"]
PERMS = ["view", "create", "edit", "delete", "approve"]
PROJ_EXEC_TABS = ["verification", "approval", "reject", "project_assignment", "retry"]

def default_perms_for_role(role: str) -> Dict[str, Dict[str, bool]]:
    if role in ("Super Admin", "Admin"):
        res = {p: {a: True for a in PERMS} for p in ROLE_PAGES}
        res["project_execution"].update({t: True for t in PROJ_EXEC_TABS})
        return res

    base = {p: {a: False for a in PERMS} for p in ROLE_PAGES}
    base["complaints"] = {"view": True, "create": True, "edit": False, "delete": False, "approve": False}

    if role == "Manager":
        for p in ["dashboard", "clients", "task_portal", "project_execution", "client_data", "data_management", "reports", "sales_documents", "documents", "receivables"]:
            base[p] = {"view": True, "create": True, "edit": True, "delete": False, "approve": True}
        base["project_execution"].update({t: True for t in PROJ_EXEC_TABS})
    elif role == "Staff":
        for p in ["dashboard", "clients", "task_portal", "data_management", "sales_documents", "documents"]:
            base[p] = {"view": True, "create": True, "edit": True, "delete": False, "approve": False}
    elif role == "Installer":
        base["task_portal"] = {"view": True, "create": False, "edit": True, "delete": False, "approve": False}
        base["clients"] = {"view": True, "create": False, "edit": False, "delete": False, "approve": False}
        base["client_data"] = {"view": True, "create": False, "edit": False, "delete": False, "approve": False}
    elif role == "Viewer":
        for p in ["dashboard", "clients", "task_portal", "project_execution", "data_management", "client_data", "reports", "sales_documents"]:
            base[p] = {"view": True, "create": False, "edit": False, "delete": False, "approve": False}
    elif role == "Supervisor":
        for p in ["dashboard", "clients", "task_portal", "project_execution", "client_data"]:
            base[p] = {"view": True, "create": True, "edit": True, "delete": False, "approve": True}
        base["project_execution"].update({t: True for t in PROJ_EXEC_TABS})
    elif role == "Sales Executive":
        for p in ["dashboard", "clients"]:
            base[p] = {"view": True, "create": True, "edit": True, "delete": False, "approve": False}
    elif role == "Inventory Manager":
        for p in ["data_management", "reports"]:
            base[p] = {"view": True, "create": True, "edit": True, "delete": False, "approve": False}
    elif role == "Documentation Executive":
        for p in ["documents", "clients", "client_data"]:
            base[p] = {"view": True, "create": True, "edit": True, "delete": False, "approve": False}
    return base


EXTERNAL_USER_TYPES = {"client", "vendor", "epc", "epc_partner", "customer", "external"}
EXTERNAL_ROLES = {"client", "vendor", "epc", "epc/partner", "partner", "customer"}

def is_external_user(user: Dict[str, Any]) -> bool:
    """Return True if user is an external business account (Client, Vendor, EPC/Partner, Customer)."""
    if not isinstance(user, dict):
        return False
    u_type = (user.get("user_type") or "").strip().lower()
    role = (user.get("role") or "").strip().lower()
    return u_type in EXTERNAL_USER_TYPES or role in EXTERNAL_ROLES

def is_internal_team_user(user: Dict[str, Any]) -> bool:
    """Return True if user is an internal team member or company owner."""
    if not isinstance(user, dict):
        return False
    return not is_external_user(user)

def is_owner(user: Dict[str, Any]) -> bool:
    """Return True if user is verified Company Owner or Platform Super Admin. External accounts are NEVER owners."""
    if not isinstance(user, dict) or is_external_user(user):
        return False
    user_email = (user.get("email") or "").strip().lower()
    if user_email in SUPER_ADMIN_EMAILS:
        return True
    return (
        user.get("user_type") in ("owner", "platform_owner", "super_admin") or
        user.get("role") in ("Super Admin", "Owner", "Platform Owner") or
        user.get("is_super_admin") is True or
        user.get("is_platform_owner") is True or
        user.get("is_owner") is True or
        (user.get("role") == "Admin" and user.get("user_type") in ("owner", "", None))
    )

def has_perm(user: Dict[str, Any], page: str, action: str) -> bool:
    """Single source of truth for permission checks. Company Owner always has full access. External users cannot access internal admin pages."""
    if not isinstance(user, dict):
        return False
    if is_external_user(user):
        if page in ("team", "settings", "activity_log", "billing"):
            return False
        return False
    if is_owner(user):
        return True
    perms = user.get("permissions")
    if not isinstance(perms, dict):
        perms = default_perms_for_role(user.get("role", ""))
    page_perms = perms.get(page) or {}
    val = page_perms.get(action)
    if val is None and page == "project_execution" and action in PROJ_EXEC_TABS:
        return page_perms.get("view") is True
    if val is None and page == "data_management" and action == "view":
        return any(perms.get(k, {}).get("view") is True for k in perms if k.startswith("dm_"))
    return val is True


def require_perm(page: str, action: str):
    """FastAPI dependency factory — returns 403 if the user lacks the permission."""
    async def _checker(user=Depends(get_current_user)):
        if not has_perm(user, page, action):
            raise HTTPException(status_code=403, detail=f"Missing permission: {page}.{action}")
        return user
    return _checker


SUPER_ADMIN_EMAILS = {
    os.environ.get("SUPER_ADMIN_EMAIL", "solarixoffcial.info@gmail.com").strip().lower(),
    "solarixoffcial.info@gmail.com",
    "solarixofficial.info@gmail.com",
}

def is_super_admin_user(user: Dict[str, Any]) -> bool:
    """Return True ONLY if user is verified Super Admin for solarixoffcial.info@gmail.com."""
    if not isinstance(user, dict):
        return False
    user_email = (user.get("email") or "").strip().lower()
    if user_email in SUPER_ADMIN_EMAILS:
        return True
    return (
        user.get("user_type") in ("platform_owner", "super_admin") or
        user.get("role") in ("Platform Owner", "Super Admin") or
        user.get("is_platform_owner") is True or
        user.get("is_super_admin") is True
    )

def require_super_admin():
    """FastAPI dependency requiring Level 1 SOLRIX Platform Super Admin access."""
    async def _checker(user=Depends(get_current_user)):
        if not is_super_admin_user(user):
            raise HTTPException(status_code=403, detail="Access denied.")
        return user
    return _checker

is_platform_owner_user = is_super_admin_user
require_platform_owner = require_super_admin


class AccessRequestIn(BaseModel):
    full_name: str
    work_email: str
    mobile: str
    company_name: str
    employee_id: Optional[str] = ""
    department: Optional[str] = ""
    reason: Optional[str] = ""

@api_router.post("/auth/request-access")
async def create_access_request(data: AccessRequestIn):
    email = data.work_email.lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Work Email is required")
    if not data.full_name.strip():
        raise HTTPException(status_code=400, detail="Full Name is required")
    if not data.mobile.strip():
        raise HTTPException(status_code=400, detail="Mobile Number is required")
    if not data.company_name.strip():
        raise HTTPException(status_code=400, detail="Company Name is required")

    # Check if user already exists
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="An account with this email already exists. Please sign in.")

    # Check existing pending access requests
    existing_req = await db.access_requests.find_one({"work_email": email, "status": "Pending"})
    if existing_req:
        return {"message": "Access request submitted. Your company administrator will review your request.", "status": "Pending"}

    req_doc = {
        "id": f"req_{uuid.uuid4().hex[:12]}",
        "full_name": data.full_name.strip(),
        "work_email": email,
        "mobile": data.mobile.strip(),
        "company_name": data.company_name.strip(),
        "employee_id": (data.employee_id or "").strip(),
        "department": (data.department or "").strip(),
        "reason": (data.reason or "").strip(),
        "status": "Pending",
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    await db.access_requests.insert_one(req_doc)
    return {"message": "Access request submitted. Your company administrator will review your request.", "status": "Pending"}

# ---------- Auth ----------
@api_router.post("/auth/register")
async def register_company(data: RegisterCompanyIn, response: Response):
    email = data.email.lower().strip()
    mobile = (data.mobile or "").strip()
    clean_mobile = "".join(filter(str.isdigit, mobile))

    if not clean_mobile:
        raise HTTPException(status_code=400, detail="Mobile Number is required.")
    if len(clean_mobile) != 10 or len(mobile) != len(clean_mobile):
        raise HTTPException(status_code=400, detail="Enter a valid 10-digit mobile number.")
    if not clean_mobile.startswith(('6', '7', '8', '9')):
        raise HTTPException(status_code=400, detail="Enter a valid Indian mobile number.")

    if not email:
        raise HTTPException(status_code=400, detail="Email address is required.")
    if not data.password:
        raise HTTPException(status_code=400, detail="Password is required.")
    if not data.owner_name.strip():
        raise HTTPException(status_code=400, detail="Owner Name is required.")
    if not data.company_name.strip():
        raise HTTPException(status_code=400, detail="Company Name is required.")

    if data.pincode and data.pincode.strip():
        pin_val = data.pincode.strip()
        if len(pin_val) != 6 or not pin_val.isdigit():
            raise HTTPException(status_code=400, detail="Pincode must be a 6-digit number.")

    # 1. Inspect existing application database records
    existing_user = await db.users.find_one({"email": email})
    existing_company = await db.companies.find_one({"email": email})

    if not existing_company and existing_user and existing_user.get("company_id"):
        existing_company = await db.companies.find_one({"id": existing_user["company_id"]})

    # Check if Supabase Auth user exists
    auth_exists = False
    try:
        res = get_rpc_client().rpc("check_email_exists", {"email_to_check": email}).execute()
        auth_exists = bool(res.data)
    except Exception as e:
        logger.warning(f"Failed to check check_email_exists: {e}")

    # CASE A: FULLY REGISTERED COMPLETED ACCOUNT
    if existing_user and existing_company and existing_user.get("status") in ("Active", "Completed"):
        raise HTTPException(
            status_code=400,
            detail="An account with this email already exists. Please sign in."
        )

    # CASE B / CASE C: INCOMPLETE SIGNUP OR BRAND NEW SIGNUP
    is_resumed = False
    auth_user_id = None
    token = ""
    refresh_token = ""

    # Synchronize / Create Supabase Auth User
    if auth_exists:
        is_resumed = True
        logger.info(f"Existing or partial Supabase Auth user found for {email}. Resuming setup.")
        try:
            sign_in_res = supabase.auth.sign_in_with_password({
                "email": email,
                "password": data.password,
            })
            if sign_in_res and sign_in_res.user:
                auth_user_id = sign_in_res.user.id
                if sign_in_res.session:
                    token = sign_in_res.session.access_token
                    refresh_token = sign_in_res.session.refresh_token
        except Exception as e:
            logger.info(f"Password sign-in during resume for {email} produced: {e}")
            if service_supabase is not None:
                try:
                    users_list = service_supabase.auth.admin.list_users()
                    target_u = next((u for u in users_list if u.email and u.email.lower() == email), None)
                    if target_u:
                        auth_user_id = target_u.id
                        service_supabase.auth.admin.update_user_by_id(auth_user_id, {"password": data.password})
                        sign_in_res = supabase.auth.sign_in_with_password({
                            "email": email,
                            "password": data.password,
                        })
                        if sign_in_res and sign_in_res.session:
                            token = sign_in_res.session.access_token
                            refresh_token = sign_in_res.session.refresh_token
                except Exception as admin_err:
                    logger.warning(f"Admin password sync failed for {email}: {admin_err}")

    if not auth_user_id:
        user_id = existing_user.get("id") if existing_user else str(uuid.uuid4())
        try:
            get_rpc_client().rpc("create_auth_user", {
                "p_id": user_id,
                "p_email": email,
                "p_password": data.password,
            }).execute()
            auth_user_id = user_id
            logger.info(f"Created new Auth user with ID: {auth_user_id}")
        except Exception as e:
            err_str = str(e).lower()
            if "already" in err_str or "duplicate" in err_str or "unique" in err_str:
                logger.info(f"create_auth_user reported existing auth user for {email}. Resuming registration.")
                is_resumed = True
                auth_user_id = user_id
            else:
                logger.error(f"create_auth_user failed: {e}")
                raise HTTPException(status_code=400, detail="Registration could not be completed. Please try again.")

    # Create or update Company document (Idempotent)
    company_id = existing_company.get("id") if existing_company else (existing_user.get("company_id") if existing_user else str(uuid.uuid4()))
    company_doc = {
        "id": company_id,
        "company_name": data.company_name,
        "owner_name": data.owner_name,
        "mobile": mobile,
        "alt_mobile": data.alt_mobile or "",
        "email": email,
        "gst_number": data.gst_number or "",
        "address": data.address or "",
        "city": data.city or "",
        "district": data.district or "",
        "state": data.state or "",
        "pincode": data.pincode or "",
        "business_type": data.business_type,
        "website": "",
        "support_number": "",
        "logo_file_id": None,
        "documents": {},
        "trial_started_at": existing_company.get("trial_started_at") if (existing_company and existing_company.get("trial_started_at")) else datetime.now(timezone.utc).isoformat(),
        "trial_ends_at": existing_company.get("trial_ends_at") if (existing_company and existing_company.get("trial_ends_at")) else (datetime.now(timezone.utc) + timedelta(days=15)).isoformat(),
        "subscription_status": existing_company.get("subscription_status") if (existing_company and existing_company.get("subscription_status")) else "trialing",
        "plan_id": existing_company.get("plan_id") if (existing_company and existing_company.get("plan_id")) else "starter",
        "billing_cycle": existing_company.get("billing_cycle") if (existing_company and existing_company.get("billing_cycle")) else "monthly",
        "cancel_at_period_end": existing_company.get("cancel_at_period_end", False) if existing_company else False,
        "created_at": existing_company.get("created_at") if existing_company else now_iso(),
        "registration_status": "COMPLETED"
    }

    try:
        await db.companies.update_one(
            {"id": company_id},
            {"$set": company_doc},
            upsert=True
        )
    except Exception as e:
        logger.error(f"Company upsert failed: {e}")
        raise HTTPException(status_code=400, detail="Registration could not be completed. Please try again.")

    # Create or update User document (Idempotent)
    user_id = auth_user_id or (existing_user.get("id") if existing_user else str(uuid.uuid4()))
    user_doc = {
        "id": user_id,
        "company_id": company_id,
        "name": data.owner_name,
        "email": email,
        "mobile": mobile,
        "role": "Admin",
        "user_type": "owner",
        "status": "Active",
        "permissions": default_perms_for_role("Admin"),
        "created_at": existing_user.get("created_at") if existing_user else now_iso(),
    }

    try:
        await db.users.update_one(
            {"id": user_id},
            {"$set": user_doc},
            upsert=True
        )
        _cache_invalidate_user(user_id)
        _test_temp_passwords[email] = data.password
        await log_activity(company_id, user_id, data.owner_name, "Company Registered", data.company_name)
    except Exception as e:
        logger.error(f"User upsert failed: {e}")
        raise HTTPException(status_code=400, detail="Registration could not be completed. Please try again.")

    # Obtain session tokens if not already acquired
    if not token:
        try:
            sign_in_res = supabase.auth.sign_in_with_password({
                "email": email,
                "password": data.password,
            })
            if sign_in_res and sign_in_res.session:
                token = sign_in_res.session.access_token
                refresh_token = sign_in_res.session.refresh_token
        except Exception as e:
            logger.warning(f"Auto sign-in after registration completion: {e}")

    if token:
        response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")

    user_doc.pop("_id", None)
    company_doc.pop("_id", None)
    return {
        "token": token,
        "refresh_token": refresh_token,
        "user": serialize_user(user_doc),
        "company": company_doc,
        "status": "completed",
        "resumed": is_resumed,
        "message": "Your registration was completed successfully." if is_resumed else "Account created successfully."
    }

def _lookup_user_for_login_sync(ident: str, raw: str):
    rpc_res = get_rpc_client().rpc("lookup_user_for_login", {
        "p_email": ident,
        "p_mobile": raw,
        "p_employee_id": raw
    }).execute()
    return rpc_res.data[0] if isinstance(rpc_res.data, list) and rpc_res.data else None

def _supabase_sign_in_sync(email: str, password: str):
    return supabase.auth.sign_in_with_password({
        "email": email,
        "password": password
    })

def _fetch_company_sync(company_id: str):
    company_rpc = get_rpc_client().rpc("get_company_by_id", {"p_company_id": company_id}).execute()
    return company_rpc.data[0] if isinstance(company_rpc.data, list) and company_rpc.data else None

def _lookup_user_profile_with_token_sync(user_id: str, token: str):
    """Fetch user profile using the user's own JWT - bypasses RLS since users can read their own row."""
    user_client = get_supabase_client(token=token)
    try:
        rpc_res = user_client.rpc("get_user_by_id", {"p_user_id": user_id}).execute()
        return rpc_res.data[0] if isinstance(rpc_res.data, list) and rpc_res.data else None
    except Exception:
        return None

def _fetch_company_with_token_sync(company_id: str, token: str):
    """Fetch company using the user's own JWT - bypasses RLS."""
    user_client = get_supabase_client(token=token)
    try:
        rpc_res = user_client.rpc("get_company_by_id", {"p_company_id": company_id}).execute()
        return rpc_res.data[0] if isinstance(rpc_res.data, list) and rpc_res.data else None
    except Exception:
        return None

@api_router.post("/auth/login")
async def login(data: LoginIn, response: Response):
    _t0 = time.time()
    def _elapsed():
        return round((time.time() - _t0) * 1000)

    raw = data.identifier.strip()
    ident = raw.lower()
    is_email = "@" in ident
    logger.info(f"[LOGIN] START ident={ident[:20]} is_email={is_email}")

    token = ""
    refresh_token = ""
    user = None
    auth_user_id = None

    if is_email:
        # ── Email login: authenticate directly with Supabase Auth (no RPC needed) ──
        logger.info(f"[LOGIN] step=auth_start elapsed={_elapsed()}ms")
        try:
            auth_res = await asyncio.to_thread(_supabase_sign_in_sync, ident, data.password)
            if not auth_res or not auth_res.session:
                raise HTTPException(status_code=401, detail="Invalid credentials")
            token = auth_res.session.access_token
            refresh_token = auth_res.session.refresh_token
            auth_user_id = auth_res.session.user.id if auth_res.session.user else None
            logger.info(f"[LOGIN] step=auth_done user_id={auth_user_id} elapsed={_elapsed()}ms")
        except HTTPException:
            raise
        except Exception as e:
            err_str = str(e).lower()
            logger.error(f"[LOGIN] step=auth_failed elapsed={_elapsed()}ms err={e}")
            db_u = await db.users.find_one({"email": ident.lower().strip()})
            if db_u and db_u.get("password_hash") and verify_password(data.password, db_u["password_hash"]):
                try:
                    if service_supabase is not None:
                        service_supabase.auth.admin.update_user_by_id(db_u["id"], {"password": data.password})
                    auth_res = await asyncio.to_thread(_supabase_sign_in_sync, ident, data.password)
                    if auth_res and auth_res.session:
                        token = auth_res.session.access_token
                        refresh_token = auth_res.session.refresh_token
                        auth_user_id = auth_res.session.user.id
                except Exception:
                    pass
                if not token:
                    token_payload = {
                        "sub": db_u["id"],
                        "id": db_u["id"],
                        "user_id": db_u["id"],
                        "email": db_u["email"],
                        "company_id": db_u["company_id"],
                        "role": db_u.get("role", "Staff"),
                        "exp": datetime.now(timezone.utc) + timedelta(days=7)
                    }
                    token = jwt.encode(token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
                    refresh_token = token
                    auth_user_id = db_u["id"]
                    user = db_u
            else:
                if "email not confirmed" in err_str:
                    raise HTTPException(status_code=401, detail="Email not confirmed. Please check your inbox.")
                raise HTTPException(status_code=401, detail="Invalid credentials")

        # Fetch user profile using the user's own JWT (bypasses RLS without needing service key)
        logger.info(f"[LOGIN] step=profile_lookup_start elapsed={_elapsed()}ms")
        if auth_user_id:
            try:
                user = await asyncio.to_thread(_lookup_user_profile_with_token_sync, auth_user_id, token)
                if user:
                    logger.info(f"[LOGIN] step=profile_via_user_jwt elapsed={_elapsed()}ms")
            except Exception:
                pass
            if not user:
                try:
                    # Fallback: try service RPC
                    rpc_res = await asyncio.to_thread(
                        lambda: get_rpc_client().rpc("get_user_by_id", {"p_user_id": auth_user_id}).execute()
                    )
                    user = rpc_res.data[0] if isinstance(rpc_res.data, list) and rpc_res.data else None
                    if user:
                        logger.info(f"[LOGIN] step=profile_via_service_rpc elapsed={_elapsed()}ms")
                except Exception:
                    pass
            if not user:
                # Last resort: try direct DB lookup via CollectionAdapter
                try:
                    user = await db.users.find_one({"id": auth_user_id}, {"_id": 0})
                    if user:
                        logger.info(f"[LOGIN] step=profile_via_db elapsed={_elapsed()}ms")
                except Exception:
                    pass

        if not user or not isinstance(user, dict):
            logger.error(f"[LOGIN] step=profile_not_found user_id={auth_user_id} elapsed={_elapsed()}ms")
            raise HTTPException(status_code=401, detail="User profile not found. Please contact admin.")

    else:
        # ── Employee ID / Mobile login: RPC lookup to get email, then sign in ──
        logger.info(f"[LOGIN] step=rpc_lookup_start elapsed={_elapsed()}ms")
        try:
            user = await asyncio.to_thread(_lookup_user_for_login_sync, ident, raw)
            logger.info(f"[LOGIN] step=rpc_lookup_done found={bool(user)} elapsed={_elapsed()}ms")
        except Exception as e:
            logger.error(f"[LOGIN] step=rpc_lookup_failed elapsed={_elapsed()}ms err={e}")
            user = None

        if not user or not isinstance(user, dict):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        email_for_auth = str(user.get("email") or "")
        logger.info(f"[LOGIN] step=auth_start elapsed={_elapsed()}ms")
        try:
            auth_res = await asyncio.to_thread(_supabase_sign_in_sync, email_for_auth, data.password)
            if not auth_res or not auth_res.session:
                raise HTTPException(status_code=401, detail="Invalid credentials")
            token = auth_res.session.access_token
            refresh_token = auth_res.session.refresh_token
            logger.info(f"[LOGIN] step=auth_done elapsed={_elapsed()}ms")
        except HTTPException:
            raise
        except Exception as e:
            err_str = str(e).lower()
            logger.error(f"[LOGIN] step=auth_failed elapsed={_elapsed()}ms err={e}")
            if "invalid login credentials" in err_str or "invalid_credentials" in err_str:
                raise HTTPException(status_code=401, detail="Invalid credentials")
            if "email not confirmed" in err_str:
                raise HTTPException(status_code=401, detail="Email not confirmed. Please check your inbox.")
            raise HTTPException(status_code=401, detail="Invalid credentials")

    if user.get("status") == "Inactive":
        raise HTTPException(status_code=403, detail="Account is inactive")

    response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")

    if user.get("role") == "Installer":
        perms = user.get("permissions")
        if not isinstance(perms, dict):
            user["permissions"] = {}
            perms = user["permissions"]
        if not isinstance(perms.get("reports"), dict):
            perms["reports"] = {}
        perms["reports"]["view"] = False

    cid = str(user.get("company_id") or "")
    logger.info(f"[LOGIN] step=company_lookup_start cid={cid} elapsed={_elapsed()}ms")
    company = _cache_get_company(cid) if cid else None
    if not company and cid:
        try:
            company = await asyncio.to_thread(_fetch_company_with_token_sync, cid, token)
            if not company:
                company = await asyncio.to_thread(_fetch_company_sync, cid)
            if not company:
                company = await db.companies.find_one({"id": cid}, {"_id": 0})
            if company and isinstance(company, dict):
                _cache_put_company(cid, company)
        except Exception as e:
            logger.error(f"[LOGIN] step=company_failed elapsed={_elapsed()}ms err={e}")
            company = None

    if not company:
        user_email = (user.get("email") or "").lower().strip()
        if user_email:
            try:
                company = await db.companies.find_one({"email": user_email}, {"_id": 0})
                if company and isinstance(company, dict):
                    user["company_id"] = company["id"]
                    _cache_put_company(company["id"], company)
                    await db.users.update_one({"id": user["id"]}, {"$set": {"company_id": company["id"]}})
            except Exception as e:
                logger.error(f"[LOGIN] step=company_by_email_failed err={e}")

    if (user.get("email") or "").strip().lower() in SUPER_ADMIN_EMAILS:
        user["is_super_admin"] = True
        user["is_platform_owner"] = True
        user["user_type"] = "super_admin"
        user["role"] = "Super Admin"

    # Proactively seed auth cache so subsequent requests (/auth/me, /clients, etc.) resolve instantly
    _cache_put_user(token, user)
    logger.info(f"[LOGIN] COMPLETE elapsed={_elapsed()}ms")
    return {"token": token, "refresh_token": refresh_token, "user": serialize_user(user), "company": company}

@api_router.post("/auth/logout")
async def logout(response: Response):
    try:
        supabase.auth.sign_out()
    except Exception:
        pass
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api_router.post("/auth/refresh")
async def refresh_token_endpoint(data: RefreshIn, response: Response):
    try:
        auth_res = supabase.auth.refresh_session(data.refresh_token)
        if not auth_res or not auth_res.session:
            raise HTTPException(status_code=401, detail="Invalid refresh token")
        token = auth_res.session.access_token
        new_refresh = auth_res.session.refresh_token
        response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
        return {
            "token": token,
            "refresh_token": new_refresh
        }
    except Exception as e:
        logger.error(f"Token refresh failed: {e}")
        raise HTTPException(status_code=401, detail="Invalid or expired refresh token")

@api_router.get("/auth/me")
async def me(user=Depends(get_current_user)):
    cid = user.get("company_id")
    company = _cache_get_company(cid) if cid else None
    if not company and cid:
        company = await db.companies.find_one({"id": cid}, {"_id": 0})
        if company and isinstance(company, dict):
            _cache_put_company(cid, company)
    if not company and user.get("email"):
        company = await db.companies.find_one({"email": user["email"].lower().strip()}, {"_id": 0})
        if company and isinstance(company, dict):
            if user.get("id"):
                await db.users.update_one({"id": user["id"]}, {"$set": {"company_id": company["id"]}})
                user["company_id"] = company["id"]
            _cache_put_company(company["id"], company)
    return {"user": user, "company": company}

@api_router.patch("/auth/me")
async def update_my_profile(data: MyProfileUpdate, user=Depends(get_current_user)):
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update:
        return user
    await db.users.update_one({"id": user["id"]}, {"$set": update})
    _cache_invalidate_user(user["id"])
    await log_activity(user["company_id"], user["id"], user["name"], "Updated Profile")
    refreshed = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0})
    return refreshed

@api_router.post("/auth/change-email")
async def change_email(data: ChangeEmailIn, request: Request, user=Depends(get_current_user)):
    try:
        supabase.auth.sign_in_with_password({
            "email": user["email"],
            "password": data.current_password
        })
    except Exception:
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    new_email = data.new_email.lower().strip()
    if new_email == user.get("email", "").lower().strip():
        raise HTTPException(status_code=400, detail="New email is the same as the current email")
    try:
        email_exists_res = get_rpc_client().rpc("check_email_exists", {"email_to_check": new_email}).execute()
        if email_exists_res.data:
            raise HTTPException(status_code=400, detail="Email already in use")
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Failed to check email exists globally: {e}")

    old_email = user["email"].lower()
    try:
        if service_supabase is not None:
            service_supabase.auth.admin.update_user_by_id(user["id"], {"email": new_email})
        else:
            token = request.cookies.get("access_token") or request.headers.get("Authorization", "").replace("Bearer ", "")
            client = get_supabase_client(token=token)
            client.auth.set_session(token, "")
            client.auth.update_user({"email": new_email})
    except Exception as e:
        logger.error(f"Supabase update_user email failed: {e}")
        raise HTTPException(status_code=400, detail="Authentication update failed. Please check inputs.")

    # Keep password sync active for new email
    if old_email in _test_temp_passwords:
        _test_temp_passwords[new_email] = _test_temp_passwords[old_email]
        _test_temp_passwords.pop(old_email, None)

    # Update user profile table ONLY (do NOT alter db.companies.email)
    await db.users.update_one({"id": user["id"]}, {"$set": {"email": new_email}})

    _cache_invalidate_user(user["id"])
    await log_activity(user["company_id"], user["id"], user["name"], "Changed Email", new_email)
    return {"ok": True, "email": new_email}

@api_router.post("/auth/change-password")
async def change_password(data: ChangePasswordIn, request: Request, user=Depends(get_current_user)):
    try:
        supabase.auth.sign_in_with_password({
            "email": user["email"],
            "password": data.current_password
        })
    except Exception:
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")

    updated_in_supabase = False
    if service_supabase is not None:
        try:
            service_supabase.auth.admin.update_user_by_id(user["id"], {"password": data.new_password})
            updated_in_supabase = True
        except Exception as se:
            logger.warning(f"service_supabase update_user_by_id failed: {se}")

    if not updated_in_supabase:
        try:
            token = request.cookies.get("access_token") or request.headers.get("Authorization", "").replace("Bearer ", "")
            client = get_supabase_client(token=token)
            client.auth.set_session(token, "")
            client.auth.update_user({"password": data.new_password})
        except Exception as e:
            logger.error(f"Supabase update_user password failed: {e}")
            raise HTTPException(status_code=400, detail="Password update failed. Please try again.")

    _test_temp_passwords[user["email"].lower()] = data.new_password
    await log_activity(user["company_id"], user["id"], user["name"], "Changed Password")
    return {"ok": True}

# ---------- Forgot Password ----------
class ForgotPasswordIn(BaseModel):
    email: EmailStr

class VerifyOtpIn(BaseModel):
    email: EmailStr
    otp: str

class ResetPasswordIn(BaseModel):
    reset_token: str
    new_password: str

# In-memory cooldown: maps email -> timestamp of last reset email sent
# Prevents repeated calls to Supabase within 60 seconds (avoids 429 rate limits)
_forgot_pw_cooldown: dict[str, float] = {}
FORGOT_PW_COOLDOWN_SECONDS = 60

@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPasswordIn):
    import time
    email = data.email.lower().strip()

    user = await db.users.find_one({"email": email})
    if not user:
        # Avoid enumeration leak
        return {"ok": True, "message": "If the email exists, an OTP has been sent."}

    # Cooldown or throttling: limit active OTP documents in DB
    count = await db.password_reset_otps.count_documents({"email": email})
    if count >= 3:
        logger.info(f"Forgot-password throttle active for {email}")
        return {"ok": True, "message": "If the email exists, an OTP has been sent."}

    # Generate 6-digit OTP
    otp_code = f"{secrets.randbelow(1000000):06d}"
    hashed_otp = bcrypt.hashpw(otp_code.encode(), bcrypt.gensalt()).decode()

    # Save to db
    otp_doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "user_id": user["id"],
        "code_hash": hashed_otp,
        "attempts": 0,
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat(),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.password_reset_otps.insert_one(otp_doc)

    # Trigger Supabase recovery just in case (for production users)
    try:
        supabase.auth.reset_password_for_email(email)
    except Exception as e:
        logger.warning(f"Failed to reset password in Supabase: {e}")

    # Send the custom OTP email
    try:
        await send_email(email, "Reset your password", render_otp_email(otp_code))
    except Exception as e:
        logger.error(f"Failed to send OTP email: {e}")

    return {"ok": True, "message": "If the email exists, an OTP has been sent."}


@api_router.post("/auth/verify-otp")
async def verify_otp(data: VerifyOtpIn):
    email = data.email.lower().strip()
    otp = (data.otp or "").strip()
    if not otp.isdigit() or len(otp) != 6:
        raise HTTPException(status_code=400, detail="OTP must be 6 digits")

    # Find the most recent active OTP document in local DB
    otp_doc = await db.password_reset_otps.find_one(
        {"email": email, "used": False},
        sort=[("created_at", -1)]
    )

    if not otp_doc:
        # Fall back to Supabase
        try:
            res = supabase.auth.verify_otp({
                "email": email,
                "token": otp,
                "type": "recovery"
            })
            if not res or not res.session:
                raise HTTPException(status_code=400, detail="OTP has expired or is invalid.")
            reset_token = res.session.access_token
            return {"reset_token": reset_token, "expires_in_minutes": 10}
        except Exception as e:
            logger.error(f"Supabase verify_otp failed: {e}")
            raise HTTPException(status_code=400, detail="OTP has expired or is invalid. Request a new one.")

    # Check if expired
    expires_at = datetime.fromisoformat(otp_doc["expires_at"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="OTP has expired or is invalid. Request a new one.")

    # Check attempts
    if otp_doc.get("attempts", 0) >= 3:
        raise HTTPException(status_code=400, detail="Too many attempts. Request a new one.")

    # Verify code
    if not bcrypt.checkpw(otp.encode(), otp_doc["code_hash"].encode()):
        await db.password_reset_otps.update_one(
            {"id": otp_doc["id"]},
            {"$inc": {"attempts": 1}}
        )
        raise HTTPException(status_code=400, detail="OTP has expired or is invalid. Request a new one.")

    # Mark as used
    await db.password_reset_otps.update_one(
        {"id": otp_doc["id"]},
        {"$set": {"used": True}}
    )

    # Generate custom reset token
    reset_token = secrets.token_urlsafe(32)
    token_doc = {
        "token": reset_token,
        "email": email,
        "user_id": otp_doc["user_id"],
        "used": False,
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.password_reset_tokens.insert_one(token_doc)

    return {"reset_token": reset_token, "expires_in_minutes": 10}

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordIn):
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")

    # Check if it's a custom reset token
    token_doc = await db.password_reset_tokens.find_one({"token": data.reset_token, "used": False})
    if token_doc:
        # Check if expired
        expires_at = datetime.fromisoformat(token_doc["expires_at"].replace("Z", "+00:00"))
        if datetime.now(timezone.utc) > expires_at:
            raise HTTPException(status_code=400, detail="Reset link expired or already used. Start again.")

        # Mark as used
        await db.password_reset_tokens.update_one({"token": token_doc["token"]}, {"$set": {"used": True}})
        user_id = token_doc["user_id"]
        user = await db.users.find_one({"id": user_id})
        if not user:
            raise HTTPException(status_code=400, detail="User not found")

        # Update password in Supabase using Admin API or temporary sign in fallback
        updated_in_supabase = False
        if service_supabase is not None:
            try:
                service_supabase.auth.admin.update_user_by_id(user_id, {"password": data.new_password})
                updated_in_supabase = True
            except Exception as se:
                logger.error(f"service_supabase update_user_by_id in reset_password failed: {se}")

        if not updated_in_supabase:
            temp_pwd = _test_temp_passwords.get(user["email"].lower())
            if default_supabase is not None and temp_pwd:
                try:
                    login_res = default_supabase.auth.sign_in_with_password({
                        "email": user["email"],
                        "password": temp_pwd
                    })
                    if login_res and login_res.session:
                        token = login_res.session.access_token
                        client = get_supabase_client(token=token)
                        client.auth.set_session(token, "")
                        client.auth.update_user({"password": data.new_password})
                except Exception as e:
                    logger.error(f"Failed to update password in Supabase auth during custom reset: {e}")

        # Update temp_password in-memory
        _test_temp_passwords[user["email"].lower()] = data.new_password
    else:
        # Fall back to Supabase JWT reset path
        try:
            client = get_supabase_client(token=data.reset_token)
            client.auth.set_session(data.reset_token, "")
            res = client.auth.update_user({"password": data.new_password})
            if not res or not res.user:
                raise HTTPException(status_code=400, detail="Reset link expired or already used.")
            user_id = res.user.id
            user = await db.users.find_one({"id": user_id})
            if user:
                _test_temp_passwords[user["email"].lower()] = data.new_password
        except Exception as e:
            logger.error(f"Supabase update_user reset password failed: {e}")
            raise HTTPException(status_code=400, detail="Reset link expired or already used. Start again.")

    if user:
        await log_activity(user["company_id"], user["id"], user["name"], "Reset Password (via OTP)")
        try:
            await send_email(user["email"], "Your Solarix password was changed", render_password_changed_email())
        except Exception:
            pass

    return {"ok": True, "email": user["email"] if user else ""}

class GoogleLoginIn(BaseModel):
    email: str
    supabase_access_token: Optional[str] = None

@api_router.post("/auth/google")
async def google_login(data: GoogleLoginIn, response: Response):
    email = (data.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Google email is required")

    # Optional: If Supabase access token is supplied, verify it against Supabase auth.get_user
    if data.supabase_access_token:
        try:
            sb_user_res = supabase.auth.get_user(data.supabase_access_token)
            if sb_user_res and sb_user_res.user and sb_user_res.user.email:
                sb_email = sb_user_res.user.email.strip().lower()
                if sb_email and sb_email != email:
                    logger.warning(f"Google OAuth email mismatch: claimed={email}, supabase={sb_email}")
                    email = sb_email
        except Exception as e:
            logger.warning(f"Failed to verify Supabase OAuth token on backend: {e}")

    # Search for an existing user record in database with matching email
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        all_users = await db.users.find({}, {"_id": 0}).to_list(10000)
        user = next((u for u in all_users if (u.get("email") or "").strip().lower() == email), None)

    if not user:
        raise HTTPException(
            status_code=403,
            detail="This Google account is not authorized for Solarix. Contact your company administrator."
        )

    if user.get("status") == "Inactive":
        raise HTTPException(status_code=403, detail="This Google account is inactive. Contact your company administrator.")

    cid = str(user.get("company_id") or "")
    if not cid:
        raise HTTPException(status_code=403, detail="Account not authorized. No valid company associated.")

    company = _cache_get_company(cid)
    if not company:
        company = await db.companies.find_one({"id": cid}, {"_id": 0})
        if company and isinstance(company, dict):
            _cache_put_company(cid, company)

    if not company:
        raise HTTPException(status_code=403, detail="Account not authorized. Company workspace not found.")

    if user.get("role") == "Installer":
        perms = user.get("permissions")
        if not isinstance(perms, dict):
            user["permissions"] = {}
            perms = user["permissions"]
        if not isinstance(perms.get("reports"), dict):
            perms["reports"] = {}
        perms["reports"]["view"] = False

    now = time.time()
    payload = {
        "sub": user["id"],
        "user_id": user["id"],
        "email": user.get("email", ""),
        "role": user.get("role", "Employee"),
        "company_id": cid,
        "exp": int(now + 604800)
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    refresh_token = jwt.encode({"sub": user["id"], "exp": int(now + 2592000)}, JWT_SECRET, algorithm=JWT_ALGORITHM)

    if (user.get("email") or "").strip().lower() in SUPER_ADMIN_EMAILS:
        user["is_super_admin"] = True
        user["is_platform_owner"] = True
        user["user_type"] = "super_admin"
        user["role"] = "Super Admin"

    response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    _cache_put_user(token, user)

    return {
        "token": token,
        "refresh_token": refresh_token,
        "user": serialize_user(user),
        "company": company
    }

# ---------- Company ----------
@api_router.get("/company")
async def get_company(user=Depends(get_current_user)):
    projection = {
        "_id": 0,
        "id": 1,
        "company_name": 1,
        "owner_name": 1,
        "mobile": 1,
        "alt_mobile": 1,
        "email": 1,
        "gst_number": 1,
        "address": 1,
        "city": 1,
        "state": 1,
        "pincode": 1,
        "business_type": 1,
        "website": 1,
        "support_number": 1,
        "logo_file_id": 1,
        "documents": 1,
        "trial_start": 1,
        "trial_end": 1,
        "plan": 1,
        "created_at": 1,
    }
    return await db.companies.find_one({"id": user["company_id"]}, projection)

@api_router.put("/company")
async def update_company(data: CompanyUpdate, request: Request, user=Depends(get_current_user)):
    if not (user.get("user_type") == "owner" or user.get("role") in ("Super Admin", "Admin") or has_perm(user, "settings", "edit")):
        raise HTTPException(status_code=403, detail="Permission required to update company details")
    
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update:
        pass
    elif "email" in update:
        new_email = update["email"].lower().strip()
        current_company = await db.companies.find_one({"id": user["company_id"]})
        current_email = current_company.get("email", "").lower().strip() if current_company else ""
        
        if new_email != current_email:
            if not re.match(r"[^@]+@[^@]+\.[^@]+", new_email):
                raise HTTPException(status_code=400, detail="Invalid email format")
            
            if await db.users.find_one({"email": new_email}):
                raise HTTPException(status_code=400, detail="Email already in use")
            
            try:
                email_exists_res = get_rpc_client().rpc("check_email_exists", {"email_to_check": new_email}).execute()
                if email_exists_res.data:
                    raise HTTPException(status_code=400, detail="Email already in use")
            except HTTPException:
                raise
            except Exception as e:
                logger.warning(f"Failed to check email exists globally: {e}")

            token = request.cookies.get("access_token") or request.headers.get("Authorization", "").replace("Bearer ", "")
            client = get_supabase_client(token=token)
            client.auth.set_session(token, "")
            try:
                client.auth.update_user({"email": new_email})
            except Exception as e:
                logger.error(f"Supabase update_user email failed from company update: {e}")
                raise HTTPException(status_code=400, detail=f"Authentication update failed: {e}")

            old_email = user["email"].lower()
            if old_email in _test_temp_passwords:
                _test_temp_passwords[new_email] = _test_temp_passwords[old_email]
                _test_temp_passwords.pop(old_email, None)

            await db.companies.update_one({"id": user["company_id"]}, {"$set": update})
            await db.users.update_one({"id": user["id"]}, {"$set": {"email": new_email}})
            _cache_invalidate_user(user["id"])
        else:
            await db.companies.update_one({"id": user["company_id"]}, {"$set": update})
    else:
        await db.companies.update_one({"id": user["company_id"]}, {"$set": update})
        
    await log_activity(user["company_id"], user["id"], user["name"], "Updated Company Profile")
    _cache_invalidate_company(user["company_id"])
    
    projection = {
        "_id": 0,
        "id": 1,
        "company_name": 1,
        "owner_name": 1,
        "mobile": 1,
        "alt_mobile": 1,
        "email": 1,
        "gst_number": 1,
        "address": 1,
        "city": 1,
        "state": 1,
        "pincode": 1,
        "business_type": 1,
        "website": 1,
        "support_number": 1,
        "logo_file_id": 1,
        "documents": 1,
        "trial_start": 1,
        "trial_end": 1,
        "plan": 1,
        "created_at": 1,
    }
    return await db.companies.find_one({"id": user["company_id"]}, projection)

@api_router.delete("/company")
async def delete_company(response: Response, user=Depends(get_current_user)):
    if user["role"] != "Admin":
        raise HTTPException(status_code=403, detail="Admin only")

    company_id = user["company_id"]
    email = user["email"].lower().strip()

    # Use service role client if available (bypasses RLS), else fallback to user's auth client
    client_to_use = service_supabase if service_supabase is not None else supabase

    # 1. Fetch all users belonging to this company
    try:
        users_cursor = db.users.find({"company_id": company_id})
        company_users = await users_cursor.to_list()
    except Exception as e:
        logger.error(f"Failed to fetch company users: {e}")
        company_users = []

    user_ids = [u["id"] for u in company_users] if company_users else []
    emails = [u["email"].lower().strip() for u in company_users if u.get("email")] if company_users else []
    if email not in emails:
        emails.append(email)
    if user["id"] not in user_ids:
        user_ids.append(user["id"])

    logger.info(f"Initiating complete deletion for company {company_id}. Associated users: {user_ids}, emails: {emails}")

    complaint_ids = []
    try:
        complaints_res = client_to_use.table("complaints").select("id").eq("company_id", company_id).execute()
        if complaints_res.data:
            complaint_ids = [c["id"] for c in complaints_res.data if isinstance(c, dict) and "id" in c]
    except Exception as e:
        logger.error(f"Failed to fetch complaints for deletion: {e}")

    task_ids = []
    try:
        tasks_res = client_to_use.table("tasks").select("id").eq("company_id", company_id).execute()
        if tasks_res.data:
            task_ids = [t["id"] for t in tasks_res.data if isinstance(t, dict) and "id" in t]
    except Exception as e:
        logger.error(f"Failed to fetch tasks for deletion: {e}")

    # 2. Deletions of child records first (linked by user, email, complaint, or task)
    # 2a. Delete password_reset_tokens
    if user_ids:
        for u_id in user_ids:
            try:
                client_to_use.table("password_reset_tokens").delete().eq("user_id", u_id).execute()
            except Exception as e:
                logger.debug(f"Failed to clear password_reset_tokens: {e}")

    # 2b. Delete password_reset_otps
    if emails:
        for em in emails:
            try:
                client_to_use.table("password_reset_otps").delete().eq("email", em).execute()
            except Exception as e:
                logger.debug(f"Failed to clear password_reset_otps: {e}")

    # 2c. Delete verifications
    if user_ids:
        for u_id in user_ids:
            try:
                client_to_use.table("verifications").delete().eq("user_id", u_id).execute()
            except Exception as e:
                logger.debug(f"Failed to clear verifications: {e}")

    # 2d. Delete complaint_comments & complaint_audit
    if complaint_ids:
        for c_id in complaint_ids:
            try:
                client_to_use.table("complaint_comments").delete().eq("complaint_id", c_id).execute()
            except Exception as e:
                logger.debug(f"Failed to clear complaint_comments: {e}")
            try:
                client_to_use.table("complaint_audit").delete().eq("complaint_id", c_id).execute()
            except Exception as e:
                logger.debug(f"Failed to clear complaint_audit: {e}")

    # 2e. Delete task_updates
    if task_ids:
        for t_id in task_ids:
            try:
                client_to_use.table("task_updates").delete().eq("task_id", t_id).execute()
            except Exception as e:
                logger.debug(f"Failed to clear task_updates: {e}")

    # 3. Clean up database tables in order to prevent foreign key constraint issues
    tables_to_clean = [
        "activity_logs",
        "notifications",
        "tasks",
        "complaints",
        "files",
        "outward_entries",
        "inward_entries",
        "material_requests",
        "products",
        "inverter_monitoring",
        "document_templates",
        "assets",
        "projects",
        "clients",
        "counters",
        "service_tickets",
        "employees",
        "users",
        "companies",
    ]

    # Delete records linked by company_id or id
    for t in tables_to_clean:
        try:
            col = "id" if t == "companies" else "company_id"
            res = client_to_use.table(t).delete().eq(col, company_id).execute()
            logger.info(f"Cleared table {t} for company {company_id}: deleted {len(res.data) if res.data else 0} rows")
        except Exception as e:
            err_msg = str(e)
            if "PGRST205" in err_msg or "schema cache" in err_msg:
                logger.info(f"Table {t} does not exist in schema, bypassing.")
            else:
                logger.error(f"Failed delete for table {t} by company_id: {e}")
                raise e

    # 4. Delete users from Supabase Auth
    if service_supabase is not None:
        for u_id in user_ids:
            try:
                service_supabase.auth.admin.delete_user(u_id)
                logger.info(f"Deleted user {u_id} from Supabase Auth")
            except Exception as e:
                logger.error(f"Failed to delete user {u_id} from Supabase Auth: {e}")
    else:
        logger.warning("SUPABASE_SERVICE_ROLE_KEY is not set. Skipping Supabase Auth deletion.")

    # 5. Invalidate cached auth states
    for u_id in user_ids:
        _cache_invalidate_user(u_id)

    # 6. Delete authentication cookie
    response.delete_cookie("access_token", path="/")

    return {"ok": True, "detail": "Company and all associated accounts/records permanently deleted."}

# ---------- Files ----------
@api_router.post("/files/upload")
async def upload_file(file: UploadFile = File(...), category: str = Form("general"), user=Depends(get_current_user)):
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "bin"
    file_id = str(uuid.uuid4())
    path = f"{APP_NAME}/{user['company_id']}/{category}/{file_id}.{ext}"
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    content_type = file.content_type or "application/octet-stream"
    result = put_object(path, data, content_type)
    doc = {
        "id": file_id, "company_id": user["company_id"], "uploader_id": user["id"],
        "storage_path": result["path"], "original_filename": file.filename,
        "content_type": content_type, "size": result.get("size", len(data)),
        "category": category, "is_deleted": False, "created_at": now_iso(),
    }
    await db.files.insert_one(doc)
    return {"id": file_id, "filename": file.filename, "content_type": content_type, "size": doc["size"]}

@api_router.get("/files/{file_id}")
async def download_file(file_id: str, request: Request, auth: Optional[str] = Query(None), download: Optional[int] = Query(None)):
    token = request.cookies.get("access_token") or auth
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]

    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    user = None
    try:
        user = await get_current_user(request)
    except Exception:
        cached = _cache_get_user(token)
        if cached:
            user = cached

    if not user or not isinstance(user, dict) or not user.get("company_id"):
        try:
            payload = jwt.decode(token, options={"verify_signature": False})
            if isinstance(payload, dict):
                company_id = (
                    payload.get("company_id") or
                    (payload.get("user_metadata") or {}).get("company_id") or
                    (payload.get("app_metadata") or {}).get("company_id")
                )
                user_id = payload.get("sub") or payload.get("user_id") or payload.get("id")
                if not company_id and user_id:
                    u_rec = await db.users.find_one({"$or": [{"id": user_id}, {"email": payload.get("email")}]}, {"_id": 0, "company_id": 1})
                    if u_rec and u_rec.get("company_id"):
                        company_id = u_rec["company_id"]
                    else:
                        c_rec = await db.companies.find_one({"owner_id": user_id}, {"_id": 0, "id": 1})
                        if c_rec:
                            company_id = c_rec["id"]
                if company_id:
                    user = {"company_id": company_id, "id": user_id or "user"}
        except Exception:
            pass

    if not user or not user.get("company_id"):
        raise HTTPException(status_code=401, detail="Invalid token or missing company context")

    company_id = user["company_id"]

    # Clean UUID/filename if a path or filename with extension was passed
    clean_id = file_id.split("/")[-1].rsplit(".", 1)[0] if "/" in file_id or "." in file_id else file_id

    rec = await db.files.find_one({
        "$or": [
            {"id": file_id},
            {"id": clean_id},
            {"storage_path": file_id},
            {"storage_path": {"$regex": clean_id}}
        ],
        "company_id": company_id,
        "is_deleted": False
    })

    if not rec:
        # Fallback check across tenant context for user uploaded document
        rec = await db.files.find_one({
            "$or": [
                {"id": file_id},
                {"id": clean_id},
                {"storage_path": file_id}
            ],
            "is_deleted": False
        })
        if rec and rec.get("company_id") and rec["company_id"] != company_id:
            raise HTTPException(status_code=403, detail="You do not have permission to access this document.")

    if not rec:
        raise HTTPException(status_code=404, detail="Document file is unavailable.")

    try:
        data, ct = get_object(rec["storage_path"])
    except Exception as e:
        logger.error(f"Error downloading object from storage {rec.get('storage_path')}: {e}")
        raise HTTPException(status_code=404, detail="Document file is unavailable in storage.")

    media_type = rec.get("content_type") or ct or "application/octet-stream"
    original_filename = rec.get("original_filename") or rec.get("filename") or f"document_{clean_id[:8]}"

    is_docx = "wordprocessingml" in media_type or original_filename.endswith(".docx")
    disposition_type = "attachment" if (download == 1 or is_docx) else "inline"
    
    headers = {
        "Content-Disposition": f'{disposition_type}; filename="{original_filename}"',
        "Access-Control-Allow-Origin": "*",
        "X-Content-Type-Options": "nosniff"
    }
    return FastAPIResponse(content=data, media_type=media_type, headers=headers)

# ---------- Clients ----------
@api_router.get("/clients")
async def list_clients(
    user=Depends(get_current_user),
    limit: int = 200,
    skip: int = 0,
    search: Optional[str] = None,
    status: Optional[str] = None,
    phase_type: Optional[str] = None,
    subsidy_eligible: Optional[bool] = None,
):
    q: Dict[str, Any] = {"company_id": user["company_id"]}
    if status and status != "All":
        q["status"] = status
    if phase_type and phase_type != "All":
        q["phase_type"] = phase_type
    if subsidy_eligible is not None:
        q["subsidy_eligible"] = subsidy_eligible
    limit = min(limit, 500)
    skip = max(0, skip)
    projection = {
        "_id": 0, "id": 1, "sol_id": 1, "full_name": 1, "mobile": 1,
        "consumer_number": 1, "status": 1, "system_kw": 1, "phase_type": 1,
        "subsidy_eligible": 1, "progress": 1, "address": 1, "city": 1,
        "created_at": 1, "updated_at": 1, "stages": 1,
    }
    if search:
        s = search.lower()
        q["$or"] = [
            {"full_name":       {"$regex": s, "$options": "i"}},
            {"mobile":          {"$regex": s}},
            {"consumer_number": {"$regex": s}},
            {"sol_id":          {"$regex": s, "$options": "i"}},
        ]
    return await db.clients.find(q, projection).sort("created_at", -1).skip(skip).to_list(limit)

@api_router.get("/clients/stats")
async def client_stats(user=Depends(get_current_user)):
    cid = user["company_id"]
    try:
        builder = supabase.table("clients").select("status,subsidy_eligible,system_kw").eq("company_id", cid)
        res = await asyncio.to_thread(builder.execute)
        rows = res.data or []
    except Exception as e:
        logger.warning(f"Failed to fetch client stats from Supabase: {e}")
        rows = await db.clients.find({"company_id": cid}, {"status": 1, "subsidy_eligible": 1, "system_kw": 1}).to_list(100000)

    total = len(rows)
    completed = sum(1 for r in rows if r.get("status") == "Handover Complete")
    pending = total - completed
    subsidy = sum(1 for r in rows if r.get("subsidy_eligible") is True)
    total_kw = sum(float(r.get("system_kw") or 0) for r in rows if r.get("status") == "Handover Complete")

    return {
        "total":     total,
        "completed": completed,
        "pending":   pending,
        "subsidy":   subsidy,
        "total_kw":  total_kw,
    }

@api_router.post("/clients")
async def create_client(data: ClientIn, user=Depends(get_current_user)):
    if not has_perm(user, "clients", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.create")

    # Check plan client limit
    from plan_config import get_plan_limits
    c_doc = await db.companies.find_one({"id": user["company_id"]}) or {}
    st = c_doc.get("subscription_status") or "trialing"
    pid = c_doc.get("plan_id") or "starter"
    is_trial = st == "trialing"
    limits = get_plan_limits(pid, is_trial=is_trial)
    curr_clients = await db.clients.count_documents({"company_id": user["company_id"]})
    if curr_clients >= limits["max_clients"]:
        raise HTTPException(
            status_code=403,
            detail=f"PLAN_LIMIT_REACHED: Your current {pid.upper()} plan allows a maximum of {limits['max_clients']} active clients/projects. Please upgrade your plan to add more clients."
        )
    client_id = str(uuid.uuid4())
    sol_id = await next_client_id(user["company_id"])
    stages = data.stages or {s: False for s in DEFAULT_STAGES}
    if data.status in ["Approved", "Installation Pending", "Installation Complete", "Handover Complete"]:
        stages["Onboarding"] = True
    payload = data.model_dump()
    payload["stages"] = stages
    full_doc = {
        "id": client_id, "sol_id": sol_id, "company_id": user["company_id"],
        "created_by": user["id"], **payload,
        "progress": calc_progress(stages),
        "notes": [], "documents": data.documents or [],
        "created_at": now_iso(), "updated_at": now_iso(),
    }

    # ── Prepare a clean Supabase-safe payload ─────────────────────────────────
    # Passing consumer_category, consumer_type, section_number, inverters etc.
    # directly as top-level columns causes PGRST204 (schema cache miss), which
    # makes insert_one silently fall back to local JSON only — the client never
    # lands in Supabase, so GET /clients/{id} returns 404 on every cold start.
    supabase_doc = _prepare_client_supabase_payload(dict(full_doc))
    supabase_doc["id"] = client_id          # always keep the primary key
    supabase_doc["sol_id"] = sol_id
    supabase_doc["company_id"] = user["company_id"]
    supabase_doc["created_by"] = user["id"]
    supabase_doc["created_at"] = full_doc["created_at"]
    supabase_doc["updated_at"] = full_doc["updated_at"]
    supabase_doc["notes"] = []
    supabase_doc["documents"] = full_doc["documents"]
    supabase_doc["progress"] = full_doc["progress"]
    supabase_doc["status"] = full_doc.get("status") or data.status

    logger.info(f"[CLIENT-CREATE DIAG] Creating client id={client_id} sol_id={sol_id} for company={user['company_id']}")
    logger.info(f"[CLIENT-CREATE DIAG] Supabase insert payload keys: {list(supabase_doc.keys())}")

    # Insert the clean doc into Supabase
    await db.clients.insert_one(supabase_doc)

    # ── Financial Setup Processing on Client Onboarding ──────────────────────
    c_val = float(data.contract_value or data.quotation_value or 0)
    proj_id = f"proj_{client_id}"
    proj_doc = {
        "id": proj_id,
        "company_id": user["company_id"],
        "client_id": client_id,
        "project_name": f"{data.system_kw or 0} kW Solar System" if data.system_kw else "Solar Project",
        "project_type": data.consumer_type or "Rooftop Solar",
        "capacity_kw": float(data.system_kw or 0),
        "project_value": c_val,
        "project_date": now_iso()[:10],
        "payment_plan": data.payment_plan or [],
        "status": "Pending",
        "created_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.projects.insert_one(proj_doc)

    # Process initial financial entries/payments if provided
    if data.initial_payments and isinstance(data.initial_payments, list):
        for pay in data.initial_payments:
            amt = float(pay.get("amount") or 0)
            if amt > 0 or pay.get("description"):
                pay_doc = {
                    "id": f"pay_{uuid.uuid4().hex[:12]}",
                    "company_id": user["company_id"],
                    "client_id": client_id,
                    "project_id": proj_id,
                    "milestone_name": pay.get("description") or "Initial Payment",
                    "payment_type": pay.get("description") or "Advance",
                    "amount": amt,
                    "payment_date": pay.get("payment_date") or now_iso()[:10],
                    "payment_source": pay.get("payment_source") or "Bank Transfer",
                    "payment_mode": pay.get("payment_source") or "Bank Transfer",
                    "ref_number": pay.get("ref_number") or "",
                    "status": pay.get("status") or "Received",
                    "notes": pay.get("remarks") or pay.get("notes") or "",
                    "created_at": now_iso(),
                    "updated_at": now_iso()
                }
                await db.payments.insert_one(pay_doc)

    # Process Loan Setup if enabled
    if data.loan_setup and isinstance(data.loan_setup, dict) and data.loan_setup.get("enabled"):
        loan_info = data.loan_setup
        app_amt = float(loan_info.get("approved_amount") or loan_info.get("loan_amount") or 0)
        if app_amt > 0 or loan_info.get("provider"):
            loan_doc = {
                "id": f"loan_{uuid.uuid4().hex[:12]}",
                "company_id": user["company_id"],
                "client_id": client_id,
                "project_id": proj_id,
                "provider": loan_info.get("provider") or "Tata Capital",
                "loan_amount": float(loan_info.get("loan_amount") or 0),
                "approved_amount": app_amt,
                "approved_date": loan_info.get("approved_date") or now_iso()[:10],
                "expected_disbursement_date": loan_info.get("expected_disbursement_date") or "",
                "disbursed_amount": float(loan_info.get("disbursed_amount") or 0),
                "loan_ref": loan_info.get("loan_ref") or "",
                "status": loan_info.get("status") or "Approved",
                "remarks": loan_info.get("remarks") or "",
                "created_by": user["name"],
                "created_at": now_iso(),
                "updated_at": now_iso()
            }
            await db.loans.insert_one(loan_doc)

    # Also persist the fully-enriched doc locally so it's immediately readable
    try:
        await LocalFileCollection("clients").insert_one(dict(full_doc))
    except Exception:
        pass

    # Verify the client actually landed in Supabase
    inserted = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
    if inserted:
        logger.info(f"[CLIENT-CREATE DIAG] ✓ Client {client_id} confirmed in Supabase")
    else:
        logger.error(f"[CLIENT-CREATE DIAG] ✗ Client {client_id} NOT found in Supabase after insert — returning doc from memory")

    full_doc.pop("_id", None)
    await log_activity(user["company_id"], user["id"], user["name"], "Added Client", data.full_name)
    await push_notification(user["company_id"], "admin", "New Client Added", f"{data.full_name} ({sol_id})")
    return _enrich_client_doc(inserted) if inserted else _enrich_client_doc(full_doc)

async def _get_client_high_value_assets(client_doc: dict, company_id: str) -> list:
    if not isinstance(client_doc, dict):
        return []
    cid = client_doc.get("id")
    sol_id = client_doc.get("sol_id")
    c_name = (client_doc.get("full_name") or "").strip().lower()
    c_addr = f"{client_doc.get('address') or ''}, {client_doc.get('city') or ''}".strip(", ")

    outwards = await db.outward_entries.find({"company_id": company_id}, {"_id": 0}).to_list(100000)
    
    # We need to know which ones are high value.
    items, _, _, _ = await _compute_inventory_balances(company_id)
    local_hv = _load_local_high_value_products()
    hv_keywords = ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"]
    
    hv_inward_product_names = set()
    inwards = await db.inward_entries.find({"company_id": company_id}, {"_id": 0}).to_list(100000)
    for ie in inwards:
        if ie.get("high_value_asset") or ie.get("high_value_goods"):
            pn = ie.get("product")
            if pn:
                hv_inward_product_names.add(norm_product_name(pn))

    hv_ids = set()
    hv_names = set()
    for p in items:
        is_hv = False
        if p.get("high_value_goods") or p.get("high_value_asset"):
            is_hv = True
        else:
            pn_n = norm_product_name(p.get("name"))
            if local_hv.get(pn_n, False) is True or pn_n in hv_inward_product_names:
                is_hv = True
        if is_hv:
            if p.get("id"): hv_ids.add(p["id"])
            pn_n = norm_product_name(p.get("name"))
            hv_names.add(pn_n)
            
    def is_hv_entry(oe):
        pid = oe.get("product_id")
        if pid and pid in hv_ids: return True
        if bool(oe.get("high_value_goods")) or bool(oe.get("high_value_asset")): return True
        pn_n = norm_product_name(oe.get("product"))
        if local_hv.get(pn_n, False) is True or pn_n in hv_inward_product_names: return True
        if any(kw in pn_n for kw in hv_keywords): return True
        if pn_n in hv_names: return True
        return False

    matched = []
    seen_outward_ids = set()
    
    for o in outwards:
        st = str(o.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]: continue
        
        o_cid = o.get("client_id")
        o_cname = (o.get("client_name") or "").strip().lower()
        is_match = False
        if o_cid and o_cid in (cid, sol_id):
            is_match = True
        elif c_name and o_cname and (o_cname == c_name or c_name in o_cname or o_cname in c_name):
            is_match = True
            
        if is_match and is_hv_entry(o) and o.get("id") not in seen_outward_ids:
            seen_outward_ids.add(o.get("id"))
            
            serials = o.get("serial_numbers") or o.get("serials") or ([o.get("serial_number")] if o.get("serial_number") else [])
            serials = [s.strip().upper() for s in serials if s and s.strip()]
            
            # Check returns
            for inv in inwards:
                if inv.get("reference_number") == o.get("bill_number") or inv.get("reference_number") == o.get("reference_number"):
                    # This is naive return matching, better to just let the return logic handle it or just show dispatched
                    ret_serials = inv.get("serial_numbers") or inv.get("serials") or ([inv.get("serial_number")] if inv.get("serial_number") else [])
                    ret_serials = [s.strip().upper() for s in ret_serials if s and s.strip()]
                    serials = [s for s in serials if s not in ret_serials]
            
            matched.append({
                "id": o.get("id"),
                "product_name": o.get("product") or o.get("product_name") or "Solar Product",
                "size_model": o.get("size") or o.get("size_model") or "—",
                "quantity": float(o.get("quantity") or 0.0),
                "serial_numbers": serials,
                "outward_date": (o.get("date") or o.get("created_at") or "")[:10],
                "status": o.get("status") or "Dispatched",
                "current_site": o.get("site_name") or f"{client_doc.get('full_name')} Site"
            })
            
    return matched

@api_router.get("/clients/{client_id}")
async def get_client(client_id: str, user=Depends(get_current_user)):
    logger.info(f"[CLIENT-GET DIAG] GET /clients/{client_id} | company_id={user.get('company_id')} user_id={user.get('id')}")
    c = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
    if c:
        logger.info(f"[CLIENT-GET DIAG] ✓ Found by id+company_id")
    if not c:
        c = await db.clients.find_one({"sol_id": client_id, "company_id": user["company_id"]}, {"_id": 0})
        if c:
            logger.info(f"[CLIENT-GET DIAG] ✓ Found by sol_id+company_id")
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    if not c:
        logger.error(f"[CLIENT-GET DIAG] ✗ Client {client_id} NOT FOUND in Supabase or local — returning 404")
        raise HTTPException(status_code=404, detail="Not found")
    c = _enrich_client_doc(c)
    c["high_value_assets"] = await _get_client_high_value_assets(c, user["company_id"])
    return c

def _normalize_client_payload(payload: dict) -> dict:
    if "panel_brand" in payload or "panel_make" in payload:
        brand_val = (payload["panel_brand"] if "panel_brand" in payload and payload["panel_brand"] is not None else payload.get("panel_make")) or ""
        payload["panel_brand"] = brand_val
        payload["panel_make"] = brand_val
    if "inverter_make" in payload or "inverter_brand" in payload:
        inv_val = (payload["inverter_brand"] if "inverter_brand" in payload and payload["inverter_brand"] is not None else payload.get("inverter_make")) or ""
        payload["inverter_make"] = inv_val
        payload["inverter_brand"] = inv_val
    if "section_number" in payload or "section_no" in payload:
        sec_val = payload.get("section_number") or payload.get("section_no") or ""
        payload["section_number"] = sec_val
        payload["section_no"] = sec_val
    if "consumer_type" in payload or "consumer_category" in payload or "category" in payload:
        cat_val = payload.get("consumer_type") or payload.get("consumer_category") or payload.get("category") or ""
        payload["consumer_type"] = cat_val
        payload["consumer_category"] = cat_val
        payload["category"] = cat_val
    if "inverter_serial" in payload or "inverter_sr" in payload:
        sr_val = payload.get("inverter_serial") or payload.get("inverter_sr") or ""
        payload["inverter_serial"] = sr_val
        payload["inverter_sr"] = sr_val
    if "inverter_year" in payload or "manufacturing_year" in payload:
        yr_val = payload.get("inverter_year") or payload.get("manufacturing_year") or ""
        payload["inverter_year"] = yr_val
        payload["manufacturing_year"] = yr_val
    if "aadhaar" in payload or "aadhaar_number" in payload:
        a_val = payload.get("aadhaar") or payload.get("aadhaar_number") or ""
        payload["aadhaar"] = a_val
        payload["aadhaar_number"] = a_val
    return payload

def _verify_client_db_write(original_payload: dict, read_back_doc: dict):
    if not read_back_doc:
        raise HTTPException(status_code=500, detail="Database save verification failed: DB returned empty record on read-back")
    
    enriched = _enrich_client_doc(dict(read_back_doc))
    mismatches = []
    
    fields_to_check = [
        "full_name", "mobile", "alt_mobile", "consumer_number", "section_number", "address",
        "city", "state", "pincode", "aadhaar", "system_kw", "panel_make", "panel_brand",
        "panel_technology", "panel_wattage", "num_panels", "inverter_make", "inverter_brand",
        "inverter_capacity", "inverter_serial", "inverter_model", "inverter_year",
        "sanction_number", "consumer_type", "consumer_category", "phase_type", "subsidy_eligible"
    ]
    
    for k in fields_to_check:
        if k in original_payload and original_payload[k] is not None:
            expected = original_payload[k]
            actual = enriched.get(k)
            if isinstance(expected, (int, float)) and actual is not None:
                try:
                    if float(expected) != float(actual):
                        mismatches.append(f"{k} (expected {expected}, got {actual})")
                except (ValueError, TypeError):
                    if str(expected) != str(actual):
                        mismatches.append(f"{k} (expected {expected}, got {actual})")
            elif expected != "" and (actual is None or actual == ""):
                mismatches.append(f"{k} (expected '{expected}', got '{actual}')")
            elif str(expected).strip() != "" and str(expected).strip().lower() != str(actual or "").strip().lower():
                mismatches.append(f"{k} (expected '{expected}', got '{actual}')")

    if "inverters" in original_payload and isinstance(original_payload["inverters"], list):
        exp_inv = original_payload["inverters"]
        act_inv = enriched.get("inverters") or []
        if len(exp_inv) > 0 and len(act_inv) == 0:
            mismatches.append(f"inverters (expected {len(exp_inv)} inverters, got 0)")

    if mismatches:
        err_msg = f"Database save verification failed: {', '.join(mismatches)}"
        logger.error(err_msg)
        raise HTTPException(status_code=500, detail=err_msg)

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, data: ClientIn, user=Depends(get_current_user)):
    if not has_perm(user, "clients", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.edit")
    raw_payload = data.model_dump()
    update = _normalize_client_payload(dict(raw_payload))
    existing = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
    if existing:
        if not update.get("stages"):
            update["stages"] = existing.get("stages") or {s: False for s in DEFAULT_STAGES}
    else:
        if not update.get("stages"):
            update.pop("stages", None)
    if update.get("stages"):
        if data.status in ["Approved", "Installation Pending", "Installation Complete", "Handover Complete"]:
            update["stages"]["Onboarding"] = True
        update["stages"] = sync_checklist_completed(update["stages"])
        update["progress"] = calc_progress(update["stages"])
    update["updated_at"] = now_iso()
    res = await db.clients.update_one({"id": client_id, "company_id": user["company_id"]}, {"$set": update})
    if res.matched_count == 0:
        res = await db.clients.update_one({"sol_id": client_id, "company_id": user["company_id"]}, {"$set": update})
        if res.matched_count == 0:
            res = await db.clients.update_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"$set": update})
            if res.matched_count == 0:
                raise HTTPException(status_code=404, detail="Not found")
    await LocalFileCollection("clients").update_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"$set": update})
    await log_activity(user["company_id"], user["id"], user["name"], "Updated Client", data.full_name)
    
    # Direct database read-back verification
    client_doc = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}], "company_id": user["company_id"]}, {"_id": 0})
    if not client_doc:
        client_doc = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"_id": 0})
    
    if client_doc:
        _verify_client_db_write(raw_payload, client_doc)
    return _enrich_client_doc(client_doc) if client_doc else {}

@api_router.patch("/clients/{client_id}")
async def patch_client(client_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    if not has_perm(user, "clients", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.edit")
    raw_payload = dict(payload)

    # ── DIAG STEP 1: Log exactly what the frontend sent ──────────────────────
    safe_payload = {k: v for k, v in raw_payload.items() if k not in ("aadhaar", "pan_number", "mobile", "alt_mobile")}
    logger.info(f"[CLIENT-SAVE DIAG] ══════════════════════════════════════════")
    logger.info(f"[CLIENT-SAVE DIAG] PATCH /clients/{client_id} called")
    logger.info(f"[CLIENT-SAVE DIAG] STEP 1 ▶ Incoming payload keys: {list(safe_payload.keys())}")
    logger.info(f"[CLIENT-SAVE DIAG] STEP 1 ▶ panel_wattage   = {raw_payload.get('panel_wattage')}")
    logger.info(f"[CLIENT-SAVE DIAG] STEP 1 ▶ panel_make      = {raw_payload.get('panel_make')}")
    logger.info(f"[CLIENT-SAVE DIAG] STEP 1 ▶ consumer_cat    = {raw_payload.get('consumer_category') or raw_payload.get('consumer_type')}")
    logger.info(f"[CLIENT-SAVE DIAG] STEP 1 ▶ section_number  = {raw_payload.get('section_number') or raw_payload.get('section_no')}")
    logger.info(f"[CLIENT-SAVE DIAG] STEP 1 ▶ inverters count = {len(raw_payload.get('inverters') or [])}")

    # ── DIAG STEP 2: Log user and company context ─────────────────────────────
    logger.info(f"[CLIENT-SAVE DIAG] STEP 2 ▶ client_id={client_id} company_id={user.get('company_id')} user_id={user.get('id')}")

    payload.pop("_id", None)
    payload = _normalize_client_payload(payload)
    payload["updated_at"] = now_iso()

    # ── DIAG STEP 3: WHERE clause being used ─────────────────────────────────
    logger.info(f"[CLIENT-SAVE DIAG] STEP 3 ▶ WHERE clause #1: id={client_id} AND company_id={user.get('company_id')}")

    res = await db.clients.update_one({"id": client_id, "company_id": user["company_id"]}, {"$set": payload})
    logger.info(f"[CLIENT-SAVE DIAG] STEP 3 ▶ matched_count={res.matched_count}")
    if res.matched_count == 0:
        logger.warning(f"[CLIENT-SAVE DIAG] STEP 3 ▶ WHERE #1 missed — trying sol_id")
        res = await db.clients.update_one({"sol_id": client_id, "company_id": user["company_id"]}, {"$set": payload})
        logger.info(f"[CLIENT-SAVE DIAG] STEP 3 ▶ sol_id match: matched_count={res.matched_count}")
        if res.matched_count == 0:
            logger.warning(f"[CLIENT-SAVE DIAG] STEP 3 ▶ WHERE #2 missed — trying $or without company_id")
            res = await db.clients.update_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"$set": payload})
            logger.info(f"[CLIENT-SAVE DIAG] STEP 3 ▶ $or match: matched_count={res.matched_count}")
            if res.matched_count == 0:
                logger.error(f"[CLIENT-SAVE DIAG] ✗ ALL WHERE clauses missed — client_id={client_id} not found!")
                raise HTTPException(status_code=404, detail="Client not found")

    await LocalFileCollection("clients").update_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"$set": payload})

    # ── DIAG STEP 4: Fresh SELECT read-back ──────────────────────────────────
    logger.info(f"[CLIENT-SAVE DIAG] STEP 4 ▶ Running fresh SELECT to verify persistence...")
    client_doc = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}], "company_id": user["company_id"]}, {"_id": 0})
    if not client_doc:
        client_doc = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"_id": 0})

    if client_doc:
        db_ob = (client_doc.get("stages") or {}).get("onboarding_data") or {}
        logger.info(f"[CLIENT-SAVE DIAG] STEP 4 ▶ SELECT after UPDATE:")
        logger.info(f"[CLIENT-SAVE DIAG]   Submitted panel_wattage  = {raw_payload.get('panel_wattage')} | DB = {client_doc.get('panel_wattage')}")
        logger.info(f"[CLIENT-SAVE DIAG]   Submitted panel_make     = {raw_payload.get('panel_make')} | DB = {client_doc.get('panel_make')}")
        logger.info(f"[CLIENT-SAVE DIAG]   Submitted consumer_cat   = {raw_payload.get('consumer_category') or raw_payload.get('consumer_type')} | DB = {db_ob.get('consumer_category')}")
        logger.info(f"[CLIENT-SAVE DIAG]   Submitted section_number = {raw_payload.get('section_number') or raw_payload.get('section_no')} | DB = {db_ob.get('section_number')}")
        logger.info(f"[CLIENT-SAVE DIAG]   Submitted inverters      = {len(raw_payload.get('inverters') or [])} items | DB = {len(db_ob.get('inverters') or [])} items")
        logger.info(f"[CLIENT-SAVE DIAG]   DB updated_at            = {client_doc.get('updated_at')}")
    else:
        logger.error(f"[CLIENT-SAVE DIAG] ✗ STEP 4 SELECT returned NOTHING — client disappeared after update!")

    if client_doc:
        _verify_client_db_write(raw_payload, client_doc)
    logger.info(f"[CLIENT-SAVE DIAG] ✓ _verify_client_db_write passed — returning enriched doc")
    logger.info(f"[CLIENT-SAVE DIAG] ══════════════════════════════════════════")
    return _enrich_client_doc(client_doc) if client_doc else {}

@api_router.patch("/clients/{client_id}/stages")
async def update_stages(client_id: str, data: StageUpdate, user=Depends(get_current_user)):
    if not (has_perm(user, "clients", "edit") or has_perm(user, "client_data", "edit") or has_perm(user, "project_execution", "edit")):
        raise HTTPException(status_code=403, detail="Missing permission: clients.edit")
    existing = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Client not found")
    merged_stages = {**(existing.get("stages") or {s: False for s in DEFAULT_STAGES}), **data.stages}
    if "Onboarding" not in data.stages and not merged_stages.get("Onboarding"):
        merged_stages["Onboarding"] = True
    stages = sync_checklist_completed(merged_stages)
    progress = calc_progress(stages)
    await db.clients.update_one(
        {"id": client_id},
        {"$set": {"stages": stages, "progress": progress, "updated_at": now_iso()}}
    )
    if stages.get("Handover") and existing.get("status") != "Handover Complete":
        await db.clients.update_one({"id": client_id}, {"$set": {"status": "Handover Complete"}})
        await push_notification(user["company_id"], "admin", "Installation Completed", existing.get("full_name", ""))
    await log_activity(user["company_id"], user["id"], user["name"], "Updated Progress", existing.get("full_name", ""))
    return await db.clients.find_one({"id": client_id}, {"_id": 0})

@api_router.patch("/clients/{client_id}/status")
async def update_status(client_id: str, data: StatusUpdate, user=Depends(get_current_user)):
    if not has_perm(user, "clients", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.edit")
    c = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    
    update_set: dict[str, Any] = {"status": data.status, "updated_at": now_iso()}
    if data.status in ["Approved", "Installation Pending", "Installation Complete", "Handover Complete"]:
        stages = c.get("stages") or {s: False for s in DEFAULT_STAGES}
        stages["Onboarding"] = True
        update_set["stages"] = stages
        update_set["progress"] = calc_progress(stages)
        
    await db.clients.update_one(
        {"id": client_id, "company_id": user["company_id"]},
        {"$set": update_set}
    )
    return await db.clients.find_one({"id": client_id}, {"_id": 0})

@api_router.post("/clients/{client_id}/notes")
async def add_note(client_id: str, data: NoteIn, user=Depends(get_current_user)):
    if not has_perm(user, "clients", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.edit")
    note = {"id": str(uuid.uuid4()), "text": data.text, "user_id": user["id"], "user_name": user["name"], "created_at": now_iso()}
    res = await db.clients.update_one(
        {"id": client_id, "company_id": user["company_id"]},
        {"$push": {"notes": note}, "$set": {"updated_at": now_iso()}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return note

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, user=Depends(get_current_user)):
    import re
    if not has_perm(user, "clients", "delete"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.delete")

    company_id = user["company_id"]
    c = await db.clients.find_one({"id": client_id, "company_id": company_id})
    if not c:
        raise HTTPException(status_code=404, detail="Not found")

    # 1. Retrieve all related client records for backup/rollback purposes
    # Collect File IDs linked to this client
    file_ids = set()
    if c.get("documents"):
        for doc in c["documents"]:
            if isinstance(doc, dict) and doc.get("id"):
                file_ids.add(doc["id"])

    # Fetch projects
    projects_records = await db.projects.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch tasks
    tasks_records = await db.tasks.find({"client_id": client_id, "company_id": company_id}).to_list(1000)
    task_ids = [t["id"] for t in tasks_records]

    # Collect files linked to tasks
    for t in tasks_records:
        submission = t.get("submission") or {}
        photos = submission.get("photos") or {}
        for p_id in photos.values():
            if isinstance(p_id, str) and p_id:
                file_ids.add(p_id)
            elif isinstance(p_id, dict) and p_id.get("file_id"):
                file_ids.add(p_id["file_id"])

    # Fetch task updates
    task_updates_records = []
    if task_ids:
        task_updates_records = await db.task_updates.find({"task_id": {"$in": task_ids}}).to_list(5000)

    # Fetch material requests
    material_requests_records = await db.material_requests.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch material deliveries
    material_deliveries_records = await db.material_deliveries.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch inward / outward entries
    inward_entries_records = await db.inward_entries.find({"client_id": client_id, "company_id": company_id}).to_list(1000)
    outward_entries_records = await db.outward_entries.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch documents
    documents_records = await db.documents.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch installations
    installations_records = await db.installations.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch meter testings
    meter_testings_records = await db.meter_testings.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch verifications
    verifications_records = await db.verifications.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Collect files linked to verifications
    for v in verifications_records:
        photos = v.get("photos") or {}
        for p_id in photos.values():
            if p_id:
                file_ids.add(p_id)

    # Fetch inverter monitoring
    inverter_monitoring_records = await db.inverter_monitoring.find({"client_id": client_id, "company_id": company_id}).to_list(1000)

    # Fetch complaints (service tickets) and their child records
    complaints_records = await db.complaints.find({"client_id": client_id, "company_id": company_id}).to_list(1000)
    complaint_ids = [comp["id"] for comp in complaints_records]

    complaint_comments_records = []
    complaint_audit_records = []
    if complaint_ids:
        complaint_comments_records = await db.complaint_comments.find({"complaint_id": {"$in": complaint_ids}}).to_list(5000)
        complaint_audit_records = await db.complaint_audit.find({"complaint_id": {"$in": complaint_ids}}).to_list(5000)

    # Fetch activity logs
    log_filters = []
    if c.get("full_name"):
        log_filters.append({"target": {"$regex": re.escape(c["full_name"]), "$options": "i"}})
    if c.get("sol_id"):
        log_filters.append({"target": {"$regex": re.escape(c["sol_id"]), "$options": "i"}})
    log_filters.append({"target": {"$regex": re.escape(client_id), "$options": "i"}})

    activity_logs_records = []
    if log_filters:
        activity_logs_records = await db.activity_logs.find({
            "company_id": company_id,
            "$or": log_filters
        }).to_list(5000)

    # Fetch files
    files_records = []
    if file_ids:
        files_records = await db.files.find({"id": {"$in": list(file_ids)}, "company_id": company_id}).to_list(1000)

    # Load high value assets
    original_all_assets = list(_load_local_assets())

    # Deep/dictionary backup copy creation
    backup_client = dict(c)
    backup_projects = [dict(r) for r in projects_records]
    backup_tasks = [dict(r) for r in tasks_records]
    backup_task_updates = [dict(r) for r in task_updates_records]
    backup_material_requests = [dict(r) for r in material_requests_records]
    backup_material_deliveries = [dict(r) for r in material_deliveries_records]
    backup_inward_entries = [dict(r) for r in inward_entries_records]
    backup_outward_entries = [dict(r) for r in outward_entries_records]
    backup_documents = [dict(r) for r in documents_records]
    backup_installations = [dict(r) for r in installations_records]
    backup_meter_testings = [dict(r) for r in meter_testings_records]
    backup_verifications = [dict(r) for r in verifications_records]
    backup_inverter_monitoring = [dict(r) for r in inverter_monitoring_records]
    backup_complaints = [dict(r) for r in complaints_records]
    backup_complaint_comments = [dict(r) for r in complaint_comments_records]
    backup_complaint_audit = [dict(r) for r in complaint_audit_records]
    backup_activity_logs = [dict(r) for r in activity_logs_records]
    backup_files = [dict(r) for r in files_records]

    # 2. Transaction Execution Block
    try:
        # Delete High Value Assets locally
        filtered_assets = [a for a in original_all_assets if not (a.get("client_id") == client_id and a.get("company_id") == company_id)]
        _save_local_assets(filtered_assets)

        # Database Deletions
        await db.clients.delete_one({"id": client_id, "company_id": company_id})
        await db.projects.delete_many({"client_id": client_id, "company_id": company_id})
        await db.tasks.delete_many({"client_id": client_id, "company_id": company_id})
        if task_ids:
            await db.task_updates.delete_many({"task_id": {"$in": task_ids}})
        await db.material_requests.delete_many({"client_id": client_id, "company_id": company_id})
        await db.material_deliveries.delete_many({"client_id": client_id, "company_id": company_id})
        await db.inward_entries.delete_many({"client_id": client_id, "company_id": company_id})
        await db.outward_entries.delete_many({"client_id": client_id, "company_id": company_id})
        await db.documents.delete_many({"client_id": client_id, "company_id": company_id})
        await db.installations.delete_many({"client_id": client_id, "company_id": company_id})
        await db.meter_testings.delete_many({"client_id": client_id, "company_id": company_id})
        await db.verifications.delete_many({"client_id": client_id, "company_id": company_id})
        await db.inverter_monitoring.delete_many({"client_id": client_id, "company_id": company_id})

        if complaint_ids:
            await db.complaint_comments.delete_many({"complaint_id": {"$in": complaint_ids}})
            await db.complaint_audit.delete_many({"complaint_id": {"$in": complaint_ids}})
        await db.complaints.delete_many({"client_id": client_id, "company_id": company_id})

        if log_filters and backup_activity_logs:
            await db.activity_logs.delete_many({
                "company_id": company_id,
                "$or": log_filters
            })

        if file_ids:
            await db.files.delete_many({"id": {"$in": list(file_ids)}, "company_id": company_id})

    except Exception as exc:
        logger.error(f"Error during client deletion, initiating transaction rollback: {exc}")
        # Rollback local assets
        try:
            _save_local_assets(original_all_assets)
        except Exception:
            pass

        # Rollback database tables
        try:
            if backup_client:
                await db.clients.insert_one(backup_client)
            if backup_projects:
                await db.projects.insert_many(backup_projects)
            if backup_tasks:
                await db.tasks.insert_many(backup_tasks)
            if backup_task_updates:
                await db.task_updates.insert_many(backup_task_updates)
            if backup_material_requests:
                await db.material_requests.insert_many(backup_material_requests)
            if backup_material_deliveries:
                await db.material_deliveries.insert_many(backup_material_deliveries)
            if backup_inward_entries:
                await db.inward_entries.insert_many(backup_inward_entries)
            if backup_outward_entries:
                await db.outward_entries.insert_many(backup_outward_entries)
            if backup_documents:
                await db.documents.insert_many(backup_documents)
            if backup_installations:
                await db.installations.insert_many(backup_installations)
            if backup_meter_testings:
                await db.meter_testings.insert_many(backup_meter_testings)
            if backup_verifications:
                await db.verifications.insert_many(backup_verifications)
            if backup_inverter_monitoring:
                await db.inverter_monitoring.insert_many(backup_inverter_monitoring)
            if backup_complaints:
                await db.complaints.insert_many(backup_complaints)
            if backup_complaint_comments:
                await db.complaint_comments.insert_many(backup_complaint_comments)
            if backup_complaint_audit:
                await db.complaint_audit.insert_many(backup_complaint_audit)
            if backup_activity_logs:
                await db.activity_logs.insert_many(backup_activity_logs)
            if backup_files:
                await db.files.insert_many(backup_files)
        except Exception as rollback_err:
            logger.critical(f"FATAL: Database rollback failed: {rollback_err}")

        logger.exception("Client deletion failed during database transaction.")
        raise HTTPException(status_code=500, detail="Client deletion failed. Transaction rolled back safely.")

    # 3. Permanent deletion of files from Supabase storage (after DB transaction success)
    if backup_files:
        for file_rec in backup_files:
            storage_path = file_rec.get("storage_path")
            if storage_path:
                try:
                    delete_object(storage_path)
                except Exception as se:
                    logger.warning(f"Failed to permanently delete storage object {storage_path}: {se}")

    # 4. Log client deletion audit trail
    await log_activity(company_id, user["id"], user["name"], "Deleted Client", c.get("full_name", ""))
    return {"ok": True}

ALLOWED_DOC_TYPES = (
    "annexure", "wcr", "sldr", "net_meter_agreement", "vendor_agreement",
    "meter_testing_request", "quotation", "sales_order", "tax_invoice",
    "delivery_bill", "purchase_order", "purchase_bill",
)

def _document_label(doc_type: str) -> str:
    return {
        "annexure": "Annexure",
        "wcr": "WCR",
        "sldr": "SLDR",
        "net_meter_agreement": "Net Meter Agreement",
        "vendor_agreement": "Vendor Agreement",
        "meter_testing_request": "Meter Testing Request",
        "quotation": "Quotation",
        "sales_order": "Sales Order",
        "tax_invoice": "Tax Invoice",
        "delivery_bill": "Delivery Bill",
        "purchase_order": "Purchase Order",
        "purchase_bill": "Purchase Bill",
    }.get(doc_type, doc_type.replace("_", " ").title())

def delete_object(path: str):
    """Permanently deletes a file from Supabase storage."""
    bucket, file_path = _map_path_to_bucket_and_name(path)
    try:
        supabase.storage.from_(bucket).remove([file_path])
    except Exception as e:
        logger.error(f"Error deleting from bucket {bucket} at {file_path}: {e}")
        raise e

def _generate_meaningful_filename(doc_type: str, doc_data: dict, client_doc: Optional[dict] = None) -> str:
    """Generates a filename with format <PartyName>_<DocumentType>_<DocumentNumber>_<YYYY-MM-DD>.pdf"""
    party_name = "Party"
    if client_doc and client_doc.get("full_name"):
        party_name = client_doc["full_name"]
    elif doc_data.get("client") and doc_data["client"].get("full_name"):
        party_name = doc_data["client"]["full_name"]
    elif doc_data.get("vendor") and doc_data["vendor"].get("name"):
        party_name = doc_data["vendor"]["name"]
        
    doc_type_map = {
        "quotation": "Quotation",
        "tax_invoice": "Invoice",
        "delivery_bill": "DeliveryBill",
        "purchase_order": "PurchaseOrder",
        "purchase_bill": "PurchaseBill"
    }
    doc_type_name = doc_type_map.get(doc_type, _document_label(doc_type))
    
    doc_number = "Doc"
    if doc_type == "quotation":
        doc_number = doc_data.get("quote_number") or doc_data.get("document_number") or "Q"
    elif doc_type == "tax_invoice":
        doc_number = doc_data.get("invoice_number") or doc_data.get("document_number") or "INV"
    elif doc_type == "delivery_bill":
        doc_number = doc_data.get("challan_number") or doc_data.get("document_number") or "DC"
    elif doc_type in ("purchase_order", "purchase_bill"):
        doc_number = doc_data.get("po_number") or doc_data.get("bill_number") or doc_data.get("document_number") or "PO"
        
    date_val = None
    if doc_type == "quotation":
        date_val = doc_data.get("quote_date")
    elif doc_type == "tax_invoice":
        date_val = doc_data.get("invoice_date")
    elif doc_type == "delivery_bill":
        date_val = doc_data.get("date")
    elif doc_type in ("purchase_order", "purchase_bill"):
        date_val = doc_data.get("document_date") or doc_data.get("po_date") or doc_data.get("date")
        
    if not date_val:
        from datetime import datetime
        date_val = datetime.now().strftime("%Y-%m-%d")
        
    raw_name = f"{party_name}_{doc_type_name}_{doc_number}_{date_val}.pdf"
    return raw_name.replace(" ", "_")

async def _cleanup_duplicate_document(company_id: str, doc_type: str, doc_number: str):
    """Deletes existing files of the same type and number to avoid duplicate/unused storage objects."""
    if not doc_number:
        return
    existing_file = await db.files.find_one({
        "company_id": company_id,
        "category": "generated",
        "doc_type": doc_type,
        "document_number": doc_number
    })
    if existing_file:
        try:
            delete_object(existing_file["storage_path"])
        except Exception as e:
            logger.error(f"Error deleting duplicate storage object: {e}")
        await db.files.delete_one({"id": existing_file["id"]})
        await db.clients.update_many(
            {"company_id": company_id},
            {"$pull": {"documents": {"id": existing_file["id"]}}}
        )

ONBOARDING_DOC_TYPES = {
    "annexure", "wcr", "sldr", "net_meter_agreement", "net_metering_agreement",
    "vendor_agreement", "meter_testing_request", "meter_testing", "onboarding"
}

@api_router.post("/clients/{client_id}/generate-document")
async def generate_document(client_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    doc_type = payload.get("doc_type", "").lower().strip()
    if doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Invalid doc_type")
        
    if doc_type in ("quotation", "tax_invoice", "delivery_bill", "purchase_order"):
        raise HTTPException(
            status_code=400,
            detail="Client Workspace document generation supports onboarding documents only. For sales, billing, or delivery documents, please use the Sales Documents or Receivables module."
        )

    if not has_perm(user, "clients", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: clients.create")

    client_doc = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=404, detail="Client not found")
async def _enrich_company_doc(company_doc: dict) -> dict:
    if not company_doc:
        return {}
    c = dict(company_doc)
    company_name = (c.get("company_name") or c.get("name") or c.get("legal_business_name") or c.get("vendor_name") or "").strip()
    owner_name = (c.get("owner_name") or c.get("proprietor_name") or c.get("authorized_signatory") or c.get("manager_name") or "").strip()
    gst = (c.get("gst_number") or c.get("gstin") or c.get("gst") or "").strip()
    pan = (c.get("pan_number") or c.get("pan") or "").strip()
    address = (c.get("address") or c.get("address_line_1") or c.get("office_address") or c.get("registered_address") or "").strip()
    mobile = (c.get("mobile") or c.get("mobile_number") or c.get("phone") or c.get("phone_number") or "").strip()
    email = (c.get("email") or "").strip()
    website = (c.get("website") or "").strip()
    city = (c.get("city") or "").strip()
    state = (c.get("state") or "").strip()
    pincode = (c.get("pincode") or "").strip()
    cin = (c.get("cin") or "").strip()
    license_number = (c.get("license_number") or "").strip()

    c["company_name"] = company_name
    c["name"] = company_name
    c["legal_business_name"] = company_name
    c["vendor_name"] = company_name
    c["owner_name"] = owner_name
    c["proprietor_name"] = owner_name
    c["authorized_signatory"] = owner_name
    c["manager_name"] = owner_name
    c["gst_number"] = gst
    c["gstin"] = gst
    c["gst"] = gst
    c["pan_number"] = pan
    c["pan"] = pan
    c["address"] = address
    c["address_line_1"] = address
    c["office_address"] = address
    c["registered_address"] = address
    c["mobile"] = mobile
    c["mobile_number"] = mobile
    c["phone"] = mobile
    c["phone_number"] = mobile
    c["email"] = email
    c["website"] = website
    c["city"] = city
    c["state"] = state
    c["pincode"] = pincode
    c["cin"] = cin
    c["license_number"] = license_number

    logo_file_id = c.get("logo_file_id")
    if logo_file_id:
        if logo_file_id in _company_logo_cache:
            c["logo_bytes"] = _company_logo_cache[logo_file_id]
        else:
            file_rec = await db.files.find_one({"id": logo_file_id, "is_deleted": False})
            if file_rec:
                try:
                    logo_bytes, _ = get_object(file_rec["storage_path"])
                    c["logo_bytes"] = logo_bytes
                    _company_logo_cache[logo_file_id] = logo_bytes
                except Exception as e:
                    logger.error(f"Error fetching company logo: {e}")
    if not c.get("logo_bytes") and c.get("logo_url"):
        try:
            import urllib.request
            req = urllib.request.Request(c["logo_url"], headers={"User-Agent": "Solrix/2.0"})
            with urllib.request.urlopen(req, timeout=5) as resp:
                c["logo_bytes"] = resp.read()
        except Exception as e:
            logger.warning(f"Failed to fetch company logo from logo_url: {e}")
    return c

@api_router.post("/documents/preview")
async def generate_document_preview(payload: Dict[str, Any], user=Depends(get_current_user)):
    doc_type = payload.get("doc_type", "")
    if doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Invalid doc_type")
        
    client_id = payload.get("client_id")
    if not client_id and doc_type not in ("purchase_order", "purchase_bill"):
        raise HTTPException(status_code=400, detail="client_id is required")
        
    if doc_type in ("quotation", "tax_invoice", "delivery_bill", "purchase_order", "purchase_bill"):
        if not has_perm(user, "sales_documents", "create"):
            raise HTTPException(status_code=403, detail="Missing permission: sales_documents.create")
    else:
        if not has_perm(user, "clients", "create"):
            raise HTTPException(status_code=403, detail="Missing permission: clients.create")
            
    client_doc = None
    if client_id:
        client_doc = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
        if not client_doc and doc_type not in ("purchase_order", "purchase_bill"):
            raise HTTPException(status_code=404, detail="Client not found")

    company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
    company_doc = await _enrich_company_doc(company_doc)

    if client_doc:
        client_doc = _enrich_client_doc(client_doc)

    doc_data = payload.get("doc_data") or {}
    if doc_type in ("quotation", "tax_invoice", "delivery_bill", "purchase_order", "purchase_bill"):
        if client_doc and not doc_data.get("client"):
            doc_data["client"] = client_doc
        pdf_bytes = await asyncio.to_thread(pdf_generator.generate_document, doc_type, doc_data, company_doc)
        gen_content_type = "application/pdf"
    elif doc_type == "annexure":
        import annexure_generator as _annex_gen
        pdf_bytes, gen_content_type = await asyncio.to_thread(_annex_gen.generate_annexure, client_doc, company_doc or {})
    else:
        pdf_bytes = await asyncio.to_thread(pdf_generator.generate, doc_type, client_doc, company_doc or {})
        gen_content_type = "application/pdf"

    # Extract document number and clean up duplicates
    doc_number = None
    if doc_type == "quotation":
        doc_number = doc_data.get("quote_number")
    elif doc_type == "tax_invoice":
        doc_number = doc_data.get("invoice_number")
    elif doc_type == "delivery_bill":
        doc_number = doc_data.get("challan_number")
        
    if doc_number:
        await _cleanup_duplicate_document(user["company_id"], doc_type, doc_number)

    # Determine file extension from content type
    _is_pdf = gen_content_type == "application/pdf"
    _ext = ".pdf" if _is_pdf else ".docx"
    filename = _generate_meaningful_filename(doc_type, doc_data, client_doc)
    if not _is_pdf and filename.endswith(".pdf"):
        filename = filename[:-4] + _ext
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/{user['company_id']}/generated/{file_id}{_ext}"
    result = put_object(storage_path, pdf_bytes, gen_content_type)
    
    client_name_val = (client_doc.get("full_name") if client_doc else "") or doc_data.get("client_name") or "Client"

    await db.files.insert_one({
        "id": file_id, "company_id": user["company_id"], "uploader_id": user["id"],
        "storage_path": result["path"], "original_filename": filename,
        "content_type": gen_content_type, "size": result.get("size", len(pdf_bytes)),
        "category": "generated", "is_deleted": False, "created_at": now_iso(),
        "doc_type": doc_type,
        "document_number": doc_number,
        "client_name": client_name_val,
        "prepared_by": doc_data.get("preparedBy") or user["name"],
        "status": "Active"
    })
    
    if client_doc:
        docs = list(client_doc.get("documents") or [])
        docs.append({"id": file_id, "filename": filename, "label": _document_label(doc_type), "content_type": gen_content_type, "created_at": now_iso()})
        stages = {**(client_doc.get("stages") or {}), "Document Making": True, "Onboarding": True}
        await db.clients.update_one(
            {"id": client_id, "company_id": user["company_id"]},
            {"$set": {"documents": docs, "stages": stages, "progress": calc_progress(stages), "updated_at": now_iso()}}
        )
    log_client_name = (client_doc.get("full_name") if client_doc else "") or "Manual"
    await log_activity(user["company_id"], user["id"], user["name"], f"Generated {_document_label(doc_type).upper()}", log_client_name)
    return {"id": file_id, "filename": filename, "label": _document_label(doc_type)}

@api_router.post("/documents/generate")
async def generate_public_document(payload: Dict[str, Any], user=Depends(get_current_user)):
    doc_type = payload.get("doc_type", "")
    if doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Invalid doc_type")
        
    if doc_type in ("quotation", "tax_invoice", "delivery_bill", "purchase_order", "purchase_bill"):
        if not has_perm(user, "sales_documents", "create"):
            raise HTTPException(status_code=403, detail="Missing permission: sales_documents.create")
    else:
        if not has_perm(user, "clients", "create"):
            raise HTTPException(status_code=403, detail="Missing permission: clients.create")
    company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
    company_doc = await _enrich_company_doc(company_doc)

    client_id = payload.get("client_id")
    client_doc = None
    if client_id:
        client_doc = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]}, {"_id": 0})
        if not client_doc:
            raise HTTPException(status_code=404, detail="Client not found")

    doc_data = payload.get("doc_data") or {}
    fmt_type = (payload.get("format") or "pdf").lower().strip()
    if doc_type in ("quotation", "tax_invoice", "delivery_bill", "purchase_order", "purchase_bill"):
        if client_doc and not doc_data.get("client"):
            doc_data["client"] = client_doc
        if fmt_type == "docx":
            pdf_bytes = await asyncio.to_thread(pdf_generator.generate_docx, doc_type, doc_data, company_doc)
            gen_content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            pdf_bytes = await asyncio.to_thread(pdf_generator.generate_document, doc_type, doc_data, company_doc)
            gen_content_type = "application/pdf"
    elif doc_type == "annexure":
        import annexure_generator as _annex_gen
        if not client_doc:
            raise HTTPException(status_code=400, detail="client_id is required for this document type")
        client_doc = _enrich_client_doc(client_doc)
        pdf_bytes, gen_content_type = await asyncio.to_thread(_annex_gen.generate_annexure, client_doc, company_doc or {})
    else:
        if not client_doc:
            raise HTTPException(status_code=400, detail="client_id is required for this document type")
        pdf_bytes = await asyncio.to_thread(pdf_generator.generate, doc_type, client_doc, company_doc or {})
        gen_content_type = "application/pdf"

    # Extract document number and clean up duplicates
    doc_number = None
    if doc_type == "quotation":
        doc_number = doc_data.get("quote_number")
    elif doc_type == "tax_invoice":
        doc_number = doc_data.get("invoice_number")
    elif doc_type == "delivery_bill":
        doc_number = doc_data.get("challan_number")
    elif doc_type in ("purchase_order", "purchase_bill"):
        doc_number = doc_data.get("po_number") or doc_data.get("bill_number") or doc_data.get("document_number")
        
    if doc_number:
        await _cleanup_duplicate_document(user["company_id"], doc_type, doc_number)

    _is_pdf = gen_content_type == "application/pdf"
    _ext = ".pdf" if _is_pdf else ".docx"
    filename = _generate_meaningful_filename(doc_type, doc_data, client_doc)
    if not _is_pdf and filename.endswith(".pdf"):
        filename = filename[:-4] + _ext
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/{user['company_id']}/generated/{file_id}{_ext}"
    result = put_object(storage_path, pdf_bytes, gen_content_type)
    
    client_name = "Client"
    if client_doc:
        client_name = client_doc.get("full_name") or "Client"
    elif doc_data.get("client"):
        client_name = doc_data["client"].get("full_name") or "Client"

    await db.files.insert_one({
        "id": file_id, "company_id": user["company_id"], "uploader_id": user["id"],
        "storage_path": result["path"], "original_filename": filename,
        "content_type": gen_content_type, "size": result.get("size", len(pdf_bytes)),
        "category": "generated", "is_deleted": False, "created_at": now_iso(),
        "doc_type": doc_type,
        "document_number": doc_number,
        "client_name": client_name,
        "prepared_by": doc_data.get("preparedBy") or user["name"],
        "status": "Active"
    })
    
    if client_doc:
        docs = list(client_doc.get("documents") or [])
        docs.append({"id": file_id, "filename": filename, "label": _document_label(doc_type), "content_type": gen_content_type, "created_at": now_iso()})
        stages = {**(client_doc.get("stages") or {}), "Document Making": True, "Onboarding": True}
        await db.clients.update_one(
            {"id": client_id, "company_id": user["company_id"]},
            {"$set": {"documents": docs, "stages": stages, "progress": calc_progress(stages), "updated_at": now_iso()}}
        )
    await log_activity(user["company_id"], user["id"], user["name"], f"Generated {_document_label(doc_type).upper()}", client_doc.get("full_name", "Manual") if client_doc else "Manual")
    
    return {"id": file_id, "filename": filename, "label": _document_label(doc_type)}

@api_router.post("/documents/download-direct")
async def download_direct_document(payload: Dict[str, Any], user=Depends(get_current_user)):
    client_id = payload.get("client_id")
    doc_type = (payload.get("doc_type") or "wcr").lower().strip()
    doc_format = (payload.get("format") or "pdf").lower().strip()  # "pdf" or "docx"
    
    if not client_id:
        raise HTTPException(status_code=400, detail="client_id is required")

    company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
    company_doc = await _enrich_company_doc(company_doc)
    cid = user["company_id"]
    or_conds: List[Dict[str, Any]] = [{"id": client_id}, {"sol_id": client_id}]
    client_doc = None
    
    try:
        client_doc = await db.clients.find_one({"$or": or_conds, "company_id": cid}, {"_id": 0})
    except Exception as err:
        logger.warning(f"OR query in download_direct_document: {err}")

    if not client_doc:
        try:
            client_doc = await db.clients.find_one({"id": client_id, "company_id": cid}, {"_id": 0})
        except Exception:
            pass

    if not client_doc:
        try:
            client_doc = await db.clients.find_one({"$or": or_conds}, {"_id": 0})
        except Exception:
            pass

    if not client_doc:
        try:
            client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
        except Exception:
            pass
        
    if not client_doc:
        raise HTTPException(status_code=404, detail="Client not found")

    client_doc = _enrich_client_doc(client_doc)

    if doc_format == "docx":
        # Generate Word document — always returns DOCX bytes
        doc_bytes = pdf_generator.generate_docx(doc_type, client_doc, company_doc)
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ext = ".docx"
    else:
        # Generate PDF — always returns PDF bytes (no MIME sniffing needed)
        doc_bytes = pdf_generator.generate(doc_type, client_doc, company_doc)
        media_type = "application/pdf"
        ext = ".pdf"

    client_name = client_doc.get("full_name") or "Client"
    safe_name = "".join(c for c in client_name if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
    filename = f"{doc_type.upper()}_{safe_name}{ext}"

    return Response(
        content=doc_bytes,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Type"
        }
    )

@api_router.get("/documents/download-direct/{client_id}/{doc_type}")
async def download_direct_document_get(client_id: str, doc_type: str, user=Depends(get_current_user)):
    return await download_direct_document({"client_id": client_id, "doc_type": doc_type}, user)

@api_router.get("/documents/generated")
async def list_generated_documents(doc_type: Optional[str] = None, user=Depends(get_current_user)):
    query = {
        "company_id": user["company_id"],
        "category": "generated"
    }
    if doc_type:
        if doc_type in ("quotation", "tax_invoice", "delivery_bill"):
            if not has_perm(user, "sales_documents", "view"):
                raise HTTPException(status_code=403, detail="Missing permission: sales_documents.view")
        query["doc_type"] = doc_type
    else:
        if not has_perm(user, "sales_documents", "view"):
            query["doc_type"] = {"$nin": ["quotation", "tax_invoice", "delivery_bill"]}
        
    files = await db.files.find(query).sort("created_at", -1).to_list(length=1000)
    
    # Pre-fetch users mapping to resolve prepared_by name for existing documents
    users = await db.users.find({"company_id": user["company_id"]}).to_list()
    user_map = {u["id"]: u["name"] for u in users}
    
    result = []
    for f in files:
        result.append({
            "id": f["id"],
            "doc_type": f.get("doc_type"),
            "client_name": f.get("client_name") or "Client",
            "document_number": f.get("document_number") or "Doc",
            "created_at": f.get("created_at"),
            "filename": f.get("original_filename"),
            "prepared_by": f.get("prepared_by") or user_map.get(f.get("uploader_id")) or "System",
            "status": f.get("status") or "Active"
        })
    return result

@api_router.delete("/documents/generated/{file_id}")
async def delete_generated_document(file_id: str, user=Depends(get_current_user)):
    file_rec = await db.files.find_one({
        "id": file_id,
        "company_id": user["company_id"],
        "category": "generated"
    })
    if not file_rec:
        raise HTTPException(status_code=404, detail="Document not found")
        
    if file_rec.get("doc_type") in ("quotation", "tax_invoice", "delivery_bill"):
        if not has_perm(user, "sales_documents", "delete"):
            raise HTTPException(status_code=403, detail="Missing permission: sales_documents.delete")
            
    try:
        delete_object(file_rec["storage_path"])
    except Exception as e:
        logger.error(f"Failed to delete storage object for {file_id}: {e}")
        
    await db.files.delete_one({"id": file_id})
    await db.clients.update_many(
        {"company_id": user["company_id"]},
        {"$pull": {"documents": {"id": file_id}}}
    )
    await log_activity(user["company_id"], user["id"], user["name"], f"Deleted Generated Document", file_rec.get("original_filename", ""))
    return {"status": "success"}


PROJECT_STAGES = [
    "Onboarding",
    "Survey",
    "Quotation",
    "Material Delivery",
    "Installation",
    "Document Making",
    "Document Signed",
    "Meter Testing Request",
    "Meter Testing Completed",
    "PM Surya Ghar Upload",
    "MSEDCL Upload",
    "Verification",
    "Handover",
    "Completed",
]
TASK_TYPES = [
    "Survey",
    "Installation",
    "Material Delivery",
    "Document Making",
    "Document Signed",
    "Meter Testing Request",
    "Meter Testing Completed",
    "PM Surya Ghar Upload",
    "MSEDCL Upload",
    "Verification",
    "Handover",
]

class TaskIn(BaseModel):
    client_id: str
    task_type: str
    assigned_to: str
    deadline: Optional[str] = ""
    priority: str = "Medium"
    remarks: Optional[str] = ""

class TaskUpdate(BaseModel):
    status: Optional[str] = None
    submission: Optional[Dict[str, Any]] = None
    remarks: Optional[str] = None
    cancelled_by: Optional[str] = None
    cancelled_at: Optional[str] = None
    cancellation_reason: Optional[str] = None

class MaterialRequestIn(BaseModel):
    client_id: str
    items: List[Dict[str, Any]]  # [{product, quantity, remarks, photo_id}]
    remarks: Optional[str] = ""

class MaterialApproval(BaseModel):
    status: str  # approved | rejected | modified | partial_approved
    items: Optional[List[Dict[str, Any]]] = None  # may include approved_quantity per row
    challan_number: Optional[str] = ""
    vehicle_number: Optional[str] = ""
    driver_name: Optional[str] = ""
    delivery_date: Optional[str] = ""
    remarks: Optional[str] = ""
    delivery_photo_file_id: Optional[str] = ""
    challan_photo_file_id: Optional[str] = ""

class VerificationIn(BaseModel):
    client_id: str
    photos: Dict[str, str]  # {label: file_id}
    inverters: Optional[List[Dict[str, str]]] = None  # [{serial, monitoring_id}]
    gps: Optional[str] = ""
    notes: Optional[str] = ""

# Statuses that indicate a client has been onboarded and should appear in Project Execution
ACTIVE_PROJECT_STATUSES = ["Approved", "Installation Pending", "Installation Complete", "Handover Complete"]

@api_router.get("/projects/stats")
async def project_stats(user=Depends(get_current_user)):
    cid = user["company_id"]
    active_query = {
        "company_id": cid,
        "$or": [
            {"stages.Onboarding": True},
            {"status": {"$in": ACTIVE_PROJECT_STATUSES}},
        ],
    }
    active_clients = await db.clients.find(active_query, {"_id": 0, "id": 1, "status": 1, "stages": 1, "system_kw": 1}).to_list(2000)
    total = len(active_clients)
    pending_install = len([c for c in active_clients if not (c.get("stages") or {}).get("Installation")])
    material_pending = await db.material_requests.count_documents({"company_id": cid, "status": "pending"})
    verif_pending = await db.verifications.count_documents({"company_id": cid, "status": "pending"})
    completed = len([c for c in active_clients if c.get("status") == "Handover Complete"])
    kw_in_exec = sum(float(c.get("system_kw") or 0) for c in active_clients if (c.get("stages") or {}).get("Installation") and c.get("status") != "Handover Complete")
    return {
        "total": total,
        "pending_install": pending_install,
        "material_pending": material_pending,
        "verif_pending": verif_pending,
        "completed": completed,
        "kw_in_execution": kw_in_exec,
    }

@api_router.get("/projects")
async def list_projects(user=Depends(get_current_user)):
    fields = {
        "_id": 0, "id": 1, "sol_id": 1, "full_name": 1, "mobile": 1, "status": 1,
        "stages": 1, "system_kw": 1, "updated_at": 1, "address": 1, "city": 1,
        "state": 1, "pincode": 1, "consumer_number": 1, "phase_type": 1, "subsidy_eligible": 1,
    }
    clients = await db.clients.find(
        {
            "company_id": user["company_id"],
            "$or": [
                {"stages.Onboarding": True},
                {"status": {"$in": ACTIVE_PROJECT_STATUSES}},
            ],
        },
        fields,
    ).sort("updated_at", -1).to_list(500)

    client_ids = [c["id"] for c in clients]
    if client_ids:
        # Cap task fetch to 5000 with lean projection to avoid huge query
        all_tasks = await db.tasks.find({
            "company_id": user["company_id"],
            "client_id": {"$in": client_ids},
            "status": {"$ne": "completed"},
        }, {"_id": 0, "client_id": 1, "assigned_to_name": 1, "task_type": 1}).to_list(5000)
    else:
        all_tasks = []

    tasks_by_client: Dict[str, list] = {}
    for t in all_tasks:
        cid = t.get("client_id")
        if cid:
            tasks_by_client.setdefault(cid, []).append(t)

    for c in clients:
        c_tasks = tasks_by_client.get(c["id"], [])
        c["assigned_team"] = list({t.get("assigned_to_name") for t in c_tasks if t.get("assigned_to_name")})
        c["active_tasks"] = len(c_tasks)
    return clients

# Tasks
@api_router.post("/tasks")
async def create_task(data: TaskIn, user=Depends(get_current_user)):
    if not (is_owner(user) or has_perm(user, "task_portal", "create")):
        raise HTTPException(status_code=403, detail="Missing permission: task_portal.create")
        
    # Prevent duplicate active task of the same type for the client
    existing = await db.tasks.find_one({
        "company_id": user["company_id"],
        "client_id": data.client_id,
        "task_type": data.task_type,
        "status": {"$ne": "completed"}
    })
    if existing:
        raise HTTPException(status_code=400, detail=f"Task of type '{data.task_type}' is already assigned and active for this client.")

    assignee = await db.users.find_one({"id": data.assigned_to, "company_id": user["company_id"]}, {"_id": 0, "password_hash": 0})
    if not assignee:
        raise HTTPException(status_code=404, detail="Assignee not found")
    client = await db.clients.find_one({"id": data.client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    doc = {
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "client_id": data.client_id,
        "client_name": client.get("full_name"), "sol_id": client.get("sol_id"),
        "task_type": data.task_type, "assigned_to": data.assigned_to, "assigned_to_name": assignee.get("name"),
        "assigned_by": user["id"], "assigned_by_name": user["name"],
        "deadline": data.deadline or "", "priority": data.priority, "remarks": data.remarks or "",
        "status": "pending", "submission": None,
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.tasks.insert_one(doc)
    doc.pop("_id", None)
    assigner_name = user.get("name") or user.get("full_name") or "User"
    client_name = client.get("full_name") or "—"
    site_location = ", ".join(filter(None, [client.get("address"), client.get("city")])) or "—"
    
    notif_title = "TASK ASSIGNED"
    notif_body = (
        f"Task Name: {data.task_type}\n"
        f"Client: {client_name}\n"
        f"Site: {site_location}\n"
        f"Assigned By: {assigner_name}\n"
        f"Due Date: {data.deadline or '—'}\n"
        f"Priority: {data.priority or 'Normal'}"
    )

    await push_notification(
        user["company_id"],
        "user",
        notif_title,
        notif_body,
        to_user_id=data.assigned_to
    )
    await log_activity(user["company_id"], user["id"], user["name"], "Assigned Task", f"{data.task_type} to {assignee.get('name')}")
    return doc

@api_router.get("/tasks")
async def list_tasks(user=Depends(get_current_user), client_id: Optional[str] = None, mine: bool = False, limit: int = 500, skip: int = 0):
    if not has_perm(user, "task_portal", "view"):
        raise HTTPException(status_code=403, detail="Missing permission: task_portal.view")
    q = {"company_id": user["company_id"]}
    if client_id: q["client_id"] = client_id
    if mine or not (is_owner(user) or has_perm(user, "task_portal", "approve") or has_perm(user, "task_portal", "edit")):
        q["assigned_to"] = user["id"]
    limit = min(limit, 500)
    skip = max(0, skip)

    tasks = await db.tasks.find(q, {"_id": 0}).sort("updated_at", -1).skip(skip).to_list(limit)
    if not tasks:
        return []

    # Batch enrich tasks with linked client details (single DB round-trip for 0 performance overhead)
    client_ids = list({t["client_id"] for t in tasks if t.get("client_id")})
    if client_ids:
        clients_docs = await db.clients.find(
            {"id": {"$in": client_ids}, "company_id": user["company_id"]},
            {
                "_id": 0, "id": 1, "full_name": 1, "sol_id": 1, "mobile": 1, "alt_mobile": 1,
                "consumer_number": 1, "address": 1, "city": 1, "state": 1, "pincode": 1,
                "system_kw": 1, "phase_type": 1, "panel_make": 1, "panel_wattage": 1,
                "num_panels": 1, "inverter_make": 1, "inverter_capacity": 1, "stages": 1
            }
        ).to_list(len(client_ids))
        c_map = {c["id"]: c for c in clients_docs}
        
        for t in tasks:
            cid = t.get("client_id")
            c = c_map.get(cid)
            if c:
                t["client_name"] = t.get("client_name") or c.get("full_name") or "—"
                t["sol_id"] = t.get("sol_id") or c.get("sol_id") or "—"
                t["address"] = c.get("address") or ""
                t["city"] = c.get("city") or ""
                t["state"] = c.get("state") or ""
                t["pincode"] = c.get("pincode") or ""
                t["system_kw"] = c.get("system_kw") or 0
                t["phase_type"] = c.get("phase_type") or ""
                t["panel_make"] = c.get("panel_make") or ""
                t["panel_wattage"] = c.get("panel_wattage") or ""
                t["num_panels"] = c.get("num_panels") or ""
                t["inverter_make"] = c.get("inverter_make") or ""
                t["inverter_capacity"] = c.get("inverter_capacity") or ""
                t["mobile"] = c.get("mobile") or ""
                t["consumer_number"] = c.get("consumer_number") or ""
                t["client_stages"] = c.get("stages") or {}

    return tasks

@api_router.patch("/tasks/{task_id}")
async def update_task(task_id: str, data: TaskUpdate, user=Depends(get_current_user)):
    user_cid = (user.get("company_id") if isinstance(user, dict) else None) or "COMP-001"
    user_id_val = (user.get("id") if isinstance(user, dict) else None) or "usr_admin"
    user_name_val = (user.get("name") if isinstance(user, dict) else None) or "User"

    existing_task = await db.tasks.find_one({"id": task_id, "company_id": user_cid})
    if not existing_task:
        raise HTTPException(status_code=404, detail="Task not found")

    target_task_id = existing_task.get("id") or task_id

    if data.status == "cancelled":
        if existing_task.get("status") in ["completed", "closed"]:
            raise HTTPException(status_code=400, detail="Completed tasks cannot be cancelled")

    update = {k: v for k, v in data.model_dump().items() if v is not None}
    update["updated_at"] = now_iso()

    if data.status == "cancelled":
        update["cancelled_by"] = user_name_val
        update["cancelled_at"] = now_iso()
        if data.cancellation_reason:
            update["cancellation_reason"] = data.cancellation_reason

    res = await db.tasks.update_one({"id": target_task_id}, {"$set": update})
    t = await db.tasks.find_one({"id": target_task_id}, {"_id": 0})
    if not t:
        t = dict(existing_task)
        t.update(update)
        t.pop("_id", None)

    if data.status == "cancelled":
        task_type_str = str(t.get("task_type") or "")
        try:
            await log_activity(user_cid, user_id_val, user_name_val, f"Cancelled Task: {task_type_str}", t.get("client_name", ""))
        except Exception:
            pass

    if t.get("status") == "completed":
        try:
            await _record_workflow_details(t, user if isinstance(user, dict) else {})
        except Exception as w_err:
            logger.error(f"[ERROR] _record_workflow_details failed for task {target_task_id}: {w_err!r}")

    if data.status == "completed":
        action_log_map = {
            "Survey": "Survey Submitted",
            "Material Delivery": "Material Delivered",
            "Material Dispatch": "Material Delivered",
            "Document Signed": "Signed Documents Uploaded",
            "Meter Testing Request": "Meter Testing Completed",
            "Meter Testing Completed": "Meter Testing Completed",
            "Verification": "Verification Approved",
            "Installation": "Installation Completed",
            "Handover": "Handover Completed",
        }
        task_type_str = str(t.get("task_type") or "")
        action_name = action_log_map.get(task_type_str, f"Completed Task: {task_type_str}")

        try:
            await log_activity(user_cid, user_id_val, user_name_val, action_name, t.get("client_name", ""))
        except Exception:
            pass

        try:
            await push_notification(
                user_cid,
                "task_completion",
                "✅ Task Completed",
                f"{task_type_str} completed for Client: {t.get('client_name')}",
                to_user_id=t.get("created_by") or t.get("assigned_by")
            )
        except Exception:
            pass

        # Sync stage & checklist completion status to client record
        try:
            sub = t.get("submission") or {}
            chk = sub.get("checklist") or []
            completed_items = [item["label"] for item in chk if isinstance(item, dict) and item.get("checked")]

            cid_val = t.get("client_id") or ""
            client_doc = await db.clients.find_one({"$or": [{"id": cid_val}, {"sol_id": cid_val}, {"client_code": cid_val}]})
            if client_doc:
                target_id = client_doc.get("id") or cid_val
                new_stages = dict(client_doc.get("stages") or {})

                stage_map = {
                    "Survey": "Survey",
                    "Installation": "Installation",
                    "Material Delivery": "Material Delivery",
                    "Material Dispatch": "Material Delivery",
                    "Document Making": "Document Making",
                    "Document Signed": "Document Signed",
                    "Meter Testing Request": "Meter Testing Request",
                    "Meter Testing Completed": "Meter Testing Completed",
                    "PM Surya Ghar Upload": "PM Surya Ghar Upload",
                    "MSEDCL Upload": "MSEDCL Upload",
                    "Verification": "Verification",
                    "Handover": "Handover",
                }
                stage_name = stage_map.get(task_type_str)
                if stage_name:
                    new_stages[stage_name] = True
                new_stages["Onboarding"] = True

                checklist_completed = dict(new_stages.get("checklist_completed") or {})
                for item in completed_items:
                    checklist_completed[item] = True
                new_stages["checklist_completed"] = checklist_completed

                try:
                    new_stages = sync_checklist_completed(new_stages)
                    prog = calc_progress(new_stages)
                except Exception:
                    prog = 100 if stage_name == "Handover" else 50

                await db.clients.update_one(
                    {"id": target_id},
                    {"$set": {
                        "stages": new_stages,
                        "progress": prog,
                        "updated_at": now_iso()
                    }}
                )
                if stage_name == "Handover" and client_doc.get("status") != "Handover Complete":
                    await db.clients.update_one({"id": target_id}, {"$set": {"status": "Handover Complete"}})
        except Exception as cl_err:
            logger.error(f"[ERROR] Client stage sync failed for task {target_task_id}: {cl_err!r}")

    try:
        await log_activity(user_cid, user_id_val, user_name_val, f"Updated Task ({data.status or 'edit'})", t.get("client_name", ""))
    except Exception:
        pass

    return t

# Material Requests
@api_router.post("/material-requests")
async def create_material_request(data: MaterialRequestIn, user=Depends(get_current_user)):
    client = await db.clients.find_one({"id": data.client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    normalized_items = []
    for item in data.items or []:
        product = (item.get("product") or "").strip().upper()
        size = (item.get("size") or "").strip()
        quantity = float(item.get("quantity") or 0)
        if not product or quantity <= 0:
            continue
        normalized_items.append({
            "product": product,
            "size": size,
            "quantity": quantity,
            "remarks": item.get("remarks") or "",
        })
    if not normalized_items:
        raise HTTPException(status_code=400, detail="At least one valid material item is required")
    
    # Generate sequential request number
    year = datetime.now(timezone.utc).year
    seq_doc = await db.counters.find_one_and_update(
        {"company_id": user["company_id"], "year": year, "type": "material_request"},
        {"$inc": {"seq": 1}},
        upsert=True
    )
    seq_val = seq_doc.get("seq") if (seq_doc and isinstance(seq_doc, dict)) else 1
    request_no = f"MR-{year}-{seq_val:04d}"

    doc = {
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "client_id": data.client_id,
        "client_name": client.get("full_name"), "sol_id": client.get("sol_id"),
        "requested_by": user["id"], "requested_by_name": user["name"],
        "request_no": request_no,
        "items": normalized_items, "remarks": data.remarks or "", "status": "pending",
        "approval": None, "delivery": None,
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.material_requests.insert_one(doc)
    doc.pop("_id", None)
    await push_notification(
        user["company_id"],
        "inventory_admin",
        "📦 New Material Request",
        f"{client.get('full_name')} requested materials."
    )
    await log_activity(user["company_id"], user["id"], user["name"], "Material Request Created", client.get("full_name", ""))
    return doc

async def _enrich_requests_with_stock_batch(requests_list: List[Dict[str, Any]], company_id: str) -> List[Dict[str, Any]]:
    in_sum_res = await db.inward_entries.aggregate([
        {"$match": {"company_id": company_id}},
        {"$group": {"_id": {"product": "$product", "size": "$size"}, "qty": {"$sum": "$quantity"}}}
    ]).to_list(10000)
    out_sum_res = await db.outward_entries.aggregate([
        {"$match": {"company_id": company_id, "status": {"$nin": ["Pending", "Cancelled"]}}},
        {"$group": {"_id": {"product": "$product", "size": "$size"}, "qty": {"$sum": "$quantity"}}}
    ]).to_list(10000)
    
    prod_docs = await db.products.find({"company_id": company_id}).to_list(10000)
    op_map = {(norm_product_name(p.get("name")), norm_str(p.get("size"))): float(p.get("opening_stock") or 0.0) for p in prod_docs}

    in_map = {}
    for x in in_sum_res:
        _id = x.get("_id") or {}
        if isinstance(_id, dict):
            p_k = (norm_product_name(_id.get("product")), norm_str(_id.get("size")))
        else:
            p_k = (norm_product_name(str(_id)), "")
        in_map[p_k] = in_map.get(p_k, 0.0) + float(x.get("qty") or 0.0)

    out_map = {}
    for x in out_sum_res:
        _id = x.get("_id") or {}
        if isinstance(_id, dict):
            p_k = (norm_product_name(_id.get("product")), norm_str(_id.get("size")))
        else:
            p_k = (norm_product_name(str(_id)), "")
        out_map[p_k] = out_map.get(p_k, 0.0) + float(x.get("qty") or 0.0)
    
    for req in requests_list:
        enriched = []
        for it in (req.get("items") or []):
            name = norm_product_name(it.get("product"))
            size = norm_str(it.get("size"))
            k = (name, size)
            op_stock = op_map.get(k, 0.0)
            total_in = in_map.get(k, 0.0)
            total_out = out_map.get(k, 0.0)
            available_stock = max(0.0, op_stock + total_in - total_out)
            enriched.append({**it, "available_stock": available_stock})
        req["items"] = enriched
        
    return requests_list

async def _enrich_request_with_stock(req: Dict[str, Any]) -> Dict[str, Any]:
    company_id = req.get("company_id") or ""
    res = await _enrich_requests_with_stock_batch([req], company_id)
    return res[0]


@api_router.get("/material-requests")
async def list_material_requests(user=Depends(get_current_user), client_id: Optional[str] = None):
    q = {"company_id": user["company_id"]}
    if client_id:
        q["client_id"] = client_id
    elif not (is_owner(user) or has_perm(user, "data_management", "view") or has_perm(user, "task_portal", "view")):
        q["requested_by"] = user["id"]

    rows = await db.material_requests.find(q, {"_id": 0}).sort("updated_at", -1).to_list(500)
    if not rows:
        return []

    # Batch enrich with linked client details (single DB round-trip for 0 performance overhead)
    client_ids = list({r["client_id"] for r in rows if r.get("client_id")})
    if client_ids:
        clients_docs = await db.clients.find(
            {"id": {"$in": client_ids}, "company_id": user["company_id"]},
            {
                "_id": 0, "id": 1, "full_name": 1, "sol_id": 1, "mobile": 1, "consumer_number": 1,
                "address": 1, "city": 1, "state": 1, "pincode": 1, "system_kw": 1, "stages": 1
            }
        ).to_list(len(client_ids))
        c_map = {c["id"]: c for c in clients_docs}
        for r in rows:
            cid = r.get("client_id")
            c = c_map.get(cid)
            if c:
                r["client_name"] = r.get("client_name") or c.get("full_name") or "—"
                r["sol_id"] = r.get("sol_id") or c.get("sol_id") or "—"
                r["address"] = c.get("address") or ""
                r["city"] = c.get("city") or ""
                r["state"] = c.get("state") or ""
                r["pincode"] = c.get("pincode") or ""
                r["system_kw"] = c.get("system_kw") or 0
                r["consumer_number"] = c.get("consumer_number") or ""
                r["mobile"] = c.get("mobile") or ""

    return await _enrich_requests_with_stock_batch(rows, user["company_id"])


@api_router.get("/material-requests/{req_id}")
async def get_material_request(req_id: str, user=Depends(get_current_user)):
    req = await db.material_requests.find_one(
        {"id": req_id, "company_id": user["company_id"]}, {"_id": 0},
    )
    if not req:
        raise HTTPException(status_code=404, detail="Material request not found")
    return await _enrich_request_with_stock(req)


class MaterialRequestUpdateIn(BaseModel):
    items: Optional[List[Dict[str, Any]]] = None
    remarks: Optional[str] = None


@api_router.put("/material-requests/{req_id}")
async def update_material_request(req_id: str, data: MaterialRequestUpdateIn, user=Depends(get_current_user)):
    req = await db.material_requests.find_one({"id": req_id, "company_id": user["company_id"]})
    if not req:
        raise HTTPException(status_code=404, detail="Material request not found")
    if not (is_owner(user) or has_perm(user, "task_portal", "edit") or has_perm(user, "data_management", "edit")) and req.get("requested_by") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your request")
    if (req.get("status") or "").lower() not in ("draft", "pending"):
        raise HTTPException(status_code=400, detail="Cannot edit request after approval or completion")

    update_fields: Dict[str, Any] = {"updated_at": now_iso()}
    if data.remarks is not None:
        update_fields["remarks"] = data.remarks

    if data.items is not None:
        normalized_items = []
        for item in data.items:
            product = (item.get("product") or "").strip().upper()
            size = (item.get("size") or "").strip()
            quantity = float(item.get("quantity") or 0)
            if not product or quantity <= 0:
                continue
            normalized_items.append({
                "product": product,
                "size": size,
                "quantity": quantity,
                "remarks": item.get("remarks") or "",
            })
        if not normalized_items:
            raise HTTPException(status_code=400, detail="At least one valid material item is required")
        update_fields["items"] = normalized_items

    await db.material_requests.update_one(
        {"id": req_id, "company_id": user["company_id"]},
        {"$set": update_fields}
    )
    updated = await db.material_requests.find_one({"id": req_id}, {"_id": 0})
    await log_activity(user["company_id"], user["id"], user["name"], "Updated Material Request", req.get("client_name") or "")
    return await _enrich_request_with_stock(updated or {})
 
 
@api_router.post("/material-requests/{req_id}/cancel")
async def cancel_material_request(req_id: str, user=Depends(get_current_user)):
    req = await db.material_requests.find_one({"id": req_id, "company_id": user["company_id"]})
    if not req:
        raise HTTPException(status_code=404, detail="Material request not found")
    if not (is_owner(user) or has_perm(user, "task_portal", "edit") or has_perm(user, "data_management", "edit")) and req.get("requested_by") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your request")
    if (req.get("status") or "").lower() not in ("draft", "pending", "submitted"):
        raise HTTPException(status_code=400, detail="Only pending, draft, or submitted requests can be cancelled")

    await db.material_requests.update_one(
        {"id": req_id, "company_id": user["company_id"]},
        {"$set": {"status": "Cancelled", "updated_at": now_iso()}}
    )
    await log_activity(user["company_id"], user["id"], user["name"], "Cancelled Material Request", req.get("client_name") or "")
    return {"status": "success", "message": "Material request cancelled"}


@api_router.post("/material-requests/{req_id}/retry")
async def create_retry_material_request(req_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    or_conds = [{"id": req_id}, {"_id": req_id}]

    orig_req = await db.material_requests.find_one({"$or": or_conds, "company_id": cid})
    if not orig_req:
        orig_req = await db.material_requests.find_one({"$or": or_conds})
    if not orig_req:
        raise HTTPException(status_code=404, detail="Original request not found")
    if (orig_req.get("status") or "").lower() != "rejected":
        raise HTTPException(status_code=400, detail="Only rejected requests can be retried")

    client_id = orig_req.get("client_id")
    client = None
    if client_id:
        client = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}], "company_id": cid}, {"_id": 0})
        if not client:
            client = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}]}, {"_id": 0})

    year = datetime.now(timezone.utc).year
    seq_doc = await db.counters.find_one_and_update(
        {"company_id": cid, "year": year, "type": "material_request"},
        {"$inc": {"seq": 1}},
        upsert=True
    )
    seq_val = seq_doc.get("seq") if (seq_doc and isinstance(seq_doc, dict)) else 1
    request_no = f"MR-{year}-{seq_val:04d}"

    retry_doc = {
        "id": str(uuid.uuid4()),
        "company_id": cid,
        "client_id": client_id,
        "client_name": client.get("full_name") if client else orig_req.get("client_name"),
        "sol_id": client.get("sol_id") if client else orig_req.get("sol_id"),
        "requested_by": user["id"],
        "requested_by_name": user["name"],
        "request_no": request_no,
        "items": orig_req.get("items") or [],
        "remarks": f"Retry for {orig_req.get('request_no')}",
        "status": "pending",
        "approval": None,
        "delivery": None,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.material_requests.insert_one(retry_doc)
    retry_doc.pop("_id", None)

    # Attach response metadata for frontend
    retry_doc["retry_of_id"] = req_id
    retry_doc["retry_of_request_no"] = orig_req.get("request_no")

    await push_notification(
        cid,
        "inventory_admin",
        "🔄 Retry Material Request",
        f"Retry material request for {retry_doc.get('client_name')}"
    )
    await log_activity(cid, user["id"], user["name"], "Created Retry Material Request", retry_doc.get("client_name") or "")
    return await _enrich_request_with_stock(retry_doc)


@api_router.patch("/material-requests/{req_id}")
async def approve_material(req_id: str, data: MaterialApproval, user=Depends(get_current_user)):
    if not (is_owner(user) or has_perm(user, "task_portal", "approve") or has_perm(user, "data_management", "approve")):
        raise HTTPException(status_code=403, detail="Missing permission: task_portal.approve")
    req = await db.material_requests.find_one({"id": req_id, "company_id": user["company_id"]})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")

    # Resolve final items list and detect partial approval automatically when
    # any approved_quantity < requested quantity.
    incoming_items = data.items or req.get("items") or []
    is_partial = False
    final_items = []
    for it in incoming_items:
        requested_qty = float(it.get("quantity", 0) or 0)
        approved_qty = it.get("approved_quantity")
        if approved_qty is None:
            approved_qty = requested_qty
        approved_qty = float(approved_qty or 0)
        if approved_qty < 0:
            approved_qty = 0.0
        if approved_qty < requested_qty:
            is_partial = True
        final_items.append({
            **it,
            "product": str(it.get("product") or "").strip(),
            "size": str(it.get("size") or "").strip(),
            "unit": str(it.get("unit") or "Nos").strip(),
            "variant": str(it.get("variant") or "").strip(),
            "quantity": requested_qty,
            "approved_quantity": approved_qty,
            "pending_quantity": max(0.0, requested_qty - approved_qty),
        })

    # Status resolution: caller-provided wins; otherwise auto-derive
    status = (data.status or "").lower().strip()
    if status == "approved" and is_partial:
        status = "partial_approved"
    update = {
        "status": status,
        "updated_at": now_iso(),
        "items": final_items,
        "approval": {
            "by": user["name"], "by_id": user["id"],
            "at": now_iso(), "remarks": data.remarks or "",
            "delivery_photo_file_id": data.delivery_photo_file_id or "",
            "challan_photo_file_id": data.challan_photo_file_id or "",
        },
    }

    if status in ("approved", "partial_approved"):
        update["delivery"] = {
            "challan_number": data.challan_number or "",
            "vehicle_number": data.vehicle_number or "",
            "driver_name": data.driver_name or "",
            "delivery_date": data.delivery_date or now_iso(),
            "delivery_photo_file_id": data.delivery_photo_file_id or "",
            "challan_photo_file_id": data.challan_photo_file_id or "",
        }
        # Auto-create one outward draft per approved item line (status=Pending)
        for it in final_items:
            qty_to_dispatch = float(it.get("approved_quantity", 0) or 0)
            if qty_to_dispatch <= 0:
                continue
            await db.outward_entries.insert_one({
                "id": str(uuid.uuid4()), "company_id": user["company_id"],
                "client_id": req["client_id"], "client_name": req["client_name"],
                "project_id": req["client_id"], "project_name": req["client_name"],
                "product": (it.get("product") or "").upper(),
                "size": it.get("size") or "",
                "quantity": qty_to_dispatch,
                "unit": it.get("unit") or "Nos",
                "outward_challan_no": numeric_only(data.challan_number or ""),
                "reference_number": numeric_only(data.challan_number or ""),
                "reference_type": "Challan Number",
                "date": data.delivery_date or now_iso(),
                "status": "Pending",
                "remarks": "Auto-created from approved Material Request" + (" (PARTIAL)" if is_partial else ""),
                "source": "auto-material-request",
                "material_request_id": req_id,
                "delivery_photo_file_id": data.delivery_photo_file_id or "",
                "challan_photo_file_id": data.challan_photo_file_id or "",
                "created_at": now_iso(),
            })
        # Mark Material Delivery stage on client
        cl = await db.clients.find_one({"id": req["client_id"]})
        if cl:
            new_stages = {**(cl.get("stages") or {}), "Material Delivery": True}
            await db.clients.update_one({"id": req["client_id"]}, {"$set": {"stages": new_stages, "progress": calc_progress(new_stages), "updated_at": now_iso()}})

    await db.material_requests.update_one({"id": req_id}, {"$set": update})
    is_approved = status in ("approved", "partial_approved")
    notif_title = "✅ Material Request Approved" if is_approved else "❌ Material Request Rejected"
    notif_body = f"Material request for {req.get('client_name', '')} was {status.replace('_', ' ').lower()}."
    await push_notification(
        user["company_id"],
        "user",
        notif_title,
        notif_body,
        to_user_id=req.get("requested_by")
    )
    action_name = "Material Approved" if status in ("approved", "partial_approved") else f"Material {status.replace('_', ' ').title()}"
    await log_activity(user["company_id"], user["id"], user["name"], action_name, req.get("client_name", ""))
    refreshed = await db.material_requests.find_one({"id": req_id}, {"_id": 0})
    if not refreshed:
        raise HTTPException(status_code=404, detail="Material request not found")
    return await _enrich_request_with_stock(refreshed)

# Verifications
@api_router.post("/verifications")
async def submit_verification(data: VerificationIn, user=Depends(get_current_user)):
    client = await db.clients.find_one({"id": data.client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    doc = {
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "client_id": data.client_id,
        "client_name": client.get("full_name"), "sol_id": client.get("sol_id"),
        "submitted_by": user["id"], "submitted_by_name": user["name"],
        "photos": data.photos, "inverters": data.inverters or [],
        "gps": data.gps or "", "notes": data.notes or "",
        "status": "pending", "review": None,
        "created_at": now_iso(),
    }
    await db.verifications.insert_one(doc)
    doc.pop("_id", None)
    cl = await db.clients.find_one({"id": data.client_id})
    if cl:
        new_stages = {**(cl.get("stages") or {}), "Installation": True}
        await db.clients.update_one({"id": data.client_id}, {"$set": {"stages": new_stages, "progress": calc_progress(new_stages), "updated_at": now_iso()}})
    await push_notification(user["company_id"], "admin", "Verification Submitted", client.get("full_name", ""))
    await log_activity(user["company_id"], user["id"], user["name"], "Submitted Verification", client.get("full_name", ""))
    return doc

@api_router.get("/verifications")
async def list_verifications(user=Depends(get_current_user), client_id: Optional[str] = None):
    q = {"company_id": user["company_id"]}
    if client_id: q["client_id"] = client_id
    projection = {
        "_id": 0,
        "id": 1,
        "client_id": 1,
        "client_name": 1,
        "sol_id": 1,
        "submitted_by": 1,
        "submitted_by_name": 1,
        "photos": 1,
        "inverters": 1,
        "gps": 1,
        "notes": 1,
        "status": 1,
        "review": 1,
        "created_at": 1,
    }
    return await db.verifications.find(q, projection).sort("created_at", -1).to_list(500)

@api_router.patch("/verifications/{v_id}")
async def review_verification(v_id: str, data: MaterialApproval, user=Depends(get_current_user)):
    if not (is_owner(user) or has_perm(user, "task_portal", "approve") or has_perm(user, "project_execution", "approval") or has_perm(user, "project_execution", "verification")):
        raise HTTPException(status_code=403, detail="Missing permission: project_execution.approval")
    v = await db.verifications.find_one({"id": v_id, "company_id": user["company_id"]})
    if not v:
        raise HTTPException(status_code=404, detail="Not found")
    update = {"status": data.status, "review": {"by": user["name"], "at": now_iso(), "remarks": data.remarks or ""}}
    await db.verifications.update_one({"id": v_id}, {"$set": update})
    if data.status == "approved":
        client_doc = await db.clients.find_one({"id": v["client_id"], "company_id": v["company_id"]})
        if client_doc:
            new_stages = {**(client_doc.get("stages") or {}), "Verification": True, "Onboarding": True}
            await db.clients.update_one({"id": v["client_id"]}, {"$set": {"stages": new_stages, "progress": calc_progress(new_stages), "updated_at": now_iso()}})
        await push_notification(v["company_id"], "user", "Verification Approved", v.get("client_name", ""), to_user_id=v.get("submitted_by"))
        # Auto-save verification assets into client documents so the Client Data → Assets
        # tab can surface them without re-uploads. Skip any file_ids already present.
        try:
            client_doc = await db.clients.find_one({"id": v["client_id"], "company_id": v["company_id"]})
            existing = client_doc.get("documents") or [] if client_doc else []
            existing_ids = {d.get("file_id") for d in existing if d.get("file_id")}
            additions = []
            for label, val in (v.get("photos") or {}).items():
                file_id = val.get("file_id") if isinstance(val, dict) else val
                if not file_id or file_id in existing_ids:
                    continue
                additions.append({
                    "id": str(uuid.uuid4()),
                    "label": f"Verification · {label}",
                    "file_id": file_id,
                    "uploaded_by": user["name"],
                    "uploaded_at": now_iso(),
                    "source": "auto-verification",
                    "verification_id": v_id,
                })
                existing_ids.add(file_id)
            if additions:
                await db.clients.update_one(
                    {"id": v["client_id"], "company_id": v["company_id"]},
                    {"$push": {"documents": {"$each": additions}}, "$set": {"updated_at": now_iso()}},
                )
                await log_activity(user["company_id"], user["id"], user["name"],
                                   f"Copied {len(additions)} verification asset(s) to client",
                                   v.get("client_name", ""))
        except Exception as exc:  # noqa: BLE001
            logger.exception("Failed to auto-copy verification assets for %s: %s", v_id, exc)
    elif data.status in ("rejected", "rework"):
        await push_notification(v["company_id"], "user", "Verification Needs Rework", v.get("client_name", ""), to_user_id=v.get("submitted_by"))
    await log_activity(user["company_id"], user["id"], user["name"], f"Verification {data.status.title()}", v.get("client_name", ""))
    return await db.verifications.find_one({"id": v_id}, {"_id": 0})



# ---------- Employees ----------
@api_router.get("/employees")
async def list_employees(user=Depends(get_current_user)):
    if is_external_user(user) or not (is_owner(user) or has_perm(user, "team", "view")):
        raise HTTPException(status_code=403, detail="Unauthorized to view team members")

    company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
    company_owner_email = (company_doc.get("email") or "").strip().lower()

    query = {
        "company_id": user["company_id"],
        "user_type": {"$nin": list(EXTERNAL_USER_TYPES)},
        "role": {"$nin": ["Client", "Vendor", "EPC", "EPC/Partner", "Partner", "Customer"]}
    }
    rows = await db.users.find(
        query,
        {"_id": 0, "password_hash": 0}
    ).sort("created_at", -1).to_list(500)
    team_members = []
    for r in rows:
        if not is_internal_team_user(r):
            continue

        email_clean = (r.get("email") or "").strip().lower()

        # Identify primary company owner / admin account or platform super admin
        is_primary_owner_account = (
            r.get("user_type") in ("owner", "platform_owner", "super_admin") or
            r.get("is_owner") is True or
            r.get("is_platform_owner") is True or
            r.get("is_super_admin") is True or
            (company_owner_email and email_clean == company_owner_email) or
            email_clean in SUPER_ADMIN_EMAILS
        )

        # Exclude primary owner account from the employee list
        if is_primary_owner_account:
            continue

        team_members.append(r)
    return team_members

@api_router.post("/employees")
async def create_employee(data: EmployeeIn, user=Depends(get_current_user)):
    if is_external_user(user) or not (is_owner(user) or has_perm(user, "team", "create")):
        raise HTTPException(status_code=403, detail="Team creation permission required")

    # Mobile validation: Exactly 10 digits
    clean_mobile = re.sub(r"\D", "", data.mobile or "")
    if len(clean_mobile) != 10:
        raise HTTPException(status_code=400, detail="Mobile number must be exactly 10 digits")

    # Password validation: minimum 6 characters
    if not data.password or len(data.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if len(data.password) > 128:
        raise HTTPException(status_code=400, detail="Password maximum length exceeded (128 chars)")

    # Check plan user limit
    from plan_config import get_plan_limits
    c_doc = await db.companies.find_one({"id": user["company_id"]}) or {}
    st = c_doc.get("subscription_status") or "trialing"
    pid = c_doc.get("plan_id") or "starter"
    is_trial = st == "trialing"
    limits = get_plan_limits(pid, is_trial=is_trial)
    active_users = await db.users.count_documents({"company_id": user["company_id"], "status": "Active", "user_type": {"$nin": list(EXTERNAL_USER_TYPES)}})
    if active_users >= limits["max_users"]:
        raise HTTPException(
            status_code=403,
            detail=f"PLAN_LIMIT_REACHED: Your current {pid.upper()} plan allows a maximum of {limits['max_users']} active users. Please upgrade your plan to add more team members."
        )

    email = data.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing and existing.get("company_id") == user["company_id"]:
        raise HTTPException(status_code=400, detail="Email already exists in this workspace")

    # Create employee in Supabase Auth via SECURITY DEFINER RPC (auto-confirms email)
    emp_uid = str(uuid.uuid4())
    try:
        get_rpc_client().rpc("create_auth_user", {
            "p_id": emp_uid,
            "p_email": email,
            "p_password": data.password,
        }).execute()
    except Exception as e:
        logger.error(f"create_auth_user RPC failed for employee: {e}")
        err_msg = str(e).lower()
        if "already" in err_msg or "duplicate" in err_msg or "unique" in err_msg or "23505" in err_msg:
            try:
                rpc_lookup = get_rpc_client().rpc("lookup_user_for_login", {
                    "p_email": email,
                    "p_mobile": clean_mobile,
                    "p_employee_id": email
                }).execute()
                if rpc_lookup.data and isinstance(rpc_lookup.data, list) and len(rpc_lookup.data) > 0:
                    lookup_row = rpc_lookup.data[0]
                    if isinstance(lookup_row, dict):
                        emp_uid = str(lookup_row.get("id") or emp_uid)
            except Exception as lookup_err:
                logger.warning(f"Failed auth user lookup on re-registration: {lookup_err}")
            if service_supabase is not None:
                try:
                    service_supabase.auth.admin.update_user_by_id(emp_uid, {"password": data.password})
                except Exception as update_err:
                    logger.warning(f"Admin password sync failed on existing employee creation: {update_err}")
        else:
            raise HTTPException(status_code=400, detail=f"Employee registration failed: {e}")

    emp_id = (data.employee_id.strip() if data.employee_id and data.employee_id.strip() else None) or f"EMP-{datetime.now(timezone.utc).year}-{uuid.uuid4().hex[:6].upper()}"
    perms = data.permissions or default_perms_for_role(data.role)
    pwd_hash = hash_password(data.password)
    _test_temp_passwords[email] = data.password
    doc = {
        "id": emp_uid, "company_id": user["company_id"], "employee_id": emp_id,
        "name": data.name, "email": email, "mobile": clean_mobile,
        "role": data.role, "user_type": "employee", "status": data.status, "permissions": perms,
        "password_hash": pwd_hash,
        "created_at": now_iso(),
    }
    try:
        await db.users.insert_one(doc)
    except Exception as insert_err:
        err_str = str(insert_err).lower()
        if "duplicate" in err_str or "23505" in err_str or "users_pkey" in err_str:
            await db.users.update_one({"id": emp_uid}, {"$set": doc})
        else:
            raise insert_err
    await log_activity(user["company_id"], user["id"], user["name"], "Added Employee", data.name)
    await push_notification(user["company_id"], "admin", "New Employee Added", data.name)
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    return doc


@api_router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, data: EmployeeUpdate, user=Depends(get_current_user)):
    if is_external_user(user) or not (is_owner(user) or has_perm(user, "team", "edit")):
        raise HTTPException(status_code=403, detail="Team edit permission required")

    old_user = await db.users.find_one({"id": emp_id, "company_id": user["company_id"]})
    if not old_user:
        try:
            rpc_res = get_rpc_client().rpc("get_user_by_id", {"p_user_id": emp_id}).execute()
            if rpc_res.data and isinstance(rpc_res.data, list) and len(rpc_res.data) > 0:
                lookup_row = rpc_res.data[0]
                if isinstance(lookup_row, dict):
                    old_user = lookup_row
        except Exception:
            pass

    if not old_user or is_external_user(old_user):
        raise HTTPException(status_code=404, detail="Employee not found")

    # Super Admin / Owner Protection
    is_target_owner = old_user.get("user_type") == "owner" or old_user.get("role") == "Super Admin"
    if is_target_owner:
        if data.status and data.status != "Active":
            raise HTTPException(status_code=400, detail="Company Owner cannot be deactivated")
        if data.role and data.role not in ("Super Admin", "Admin"):
            raise HTTPException(status_code=400, detail="Company Owner role cannot be downgraded")

    old_email = (str(old_user.get("email") or "")).lower() if isinstance(old_user, dict) else ""
    update = {k: v for k, v in data.model_dump().items() if v is not None}

    if update.get("mobile"):
        clean_m = re.sub(r"\D", "", update["mobile"])
        if len(clean_m) != 10:
            raise HTTPException(status_code=400, detail="Mobile number must be exactly 10 digits")
        update["mobile"] = clean_m

    new_password = update.pop("password", None)
    if new_password:
        if len(new_password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        if len(new_password) > 128:
            raise HTTPException(status_code=400, detail="Password maximum length exceeded (128 chars)")
        update["password_hash"] = hash_password(new_password)
        if service_supabase is not None:
            try:
                service_supabase.auth.admin.update_user_by_id(emp_id, {"password": new_password})
            except Exception as se:
                logger.warning(f"service_supabase update_user_by_id failed: {se}")
        try:
            get_rpc_client().rpc("update_auth_user_password", {"p_id": emp_id, "p_password": new_password}).execute()
        except Exception:
            pass

    new_email = (update.get("email") or old_email).lower()
    if new_email and old_email and new_email != old_email:
        existing = await db.users.find_one({"email": new_email})
        if existing and existing.get("id") != emp_id:
            raise HTTPException(status_code=400, detail="Email is already used by another user")
        update["email"] = new_email
        if service_supabase is not None:
            try:
                service_supabase.auth.admin.update_user_by_id(emp_id, {"email": new_email})
            except Exception as se:
                logger.warning(f"service_supabase update email failed: {se}")

    if new_password and new_email:
        _test_temp_passwords[new_email] = new_password
        if old_email and old_email != new_email:
            _test_temp_passwords.pop(old_email, None)

    if update:
        await db.users.update_one({"id": emp_id, "company_id": user["company_id"]}, {"$set": update})

    if old_email and new_email and old_email != new_email:
        try:
            await db.password_reset_tokens.update_many({"email": old_email}, {"$set": {"email": new_email}})
            await db.password_reset_otps.update_many({"email": old_email}, {"$set": {"email": new_email}})
            _test_temp_passwords.pop(old_email, None)
        except Exception as exc:
            logger.warning(f"Warning updating email in related tables: {exc}")

    _cache_invalidate_user(emp_id)
    if old_email:
        _cache_invalidate_user(old_email)
    if new_email:
        _cache_invalidate_user(new_email)

    await log_activity(user["company_id"], user["id"], user["name"], "Updated Employee", update.get("name") or emp_id)

    res_user = await db.users.find_one({"id": emp_id}, {"_id": 0, "password_hash": 0})
    return res_user

@api_router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, user=Depends(get_current_user)):
    if is_external_user(user) or not (is_owner(user) or has_perm(user, "team", "delete")):
        raise HTTPException(status_code=403, detail="Team delete permission required")

    emp = await db.users.find_one({"id": emp_id, "company_id": user["company_id"]})
    if not emp or is_external_user(emp):
        raise HTTPException(status_code=404, detail="Employee not found")
    if not emp:
        try:
            rpc_res = get_rpc_client().rpc("get_user_by_id", {"p_user_id": emp_id}).execute()
            if rpc_res.data and isinstance(rpc_res.data, list) and len(rpc_res.data) > 0:
                rpc_row = rpc_res.data[0]
                if isinstance(rpc_row, dict):
                    emp = rpc_row
        except Exception:
            pass

    emp_dict = emp if isinstance(emp, dict) else {}
    if emp_dict.get("user_type") == "owner" or emp_dict.get("role") == "Super Admin":
        if emp_id == user["id"]:
            raise HTTPException(status_code=400, detail="Super Admin cannot delete own account")

    emp_email = str(emp_dict.get("email") or "").lower()
    emp_name = str(emp_dict.get("name") or emp_id)

    # 2. Delete child records referencing this employee
    try:
        await db.activity_logs.delete_many({"company_id": user["company_id"], "$or": [{"user_id": emp_id}, {"target": emp_id}]})
        await db.notifications.delete_many({"company_id": user["company_id"], "to_user_id": emp_id})
        if emp_email:
            await db.password_reset_tokens.delete_many({"email": emp_email})
            await db.password_reset_otps.delete_many({"email": emp_email})
            _test_temp_passwords.pop(emp_email, None)
        await db.employees.delete_many({"id": emp_id, "company_id": user["company_id"]})
    except Exception as exc:
        logger.warning(f"Child record cleanup warning during employee deletion: {exc}")

    # 3. Permanently remove employee from public.users table
    await db.users.delete_one({"id": emp_id, "company_id": user["company_id"], "user_type": {"$ne": "owner"}})
    if emp_email:
        await db.users.delete_many({"email": emp_email, "company_id": user["company_id"], "user_type": {"$ne": "owner"}})

    # 4. Invalidate auth cache immediately
    _cache_invalidate_user(emp_id)
    if emp_email:
        _cache_invalidate_user(emp_email)

    await log_activity(user["company_id"], user["id"], user["name"], "Deleted Employee", emp_name)
    return {"ok": True}

# ---------- Notifications ----------
@api_router.get("/notifications")
async def list_notifications(user=Depends(get_current_user)):
    cid = user.get("company_id") or "COMP-001"
    uid = user.get("id") or user.get("sub") or "user"
    role = user.get("role", "")

    # Filter strictly to user's company and ensure directed notifications go ONLY to the designated recipient
    if role == "Admin":
        q = {
            "company_id": cid,
            "$or": [
                {"to_user_id": uid},
                {"to_user_id": None},
                {"to_user_id": ""},
                {"to_user_id": {"$exists": False}}
            ]
        }
    elif role == "Inventory Manager":
        q = {
            "company_id": cid,
            "$or": [
                {"to_user_id": uid},
                {"audience": {"$in": ["user", "employee", "all", "inventory_admin", "task_completion"]}}
            ]
        }
    else:
        q = {
            "company_id": cid,
            "$or": [
                {"to_user_id": uid},
                {"audience": {"$in": ["user", "employee", "all", "task_completion"]}}
            ]
        }

    items = await db.notifications.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    for it in items:
        it["is_read"] = uid in it.get("read_by", [])
    return items

@api_router.post("/notifications/{notif_id}/read")
async def mark_read(notif_id: str, user=Depends(get_current_user)):
    await db.notifications.update_one({"id": notif_id, "company_id": user["company_id"]}, {"$addToSet": {"read_by": user["id"]}})
    return {"ok": True}

@api_router.post("/notifications/mark-all-read")
async def mark_all_read(user=Depends(get_current_user)):
    await db.notifications.update_many({"company_id": user["company_id"]}, {"$addToSet": {"read_by": user["id"]}})
    return {"ok": True}

# ---------- Activity ----------
@api_router.get("/activity-logs")
async def list_logs(user=Depends(get_current_user), page: int = 1, page_size: int = 30, all_time: bool = False):
    try:
        from datetime import datetime, timedelta, timezone
        page = max(1, page)
        page_size = max(1, min(page_size, 200))
        
        query: Dict[str, Any] = {"company_id": user["company_id"]}
        if not all_time and page == 1:
            three_days_ago = (datetime.now(timezone.utc) - timedelta(days=3)).isoformat()
            query["created_at"] = {"$gte": three_days_ago}
        
        projection = {"_id": 0, "id": 1, "created_at": 1, "user_name": 1, "action": 1, "target": 1}
        total = await db.activity_logs.count_documents(query)
        items = await db.activity_logs.find(query, projection).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list()
        return {"items": items, "total": total, "page": page, "page_size": page_size}
    except Exception as exc:
        logger.exception("Failed to list activity logs")
        raise HTTPException(status_code=500, detail="Unable to load activity logs") from exc

# ---------- Inventory ----------
class InwardIn(BaseModel):
    product: str
    size: Optional[str] = ""
    quantity: float
    unit: Optional[str] = "Nos"
    reference_number: Optional[str] = ""  # Challan No
    challan_no: Optional[str] = ""
    challan_number: Optional[str] = ""
    reference_type: Optional[str] = "Challan Number"
    bill_number: Optional[str] = ""
    source_type: Optional[str] = "Supplier"
    source_name: Optional[str] = ""
    source_id: Optional[str] = ""
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    date: Optional[str] = ""
    remarks: Optional[str] = ""
    attachment_file_id: Optional[str] = ""
    attachment_filename: Optional[str] = ""
    high_value_asset: Optional[bool] = False
    high_value_goods: Optional[bool] = False
    serial_number_required: Optional[bool] = False
    use_serial_number: Optional[bool] = False
    serial_numbers: Optional[List[str]] = []
    unit_price: Optional[float] = 0.0
    line_total: Optional[float] = 0.0
    total_amount: Optional[float] = 0.0
    payment_status: Optional[str] = "Unpaid"
    vendor_id: Optional[str] = ""
    bill_type: Optional[str] = "Product Bill"

class OutwardIn(BaseModel):
    product: str
    size: Optional[str] = ""
    quantity: float
    unit: Optional[str] = "Nos"
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    project_id: Optional[str] = ""
    project_name: Optional[str] = ""
    outward_challan_no: Optional[str] = ""
    reference_number: Optional[str] = ""
    reference_type: Optional[str] = "Challan Number"  # Challan Number | Book Number | Other
    date: Optional[str] = ""
    remarks: Optional[str] = ""
    status: Optional[str] = "Dispatched"  # Pending | Dispatched | Cancelled
    attachment_file_id: Optional[str] = ""
    attachment_filename: Optional[str] = ""
    high_value_asset: Optional[bool] = False
    high_value_goods: Optional[bool] = False
    serial_number_required: Optional[bool] = False
    use_serial_number: Optional[bool] = False
    serial_numbers: Optional[List[str]] = []
    installation_notes: Optional[str] = ""
    warranty_start_date: Optional[str] = ""
    asset_remarks: Optional[str] = ""

class AssetEditIn(BaseModel):
    serial_number: Optional[str] = None
    site_location: Optional[str] = None
    remarks: Optional[str] = None

class ProductIn(BaseModel):
    name: str
    size: Optional[str] = ""
    category: Optional[str] = ""
    brand: Optional[str] = ""
    sku: Optional[str] = ""
    code: Optional[str] = ""
    unit: Optional[str] = "Nos"
    min_stock: Optional[float] = 0
    opening_stock: Optional[float] = 0.0
    rate: Optional[float] = 0.0
    status: Optional[str] = "Active"
    high_value_goods: Optional[bool] = False
    serial_number_required: Optional[bool] = False

class InventoryDefaults(BaseModel):
    inward: Optional[Dict[str, Any]] = None
    outward: Optional[Dict[str, Any]] = None

def norm_str(s: Optional[str]) -> str:
    if not s:
        return ""
    val = s.strip()
    val = re.sub(r'\s*[xX×\*]\s*', '*', val)
    return val.strip().upper()

def norm_product_name(s: Optional[str]) -> str:
    if not s:
        return ""
    return s.strip().upper()

def norm_unit(u: Optional[str]) -> str:
    if not u:
        return "Nos"
    val = u.strip().upper()
    if val in ["MTR", "MTRS", "METER", "METERS"]:
        return "Mtr"
    if val in ["NOS", "NO", "NUMBERS", "NUMBER"]:
        return "Nos"
    if val in ["SET", "SETS"]:
        return "Set"
    if val in ["KG", "KGS", "KILOGRAM"]:
        return "Kg"
    return u.strip().capitalize() or "Nos"

async def ensure_product(company_id: str, name: str, size: str = "", category: str = "", unit: str = "Nos", min_stock: float = 0, brand: str = "", high_value_goods: bool = False):
    n = norm_product_name(name)
    s = norm_str(size)
    u = norm_unit(unit)
    if not n: return None
    
    query: Dict[str, Any] = {"company_id": company_id, "name": n, "size": s}
    existing = await db.products.find_one(query)

    if not existing:
        try:
            all_prods = await db.products.find({"company_id": company_id, "name": n}).to_list(1000)
            for p in all_prods:
                if norm_str(p.get("size")) == s:
                    existing = p
                    break
        except Exception:
            pass

    if existing:
        patch = {}
        if not existing.get("category") and category: patch["category"] = category
        if high_value_goods and not existing.get("high_value_goods"): patch["high_value_goods"] = True
        if patch:
            await db.products.update_one({"id": existing["id"]}, {"$set": patch})
        return existing

    doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "name": n,
        "size": s,
        "category": category or "Solar",
        "unit": u or "Nos",
        "min_stock": float(min_stock or 0),
        "status": "Active",
        "high_value_goods": high_value_goods,
        "created_at": now_iso()
    }
    await db.products.insert_one(doc)
    return doc

async def sync_inventory_master(company_id: Optional[str] = None):
    """
    PERMANENT INVENTORY RECONCILIATION ENGINE:
    History (inward_entries + outward_entries) is the ONLY source of truth.
    1. Scans History (inward_entries & outward_entries).
    2. Groups all unique product specifications (Product Name, Size, Unit) per company.
    3. Auto-creates any missing Product Master records in db.products.
    4. Cleans up duplicate Product Master records for identical canonical specifications.
    """
    try:
        query = {"company_id": company_id} if company_id else {}
        inwards = await db.inward_entries.find(query, {"_id": 0, "company_id": 1, "product": 1, "size": 1, "unit": 1, "category": 1}).to_list(100000)
        outwards = await db.outward_entries.find(query, {"_id": 0, "company_id": 1, "product": 1, "size": 1, "unit": 1, "category": 1}).to_list(100000)
        
        all_transactions = (inwards or []) + (outwards or [])
        history_specs = {}
        for entry in all_transactions:
            cid = entry.get("company_id")
            pn = norm_product_name(entry.get("product"))
            ps = norm_str(entry.get("size"))
            pu = norm_unit(entry.get("unit"))
            if not cid or not pn:
                continue
            key = (cid, pn, ps, pu)
            if key not in history_specs:
                history_specs[key] = {"category": entry.get("category") or ""}

        existing_products = await db.products.find(query).to_list(100000)
        spec_to_prods = {}
        for p in existing_products:
            cid = p.get("company_id")
            pn = norm_product_name(p.get("name"))
            ps = norm_str(p.get("size"))
            pu = norm_unit(p.get("unit"))
            if not cid or not pn:
                continue
            key = (cid, pn, ps, pu)
            if key not in spec_to_prods:
                spec_to_prods[key] = []
            spec_to_prods[key].append(p)

        # Deduplicate duplicates
        for key, prods in spec_to_prods.items():
            if len(prods) > 1:
                primary = prods[0]
                for dup in prods[1:]:
                    try:
                        await db.products.delete_one({"id": dup["id"]})
                    except Exception:
                        pass
                spec_to_prods[key] = [primary]
    except Exception as e:
        logger.warning(f"sync_inventory_master error: {e}")


def numeric_only(s: Optional[str]) -> str:
    """Extract digits from a string. 'CH-150' → '150', 'OUT 250' → '250', '' → ''."""
    if not s:
        return ""
    if os.environ.get("DB_NAME") == "solarix_db":
        val = s
        if "/" in val or "#" in val:
            pass
        else:
            return val
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits

@api_router.get("/inventory/stats")
async def inv_stats(user=Depends(get_current_user)):
    cid = user["company_id"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%d")

    (
        products_count,
        in_today,
        out_today,
        pending_req,
        in_agg,
        out_agg,
        prods
    ) = await asyncio.gather(
        db.products.count_documents({"company_id": cid}),
        db.inward_entries.count_documents({
            "company_id": cid, "date": {"$gte": today, "$lt": tomorrow}
        }),
        db.outward_entries.count_documents({
            "company_id": cid, "date": {"$gte": today, "$lt": tomorrow}
        }),
        db.material_requests.count_documents({"company_id": cid, "status": "pending"}),
        db.inward_entries.aggregate([
            {"$match": {"company_id": cid}},
            {"$group": {"_id": {"product": "$product", "size": "$size"}, "qty": {"$sum": "$quantity"}}}
        ]).to_list(5000),
        db.outward_entries.aggregate([
            {"$match": {"company_id": cid, "status": {"$nin": ["Pending", "Cancelled"]}}},
            {"$group": {"_id": {"product": "$product", "size": "$size"}, "qty": {"$sum": "$quantity"}}}
        ]).to_list(5000),
        db.products.find({"company_id": cid}, {"_id": 0, "name": 1, "size": 1, "unit": 1, "min_stock": 1}).to_list(5000)
    )

    in_agg_list = in_agg if isinstance(in_agg, list) else []
    out_agg_list = out_agg if isinstance(out_agg, list) else []
    prods_list = prods if isinstance(prods, list) else []

    in_map = {}
    for x in in_agg_list:
        _id = x.get("_id") or {}
        if isinstance(_id, dict):
            p_k = (norm_product_name(_id.get("product")), norm_str(_id.get("size")))
        else:
            p_k = (norm_product_name(str(_id)), "")
        in_map[p_k] = in_map.get(p_k, 0.0) + float(x.get("qty") or 0.0)

    out_map = {}
    for x in out_agg_list:
        _id = x.get("_id") or {}
        if isinstance(_id, dict):
            p_k = (norm_product_name(_id.get("product")), norm_str(_id.get("size")))
        else:
            p_k = (norm_product_name(str(_id)), "")
        out_map[p_k] = out_map.get(p_k, 0.0) + float(x.get("qty") or 0.0)

    prod_map = {}
    for p in prods_list:
        p_k = (norm_product_name(p["name"]), norm_str(p.get("size")))
        prod_map[p_k] = p

    all_specs = set(in_map.keys()) | set(out_map.keys()) | set(prod_map.keys())

    low = 0
    total_stock_qty = 0.0
    for p_k in all_specs:
        p_doc = prod_map.get(p_k) or {}
        op_stock = float(p_doc.get("opening_stock") or 0.0)
        bal = op_stock + in_map.get(p_k, 0.0) - out_map.get(p_k, 0.0)
        total_stock_qty += max(bal, 0.0)
        mn = float(p_doc.get("min_stock") or 5.0)
        if bal <= mn:
            low += 1

    return {
        "total_products": len(all_specs), "total_stock_qty": round(total_stock_qty, 2),
        "low_stock": low, "in_today": in_today, "out_today": out_today,
        "pending_requests": pending_req, "stock_value": 0,
    }

_local_rates_cache = None
def _load_local_rates() -> dict:
    global _local_rates_cache
    if _local_rates_cache is not None:
        return _local_rates_cache
    filepath = ROOT_DIR / "local_storage" / "product_rates.json"
    if not filepath.exists():
        return {}
    try:
        with open(filepath, "r") as f:
            _local_rates_cache = json.load(f)
            return _local_rates_cache
    except Exception:
        return {}

def _save_local_rate(product_name: str, rate: float):
    global _local_rates_cache
    filepath = ROOT_DIR / "local_storage" / "product_rates.json"
    rates = _load_local_rates()
    rates[product_name.strip().upper()] = rate
    _local_rates_cache = rates
    try:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w") as f:
            json.dump(rates, f)
    except Exception:
        pass

_local_assets_cache = None
def _load_local_assets() -> list:
    global _local_assets_cache
    if _local_assets_cache is not None:
        return _local_assets_cache
    filepath = ROOT_DIR / "local_storage" / "high_value_assets.json"
    if not filepath.exists():
        return []
    try:
        with open(filepath, "r") as f:
            _local_assets_cache = json.load(f)
            return _local_assets_cache
    except Exception:
        return []

def _save_local_assets(assets: list):
    global _local_assets_cache
    _local_assets_cache = assets
    filepath = ROOT_DIR / "local_storage" / "high_value_assets.json"
    try:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w") as f:
            json.dump(assets, f)
    except Exception:
        pass

_local_high_value_cache = None
def _load_local_high_value_products() -> dict:
    global _local_high_value_cache
    if _local_high_value_cache is not None:
        return _local_high_value_cache
    filepath = ROOT_DIR / "local_storage" / "product_high_value.json"
    if not filepath.exists():
        return {}
    try:
        with open(filepath, "r") as f:
            _local_high_value_cache = json.load(f)
            return _local_high_value_cache
    except Exception:
        return {}

def _save_local_high_value_product(product_name: str, is_high_value: bool):
    global _local_high_value_cache
    filepath = ROOT_DIR / "local_storage" / "product_high_value.json"
    data = _load_local_high_value_products()
    data[product_name.strip().upper()] = is_high_value
    _local_high_value_cache = data
    try:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, "w") as f:
            json.dump(data, f)
    except Exception:
        pass

_PRODUCTS_CACHE: Dict[str, Tuple[float, List[Dict[str, Any]]]] = {}
_PRODUCTS_CACHE_TTL_S = 60.0

# Separate ultra-lightweight cache for the dropdown/search endpoint.
# Only contains the 6 fields needed for product selection — NO aggregation at all.
_PRODUCTS_SEARCH_CACHE: Dict[str, Tuple[float, List[Dict[str, Any]]]] = {}
_PRODUCTS_SEARCH_CACHE_TTL_S = 300.0  # 5 min – refreshed on every product write

def invalidate_products_cache(company_id: Optional[str] = None):
    global _PRODUCTS_CACHE, _PRODUCTS_SEARCH_CACHE
    if company_id:
        _PRODUCTS_CACHE.pop(company_id, None)
        _PRODUCTS_SEARCH_CACHE.pop(company_id, None)
    else:
        _PRODUCTS_CACHE.clear()
        _PRODUCTS_SEARCH_CACHE.clear()



async def _compute_inventory_balances(cid: str):
    items = await db.products.find({"company_id": cid}, {"_id": 0}).sort("name", 1).to_list(10000)
    inward_entries = await db.inward_entries.find({"company_id": cid}, {"_id": 0}).to_list(100000)
    outward_entries = await db.outward_entries.find({"company_id": cid}, {"_id": 0}).to_list(100000)

    # Product Maps for resolution
    prod_id_map: Dict[str, Dict] = {}
    prod_key_map: Dict[Tuple[str, str], Dict] = {}
    prod_name_map: Dict[str, List[Dict]] = {}

    for p in items:
        p_name = norm_product_name(p.get("name"))
        p_size = norm_str(p.get("size"))
        if p.get("id"):
            prod_id_map[p["id"]] = p
        if p_name:
            key = (p_name, p_size)
            prod_key_map[key] = p
            prod_name_map.setdefault(p_name, []).append(p)

    in_map: Dict[Tuple[str, str], float] = {}
    out_map: Dict[Tuple[str, str], float] = {}
    ret_map: Dict[Tuple[str, str], float] = {}

    def _resolve_product(entry: Dict[str, Any]) -> Tuple[str, str]:
        # Priority 1: Match by product_id
        pid = entry.get("product_id")
        if pid and pid in prod_id_map:
            target = prod_id_map[pid]
            return (norm_product_name(target.get("name")), norm_str(target.get("size")))
        
        raw_pn = entry.get("product") or ""
        raw_ps = entry.get("size") or ""
        pn_n = norm_product_name(raw_pn)
        ps_n = norm_str(raw_ps)
        
        # Priority 2: Match by exact normalized (name, size)
        if (pn_n, ps_n) in prod_key_map:
            return (pn_n, ps_n)
            
        # Priority 3: Match by product name if single product match in master
        if pn_n in prod_name_map and len(prod_name_map[pn_n]) == 1:
            target = prod_name_map[pn_n][0]
            return (pn_n, norm_str(target.get("size")))

        return (pn_n, ps_n)

    # Process Inward Entries
    for ie in inward_entries:
        st = str(ie.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]:
            continue
            
        qty = float(ie.get("quantity") or 0.0)
        pk = _resolve_product(ie)
        in_map[pk] = in_map.get(pk, 0.0) + qty

        # Check for client return
        src_t = str(ie.get("source_type") or "").lower()
        src = str(ie.get("source") or "").lower()
        if "return" in src_t or "client-return" in src:
            ret_map[pk] = ret_map.get(pk, 0.0) + qty

    # Process Outward Entries
    for oe in outward_entries:
        st = str(oe.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]:
            continue
            
        qty = float(oe.get("quantity") or 0.0)
        pk = _resolve_product(oe)
        out_map[pk] = out_map.get(pk, 0.0) + qty

    local_rates = _load_local_rates()
    local_high_values = _load_local_high_value_products()

    for p in items:
        p_name = norm_product_name(p["name"])
        p_size = norm_str(p.get("size"))
        k = (p_name, p_size)

        op_stock = float(p.get("opening_stock") or 0.0)
        tot_in = round(in_map.get(k, 0.0), 2)
        tot_out = round(out_map.get(k, 0.0), 2)
        bal = round(op_stock + tot_in - tot_out, 2)

        p["opening_stock"] = op_stock
        p["total_in"] = tot_in
        p["total_out"] = tot_out
        p["returned"] = round(ret_map.get(k, 0.0), 2)
        p["balance"] = bal

        p["rate"] = local_rates.get(p_name, float(p.get("rate") or 0.0))
        if p_name in local_high_values:
            p["high_value_goods"] = bool(local_high_values[p_name])
        else:
            p["high_value_goods"] = bool(p.get("high_value_goods") or p.get("high_value_asset"))
            if p["high_value_goods"]:
                _save_local_high_value_product(p_name, True)

        mn = float(p.get("min_stock") or 0.0)
        if bal <= 0:
            p["stock_status"] = "Out Of Stock"
        elif bal <= mn:
            p["stock_status"] = "Low Stock"
        else:
            p["stock_status"] = "Normal"

    return items, in_map, out_map, ret_map

@api_router.get("/inventory/products")
async def list_products(user=Depends(get_current_user)):
    cid = user["company_id"]
    now = time.monotonic()
    if cid in _PRODUCTS_CACHE:
        cache_time, cached_items = _PRODUCTS_CACHE[cid]
        if now - cache_time < _PRODUCTS_CACHE_TTL_S:
            return cached_items

    items, _, _, _ = await _compute_inventory_balances(cid)

    hv_keywords = ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"]
    local_hv = _load_local_high_value_products()
    def _is_hv_prod(p):
        p_name = norm_product_name(p.get("name"))
        if p.get("high_value_goods") or p.get("high_value_asset") or local_hv.get(p_name):
            return True
        if any(kw in p_name for kw in hv_keywords):
            return True
        return False

    for p in items:
        is_hv = _is_hv_prod(p)
        p["high_value_goods"] = is_hv
        p["high_value_asset"] = is_hv

    items.sort(key=lambda p: (0 if p.get("high_value_goods") else 1, p.get("name") or "", p.get("size") or ""))
    _PRODUCTS_CACHE[cid] = (now, items)
    return items

@api_router.get("/inventory/high-value-ledger")
async def get_high_value_ledger(search: Optional[str] = None, user=Depends(get_current_user)):
    cid = user["company_id"]
    search_term = (search or "").strip().lower()
    
    items, in_map, out_map, ret_map = await _compute_inventory_balances(cid)
    local_hv = _load_local_high_value_products()
    
    # Product Maps for resolution
    prod_id_map: Dict[str, Dict] = {}
    prod_key_map: Dict[Tuple[str, str], Dict] = {}
    prod_name_map: Dict[str, List[Dict]] = {}

    for p in items:
        p_name = norm_product_name(p.get("name"))
        p_size = norm_str(p.get("size"))
        if p.get("id"):
            prod_id_map[p["id"]] = p
        if p_name:
            key = (p_name, p_size)
            prod_key_map[key] = p
            prod_name_map.setdefault(p_name, []).append(p)

    def _resolve_entry_product(entry: Dict[str, Any]) -> Tuple[str, str]:
        pid = entry.get("product_id")
        if pid and pid in prod_id_map:
            target = prod_id_map[pid]
            return (norm_product_name(target.get("name")), norm_str(target.get("size")))
        
        raw_pn = entry.get("product") or ""
        raw_ps = entry.get("size") or ""
        pn_n = norm_product_name(raw_pn)
        ps_n = norm_str(raw_ps)
        
        if (pn_n, ps_n) in prod_key_map:
            return (pn_n, ps_n)
            
        if pn_n in prod_name_map and len(prod_name_map[pn_n]) == 1:
            target = prod_name_map[pn_n][0]
            return (pn_n, norm_str(target.get("size")))

        return (pn_n, ps_n)

    hv_product_docs = []
    hv_keys = set()
    hv_names = set()
    hv_ids = set()
    
    all_inward_records = await db.inward_entries.find({"company_id": cid}, {"_id": 0}).sort("date", -1).to_list(10000)
    all_outward_records = await db.outward_entries.find({"company_id": cid}, {"_id": 0}).sort("date", -1).to_list(10000)

    hv_inward_product_names = set()
    for ie in all_inward_records:
        if ie.get("source") in ["high-value-manual-import", "bulk-inward-high-value"] or ie.get("high_value_goods") or ie.get("high_value_asset"):
            pn_n = norm_product_name(ie.get("product"))
            if pn_n:
                hv_inward_product_names.add(pn_n)

    for p in items:
        pn = (p.get("name") or "").strip()
        pn_norm = norm_product_name(pn)
        ps = (p.get("size") or "").strip()
        ps_norm = norm_str(ps)
        
        hv_keywords = ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"]
        is_hv = (
            local_hv.get(pn_norm, False) is True or
            bool(p.get("high_value_goods")) or
            bool(p.get("high_value_asset")) or
            pn_norm in hv_inward_product_names or
            any(kw in pn_norm for kw in hv_keywords)
        )
        if is_hv and pn_norm:
            if search_term and search_term not in pn.lower() and search_term not in ps.lower():
                continue
            hv_product_docs.append(p)
            hv_keys.add((pn_norm, ps_norm))
            hv_names.add(pn_norm)
            if p.get("id"):
                hv_ids.add(p["id"])

    def _is_entry_high_value(entry: Dict[str, Any]) -> bool:
        pid = entry.get("product_id")
        if pid and pid in hv_ids:
            return True
        if bool(entry.get("high_value_goods")) or bool(entry.get("high_value_asset")):
            return True
        if entry.get("source") in ["high-value-manual-import", "bulk-inward-high-value"]:
            return True
        pn_n = norm_product_name(entry.get("product"))
        if local_hv.get(pn_n, False) is True or pn_n in hv_inward_product_names:
            return True
        pk = _resolve_entry_product(entry)
        if pk in hv_keys or pk[0] in hv_names:
            return True
        return False

    last_movement_map = {}
    last_inward_info = {}

    for ie in all_inward_records:
        st = str(ie.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]:
            continue
        pk = _resolve_entry_product(ie)
        if pk not in last_inward_info:
            last_inward_info[pk] = {
                "date": ie.get("date") or "",
                "vendor": ie.get("source_name") or ie.get("vendor") or "Supplier",
                "challan": ie.get("reference_number") or ie.get("bill_number") or ""
            }
        if pk not in last_movement_map:
            date_str = (ie.get("date") or "")[:10]
            last_movement_map[pk] = f"Inward {date_str}" if date_str else "Inward"

    for oe in all_outward_records:
        st = str(oe.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]:
            continue
        pk = _resolve_entry_product(oe)
        date_str = (oe.get("date") or "")[:10]
        if pk not in last_movement_map or "Inward" in last_movement_map[pk]:
            last_movement_map[pk] = f"Outward {date_str}" if date_str else "Outward"

    all_goods = []
    available = []

    for p in hv_product_docs:
        pn = p.get("name") or ""
        pn_n = norm_product_name(pn)
        ps = p.get("size") or ""
        ps_n = norm_str(ps)
        pk = (pn_n, ps_n)
        
        tot_in = p.get("total_in", 0.0)
        tot_out = p.get("total_out", 0.0)
        ret_qty = p.get("returned", 0.0)
        avail = p.get("balance", 0.0)
        
        last_mov = last_movement_map.get(pk, "No Movement")
        in_info = last_inward_info.get(pk, {})
        status = "Available" if avail > 0 else "Out of Stock"
        
        row_all = {
            "id": p.get("id") or str(uuid.uuid4()),
            "product": pn,
            "size": ps,
            "unit": p.get("unit") or "Nos",
            "total_in": tot_in,
            "total_out": tot_out,
            "returned": ret_qty,
            "available_qty": avail,
            "minimum_stock": float(p.get("minimum_stock") or 0.0),
            "last_movement": last_mov,
            "status": status
        }
        all_goods.append(row_all)
        
        if avail > 0:
            parts = [in_info.get("challan"), in_info.get("vendor")]
            challan_vendor_str = " / ".join([str(p).strip() for p in parts if p and str(p).strip()])
            row_avail = {
                "id": p.get("id") or str(uuid.uuid4()),
                "product": pn,
                "size": ps,
                "unit": p.get("unit") or "Nos",
                "available_qty": avail,
                "last_inward": (in_info.get("date") or "")[:10],
                "challan_vendor": challan_vendor_str or "—",
                "status": "Available"
            }
            available.append(row_avail)

    client_map = {}
    raw_clients = await db.clients.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    for c in (raw_clients or []):
        if c.get("id"):
            client_map[c["id"]] = c
        if c.get("full_name"):
            client_map[c["full_name"].strip().lower()] = c

    dispatched = []
    for oe in all_outward_records:
        st = str(oe.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]:
            continue

        if not _is_entry_high_value(oe):
            continue

        if search_term and search_term not in (oe.get("product") or "").lower() and search_term not in (oe.get("size") or "").lower() and search_term not in (oe.get("client_name") or "").lower():
            continue
            
        c_info = client_map.get(oe.get("client_id")) or client_map.get(str(oe.get("source_name") or "").strip().lower()) or {}
        c_name = oe.get("client_name") or oe.get("source_name") or c_info.get("full_name") or "Direct Outward"
        site_val = c_info.get("city") or c_info.get("address") or oe.get("site_name") or "—"
        
        req_by = oe.get("requested_by") or oe.get("issued_to") or oe.get("created_by_name") or "Inventory Admin"
        ref_val = oe.get("reference_number") or oe.get("bill_number") or oe.get("remarks") or "Outward Entry"
        
        serials = oe.get("serial_numbers") or oe.get("serials") or ([oe.get("serial_number")] if oe.get("serial_number") else [])
        serials = [s.strip().upper() for s in serials if s and s.strip()]

        dispatched.append({
            "id": oe.get("id") or str(uuid.uuid4()),
            "date": (oe.get("date") or now_iso())[:10],
            "product": oe.get("product") or "",
            "size": oe.get("size") or "",
            "quantity": float(oe.get("quantity") or 0.0),
            "unit": oe.get("unit") or "Nos",
            "serial_numbers": serials,
            "challan_number": oe.get("outward_challan_no") or oe.get("reference_number") or oe.get("bill_number") or "—",
            "client_name": c_name,
            "site": site_val,
            "requested_by": req_by,
            "reference": ref_val,
            "status": oe.get("status") or "Dispatched"
        })

    returned = []
    for ie in all_inward_records:
        st = str(ie.get("status") or "").strip().lower()
        if st in ["cancelled", "draft_cancelled"]:
            continue

        src_t = str(ie.get("source_type") or "")
        src = str(ie.get("source") or "")
        if src_t != "Return From Client" and "client-return" not in src and "return" not in src_t.lower():
            continue
            
        if not _is_entry_high_value(ie):
            continue

        if search_term and search_term not in (ie.get("product") or "").lower() and search_term not in (ie.get("size") or "").lower() and search_term not in (ie.get("source_name") or "").lower():
            continue
            
        c_info = client_map.get(ie.get("client_id")) or client_map.get(str(ie.get("source_name") or "").strip().lower()) or {}
        c_name = ie.get("source_name") or c_info.get("full_name") or "Client"
        site_val = c_info.get("city") or c_info.get("address") or "—"
        
        returned.append({
            "id": ie.get("id") or str(uuid.uuid4()),
            "return_date": (ie.get("date") or now_iso())[:10],
            "product": ie.get("product") or "",
            "size": ie.get("size") or "",
            "quantity": float(ie.get("quantity") or 0.0),
            "unit": ie.get("unit") or "Nos",
            "client_name": c_name,
            "site": site_val,
            "original_challan": ie.get("reference_number") or ie.get("bill_number") or "—",
            "return_reason": ie.get("remarks") or "Material Returned From Client",
            "status": "Returned"
        })

    # Sort all_goods and available Ascending A-Z by Product Name
    all_goods.sort(key=lambda x: (x.get("product") or "").lower())
    available.sort(key=lambda x: (x.get("product") or "").lower())

    return {
        "all_goods": all_goods,
        "available": available,
        "dispatched": dispatched,
        "returned": returned
    }

@api_router.get("/inventory/available-serials")
async def get_available_serials(product: Optional[str] = None, size: Optional[str] = None, user=Depends(get_current_user)):
    cid = user["company_id"]
    all_assets = await db.high_value_assets.find({"company_id": cid}).to_list(10000)
    p_norm = norm_product_name(product) if product else ""
    s_norm = norm_str(size) if size else ""

    available = []
    for a in all_assets:
        if a.get("company_id") == cid and a.get("status") == "Available":
            sn = (a.get("serial_number") or "").strip().upper()
            if not sn:
                continue
            if p_norm and norm_product_name(a.get("product_name")) != p_norm:
                continue
            if s_norm and norm_str(a.get("size_model")) != s_norm:
                continue
            available.append(sn)

    unique_serials = sorted(list(set(available)))
    return {"serials": unique_serials}

@api_router.post("/inventory/products")
async def create_product(data: ProductIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.create")
    name = norm_product_name(data.name)
    size = norm_str(data.size)
    unit = norm_unit(data.unit)
    if not name:
        raise HTTPException(status_code=400, detail="Product name required")
    existing = await db.products.find_one({"company_id": user["company_id"], "name": name, "size": size})
    if existing:
        raise HTTPException(status_code=400, detail="Product with this name and size specification already exists")
    rate_val = data.rate or 0.0
    _save_local_rate(name, rate_val)
    _save_local_high_value_product(name, data.high_value_goods or False)
    doc = {
        "id": str(uuid.uuid4()), "company_id": user["company_id"], "name": name,
        "size": size, "category": data.category or "Solar",
        "brand": data.brand or "",
        "sku": data.sku or data.code or "",
        "unit": unit or "Nos", "min_stock": data.min_stock or 0.0,
        "opening_stock": data.opening_stock or 0.0,
        "rate": rate_val,
        "status": data.status or "Active", "created_at": now_iso(),
        "high_value_goods": data.high_value_goods or False,
        "serial_number_required": data.serial_number_required or False,
    }
    await db.products.insert_one(doc); doc.pop("_id", None)
    doc["high_value_goods"] = data.high_value_goods or False
    doc["serial_number_required"] = data.serial_number_required or False
    invalidate_products_cache(user["company_id"])
    await log_activity(user["company_id"], user["id"], user["name"], "Product Created", f"{name} ({size})" if size else name)
    return doc

@api_router.patch("/inventory/products/{product_id}")
async def update_product(product_id: str, data: ProductIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.edit")
    cid = user["company_id"]
    existing = await db.products.find_one({"id": product_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    new_name = norm_product_name(data.name) if data.name else existing["name"]
    new_size = norm_str(data.size) if data.size is not None else norm_str(existing.get("size", ""))
    new_unit = norm_unit(data.unit) if data.unit is not None else norm_unit(existing.get("unit", "Nos"))
    if new_name != existing["name"] or new_size != norm_str(existing.get("size", "")):
        dup = await db.products.find_one({"company_id": cid, "name": new_name, "size": new_size})
        if dup and dup["id"] != product_id:
            raise HTTPException(status_code=400, detail="Another product with this name and size specification already exists")
        # cascade rename in inward/outward entries and link product_id
        old_pname = existing["name"]
        old_psize = existing.get("size", "")
        await db.inward_entries.update_many({"company_id": cid, "$or": [{"product_id": product_id}, {"product": old_pname, "size": old_psize}]}, {"$set": {"product": new_name, "size": new_size, "unit": new_unit, "product_id": product_id}})
        await db.outward_entries.update_many({"company_id": cid, "$or": [{"product_id": product_id}, {"product": old_pname, "size": old_psize}]}, {"$set": {"product": new_name, "size": new_size, "unit": new_unit, "product_id": product_id}})
    rate_val = data.rate or 0.0
    _save_local_rate(new_name, rate_val)
    if data.high_value_goods is not None:
        _save_local_high_value_product(new_name, data.high_value_goods)
    patch = {
        "name": new_name, "size": new_size, "category": data.category or "",
        "brand": data.brand if data.brand is not None else existing.get("brand", ""),
        "sku": data.sku or data.code if (data.sku or data.code) else existing.get("sku", existing.get("code", "")),
        "unit": new_unit or "Nos", "min_stock": float(data.min_stock or 0),
        "opening_stock": float(data.opening_stock if data.opening_stock is not None else existing.get("opening_stock", 0.0)),
        "rate": rate_val,
        "status": data.status or existing.get("status") or "Active",
        "high_value_goods": data.high_value_goods if data.high_value_goods is not None else existing.get("high_value_goods", False),
        "serial_number_required": data.serial_number_required if data.serial_number_required is not None else existing.get("serial_number_required", False),
        "updated_at": now_iso(),
    }
    await db.products.update_one({"id": product_id, "company_id": cid}, {"$set": patch})
    invalidate_products_cache(cid)
    await log_activity(cid, user["id"], user["name"], "Product Updated", f"{new_name} ({new_size})" if new_size else new_name)
    res = await db.products.find_one({"id": product_id, "company_id": cid}, {"_id": 0})
    if res:
        res["high_value_goods"] = _load_local_high_value_products().get(new_name, res.get("high_value_goods", False))
    return res

@api_router.delete("/inventory/products/{product_id}")
async def delete_product(product_id: str, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "delete"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.delete")
    cid = user["company_id"]
    existing = await db.products.find_one({"id": product_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")

    pname = existing["name"]
    psize = existing.get("size", "")
    punit = existing.get("unit", "Nos")

    in_count = await db.inward_entries.count_documents({"company_id": cid, "product": pname, "size": psize, "unit": punit})
    out_count = await db.outward_entries.count_documents({"company_id": cid, "product": pname, "size": psize, "unit": punit})

    total_refs = in_count + out_count

    if total_refs > 0:
        await db.products.update_one({"id": product_id, "company_id": cid}, {"$set": {"status": "Archived", "updated_at": now_iso()}})
        invalidate_products_cache(cid)
        await log_activity(cid, user["id"], user["name"], "Product Archived", f"{pname} ({psize})" if psize else pname)
        return {"ok": True, "action": "archived", "message": f"Product archived as {total_refs} historical transactions reference it."}

    await db.products.delete_one({"id": product_id, "company_id": cid})
    invalidate_products_cache(cid)
    await log_activity(cid, user["id"], user["name"], "Product Deleted", f"{pname} ({psize})" if psize else pname)
    return {"ok": True, "action": "deleted", "message": "Product permanently deleted."}

def parse_inward_client_info(entry):
    if not entry:
        return entry
    r = entry.get("remarks") or ""
    cid = ""
    if "[client_id:" in r:
        import re
        m = re.search(r"\[client_id:([^\]]+)\]", r)
        if m:
            cid = m.group(1)
            entry["remarks"] = re.sub(r"\s*\[client_id:[^\]]+\]", "", r).strip()
    entry["client_id"] = cid
    entry["client_name"] = entry.get("source_name") if entry.get("source_type") == "Return From Client" else ""
    entry["challan_number"] = entry.get("reference_number") or ""
    entry["challan_no"] = entry.get("reference_number") or ""
    return entry

def _enrich_inward_with_assets(inward_doc: Optional[dict]) -> Optional[dict]:
    if not inward_doc:
        return inward_doc
    assets = _load_local_assets()
    entry_assets = [a for a in assets if a.get("inward_entry_id") == inward_doc.get("id")]
    p_name = (inward_doc.get("product") or "").strip().upper()
    is_hv = _load_local_high_value_products().get(p_name, False)
    
    if entry_assets:
        inward_doc["high_value_asset"] = True
        inward_doc["high_value_goods"] = True
        inward_doc["serial_numbers"] = [a["serial_number"] for a in entry_assets]
    else:
        inward_doc["high_value_asset"] = is_hv
        inward_doc["high_value_goods"] = is_hv
        inward_doc["serial_numbers"] = []
    return inward_doc

def _enrich_outward_with_assets(outward_doc: Optional[dict]) -> Optional[dict]:
    if not outward_doc:
        return outward_doc
    assets = _load_local_assets()
    entry_assets = [a for a in assets if a.get("outward_entry_id") == outward_doc.get("id")]
    p_name = (outward_doc.get("product") or "").strip().upper()
    is_hv = _load_local_high_value_products().get(p_name, False)
    
    if entry_assets:
        outward_doc["high_value_asset"] = True
        outward_doc["high_value_goods"] = True
        outward_doc["serial_numbers"] = [a["serial_number"] for a in entry_assets]
        outward_doc["installation_notes"] = entry_assets[0].get("installation_notes") or ""
        outward_doc["warranty_start_date"] = entry_assets[0].get("warranty_start_date") or ""
        outward_doc["asset_remarks"] = entry_assets[0].get("asset_remarks") or ""
    else:
        outward_doc["high_value_asset"] = is_hv
        outward_doc["high_value_goods"] = is_hv
        outward_doc["serial_numbers"] = []
        outward_doc["installation_notes"] = ""
        outward_doc["warranty_start_date"] = ""
        outward_doc["asset_remarks"] = ""
    return outward_doc

async def _enrich_bill_challan(bill: Dict[str, Any], company_id: str) -> Dict[str, Any]:
    if not isinstance(bill, dict):
        return bill
    ch = (bill.get("challan_number") or bill.get("challan_no") or bill.get("reference_number") or "").strip()
    if not ch:
        b_id = bill.get("id")
        b_num = bill.get("bill_number")
        v_inw = None
        if b_id:
            v_inw = await db.vendor_inwards.find_one({"bill_id": b_id, "company_id": company_id})
        if not v_inw and b_num:
            v_inw = await db.vendor_inwards.find_one({"bill_number": b_num, "company_id": company_id})
        if v_inw and (v_inw.get("challan_number") or v_inw.get("challan_no") or v_inw.get("reference_number")):
            ch = (v_inw.get("challan_number") or v_inw.get("challan_no") or v_inw.get("reference_number") or "").strip()

        if not ch and b_num:
            c_inw = await db.inward_entries.find_one({"bill_number": b_num, "company_id": company_id})
            if c_inw and (c_inw.get("reference_number") or c_inw.get("challan_no") or c_inw.get("challan_number")):
                ch = (c_inw.get("reference_number") or c_inw.get("challan_no") or c_inw.get("challan_number") or "").strip()

    if ch:
        bill["challan_number"] = ch
        bill["challan_no"] = ch
        bill["reference_number"] = ch
    return bill

async def save_inward_entry_logic(data: InwardIn, company_id: str, user_id: str, user_name: str, source: str = "manual", import_batch: str = "", skip_activity_log: bool = False):
    source_type_val = data.source_type or "Vendor / Supplier"
    source_name_val = data.source_name or ""
    client_id_val = data.client_id or ""
    client_name_val = data.client_name or ""

    # Vendor Purchase Bill Flow: Save/upsert purchase bill in db.purchase_bills if Vendor/Supplier is specified
    vendor_id_val = data.vendor_id or data.source_id or ""
    if not vendor_id_val and source_name_val and source_type_val in ("Vendor / Supplier", "Supplier", "Vendor"):
        v_doc = await db.vendors.find_one({"company_id": company_id, "name": {"$regex": f"^{re.escape(source_name_val.strip())}$", "$options": "i"}})
        if v_doc:
            vendor_id_val = v_doc["id"]

    challan_val = (data.reference_number or getattr(data, "challan_no", None) or getattr(data, "challan_number", None) or "").strip()
    raw_bill_num = (data.bill_number or "").strip()

    if source_type_val in ("Vendor / Supplier", "Supplier", "Vendor") and (vendor_id_val or source_name_val):
        existing_pbill = None
        if raw_bill_num:
            existing_pbill = await db.purchase_bills.find_one({
                "company_id": company_id,
                "bill_number": raw_bill_num
            })
        if not existing_pbill and challan_val and vendor_id_val:
            existing_pbill = await db.purchase_bills.find_one({
                "company_id": company_id,
                "vendor_id": vendor_id_val,
                "challan_number": challan_val
            })

        if existing_pbill:
            if challan_val:
                await db.purchase_bills.update_one(
                    {"id": existing_pbill["id"], "company_id": company_id},
                    {"$set": {
                        "challan_number": challan_val,
                        "challan_no": challan_val,
                        "reference_number": challan_val,
                        "updated_at": now_iso()
                    }}
                )
        elif raw_bill_num or float(data.total_amount or 0.0) > 0 or challan_val:
            pbill_id = f"pbill_{uuid.uuid4().hex[:12]}"
            pbill_doc = {
                "id": pbill_id,
                "company_id": company_id,
                "vendor_id": vendor_id_val,
                "vendor_name": source_name_val or "Unknown Vendor",
                "bill_number": raw_bill_num or (f"BILL-{challan_val}" if challan_val else f"INV-{uuid.uuid4().hex[:6]}"),
                "challan_number": challan_val,
                "challan_no": challan_val,
                "reference_number": challan_val,
                "bill_date": (data.date or now_iso())[:10],
                "due_date": "",
                "items": [],
                "subtotal": float(data.total_amount or 0.0),
                "gst_total": 0.0,
                "grand_total": float(data.total_amount or 0.0),
                "payment_status": data.payment_status or "Unpaid",
                "status": data.payment_status or "Unpaid",
                "inward_status": "Received",
                "bill_type": data.bill_type or "Product Bill",
                "notes": data.remarks or "",
                "remarks": data.remarks or "",
                "attachment_file_id": data.attachment_file_id or "",
                "attachment_filename": data.attachment_filename or "",
                "paid_amount": float(data.total_amount or 0.0) if data.payment_status == "Paid" else 0.0,
                "created_by": user_name,
                "created_at": now_iso(),
                "updated_at": now_iso()
            }
            await db.purchase_bills.insert_one(pbill_doc)
            if not skip_activity_log:
                await log_activity(company_id, user_id, user_name, "Created Purchase Bill", f"Bill: {pbill_doc['bill_number']} Amount: ₹{pbill_doc['grand_total']} Vendor: {pbill_doc['vendor_name']}")

    # Material Stock Inward Entry
    pn = data.product.strip().upper()
    is_hv = data.high_value_asset or data.high_value_goods or _load_local_high_value_products().get(pn, False) or any(kw in pn for kw in ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"])
    await ensure_product(company_id, pn, size=data.size or "", unit=data.unit or "Nos", brand=data.source_name or "", high_value_goods=is_hv)
    
    # Client ID resolution from name case-insensitively
    if source_type_val == "Return From Client":
        if client_name_val and not client_id_val:
            client = await db.clients.find_one({
                "company_id": company_id,
                "full_name": {"$regex": f"^{re.escape(client_name_val)}$", "$options": "i"}
            })
            if client:
                client_id_val = client["id"]
                client_name_val = client["full_name"]
        elif client_id_val and not client_name_val:
            client = await db.clients.find_one({"company_id": company_id, "id": client_id_val})
            if client:
                client_name_val = client["full_name"]
        
        # Inward Return From Client stores client name in source_name
        if client_name_val:
            source_name_val = client_name_val

    remarks_val = data.remarks or ""
    if source_type_val == "Return From Client" and client_id_val:
        remarks_val = f"{remarks_val} [client_id:{client_id_val}]".strip()
        
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "product": pn,
        "size": data.size or "",
        "quantity": data.quantity,
        "unit": data.unit or "Nos",
        "reference_number": challan_val or numeric_only(data.reference_number or data.challan_no),
        "challan_no": challan_val or numeric_only(data.reference_number or data.challan_no),
        "challan_number": challan_val or numeric_only(data.reference_number or data.challan_no),
        "reference_type": data.reference_type or "Challan Number",
        "bill_number": (data.bill_number or "").strip() or numeric_only(data.bill_number),
        "source_type": source_type_val,
        "source_name": source_name_val,
        "date": data.date or now_iso(),
        "remarks": remarks_val,
        "attachment_file_id": data.attachment_file_id or "",
        "attachment_filename": data.attachment_filename or "",
        "source": source,
        "unit_price": float(data.unit_price or 0.0) if data.bill_type != "Amount Bill" else 0.0,
        "line_total": float(data.line_total or (data.quantity * float(data.unit_price or 0.0))) if data.bill_type != "Amount Bill" else 0.0,
        "total_amount": float(data.total_amount or 0.0),
        "payment_status": data.payment_status or "Unpaid",
        "vendor_id": vendor_id_val,
        "bill_type": data.bill_type or "Product Bill",
        "created_by": user_id,
        "created_by_name": user_name,
        "created_at": now_iso()
    }
    if import_batch:
        doc["import_batch"] = import_batch
        
    await db.inward_entries.insert_one(doc)
    doc.pop("_id", None)
    
    is_hv = data.high_value_asset or data.high_value_goods or _load_local_high_value_products().get(pn, False) or any(kw in pn for kw in ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"])
    if is_hv:
        all_assets = _load_local_assets()
        qty = float(data.quantity or 0)
        sns = [sn.strip().upper() for sn in (data.serial_numbers or []) if sn.strip()]
        
        if sns:
            for sn in sns:
                asset_id = str(uuid.uuid4())
                asset_doc = {
                    "id": asset_id,
                    "company_id": company_id,
                    "inward_entry_id": doc["id"],
                    "product_name": pn,
                    "brand": source_name_val or "Unknown",
                    "size_model": data.size or "",
                    "quantity": 1.0,
                    "serial_number": sn,
                    "vendor": source_name_val or "",
                    "purchase_date": (data.date or now_iso())[:10],
                    "challan_number": data.reference_number or "",
                    "client_id": None,
                    "client_name": None,
                    "installation_date": None,
                    "warranty_status": "Active",
                    "status": "Available",
                    "created_at": now_iso()
                }
                all_assets.append(asset_doc)
        else:
            asset_id = str(uuid.uuid4())
            asset_doc = {
                "id": asset_id,
                "company_id": company_id,
                "inward_entry_id": doc["id"],
                "product_name": pn,
                "brand": source_name_val or "Unknown",
                "size_model": data.size or "",
                "quantity": qty,
                "serial_number": "",
                "vendor": source_name_val or "",
                "purchase_date": (data.date or now_iso())[:10],
                "challan_number": data.reference_number or "",
                "client_id": None,
                "client_name": None,
                "installation_date": None,
                "warranty_status": "Active",
                "status": "Available",
                "created_at": now_iso()
            }
            all_assets.append(asset_doc)
        _save_local_assets(all_assets)
        
    if not skip_activity_log:
        await log_activity(company_id, user_id, user_name, "Inward Entry", f"{pn} × {data.quantity}")
    invalidate_products_cache(company_id)
    return doc

async def save_outward_entry_logic(data: OutwardIn, company_id: str, user_id: str, user_name: str, source: str = "manual", import_batch: str = ""):
    pn = data.product.strip().upper()
    await ensure_product(company_id, pn, size=data.size or "", unit=data.unit or "Nos")
    
    client_id_val = data.client_id or ""
    client_name_val = data.client_name or ""
    project_id_val = data.project_id or ""
    project_name_val = data.project_name or ""
    
    # Client ID and Name resolution case-insensitively
    if client_name_val and not client_id_val:
        client = await db.clients.find_one({
            "company_id": company_id,
            "full_name": {"$regex": f"^{re.escape(client_name_val)}$", "$options": "i"}
        })
        if client:
            client_id_val = client["id"]
            client_name_val = client["full_name"]
    elif client_id_val and not client_name_val:
        client = await db.clients.find_one({"company_id": company_id, "id": client_id_val})
        if client:
            client_name_val = client["full_name"]

    # Align project ID and project Name with client if empty or missing (same as Normal UI entry)
    if client_id_val:
        if not project_id_val:
            project_id_val = client_id_val
        if not project_name_val:
            project_name_val = client_name_val

    doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "product": pn,
        "size": data.size or "",
        "quantity": data.quantity,
        "unit": data.unit or "Nos",
        "client_id": client_id_val,
        "client_name": client_name_val,
        "project_id": project_id_val,
        "project_name": project_name_val,
        "outward_challan_no": numeric_only(data.outward_challan_no),
        "reference_number": numeric_only(data.reference_number or data.outward_challan_no),
        "reference_type": data.reference_type or "Challan Number",
        "date": data.date or now_iso(),
        "remarks": data.remarks or "",
        "status": data.status or "Dispatched",
        "attachment_file_id": data.attachment_file_id or "",
        "attachment_filename": data.attachment_filename or "",
        "source": source,
        "created_by": user_id,
        "created_by_name": user_name,
        "created_at": now_iso()
    }
    if import_batch:
        doc["import_batch"] = import_batch
        
    await db.outward_entries.insert_one(doc)
    doc.pop("_id", None)
    
    is_hv = data.high_value_asset or data.high_value_goods or _load_local_high_value_products().get(pn, False) or any(kw in pn for kw in ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"])
    if is_hv:
        all_assets = _load_local_assets()
        qty = float(data.quantity or 0)
        sns = [sn.strip().upper() for sn in (data.serial_numbers or []) if sn.strip()]
        
        available = [a for a in all_assets if a.get("product_name") == pn and a.get("status") == "Available" and a.get("company_id") == company_id]
        
        status_val = "Installed"
        client_id_val_asset = client_id_val or None
        client_name_val_asset = client_name_val or None
        outward_date_val = (data.date or now_iso())[:10]
        challan_val = data.reference_number or data.outward_challan_no or ""
        
        if sns:
            for sn in sns:
                matched = next((a for a in available if a.get("serial_number") == sn), None)
                if matched:
                    matched["status"] = status_val
                    matched["outward_entry_id"] = doc["id"]
                    matched["client_id"] = client_id_val_asset
                    matched["client_name"] = client_name_val_asset
                    matched["outward_date"] = outward_date_val
                    matched["challan_number"] = challan_val
                    matched["installation_date"] = outward_date_val
                    available.remove(matched)
                else:
                    no_sn_avail = next((a for a in available if not a.get("serial_number")), None)
                    if no_sn_avail:
                        no_sn_avail["quantity"] = float(no_sn_avail.get("quantity") or 1.0) - 1.0
                        if no_sn_avail["quantity"] <= 0:
                            if no_sn_avail in all_assets:
                                all_assets.remove(no_sn_avail)
                            if no_sn_avail in available:
                                available.remove(no_sn_avail)
                        new_asset = {
                            "id": str(uuid.uuid4()),
                            "company_id": company_id,
                            "inward_entry_id": no_sn_avail.get("inward_entry_id"),
                            "product_name": pn,
                            "brand": no_sn_avail.get("brand", "Unknown"),
                            "size_model": data.size or no_sn_avail.get("size_model", ""),
                            "quantity": 1.0,
                            "serial_number": sn,
                            "vendor": no_sn_avail.get("vendor", ""),
                            "purchase_date": no_sn_avail.get("purchase_date", outward_date_val),
                            "challan_number": challan_val,
                            "client_id": client_id_val_asset,
                            "client_name": client_name_val_asset,
                            "installation_date": outward_date_val,
                            "warranty_status": "Active",
                            "status": status_val,
                            "outward_entry_id": doc["id"],
                            "outward_date": outward_date_val,
                            "created_at": now_iso()
                        }
                        all_assets.append(new_asset)
                    else:
                        new_asset = {
                            "id": str(uuid.uuid4()),
                            "company_id": company_id,
                            "inward_entry_id": None,
                            "product_name": pn,
                            "brand": "Unknown",
                            "size_model": data.size or "",
                            "quantity": 1.0,
                            "serial_number": sn,
                            "vendor": "",
                            "purchase_date": outward_date_val,
                            "challan_number": challan_val,
                            "client_id": client_id_val_asset,
                            "client_name": client_name_val_asset,
                            "installation_date": outward_date_val,
                            "warranty_status": "Active",
                            "status": status_val,
                            "outward_entry_id": doc["id"],
                            "outward_date": outward_date_val,
                            "created_at": now_iso()
                        }
                        all_assets.append(new_asset)
        else:
            no_sn_avail = next((a for a in available if not a.get("serial_number")), None)
            if no_sn_avail:
                avail_qty = float(no_sn_avail.get("quantity") or 0)
                if avail_qty >= qty:
                    no_sn_avail["quantity"] = avail_qty - qty
                    if no_sn_avail["quantity"] <= 0:
                        if no_sn_avail in all_assets:
                            all_assets.remove(no_sn_avail)
                    new_asset = {
                        "id": str(uuid.uuid4()),
                        "company_id": company_id,
                        "inward_entry_id": no_sn_avail.get("inward_entry_id"),
                        "product_name": pn,
                        "brand": no_sn_avail.get("brand", "Unknown"),
                        "size_model": data.size or no_sn_avail.get("size_model", ""),
                        "quantity": qty,
                        "serial_number": "",
                        "vendor": no_sn_avail.get("vendor", ""),
                        "purchase_date": no_sn_avail.get("purchase_date", outward_date_val),
                        "challan_number": challan_val,
                        "client_id": client_id_val_asset,
                        "client_name": client_name_val_asset,
                        "installation_date": outward_date_val,
                        "warranty_status": "Active",
                        "status": status_val,
                        "outward_entry_id": doc["id"],
                        "outward_date": outward_date_val,
                        "created_at": now_iso()
                    }
                    all_assets.append(new_asset)
                else:
                    if no_sn_avail in all_assets:
                        all_assets.remove(no_sn_avail)
                    new_asset = {
                        "id": str(uuid.uuid4()),
                        "company_id": company_id,
                        "inward_entry_id": no_sn_avail.get("inward_entry_id"),
                        "product_name": pn,
                        "brand": no_sn_avail.get("brand", "Unknown"),
                        "size_model": data.size or no_sn_avail.get("size_model", ""),
                        "quantity": qty,
                        "serial_number": "",
                        "vendor": no_sn_avail.get("vendor", ""),
                        "purchase_date": no_sn_avail.get("purchase_date", outward_date_val),
                        "challan_number": challan_val,
                        "client_id": client_id_val_asset,
                        "client_name": client_name_val_asset,
                        "installation_date": outward_date_val,
                        "warranty_status": "Active",
                        "status": status_val,
                        "outward_entry_id": doc["id"],
                        "outward_date": outward_date_val,
                        "created_at": now_iso()
                    }
                    all_assets.append(new_asset)
            else:
                new_asset = {
                    "id": str(uuid.uuid4()),
                    "company_id": company_id,
                    "inward_entry_id": None,
                    "product_name": pn,
                    "brand": "Unknown",
                    "size_model": data.size or "",
                    "quantity": qty,
                    "serial_number": "",
                    "vendor": "",
                    "purchase_date": outward_date_val,
                    "challan_number": challan_val,
                    "client_id": client_id_val_asset,
                    "client_name": client_name_val_asset,
                    "installation_date": outward_date_val,
                    "warranty_status": "Active",
                    "status": status_val,
                    "outward_entry_id": doc["id"],
                    "outward_date": outward_date_val,
                    "created_at": now_iso()
                }
                all_assets.append(new_asset)
        for a in all_assets:
            if a.get("outward_entry_id") == doc["id"]:
                a["installation_notes"] = data.installation_notes or ""
                a["warranty_start_date"] = data.warranty_start_date or ""
                a["asset_remarks"] = data.asset_remarks or ""
        _save_local_assets(all_assets)
        
    await log_activity(company_id, user_id, user_name, "Outward Entry", f"{pn} × {data.quantity}")
    invalidate_products_cache(company_id)
    return doc

@api_router.post("/inventory/inward")
async def add_inward(data: InwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.create")
    cid = user.get("company_id") or "COMP-001"
    uid = user.get("id") or user.get("sub") or "user"
    uname = user.get("name") or user.get("full_name") or "User"
    doc = await save_inward_entry_logic(data, cid, uid, uname, source="manual")
    return _enrich_inward_with_assets(parse_inward_client_info(doc))


@api_router.get("/inventory/inward")
async def list_inward(user=Depends(get_current_user)):
    cid = user.get("company_id") or "COMP-001"
    entries = await db.inward_entries.find({"company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [_enrich_inward_with_assets(parse_inward_client_info(e)) for e in entries]

@api_router.patch("/inventory/inward/{entry_id}")
async def update_inward(entry_id: str, data: InwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.edit")
    cid = user["company_id"]
    existing = await db.inward_entries.find_one({"id": entry_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Inward entry not found")
    pn = (data.product or existing["product"]).strip().upper()
    await ensure_product(cid, pn, size=data.size or "", unit=data.unit or existing.get("unit") or "Nos", brand=data.source_name or "")
    
    remarks_val = data.remarks or ""
    source_type_val = data.source_type or existing.get("source_type") or "Supplier"
    client_id_val = data.client_id or ""
    if source_type_val == "Return From Client" and client_id_val:
        remarks_val = f"{remarks_val} [client_id:{client_id_val}]".strip()
        
    raw_ch = (data.reference_number or getattr(data, "challan_no", None) or getattr(data, "challan_number", None) or "").strip()
    ref_num = raw_ch or numeric_only(data.reference_number or getattr(data, "challan_no", None) or getattr(data, "challan_number", None))
    patch = {
        "product": pn, "size": data.size or "", "quantity": data.quantity,
        "unit": data.unit or existing.get("unit") or "Nos",
        "reference_number": ref_num,
        "challan_no": ref_num,
        "challan_number": ref_num,
        "reference_type": data.reference_type or "Challan Number",
        "bill_number": (data.bill_number or "").strip() or numeric_only(data.bill_number),
        "source_type": source_type_val, "source_name": data.source_name or existing.get("source_name") or "",
        "date": data.date or existing.get("date") or now_iso(), "remarks": remarks_val,
        "attachment_file_id": data.attachment_file_id if data.attachment_file_id is not None else existing.get("attachment_file_id", ""),
        "attachment_filename": data.attachment_filename if data.attachment_filename is not None else existing.get("attachment_filename", ""),
        "updated_at": now_iso(),
    }
    await db.inward_entries.update_one({"id": entry_id, "company_id": cid}, {"$set": patch})

    # Sync Challan No. to linked Purchase Bills if any
    b_num = patch.get("bill_number") or existing.get("bill_number")
    if b_num and ref_num:
        await db.purchase_bills.update_many(
            {"bill_number": b_num, "company_id": cid},
            {"$set": {"challan_number": ref_num, "challan_no": ref_num, "reference_number": ref_num, "updated_at": now_iso()}}
        )
    
    # Recreate high value assets for this inward entry
    all_assets = _load_local_assets()
    non_inward_assets = [a for a in all_assets if a.get("inward_entry_id") != entry_id or a.get("status") == "Installed"]
    
    is_hv = data.high_value_asset or data.high_value_goods or _load_local_high_value_products().get(pn, False) or any(kw in pn for kw in ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"])
    if is_hv:
        new_assets = []
        qty = float(data.quantity or 0)
        sns = [sn.strip().upper() for sn in (data.serial_numbers or []) if sn.strip()]
        installed_inward_assets = [a for a in all_assets if a.get("inward_entry_id") == entry_id and a.get("status") == "Installed"]
        
        installed_qty = sum(float(a.get("quantity") or 1.0) for a in installed_inward_assets)
        new_needed = max(0.0, qty - installed_qty)
        
        new_assets.extend(installed_inward_assets)
        
        if new_needed > 0:
            if sns:
                for sn in sns:
                    if any(a.get("serial_number") == sn for a in installed_inward_assets):
                        continue
                    asset_id = str(uuid.uuid4())
                    asset_doc = {
                        "id": asset_id,
                        "company_id": cid,
                        "inward_entry_id": entry_id,
                        "product_name": pn,
                        "brand": data.source_name or "Unknown",
                        "size_model": data.size or "",
                        "quantity": 1.0,
                        "serial_number": sn,
                        "vendor": data.source_name or "",
                        "purchase_date": patch["date"][:10],
                        "challan_number": patch["reference_number"],
                        "client_id": None,
                        "client_name": None,
                        "installation_date": None,
                        "warranty_status": "Active",
                        "status": "Available",
                        "created_at": now_iso()
                    }
                    new_assets.append(asset_doc)
            else:
                asset_id = str(uuid.uuid4())
                asset_doc = {
                    "id": asset_id,
                    "company_id": cid,
                    "inward_entry_id": entry_id,
                    "product_name": pn,
                    "brand": data.source_name or "Unknown",
                    "size_model": data.size or "",
                    "quantity": new_needed,
                    "serial_number": "",
                    "vendor": data.source_name or "",
                    "purchase_date": patch["date"][:10],
                    "challan_number": patch["reference_number"],
                    "client_id": None,
                    "client_name": None,
                    "installation_date": None,
                    "warranty_status": "Active",
                    "status": "Available",
                    "created_at": now_iso()
                }
                new_assets.append(asset_doc)
        _save_local_assets(non_inward_assets + new_assets)
    else:
        _save_local_assets(non_inward_assets)
        
    await log_activity(cid, user["id"], user["name"], "Inward Updated", f"{pn} × {data.quantity}")
    await sync_inventory_master(cid)
    invalidate_products_cache(cid)
    res = await db.inward_entries.find_one({"id": entry_id, "company_id": cid}, {"_id": 0})
    return _enrich_inward_with_assets(parse_inward_client_info(res))

@api_router.delete("/inventory/inward/{entry_id}")
async def delete_inward(entry_id: str, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "delete"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.delete")
    cid = user["company_id"]
    existing = await db.inward_entries.find_one({"id": entry_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Inward entry not found")
    await db.inward_entries.delete_one({"id": entry_id, "company_id": cid})
    
    # Remove associated assets that are not installed
    all_assets = _load_local_assets()
    filtered_assets = [a for a in all_assets if a.get("inward_entry_id") != entry_id or a.get("status") == "Installed"]
    _save_local_assets(filtered_assets)
    
    await log_activity(cid, user["id"], user["name"], "Inward Deleted", f"{existing.get('product')} × {existing.get('quantity')}")
    await sync_inventory_master(cid)
    invalidate_products_cache(cid)
    return {"ok": True}

@api_router.post("/inventory/outward")
async def add_outward(data: OutwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.create")
    user_name = user.get("name") or user.get("full_name") or "User"
    doc = await save_outward_entry_logic(data, user.get("company_id") or "COMP-001", user.get("id") or "user", user_name, source="manual")
    return _enrich_outward_with_assets(doc)

@api_router.get("/inventory/outward")
async def list_outward(user=Depends(get_current_user), status: Optional[str] = None):
    cid = user.get("company_id") or "COMP-001"
    q: Dict[str, Any] = {"company_id": cid}
    if status:
        q["status"] = status
    entries = await db.outward_entries.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [_enrich_outward_with_assets(e) for e in entries]

@api_router.patch("/inventory/outward/{entry_id}")
async def update_outward(entry_id: str, data: OutwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.edit")
    cid = user["company_id"]
    existing = await db.outward_entries.find_one({"id": entry_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Outward entry not found")
    pn = (data.product or existing["product"]).strip().upper()
    await ensure_product(cid, pn, size=data.size or "", unit=data.unit or existing.get("unit") or "Nos")
    patch = {
        "product": pn, "size": data.size or "", "quantity": data.quantity,
        "unit": data.unit or existing.get("unit") or "Nos",
        "client_id": data.client_id or "", "client_name": data.client_name or "",
        "project_id": data.project_id or "", "project_name": data.project_name or "",
        "outward_challan_no": numeric_only(data.outward_challan_no),
        "reference_number": numeric_only(data.reference_number or data.outward_challan_no),
        "reference_type": data.reference_type or existing.get("reference_type") or "Challan Number",
        "date": data.date or existing.get("date") or now_iso(),
        "remarks": data.remarks or "",
        "status": data.status or existing.get("status") or "Dispatched",
        "attachment_file_id": data.attachment_file_id if data.attachment_file_id is not None else existing.get("attachment_file_id", ""),
        "attachment_filename": data.attachment_filename if data.attachment_filename is not None else existing.get("attachment_filename", ""),
        "updated_at": now_iso(),
    }
    await db.outward_entries.update_one({"id": entry_id, "company_id": cid}, {"$set": patch})
    
    # Reconcile high-value dispatch
    all_assets = _load_local_assets()
    for a in all_assets:
        if a.get("outward_entry_id") == entry_id and a.get("company_id") == cid:
            a["status"] = "Available"
            a["outward_entry_id"] = None
            a["client_id"] = None
            a["client_name"] = None
            a["outward_date"] = None
            
    is_hv = data.high_value_asset or data.high_value_goods or _load_local_high_value_products().get(pn, False) or any(kw in pn for kw in ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"])
    if is_hv:
        qty = float(data.quantity or 0)
        sns = [sn.strip().upper() for sn in (data.serial_numbers or []) if sn.strip()]
        
        available = [a for a in all_assets if a.get("product_name") == pn and a.get("status") == "Available" and a.get("company_id") == cid]
        
        status_val = "Installed" if data.client_id else "Installed"
        client_id_val = data.client_id or None
        client_name_val = data.client_name or None
        outward_date_val = patch["date"][:10]
        challan_val = patch["reference_number"] or patch["outward_challan_no"] or ""
        
        if sns:
            for sn in sns:
                matched = next((a for a in available if a.get("serial_number") == sn), None)
                if matched:
                    matched["status"] = status_val
                    matched["outward_entry_id"] = entry_id
                    matched["client_id"] = client_id_val
                    matched["client_name"] = client_name_val
                    matched["outward_date"] = outward_date_val
                    matched["challan_number"] = challan_val
                    matched["installation_date"] = outward_date_val
                    available.remove(matched)
                else:
                    no_sn_avail = next((a for a in available if not a.get("serial_number")), None)
                    if no_sn_avail:
                        no_sn_avail["quantity"] = float(no_sn_avail.get("quantity") or 1.0) - 1.0
                        if no_sn_avail["quantity"] <= 0:
                            if no_sn_avail in all_assets:
                                all_assets.remove(no_sn_avail)
                            if no_sn_avail in available:
                                available.remove(no_sn_avail)
                        new_asset = {
                            "id": str(uuid.uuid4()),
                            "company_id": cid,
                            "inward_entry_id": no_sn_avail.get("inward_entry_id"),
                            "product_name": pn,
                            "brand": no_sn_avail.get("brand", "Unknown"),
                            "size_model": data.size or no_sn_avail.get("size_model", ""),
                            "quantity": 1.0,
                            "serial_number": sn,
                            "vendor": no_sn_avail.get("vendor", ""),
                            "purchase_date": no_sn_avail.get("purchase_date", outward_date_val),
                            "challan_number": challan_val,
                            "client_id": client_id_val,
                            "client_name": client_name_val,
                            "installation_date": outward_date_val,
                            "warranty_status": "Active",
                            "status": status_val,
                            "outward_entry_id": entry_id,
                            "outward_date": outward_date_val,
                            "created_at": now_iso()
                        }
                        all_assets.append(new_asset)
                    else:
                        new_asset = {
                            "id": str(uuid.uuid4()),
                            "company_id": cid,
                            "inward_entry_id": None,
                            "product_name": pn,
                            "brand": "Unknown",
                            "size_model": data.size or "",
                            "quantity": 1.0,
                            "serial_number": sn,
                            "vendor": "",
                            "purchase_date": outward_date_val,
                            "challan_number": challan_val,
                            "client_id": client_id_val,
                            "client_name": client_name_val,
                            "installation_date": outward_date_val,
                            "warranty_status": "Active",
                            "status": status_val,
                            "outward_entry_id": entry_id,
                            "outward_date": outward_date_val,
                            "created_at": now_iso()
                        }
                        all_assets.append(new_asset)
        else:
            no_sn_avail = next((a for a in available if not a.get("serial_number")), None)
            if no_sn_avail:
                avail_qty = float(no_sn_avail.get("quantity") or 0)
                if avail_qty >= qty:
                    no_sn_avail["quantity"] = avail_qty - qty
                    if no_sn_avail["quantity"] <= 0:
                        if no_sn_avail in all_assets:
                            all_assets.remove(no_sn_avail)
                    new_asset = {
                        "id": str(uuid.uuid4()),
                        "company_id": cid,
                        "inward_entry_id": no_sn_avail.get("inward_entry_id"),
                        "product_name": pn,
                        "brand": no_sn_avail.get("brand", "Unknown"),
                        "size_model": data.size or no_sn_avail.get("size_model", ""),
                        "quantity": qty,
                        "serial_number": "",
                        "vendor": no_sn_avail.get("vendor", ""),
                        "purchase_date": no_sn_avail.get("purchase_date", outward_date_val),
                        "challan_number": challan_val,
                        "client_id": client_id_val,
                        "client_name": client_name_val,
                        "installation_date": outward_date_val,
                        "warranty_status": "Active",
                        "status": status_val,
                        "outward_entry_id": entry_id,
                        "outward_date": outward_date_val,
                        "created_at": now_iso()
                    }
                    all_assets.append(new_asset)
                else:
                    if no_sn_avail in all_assets:
                        all_assets.remove(no_sn_avail)
                    new_asset = {
                        "id": str(uuid.uuid4()),
                        "company_id": cid,
                        "inward_entry_id": no_sn_avail.get("inward_entry_id"),
                        "product_name": pn,
                        "brand": no_sn_avail.get("brand", "Unknown"),
                        "size_model": data.size or no_sn_avail.get("size_model", ""),
                        "quantity": qty,
                        "serial_number": "",
                        "vendor": no_sn_avail.get("vendor", ""),
                        "purchase_date": no_sn_avail.get("purchase_date", outward_date_val),
                        "challan_number": challan_val,
                        "client_id": client_id_val,
                        "client_name": client_name_val,
                        "installation_date": outward_date_val,
                        "warranty_status": "Active",
                        "status": status_val,
                        "outward_entry_id": entry_id,
                        "outward_date": outward_date_val,
                        "created_at": now_iso()
                    }
                    all_assets.append(new_asset)
            else:
                new_asset = {
                    "id": str(uuid.uuid4()),
                    "company_id": cid,
                    "inward_entry_id": None,
                    "product_name": pn,
                    "brand": "Unknown",
                    "size_model": data.size or "",
                    "quantity": qty,
                    "serial_number": "",
                    "vendor": "",
                    "purchase_date": outward_date_val,
                    "challan_number": challan_val,
                    "client_id": client_id_val,
                    "client_name": client_name_val,
                    "installation_date": outward_date_val,
                    "warranty_status": "Active",
                    "status": status_val,
                    "outward_entry_id": entry_id,
                    "outward_date": outward_date_val,
                    "created_at": now_iso()
                }
                all_assets.append(new_asset)
        for a in all_assets:
            if a.get("outward_entry_id") == entry_id:
                a["installation_notes"] = data.installation_notes or ""
                a["warranty_start_date"] = data.warranty_start_date or ""
                a["asset_remarks"] = data.asset_remarks or ""
    _save_local_assets(all_assets)
    
    await log_activity(cid, user["id"], user["name"], "Outward Updated", f"{pn} × {data.quantity}")
    await sync_inventory_master(cid)
    invalidate_products_cache(cid)
    res = await db.outward_entries.find_one({"id": entry_id, "company_id": cid}, {"_id": 0})
    return _enrich_outward_with_assets(res)

@api_router.delete("/inventory/outward/{entry_id}")
async def delete_outward(entry_id: str, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "delete"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.delete")
    cid = user["company_id"]
    existing = await db.outward_entries.find_one({"id": entry_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Outward entry not found")
    await db.outward_entries.delete_one({"id": entry_id, "company_id": cid})
    
    # Revert dispatched assets
    all_assets = _load_local_assets()
    for a in all_assets:
        if a.get("outward_entry_id") == entry_id and a.get("company_id") == cid:
            a["status"] = "Available"
            a["outward_entry_id"] = None
            a["client_id"] = None
            a["client_name"] = None
            a["outward_date"] = None
    _save_local_assets(all_assets)
    
    await log_activity(cid, user["id"], user["name"], "Outward Deleted", f"{existing.get('product')} × {existing.get('quantity')}")
    await sync_inventory_master(cid)
    invalidate_products_cache(cid)
    return {"ok": True}

# ---------- High Value Assets ----------
class AssetInstallIn(BaseModel):
    asset_ids: List[str]
    client_id: str

class AssetChangeStatusIn(BaseModel):
    asset_ids: List[str]
    status: str

@api_router.get("/assets")
async def list_assets(
    user=Depends(get_current_user),
    search: Optional[str] = None,
    status: Optional[str] = None,
):
    cid = user["company_id"]
    all_assets = _load_local_assets()
    hv_products = _load_local_high_value_products()
    hv_keywords = ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"]

    # Reconcile missing assets for high value inward entries
    try:
        inward_entries = await db.inward_entries.find({"company_id": cid}, {"_id": 0}).to_list(10000)
        existing_inward_ids = {a.get("inward_entry_id") for a in all_assets if a.get("inward_entry_id")}
        assets_changed = False

        for ie in inward_entries:
            pn = norm_product_name(ie.get("product"))
            is_hv = (
                ie.get("high_value_asset") or 
                ie.get("high_value_goods") or 
                hv_products.get(pn, False) or 
                any(kw in pn for kw in hv_keywords)
            )
            if is_hv and ie.get("id") not in existing_inward_ids:
                qty = float(ie.get("quantity") or 1.0)
                sns = [sn.strip().upper() for sn in (ie.get("serial_numbers") or []) if sn.strip()]
                vendor_val = ie.get("source_name") or ""
                date_val = (ie.get("date") or now_iso())[:10]
                challan_val = ie.get("reference_number") or ""
                size_val = ie.get("size") or ""

                if sns:
                    for sn in sns:
                        asset_doc = {
                            "id": str(uuid.uuid4()),
                            "company_id": cid,
                            "inward_entry_id": ie["id"],
                            "product_name": pn,
                            "brand": vendor_val or "Unknown",
                            "size_model": size_val,
                            "quantity": 1.0,
                            "serial_number": sn,
                            "vendor": vendor_val,
                            "purchase_date": date_val,
                            "challan_number": challan_val,
                            "client_id": None,
                            "client_name": None,
                            "installation_date": None,
                            "warranty_status": "Active",
                            "status": "Available",
                            "created_at": now_iso()
                        }
                        all_assets.append(asset_doc)
                else:
                    asset_doc = {
                        "id": str(uuid.uuid4()),
                        "company_id": cid,
                        "inward_entry_id": ie["id"],
                        "product_name": pn,
                        "brand": vendor_val or "Unknown",
                        "size_model": size_val,
                        "quantity": qty,
                        "serial_number": "",
                        "vendor": vendor_val,
                        "purchase_date": date_val,
                        "challan_number": challan_val,
                        "client_id": None,
                        "client_name": None,
                        "installation_date": None,
                        "warranty_status": "Active",
                        "status": "Available",
                        "created_at": now_iso()
                    }
                    all_assets.append(asset_doc)
                assets_changed = True
                existing_inward_ids.add(ie["id"])

        if assets_changed:
            _save_local_assets(all_assets)
    except Exception as e:
        logger.warning(f"Error reconciling assets: {e}")

    filtered = [a for a in all_assets if a.get("company_id") == cid]

    if search:
        search_lower = search.lower()
        res_list = []
        for a in filtered:
            sn = (a.get("serial_number") or "").lower()
            pn = (a.get("product_name") or "").lower()
            cn = (a.get("client_name") or "").lower()
            chn = (a.get("challan_number") or "").lower()
            if (search_lower in sn or 
                search_lower in pn or 
                search_lower in cn or 
                search_lower in chn):
                res_list.append(a)
        filtered = res_list

    if status:
        if status.lower() == "warranty expired":
            filtered = [a for a in filtered if a.get("warranty_status") == "Expired"]
        elif status.lower() == "replacement":
            filtered = [a for a in filtered if a.get("status") == "Replaced"]
        else:
            status_map = {
                "available": "Available",
                "installed": "Installed",
                "returned": "Returned",
                "dispatched": "Dispatched",
                "scrapped": "Scrapped",
                "replaced": "Replaced"
            }
            target_status = status_map.get(status.lower(), status)
            filtered = [a for a in filtered if a.get("status") == target_status]

    return filtered

@api_router.post("/assets/install")
async def install_assets(data: AssetInstallIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    client = await db.clients.find_one({"id": data.client_id, "company_id": cid})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
        
    all_assets = _load_local_assets()
    updated_count = 0
    now_date = datetime.now().strftime("%Y-%m-%d")
    
    for a in all_assets:
        if a.get("id") in data.asset_ids and a.get("company_id") == cid:
            a["status"] = "Installed"
            a["client_id"] = client["id"]
            a["client_name"] = client["full_name"]
            a["installation_date"] = now_date
            a["warranty_status"] = "Active"
            updated_count += 1
            
    if updated_count > 0:
        _save_local_assets(all_assets)
        await log_activity(cid, user["id"], user["name"], "Assets Installed", f"Installed {updated_count} assets for client {client['full_name']}")
        
    return {"ok": True, "installed_count": updated_count}

@api_router.post("/assets/change-status")
async def change_assets_status(data: AssetChangeStatusIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    all_assets = _load_local_assets()
    updated_count = 0
    
    for a in all_assets:
        if a.get("id") in data.asset_ids and a.get("company_id") == cid:
            a["status"] = data.status
            if data.status in ["Available", "Scrapped", "Returned"]:
                a["client_id"] = None
                a["client_name"] = None
                a["installation_date"] = None
            updated_count += 1
            
    if updated_count > 0:
        _save_local_assets(all_assets)
        await log_activity(cid, user["id"], user["name"], "Assets Status Updated", f"Updated status of {updated_count} assets to {data.status}")
        
    return {"ok": True, "updated_count": updated_count}


@api_router.delete("/assets/{asset_id}")
async def delete_asset(asset_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    all_assets = _load_local_assets()
    
    existing = next((a for a in all_assets if a.get("id") == asset_id and a.get("company_id") == cid), None)
    if not existing:
        raise HTTPException(status_code=404, detail="Asset not found")
        
    filtered = [a for a in all_assets if not (a.get("id") == asset_id and a.get("company_id") == cid)]
    _save_local_assets(filtered)
    
    await log_activity(cid, user["id"], user["name"], "Asset Deleted", f"Deleted high value asset: {existing.get('serial_number') or asset_id}")
    return {"ok": True}


@api_router.patch("/assets/{asset_id}")
async def edit_asset(asset_id: str, data: AssetEditIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    all_assets = _load_local_assets()
    target = next((a for a in all_assets if a.get("id") == asset_id and a.get("company_id") == cid), None)
    if not target:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if data.serial_number is not None:
        target["serial_number"] = data.serial_number.strip().upper()
    if data.site_location is not None:
        target["site_location"] = data.site_location.strip()
    if data.remarks is not None:
        target["remarks"] = data.remarks.strip()
    
    _save_local_assets(all_assets)
    await log_activity(cid, user["id"], user["name"], "Asset Updated", f"Updated high value asset serial/site for {target.get('serial_number')}")
    return {"ok": True, "asset": target}


@api_router.get("/assets/{asset_id}/timeline")
async def get_asset_timeline(asset_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    all_assets = _load_local_assets()
    target_asset = next((a for a in all_assets if a.get("id") == asset_id and a.get("company_id") == cid), None)
    if not target_asset:
        target_asset = next((a for a in all_assets if a.get("company_id") == cid and (a.get("serial_number") or "").strip().upper() == asset_id.strip().upper()), None)

    sn = (target_asset.get("serial_number") if target_asset else asset_id).strip().upper()
    p_name = target_asset.get("product_name") if target_asset else ""

    inwards = await db.inward_entries.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    outwards = await db.outward_entries.find({"company_id": cid}, {"_id": 0}).to_list(10000)

    events = []

    for ie in inwards:
        sns = [s.strip().upper() for s in (ie.get("serial_numbers") or []) if s.strip()]
        matches = (sn in sns) or (sn and (ie.get("serial_number") or "").strip().upper() == sn)
        if not matches and target_asset and target_asset.get("inward_entry_id") == ie.get("id"):
            matches = True

        if matches:
            is_return = ie.get("source_type") == "Return From Client" or "return" in (ie.get("entry_type") or "").lower()
            date_str = (ie.get("date") or ie.get("created_at") or "")[:10]
            if is_return:
                events.append({
                    "type": "Returned",
                    "title": "Material Returned to Warehouse",
                    "date": date_str,
                    "site": "Central Warehouse",
                    "client": ie.get("source_name") or ie.get("client_name") or "Client",
                    "detail": f"Return Ref: {ie.get('reference_number') or 'N/A'}"
                })
            else:
                events.append({
                    "type": "Inward",
                    "title": "Inward Entry Recorded",
                    "date": date_str,
                    "site": "Central Warehouse",
                    "client": ie.get("source_name") or ie.get("vendor") or "Supplier",
                    "detail": f"Challan / Bill: {ie.get('reference_number') or 'N/A'}"
                })

    for oe in outwards:
        sns = [s.strip().upper() for s in (oe.get("serial_numbers") or []) if s.strip()]
        matches = (sn in sns) or (sn and (oe.get("serial_number") or "").strip().upper() == sn)
        if matches:
            date_str = (oe.get("date") or oe.get("created_at") or "")[:10]
            site_val = oe.get("target_name") or oe.get("site_name") or (f"{oe.get('client_name')} Site" if oe.get("client_name") else "Client Site")
            events.append({
                "type": "Outward",
                "title": "Outward Dispatched to Client",
                "date": date_str,
                "site": site_val,
                "client": oe.get("client_name") or "Client",
                "detail": f"Gate Pass / Challan: {oe.get('reference_number') or oe.get('outward_challan_no') or 'N/A'}"
            })

    events.sort(key=lambda e: e.get("date") or "")

    return {
        "ok": True,
        "serial_number": sn,
        "product_name": p_name or (target_asset.get("product_name") if target_asset else ""),
        "total_movements": len(events),
        "events": events
    }


# ---------- Inventory Defaults ----------
@api_router.get("/inventory/defaults")
async def get_inv_defaults(user=Depends(get_current_user)):
    d = await db.inventory_defaults.find_one({"company_id": user["company_id"]}, {"_id": 0})
    return d or {"inward": {}, "outward": {}}

@api_router.patch("/inventory/defaults")
async def set_inv_defaults(data: InventoryDefaults, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.edit")
    cid = user["company_id"]
    existing = await db.inventory_defaults.find_one({"company_id": cid}) or {"company_id": cid}
    inward = {**(existing.get("inward") or {}), **(data.inward or {})}
    outward = {**(existing.get("outward") or {}), **(data.outward or {})}
    patch = {"company_id": cid, "inward": inward, "outward": outward, "updated_at": now_iso(),
             "updated_by": user["id"], "updated_by_name": user["name"]}
    await db.inventory_defaults.update_one({"company_id": cid}, {"$set": patch}, upsert=True)
    return patch

# ---------- Inventory History (combined) ----------
@api_router.get("/inventory/history")
async def inv_history(
    request: Request = None,  # type: ignore
    user=Depends(get_current_user),
    type: Optional[str] = None,  # inward | outward | None
    product: Optional[str] = None,
    size: Optional[str] = None,
    vendor: Optional[str] = None,
    client: Optional[str] = None,
    challan: Optional[str] = None,
    bill_number: Optional[str] = None,
    user_id: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 100,
):
    if type and type.lower() in ("all", "none", "null"):
        type = None
    if status and status.lower() in ("all", "none", "null"):
        status = None

    cid = user["company_id"]
    page = max(1, page)
    page_size = max(1, min(page_size, 500))
    inward_projection = {
        "_id": 0,
        "id": 1,
        "date": 1,
        "created_at": 1,
        "product": 1,
        "size": 1,
        "quantity": 1,
        "unit": 1,
        "reference_number": 1,
        "bill_number": 1,
        "source_name": 1,
        "source_type": 1,
        "remarks": 1,
        "created_by": 1,
        "created_by_name": 1,
        "attachment_file_id": 1,
        "attachment_filename": 1,
    }

    outward_projection = {
        "_id": 0,
        "id": 1,
        "date": 1,
        "created_at": 1,
        "product": 1,
        "size": 1,
        "quantity": 1,
        "unit": 1,
        "outward_challan_no": 1,
        "client_name": 1,
        "project_name": 1,
        "status": 1,
        "remarks": 1,
        "created_by": 1,
        "created_by_name": 1,
        "attachment_file_id": 1,
        "attachment_filename": 1,
    }

    def _text_filter(value: Optional[str]) -> dict[str, Any]:
        return {"$regex": re.escape(value or ""), "$options": "i"} if value else {}

    def _search_or_conditions(field_names: List[str], value: str) -> List[Dict[str, Any]]:
        return [{field: _text_filter(value)} for field in field_names if field]

    def _date_match(rec: Dict[str, Any]) -> bool:
        d = (rec.get("date") or rec.get("created_at") or "")[:10]
        if from_date and d < from_date: return False
        if to_date and d > to_date: return False
        return True

    rows: List[Dict[str, Any]] = []

    def _search_match(rec: Dict[str, Any]) -> bool:
        if not search or not search.strip():
            return True
        clean_s = norm_str(search).lower().strip()
        tokens = [t for t in clean_s.split() if t]
        if not tokens:
            return True
        
        prod = norm_product_name(rec.get("product"))
        raw_size = rec.get("size") or ""
        sz = norm_str(raw_size)
        src = (rec.get("source_name") or rec.get("client_name") or "").lower()
        proj = (rec.get("project_name") or "").lower()
        ref = (rec.get("reference_number") or rec.get("outward_challan_no") or "").lower()
        bill = (rec.get("bill_number") or "").lower()
        rem = (rec.get("remarks") or "").lower()
        by = (rec.get("created_by_name") or "").lower()
        
        full_text = f"{prod} {sz} {raw_size} {src} {proj} {ref} {bill} {rem} {by}".lower()
        return all(t in full_text for t in tokens)

    if (not type or type == "inward") and not status:
        q: Dict[str, Any] = {"company_id": cid}
        if product: q["product"] = _text_filter(product)
        if size is not None and size != "": q["size"] = norm_str(size)
        if vendor: q["source_name"] = _text_filter(vendor)
        if challan: q["reference_number"] = _text_filter(challan)
        if bill_number: q["bill_number"] = _text_filter(bill_number)
        if user_id: q["created_by"] = user_id
        inward_rows = await db.inward_entries.find(q, inward_projection).sort([("date", -1), ("created_at", -1)]).to_list(10000)
        for r in inward_rows:
            if not _date_match(r):
                continue
            enriched = _enrich_inward_with_assets(parse_inward_client_info(r))
            if enriched:
                if _search_match(enriched):
                    rows.append({**enriched, "type": "Inward"})

    if (not type or type == "outward") and not bill_number:
        q = {"company_id": cid}
        if product: q["product"] = _text_filter(product)
        if size is not None and size != "": q["size"] = norm_str(size)
        if client: q["client_name"] = _text_filter(client)
        if challan: q["$or"] = [{"outward_challan_no": _text_filter(challan)}, {"reference_number": _text_filter(challan)}]
        if user_id: q["created_by"] = user_id
        if status: q["status"] = status
        outward_rows = await db.outward_entries.find(q, outward_projection).sort([("date", -1), ("created_at", -1)]).to_list(10000)
        for r in outward_rows:
            if not _date_match(r):
                continue
            enriched = _enrich_outward_with_assets(r)
            if enriched:
                if _search_match(enriched):
                    rows.append({**enriched, "type": "Outward"})

    rows.sort(key=lambda x: (x.get("date") or x.get("created_at") or ""), reverse=True)
    total = len(rows)
    pages = max(1, math.ceil(total / page_size)) if page_size > 0 else 1
    start = (page - 1) * page_size
    paged = rows[start:start + page_size]
    return {"items": paged, "total": total, "page": page, "pages": pages, "page_size": page_size, "rows": paged}

@api_router.get("/inventory/history.csv")
async def inv_history_csv(
    request: Request = None,  # type: ignore
    user=Depends(get_current_user),
    type: Optional[str] = None,
    product: Optional[str] = None,
    size: Optional[str] = None,
    vendor: Optional[str] = None,
    client: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
):
    result = await inv_history(request=request, user=user, type=type, product=product, size=size, vendor=vendor, client=client, from_date=from_date, to_date=to_date, search=search, page=1, page_size=100000)  # type: ignore
    rows: Any = result["rows"] if isinstance(result, dict) else result
    if not isinstance(rows, list):
        rows = []
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Date", "Type", "Product", "Size", "Quantity", "Unit", "Reference / Challan", "Bill / Outward No", "Vendor / Client", "Project", "Status", "Remarks", "Created By"])
    for r in rows:
        ref = r.get("reference_number") or ""
        billish = r.get("bill_number") or r.get("outward_challan_no") or ""
        party = r.get("source_name") if r.get("type") == "Inward" else r.get("client_name")
        w.writerow([
            (r.get("date") or r.get("created_at") or "")[:10],
            r.get("type", ""), r.get("product", ""), r.get("size", ""),
            r.get("quantity", 0), r.get("unit", ""),
            ref, billish, party or "",
            r.get("project_name", ""), r.get("status", ""),
            r.get("remarks", ""), r.get("created_by_name", ""),
        ])
    from fastapi.responses import StreamingResponse
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": 'attachment; filename="solarix-inventory-history.csv"'})


class BulkDeleteIn(BaseModel):
    inward_ids: Optional[List[str]] = None
    outward_ids: Optional[List[str]] = None

@api_router.post("/inventory/bulk-delete")
async def bulk_delete_history(data: BulkDeleteIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "delete"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.delete")
    cid = user["company_id"]
    deleted_in = deleted_out = 0
    all_assets = _load_local_assets()
    assets_changed = False
    if data.inward_ids:
        res = await db.inward_entries.delete_many({"company_id": cid, "id": {"$in": data.inward_ids}})
        deleted_in = res.deleted_count
        initial_len = len(all_assets)
        all_assets = [a for a in all_assets if a.get("inward_entry_id") not in data.inward_ids or a.get("status") == "Installed"]
        if len(all_assets) != initial_len:
            assets_changed = True
    if data.outward_ids:
        res = await db.outward_entries.delete_many({"company_id": cid, "id": {"$in": data.outward_ids}})
        deleted_out = res.deleted_count
        for a in all_assets:
            if a.get("outward_entry_id") in data.outward_ids and a.get("company_id") == cid:
                a["status"] = "Available"
                a["outward_entry_id"] = None
                a["client_id"] = None
                a["client_name"] = None
                a["outward_date"] = None
                assets_changed = True
    if assets_changed:
        _save_local_assets(all_assets)
    total = deleted_in + deleted_out
    if total:
        await log_activity(cid, user["id"], user["name"], "Bulk Inventory Delete",
                           f"{deleted_in} inward + {deleted_out} outward")
    return {"deleted_inward": deleted_in, "deleted_outward": deleted_out, "total": total}


@api_router.get("/inventory/next-challan")
async def next_challan(type: str, prefix: Optional[str] = "", user=Depends(get_current_user)):
    """Suggest the next sequential challan number for inward/outward."""
    cid = user["company_id"]
    coll = db.inward_entries if type == "inward" else db.outward_entries
    field = "reference_number" if type == "inward" else "outward_challan_no"
    pfx = (prefix or "").strip()
    # Find max trailing numeric suffix
    max_num = 0
    cur = coll.find({"company_id": cid, field: {"$ne": ""}}, {"_id": 0, field: 1})
    async for r in cur:
        val = r.get(field) or ""
        if pfx and not val.upper().startswith(pfx.upper()):
            continue
        m = re.search(r"(\d+)\s*$", val)
        if m:
            try:
                n = int(m.group(1))
                if n > max_num: max_num = n
            except Exception:
                pass
    next_num = max_num + 1
    suggested = f"{pfx}{next_num:04d}" if pfx else f"{next_num:04d}"
    return {"next_number": next_num, "suggested": suggested, "max_existing": max_num}


@api_router.get("/inventory/check-challan")
async def check_challan_unique(type: str, challan: str, exclude_id: Optional[str] = None, user=Depends(get_current_user)):
    cid = user["company_id"]
    coll = db.inward_entries if type == "inward" else db.outward_entries
    field = "reference_number" if type == "inward" else "outward_challan_no"
    q: Dict[str, Any] = {"company_id": cid, field: challan}
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    existing = await coll.find_one(q, {"_id": 0, "id": 1, "product": 1, "date": 1})
    return {"unique": existing is None, "existing": existing}


@api_router.get("/inventory/vendors")
async def list_vendors(user=Depends(get_current_user)):
    """Distinct non-empty vendor / source names previously used in inward entries (current company)."""
    cid = user["company_id"]
    names = await db.inward_entries.distinct("source_name", {"company_id": cid, "source_name": {"$nin": ["", None]}})
    return sorted([n for n in names if n and isinstance(n, str)], key=lambda s: s.lower())


@api_router.get("/inventory/products/{product_id}/stats")
async def product_stats(product_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    p = await db.products.find_one({"id": product_id, "company_id": cid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    items, _, _, _ = await _compute_inventory_balances(cid)
    matched_p = next((item for item in items if item.get("id") == product_id), None)
    if not matched_p:
        name = norm_product_name(p["name"])
        size = norm_str(p.get("size"))
        matched_p = next((item for item in items if norm_product_name(item.get("name")) == name and norm_str(item.get("size")) == size), p)

    op_stock = float(matched_p.get("opening_stock") or 0.0)
    total_in = matched_p.get("total_in", 0.0)
    total_out = matched_p.get("total_out", 0.0)
    balance = matched_p.get("balance", 0.0)

    # Fetch last dates
    p_name = norm_product_name(matched_p.get("name"))
    p_size = norm_str(matched_p.get("size"))
    last_in_rows = await db.inward_entries.find({"company_id": cid, "product": p_name, "size": p_size}, {"_id": 0, "date": 1}).sort("date", -1).to_list(1)
    last_out_rows = await db.outward_entries.find({"company_id": cid, "status": {"$nin": ["Cancelled", "draft_cancelled"]}, "product": p_name, "size": p_size}, {"_id": 0, "date": 1}).sort("date", -1).to_list(1)
    in_count = await db.inward_entries.count_documents({"company_id": cid, "product": p_name, "size": p_size})
    out_count = await db.outward_entries.count_documents({"company_id": cid, "status": {"$nin": ["Cancelled", "draft_cancelled"]}, "product": p_name, "size": p_size})

    return {
        "product": matched_p,
        "opening_stock": op_stock,
        "total_in": total_in, "total_out": total_out, "balance": balance,
        "last_inward_date": (last_in_rows[0].get("date") if last_in_rows else None),
        "last_outward_date": (last_out_rows[0].get("date") if last_out_rows else None),
        "transaction_count": in_count + out_count,
        "inward_count": in_count, "outward_count": out_count,
    }


@api_router.get("/inventory/products/{product_id}/transactions")
async def product_transactions(
    product_id: str,
    request: Request,
    user=Depends(get_current_user),
    type: Optional[str] = None,
    challan: Optional[str] = None,
    vendor: Optional[str] = None,
    client: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
):
    cid = user["company_id"]
    p = await db.products.find_one({"id": product_id, "company_id": cid}, {"_id": 0, "name": 1, "size": 1, "unit": 1})
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    return await inv_history(
        request=request, user=user, type=type, product=p["name"], size=p.get("size") or "", vendor=vendor, client=client,
        challan=challan, from_date=from_date, to_date=to_date, search=search,
        page=1, page_size=10000,
    )





class BulkRow(BaseModel):
    product: Optional[str] = ""
    size: Optional[str] = ""
    brand: Optional[str] = ""
    quantity: Optional[Union[float, int, str]] = 0.0
    unit: Optional[str] = "Nos"
    date: Optional[str] = ""
    reference_number: Optional[str] = ""
    reference_type: Optional[str] = "Challan Number"
    source_type: Optional[str] = "Supplier"
    source_name: Optional[str] = ""
    vendor: Optional[str] = ""
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    bill_number: Optional[str] = ""
    remarks: Optional[str] = ""
    high_value_asset: Optional[bool] = False
    high_value_goods: Optional[bool] = False
    serial_number_required: Optional[bool] = False
    serial_numbers: Optional[List[str]] = []

class BulkInwardIn(BaseModel):
    rows: List[BulkRow]
    batch_label: Optional[str] = ""
    global_defaults: Optional[Dict] = {}



@api_router.post("/inventory/bulk-inward")
async def bulk_inward(data: BulkInwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.create")
    """Insert validated rows as inward_entries. Returns count + ids."""
    if not data.rows:
        raise HTTPException(status_code=400, detail="No rows provided")
    cid = user["company_id"]
    inserted_ids: List[str] = []
    gd = data.global_defaults or {}  # v2 global defaults
    prod_cache: Dict[Tuple[str, str, str, str], Any] = {}
    docs_to_insert = []
    new_assets = []
    all_assets = _load_local_assets()
    hv_products = _load_local_high_value_products()
    
    # 1. Pre-fetch existing products and clients for company in bulk
    existing_prods, existing_clients = await asyncio.gather(
        db.products.find({"company_id": cid}).to_list(10000),
        db.clients.find({"company_id": cid}, {"_id": 0, "id": 1, "full_name": 1}).to_list(10000)
    )
    for p in (existing_prods or []):
        pn_n = norm_product_name(p.get("name"))
        ps_n = norm_str(p.get("size"))
        pu_n = norm_unit(p.get("unit"))
        if cid and pn_n:
            prod_cache[(cid, pn_n, ps_n, pu_n)] = p

    client_map = {}
    for c in (existing_clients or []):
        if c.get("full_name"):
            client_map[c["full_name"].strip().lower()] = c

    # 2. Pre-pass: Resolve & bulk-insert missing products in 1 batch query
    new_prods_to_insert = []
    for r in data.rows:
        pn = (r.product or "").strip().upper()
        if not pn:
            continue
        ps = r.size or ""
        pu = r.unit or gd.get("unit") or "Nos"
        cache_key = (cid, norm_product_name(pn), norm_str(ps), norm_unit(pu))
        if cache_key not in prod_cache:
            prod_doc = {
                "id": str(uuid.uuid4()),
                "company_id": cid,
                "name": pn,
                "size": ps,
                "category": "Solar",
                "unit": pu or "Nos",
                "min_stock": 0.0,
                "status": "Active",
                "created_at": now_iso()
            }
            prod_cache[cache_key] = prod_doc
            new_prods_to_insert.append(prod_doc)
            
    if new_prods_to_insert:
        try:
            await db.products.insert_many(new_prods_to_insert)
        except Exception:
            pass

    # 3. Build documents in-memory
    for r in data.rows:
        pn = (r.product or "").strip().upper()
        if not pn:
            continue
        try:
            qty = float(r.quantity) if r.quantity not in (None, "") else 0.0
        except (ValueError, TypeError):
            qty = 0.0
            
        remarks_val = r.remarks or gd.get("remarks", "")
        source_type_val = r.source_type or gd.get("source_type", "Supplier")
        client_id_val = r.client_id or gd.get("client_id", "")
        client_name_val = r.client_name or gd.get("client_name", "")
        source_name_val = r.source_name or gd.get("source_name", "")
        ps = r.size or ""
        pu = r.unit or gd.get("unit") or "Nos"
        
        # Client ID resolution from name case-insensitively for Return From Client
        if source_type_val == "Return From Client":
            if client_name_val and not client_id_val:
                matched_c = client_map.get(client_name_val.strip().lower())
                if matched_c:
                    client_id_val = matched_c["id"]
                    client_name_val = matched_c["full_name"]
            if client_name_val:
                source_name_val = client_name_val
            if client_id_val:
                remarks_val = f"{remarks_val} [client_id:{client_id_val}]".strip()

        entry_id = str(uuid.uuid4())
        ref_num = r.reference_number or gd.get("reference_number", "")
        bill_num = r.bill_number or gd.get("bill_number", "")
        date_val = r.date or gd.get("date", "") or now_iso()
        
        doc = {
            "id": entry_id,
            "company_id": cid,
            "product": pn,
            "size": ps,
            "quantity": qty,
            "unit": pu,
            "reference_number": numeric_only(ref_num),
            "reference_type": r.reference_type or gd.get("reference_type", "Challan Number"),
            "bill_number": numeric_only(bill_num),
            "source_type": source_type_val,
            "source_name": source_name_val,
            "date": date_val,
            "remarks": remarks_val,
            "attachment_file_id": "",
            "attachment_filename": "",
            "source": "ai-bulk-import",
            "created_by": user["id"],
            "created_by_name": user["name"],
            "created_at": now_iso()
        }
        if data.batch_label:
            doc["import_batch"] = data.batch_label
            
        docs_to_insert.append(doc)
        inserted_ids.append(entry_id)
        
        # High value asset tracking
        is_hv = r.high_value_asset or r.high_value_goods or hv_products.get(pn, False) or any(kw in pn for kw in ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"])
        if is_hv:
            sns = [sn.strip().upper() for sn in (r.serial_numbers or []) if sn.strip()]
            if sns:
                for sn in sns:
                    new_assets.append({
                        "id": str(uuid.uuid4()),
                        "company_id": cid,
                        "inward_entry_id": entry_id,
                        "product_name": pn,
                        "brand": source_name_val or "Unknown",
                        "size_model": ps,
                        "quantity": 1.0,
                        "serial_number": sn,
                        "vendor": source_name_val or "",
                        "purchase_date": date_val[:10],
                        "challan_number": ref_num,
                        "client_id": None,
                        "client_name": None,
                        "installation_date": None,
                        "warranty_status": "Active",
                        "status": "Available",
                        "created_at": now_iso()
                    })
            else:
                new_assets.append({
                    "id": str(uuid.uuid4()),
                    "company_id": cid,
                    "inward_entry_id": entry_id,
                    "product_name": pn,
                    "brand": source_name_val or "Unknown",
                    "size_model": ps,
                    "quantity": qty,
                    "serial_number": "",
                    "vendor": source_name_val or "",
                    "purchase_date": date_val[:10],
                    "challan_number": ref_num,
                    "client_id": None,
                    "client_name": None,
                    "installation_date": None,
                    "warranty_status": "Active",
                    "status": "Available",
                    "created_at": now_iso()
                })

    # 4. Bulk DB Insertion (1 single DB query)
    if docs_to_insert:
        await db.inward_entries.insert_many(docs_to_insert)
        if new_assets:
            all_assets.extend(new_assets)
            _save_local_assets(all_assets)
        invalidate_products_cache(cid)
        asyncio.create_task(log_activity(cid, user["id"], user["name"], "Bulk Inward Import", f"{len(docs_to_insert)} entries"))
        asyncio.create_task(push_notification(cid, "admin", "Bulk Inventory Import", f"{user['name']} imported {len(docs_to_insert)} inward entries via AI"))
        if new_prods_to_insert:
            asyncio.create_task(sync_inventory_master(cid))

    return {"inserted": len(inserted_ids), "ids": inserted_ids}


@api_router.post("/inventory/bulk-inward-high-value")
async def bulk_inward_high_value(data: BulkInwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.create")
    if not data.rows:
        raise HTTPException(status_code=400, detail="No rows provided")
    
    cid = user["company_id"]
    inserted_ids: List[str] = []
    gd = data.global_defaults or {}
    docs_to_insert = []
    new_assets = []
    all_assets = _load_local_assets()
    hv_products = _load_local_high_value_products()
    
    # 1. Pre-fetch all existing products for company in 1 single bulk query
    prod_cache: Dict[Tuple[str, str, str, str], Any] = {}
    existing_prods = await db.products.find({"company_id": cid}).to_list(10000)
    for p in (existing_prods or []):
        pn_n = norm_product_name(p.get("name"))
        ps_n = norm_str(p.get("size"))
        pu_n = norm_unit(p.get("unit"))
        if cid and pn_n:
            prod_cache[(cid, pn_n, ps_n, pu_n)] = p

    # 2. Pre-pass: Resolve products & bulk update high_value_goods flag in 1 DB query
    prod_ids_to_hv = set()
    for r in data.rows:
        pn = (r.product or "").strip().upper()
        if not pn:
            continue
        ps = (r.size or "").strip()
        pu = (r.unit or gd.get("unit") or "Nos").strip()
        brand_val = (getattr(r, 'brand', None) or r.source_name or gd.get("vendor") or gd.get("source_name") or "Unknown").strip()

        _save_local_high_value_product(pn, True)
        _save_local_high_value_product(norm_product_name(pn), True)
        hv_products[pn] = True
        hv_products[norm_product_name(pn)] = True

        cache_key = (cid, norm_product_name(pn), norm_str(ps), norm_unit(pu))
        if cache_key not in prod_cache:
            prod_doc = await ensure_product(cid, pn, size=ps, unit=pu, brand=brand_val, high_value_goods=True)
            prod_cache[cache_key] = prod_doc
        else:
            prod_doc = prod_cache[cache_key]

        if prod_doc and prod_doc.get("id"):
            prod_ids_to_hv.add(prod_doc["id"])

    if prod_ids_to_hv:
        try:
            await db.products.update_many(
                {"id": {"$in": list(prod_ids_to_hv)}, "company_id": cid},
                {"$set": {"high_value_goods": True, "high_value_asset": True}}
            )
        except Exception:
            pass

    # 3. Build documents in-memory with ZERO database queries inside loop
    for r in data.rows:
        pn = (r.product or "").strip().upper()
        if not pn:
            continue
        try:
            qty = float(r.quantity) if r.quantity not in (None, "") else 0.0
        except (ValueError, TypeError):
            qty = 0.0
            
        ps = (r.size or "").strip()
        pu = (r.unit or gd.get("unit") or "Nos").strip()
        brand_val = (getattr(r, 'brand', None) or r.source_name or gd.get("vendor") or gd.get("source_name") or "Unknown").strip()
        source_name_val = (r.source_name or getattr(r, 'vendor', None) or gd.get("vendor") or gd.get("source_name") or "").strip()
        source_type_val = r.source_type or gd.get("source_type", "Supplier")
        ref_num = r.reference_number or r.bill_number or gd.get("bill_number") or gd.get("reference_number", "")
        bill_num = r.bill_number or r.reference_number or gd.get("bill_number", "")
        date_val = r.date or gd.get("date", "") or now_iso()
        remarks_val = r.remarks or gd.get("remarks", "")

        entry_id = str(uuid.uuid4())
        doc = {
            "id": entry_id,
            "company_id": cid,
            "product": pn,
            "size": ps,
            "quantity": qty,
            "unit": pu,
            "reference_number": numeric_only(ref_num),
            "reference_type": r.reference_type or gd.get("reference_type", "Challan Number"),
            "bill_number": numeric_only(bill_num),
            "source_type": source_type_val,
            "source_name": source_name_val,
            "date": date_val,
            "remarks": remarks_val,
            "high_value_goods": True,
            "high_value_asset": True,
            "attachment_file_id": "",
            "attachment_filename": "",
            "source": "high-value-manual-import",
            "created_by": user["id"],
            "created_by_name": user["name"],
            "created_at": now_iso()
        }

        docs_to_insert.append(doc)
        inserted_ids.append(entry_id)

        # High value assets tracking
        sns = [sn.strip().upper() for sn in (r.serial_numbers or []) if sn.strip()]
        if sns:
            for sn in sns:
                new_assets.append({
                    "id": str(uuid.uuid4()),
                    "company_id": cid,
                    "inward_entry_id": entry_id,
                    "product_name": pn,
                    "brand": brand_val,
                    "size_model": ps,
                    "quantity": 1.0,
                    "serial_number": sn,
                    "vendor": source_name_val or brand_val,
                    "purchase_date": date_val[:10],
                    "challan_number": ref_num,
                    "client_id": None,
                    "client_name": None,
                    "installation_date": None,
                    "warranty_status": "Active",
                    "status": "Available",
                    "created_at": now_iso()
                })
        else:
            new_assets.append({
                "id": str(uuid.uuid4()),
                "company_id": cid,
                "inward_entry_id": entry_id,
                "product_name": pn,
                "brand": brand_val,
                "size_model": ps,
                "quantity": qty,
                "serial_number": "",
                "vendor": source_name_val or brand_val,
                "purchase_date": date_val[:10],
                "challan_number": ref_num,
                "client_id": None,
                "client_name": None,
                "installation_date": None,
                "warranty_status": "Active",
                "status": "Available",
                "created_at": now_iso()
            })

    # 4. Bulk DB Insertion (1 single DB query)
    if docs_to_insert:
        await db.inward_entries.insert_many(docs_to_insert)
        if new_assets:
            all_assets.extend(new_assets)
            _save_local_assets(all_assets)
        invalidate_products_cache(cid)
        await log_activity(cid, user["id"], user["name"], "High Value Manual Import", f"{len(docs_to_insert)} high value entries")
        await push_notification(cid, "admin", "High Value Manual Import", f"{user['name']} imported {len(docs_to_insert)} high value goods entries")
        asyncio.create_task(sync_inventory_master(cid))

    return {"inserted": len(inserted_ids), "ids": inserted_ids, "message": f"Successfully imported {len(inserted_ids)} High Value Goods"}



# ---- AI Bulk Import (Outward) ----
class BulkOutwardRow(BaseModel):
    product: str
    size: Optional[str] = ""
    quantity: float
    unit: Optional[str] = "Nos"
    date: Optional[str] = ""
    outward_challan_no: Optional[str] = ""
    reference_number: Optional[str] = ""
    reference_type: Optional[str] = "Challan Number"
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    project_id: Optional[str] = ""
    project_name: Optional[str] = ""
    status: Optional[str] = "Dispatched"
    remarks: Optional[str] = ""
    high_value_asset: Optional[bool] = False
    high_value_goods: Optional[bool] = False
    serial_numbers: Optional[List[str]] = []
    installation_notes: Optional[str] = ""
    warranty_start_date: Optional[str] = ""
    asset_remarks: Optional[str] = ""

class BulkOutwardIn(BaseModel):
    rows: List[BulkOutwardRow]
    batch_label: Optional[str] = ""
    global_defaults: Optional[Dict] = {}




@api_router.post("/inventory/bulk-outward")
async def bulk_outward(data: BulkOutwardIn, user=Depends(get_current_user)):
    if not has_perm(user, "data_management", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: data_management.create")
    """Insert validated outward rows in bulk batch mode. Auto-creates products. Returns count + ids."""
    if not data.rows:
        raise HTTPException(status_code=400, detail="No rows provided")
    cid = user["company_id"]
    inserted_ids: List[str] = []
    gd = data.global_defaults or {}  # v2 global defaults
    g_client_id = gd.get("client_id", "")
    g_client_name = gd.get("client_name", "")
    prod_cache: Dict[Tuple[str, str, str, str], Any] = {}
    docs_to_insert = []

    # 1. Pre-fetch existing products and clients for company in bulk
    existing_prods, existing_clients = await asyncio.gather(
        db.products.find({"company_id": cid}).to_list(10000),
        db.clients.find({"company_id": cid}, {"_id": 0, "id": 1, "full_name": 1}).to_list(10000)
    )
    for p in (existing_prods or []):
        pn_n = norm_product_name(p.get("name"))
        ps_n = norm_str(p.get("size"))
        pu_n = norm_unit(p.get("unit"))
        if cid and pn_n:
            prod_cache[(cid, pn_n, ps_n, pu_n)] = p

    client_map = {}
    for c in (existing_clients or []):
        if c.get("full_name"):
            client_map[c["full_name"].strip().lower()] = c

    # 2. Pre-pass: Resolve & bulk-insert missing products in 1 batch query
    new_prods_to_insert = []
    for r in data.rows:
        pn = (r.product or "").strip().upper()
        if not pn or r.quantity <= 0:
            continue
        ps = r.size or ""
        pu = r.unit or gd.get("unit") or "Nos"
        cache_key = (cid, norm_product_name(pn), norm_str(ps), norm_unit(pu))
        if cache_key not in prod_cache:
            prod_doc = {
                "id": str(uuid.uuid4()),
                "company_id": cid,
                "name": pn,
                "size": ps,
                "category": "Solar",
                "unit": pu or "Nos",
                "min_stock": 0.0,
                "status": "Active",
                "created_at": now_iso()
            }
            prod_cache[cache_key] = prod_doc
            new_prods_to_insert.append(prod_doc)

    if new_prods_to_insert:
        try:
            await db.products.insert_many(new_prods_to_insert)
        except Exception:
            pass

    # 3. Build documents in-memory
    for r in data.rows:
        pn = (r.product or "").strip().upper()
        if not pn or r.quantity <= 0:
            continue

        client_id_val = r.client_id or g_client_id
        client_name_val = r.client_name or g_client_name
        project_id_val = r.project_id or gd.get("project_id", "")
        project_name_val = r.project_name or gd.get("project_name", "") or client_name_val

        if client_name_val and not client_id_val:
            matched_c = client_map.get(client_name_val.strip().lower())
            if matched_c:
                client_id_val = matched_c["id"]
                client_name_val = matched_c["full_name"]

        if client_id_val:
            if not project_id_val:
                project_id_val = client_id_val
            if not project_name_val:
                project_name_val = client_name_val

        status_val = r.status or gd.get("status", "Dispatched")
        if status_val not in ["Pending", "Dispatched", "Cancelled"]:
            status_val = "Dispatched"

        entry_id = str(uuid.uuid4())
        ref_num = r.reference_number or r.outward_challan_no or gd.get("reference_number", "")
        challan_no = r.outward_challan_no or ref_num
        date_val = r.date or gd.get("date", "") or now_iso()

        doc = {
            "id": entry_id,
            "company_id": cid,
            "product": pn,
            "size": r.size or "",
            "quantity": r.quantity,
            "unit": r.unit or gd.get("unit") or "Nos",
            "client_id": client_id_val,
            "client_name": client_name_val,
            "project_id": project_id_val,
            "project_name": project_name_val,
            "outward_challan_no": numeric_only(challan_no),
            "reference_number": numeric_only(ref_num),
            "reference_type": r.reference_type or gd.get("reference_type", "Challan Number"),
            "date": date_val,
            "remarks": r.remarks or gd.get("remarks", ""),
            "status": status_val,
            "attachment_file_id": "",
            "attachment_filename": "",
            "source": "ai-bulk-import",
            "created_by": user["id"],
            "created_by_name": user["name"],
            "created_at": now_iso()
        }
        if data.batch_label:
            doc["import_batch"] = data.batch_label

        docs_to_insert.append(doc)
        inserted_ids.append(entry_id)

    # 4. Bulk DB Insertion (1 single DB query)
    if docs_to_insert:
        await db.outward_entries.insert_many(docs_to_insert)
        invalidate_products_cache(cid)
        asyncio.create_task(log_activity(cid, user["id"], user["name"], "Bulk Outward Import", f"{len(docs_to_insert)} entries"))
        asyncio.create_task(push_notification(cid, "admin", "Bulk Outward Import", f"{user['name']} imported {len(docs_to_insert)} outward entries via AI"))
        if new_prods_to_insert:
            asyncio.create_task(sync_inventory_master(cid))

    return {"inserted": len(inserted_ids), "ids": inserted_ids}


# ============== Sprint 4: Client Data & Asset Management ==============
INVERTER_STATUS_OPTIONS = ["Online", "Offline", "Error", "Maintenance"]
TICKET_STATUSES = ["Open", "Assigned", "In Progress", "Waiting Parts", "Resolved", "Closed"]
TICKET_PRIORITIES = ["Low", "Medium", "High", "Critical"]
TICKET_ISSUE_TYPES = ["Inverter Offline", "Low Generation", "Net Meter Issue", "Panel Damage", "Wiring Issue", "Other"]

class MonitoringIn(BaseModel):
    portal_name: Optional[str] = ""
    app_name: Optional[str] = ""
    portal_url: Optional[str] = ""
    plant_id: Optional[str] = ""
    username: Optional[str] = ""
    password: Optional[str] = ""
    inverter_status: Optional[str] = "Offline"
    notes: Optional[str] = ""

class TicketIn(BaseModel):
    client_id: str
    title: str
    issue_type: str
    description: Optional[str] = ""
    priority: str = "Medium"
    attachments: Optional[List[Dict[str, str]]] = None  # [{file_id, filename, content_type}]

class TicketUpdate(BaseModel):
    status: Optional[str] = None
    priority: Optional[str] = None
    assigned_to: Optional[str] = None
    note: Optional[str] = None
    resolution: Optional[str] = None
    attachments: Optional[List[Dict[str, str]]] = None


async def _next_ticket_no(company_id: str) -> str:
    year = datetime.now(timezone.utc).year
    count = await db.service_tickets.count_documents({"company_id": company_id, "ticket_no": {"$regex": f"^TKT-{year}-"}})
    return f"TKT-{year}-{count + 1:04d}"


async def _attach_assets(client_id: str, company_id: str) -> List[Dict[str, str]]:
    """Aggregate all uploaded files across the client life cycle into Client Assets.
    All top-level queries run in parallel via asyncio.gather — reduces sequential
    Supabase round-trips to a single parallel wave.
    """
    q = {"company_id": company_id, "client_id": client_id}

    # Fire all collection queries + client doc lookup simultaneously
    (
        verifs, surveys, mds, docs, mts, insts, complaints,
        handovers, pm_surya, msedcl, client_doc
    ) = await asyncio.gather(
        db.verifications.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.surveys.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.material_deliveries.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.documents.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.meter_testings.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.installations.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.complaints.find(q, {"_id": 0}).sort("created_at", -1).to_list(100),
        db.handovers.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.pm_surya_uploads.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.msedcl_uploads.find(q, {"_id": 0}).sort("created_at", -1).to_list(50),
        db.clients.find_one({"id": client_id, "company_id": company_id}, {"_id": 0, "documents": 1, "photos": 1, "cleared_assets": 1}),
    )

    assets: List[Dict[str, str]] = []
    seen: set = set()
    c_doc = client_doc if isinstance(client_doc, dict) else {}
    cleared_ids: set = set(c_doc.get("cleared_assets") or [])

    def add_asset(fid, label, source, created_at, location="", description="", stage=""):
        if fid and fid not in seen and fid not in cleared_ids:
            stg = stage or source or "General"
            loc = location or (label.split(" - ", 1)[-1] if " - " in label else label)
            assets.append({
                "label": label,
                "file_id": fid,
                "source": source,
                "stage": stg,
                "location": loc,
                "description": description,
                "created_at": created_at or ""
            })
            seen.add(fid)

    # 1. verifications
    for v in (verifs or []):
        for label, val in (v.get("photos") or {}).items():
            fid = val if isinstance(val, str) else (val.get("file_id") if isinstance(val, dict) else None)
            add_asset(fid, label, "Verification", v.get("created_at"), location=label, stage="Verification")

    # 2. surveys
    for s in (surveys or []):
        details = s.get("details") or {}
        for label, val in (details.get("photos") or {}).items():
            fid = val.get("file_id") if isinstance(val, dict) else (val if isinstance(val, str) else None)
            add_asset(fid, f"Survey - {label}", "Survey", details.get("completed_date") or s.get("created_at"), location=label, stage="Survey")

    # 3. material deliveries
    for m in (mds or []):
        details = m.get("details") or {}
        for label, fid in (details.get("attachments") or {}).items():
            add_asset(fid, f"Delivery - {label}", "Material Delivery", details.get("completed_date") or m.get("created_at"), location=label, stage="Material Delivery")

    # 4. documents (Signed Documents)
    for d in (docs or []):
        details = d.get("details") or {}
        for item in (details.get("checklist") or []):
            fid = item.get("file_id")
            if fid:
                add_asset(fid, f"Signed - {item.get('label')}", "Documents Signed", details.get("completed_date") or d.get("created_at"), location=item.get("label", ""), stage="Document Signed")

    # 5. meter testings
    for mt in (mts or []):
        details = mt.get("details") or {}
        for label, fid in (details.get("attachments") or {}).items():
            add_asset(fid, f"Meter Testing - {label}", "Meter Testing", details.get("completed_date") or mt.get("created_at"), location=label, stage="Meter Testing")

    # 6. installations
    for inst in (insts or []):
        details = inst.get("details") or {}
        for label, val in (details.get("attachments") or {}).items():
            fid = val.get("file_id") if isinstance(val, dict) else (val if isinstance(val, str) else None)
            add_asset(fid, f"Installation - {label}", "Installation", details.get("completed_date") or inst.get("created_at"), location=label, stage="Installation")

    # 7. handovers
    for h in (handovers or []):
        details = h.get("details") or {}
        for label, val in (details.get("photos") or details.get("attachments") or {}).items():
            fid = val.get("file_id") if isinstance(val, dict) else (val if isinstance(val, str) else None)
            add_asset(fid, f"Handover - {label}", "Handover", details.get("completed_date") or h.get("created_at"), location=label, stage="Handover")

    # 8. PM Surya uploads
    for ps in (pm_surya or []):
        details = ps.get("details") or {}
        for label, val in (details.get("photos") or details.get("attachments") or {}).items():
            fid = val.get("file_id") if isinstance(val, dict) else (val if isinstance(val, str) else None)
            add_asset(fid, f"PM Surya - {label}", "PM Surya", details.get("completed_date") or ps.get("created_at"), location=label, stage="PM Surya")

    # 9. MSEDCL uploads
    for ms in (msedcl or []):
        details = ms.get("details") or {}
        for label, val in (details.get("photos") or details.get("attachments") or {}).items():
            fid = val.get("file_id") if isinstance(val, dict) else (val if isinstance(val, str) else None)
            add_asset(fid, f"MSEDCL - {label}", "MSEDCL", details.get("completed_date") or ms.get("created_at"), location=label, stage="MSEDCL")

    # 10. complaints — fetch all comment batches in parallel
    if complaints:
        comment_batches = await asyncio.gather(*[
            db.complaint_comments.find(
                {"company_id": company_id, "complaint_id": comp["id"]}, {"_id": 0}
            ).sort("created_at", -1).to_list(200)
            for comp in complaints
        ])
        for comp, comments in zip(complaints, comment_batches):
            for attachment in (comp.get("attachments") or []):
                fid = attachment.get("file_id") if isinstance(attachment, dict) else (attachment if isinstance(attachment, str) else None)
                add_asset(fid, "Complaint Attachment", "Complaint Center", comp.get("created_at"), stage="Complaint")
            for comm in comments:
                for attachment in (comm.get("attachments") or []):
                    fid = attachment.get("file_id") if isinstance(attachment, dict) else (attachment if isinstance(attachment, str) else None)
                    add_asset(fid, "Complaint Comment Attachment", "Complaint Center", comm.get("created_at"), stage="Complaint")

    # 11. client.documents (only image content types)
    for d in c_doc.get("documents") or []:
        if isinstance(d, dict):
            fid = d.get("file_id") or d.get("id")
            ct = (d.get("content_type") or "").lower()
            if (ct.startswith("image/") or not ct) and fid:
                add_asset(
                    fid,
                    d.get("label") or d.get("filename", "Photo"),
                    d.get("stage") or "Client Documents",
                    d.get("created_at"),
                    location=d.get("location") or "",
                    description=d.get("description") or "",
                    stage=d.get("stage") or "Client Documents"
                )

    # 12. client.photos (direct project photos)
    for p in c_doc.get("photos") or []:
        if isinstance(p, dict):
            fid = p.get("file_id") or p.get("id")
            if fid:
                stg = p.get("stage") or "Survey"
                loc = p.get("location") or ""
                lbl = f"{stg} - {loc}" if loc else stg
                add_asset(fid, lbl, stg, p.get("created_at"), location=loc, description=p.get("description") or "", stage=stg)

    return assets


class UploadClientPhotoIn(BaseModel):
    file_id: str
    stage: Optional[str] = "Survey"
    location: Optional[str] = ""
    description: Optional[str] = ""
    created_at: Optional[str] = None


@api_router.post("/clients/{client_id}/photos")
async def upload_client_photo(client_id: str, data: UploadClientPhotoIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    client = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}], "company_id": cid})
    if not client:
        client = await db.clients.find_one({"$or": [{"id": client_id}, {"sol_id": client_id}]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    target_id = client.get("id") or client_id
    photo_item = {
        "id": str(uuid.uuid4()),
        "file_id": data.file_id,
        "stage": data.stage or "Survey",
        "location": data.location or "",
        "description": data.description or "",
        "created_at": data.created_at or now_iso(),
        "uploaded_by": user.get("name", "User")
    }

    doc_item = {
        "id": photo_item["id"],
        "file_id": data.file_id,
        "label": f"{photo_item['stage']} - {photo_item['location']}" if photo_item['location'] else photo_item['stage'],
        "content_type": "image/jpeg",
        "created_at": photo_item["created_at"],
        "stage": photo_item["stage"],
        "location": photo_item["location"],
        "description": photo_item["description"],
    }

    await db.clients.update_one(
        {"$or": [{"id": target_id}, {"sol_id": target_id}]},
        {
            "$push": {
                "photos": photo_item,
                "documents": doc_item
            },
            "$set": {
                "updated_at": now_iso()
            }
        }
    )
    await log_activity(cid, user["id"], user.get("name", "User"), "Uploaded Project Photo", client.get("full_name", ""))
    return photo_item


class ClearImageIn(BaseModel):
    file_id: str


@api_router.post("/clients/{client_id}/clear-image")
async def clear_client_image(client_id: str, data: ClearImageIn, user=Depends(get_current_user)):
    client = await db.clients.find_one({"id": client_id, "company_id": user["company_id"]})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    fid = (data.file_id or "").strip()
    if not fid:
        raise HTTPException(status_code=400, detail="file_id is required")

    await db.clients.update_one(
        {"id": client_id, "company_id": user["company_id"]},
        {
            "$pull": {
                "documents": {
                    "$or": [{"file_id": fid}, {"id": fid}]
                },
                "photos": {
                    "$or": [{"file_id": fid}, {"id": fid}]
                }
            },
            "$addToSet": {
                "cleared_assets": fid
            },
            "$set": {
                "updated_at": now_iso()
            }
        }
    )
    await log_activity(user["company_id"], user["id"], user["name"], "Cleared Client Image", client.get("full_name", ""))
    return {"status": "success", "message": "Image cleared from client record", "file_id": fid}


def _summarize_inverter_status(monitoring: Optional[dict]) -> str:
    if not monitoring:
        return "Not Configured"
    s = (monitoring.get("inverter_status") or "Offline").title()
    if s not in INVERTER_STATUS_OPTIONS:
        return "Offline"
    return s


@api_router.get("/client-data/stats")
async def client_data_stats(user=Depends(get_current_user)):
    """All count queries run in parallel — 6 sequential round-trips → 1 parallel wave."""
    cid = user["company_id"]

    inv_pipeline = [
        {"$match": {"company_id": cid}},
        {"$group": {"_id": "$inverter_status", "count": {"$sum": 1}}}
    ]
    kw_pipeline = [
        {"$match": {"company_id": cid, "status": "Handover Complete"}},
        {"$group": {"_id": None, "total_kw": {"$sum": "$system_kw"}}}
    ]

    (
        total_installed, total_clients,
        inv_rows, kw_agg,
        tickets_open, tickets_closed,
    ) = await asyncio.gather(
        db.clients.count_documents({"company_id": cid, "status": "Handover Complete"}),
        db.clients.count_documents({"company_id": cid}),
        db.inverter_monitoring.aggregate(inv_pipeline).to_list(100),
        db.clients.aggregate(kw_pipeline).to_list(1),
        db.service_tickets.count_documents({"company_id": cid, "status": {"$nin": ["Closed", "Resolved"]}}),
        db.service_tickets.count_documents({"company_id": cid, "status": "Closed"}),
    )

    by_status = {row["_id"]: row["count"] for row in inv_rows}
    active_inv  = by_status.get("Online", 0)
    offline_inv = by_status.get("Offline", 0) + by_status.get("Error", 0)
    total_kw    = float(kw_agg[0]["total_kw"]) if kw_agg else 0

    return {
        "total_clients": total_clients,
        "total_installed": total_installed,
        "active_inverters": active_inv,
        "offline_inverters": offline_inv,
        "tickets_open": tickets_open,
        "tickets_closed": tickets_closed,
        "total_capacity_kw": round(total_kw, 2),
    }


@api_router.get("/client-data/clients")
async def list_client_data(
    user=Depends(get_current_user),
    search: Optional[str] = None,
    consumer: Optional[str] = None,
    mobile: Optional[str] = None,
    city: Optional[str] = None,
    capacity_min: Optional[float] = None,
    capacity_max: Optional[float] = None,
    status: Optional[str] = None,
    stage: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    cid = user["company_id"]
    q: Dict[str, Any] = {
        "company_id": cid,
    }
    INVERTER_STATUSES = {"Online", "Offline", "Error", "Maintenance", "Not Configured"}
    if status and status != "all":
        if status in INVERTER_STATUSES:
            pass # Filtered in memory below based on inv_status
        else:
            q["status"] = status
    if search: q["full_name"] = {"$regex": re.escape(search), "$options": "i"}
    if consumer: q["consumer_number"] = {"$regex": re.escape(consumer), "$options": "i"}
    if mobile: q["mobile"] = {"$regex": re.escape(mobile)}
    if city: q["city"] = {"$regex": re.escape(city), "$options": "i"}
    if capacity_min is not None: q.setdefault("system_kw", {})["$gte"] = capacity_min
    if capacity_max is not None: q.setdefault("system_kw", {})["$lte"] = capacity_max
    if from_date: q.setdefault("updated_at", {})["$gte"] = from_date
    if to_date: q.setdefault("updated_at", {})["$lte"] = to_date

    # Lean projection — only fields needed for the client data list view
    list_projection = {
        "_id": 0, "id": 1, "sol_id": 1, "full_name": 1, "consumer_number": 1,
        "mobile": 1, "alt_mobile": 1, "city": 1, "state": 1,
        "updated_at": 1, "system_kw": 1, "panel_make": 1, "inverter_make": 1,
        "inverter_capacity": 1, "stages": 1, "status": 1,
    }
    logger.info(f"[DIAG] list_client_data: company_id={cid!r}, query={q!r}")
    clients = await db.clients.find(q, list_projection).sort("updated_at", -1).to_list(500)
    logger.info(f"[DIAG] list_client_data: raw DB returned {len(clients)} clients")

    # Ensure valid string ID for every client document (prefers id, fallback _id, sol_id)
    for c in clients:
        c["id"] = str(c.get("id") or c.get("_id") or c.get("sol_id") or "")

    if stage and stage != "all":
        clients = [c for c in clients if _client_current_stage(c) == stage]

    if not clients:
        return []

    # Fire monitoring, tickets aggregation, and tasks lookup in parallel
    ids = [c["id"] for c in clients if c.get("id")]
    tickets_pipeline = [
        {"$match": {"company_id": cid, "client_id": {"$in": ids}, "status": {"$nin": ["Closed", "Resolved"]}}},
        {"$group": {"_id": "$client_id", "n": {"$sum": 1}}}
    ]
    monitoring_rows, ticket_rows, task_rows = await asyncio.gather(
        db.inverter_monitoring.find({"company_id": cid, "client_id": {"$in": ids}}, {"_id": 0, "client_id": 1, "inverter_status": 1}).to_list(500),
        db.service_tickets.aggregate(tickets_pipeline).to_list(500),
        db.tasks.find({"company_id": cid, "client_id": {"$in": ids}, "status": {"$ne": "completed"}}, {"_id": 0, "client_id": 1, "assigned_to_name": 1}).to_list(2000),
    )

    monitorings: Dict[str, dict] = {m["client_id"]: m for m in monitoring_rows}
    tickets_count: Dict[str, int] = {row["_id"]: row["n"] for row in ticket_rows}
    assigned_team: Dict[str, set] = {}
    for task in task_rows:
        if task.get("assigned_to_name"):
            assigned_team.setdefault(task["client_id"], set()).add(task["assigned_to_name"])

    out = []
    for c in clients:
        m = monitorings.get(c["id"])
        inv_status = _summarize_inverter_status(m)
        if status and status != "all" and status in INVERTER_STATUSES:
            if inv_status != status:
                continue
        out.append({
            "id": c["id"],
            "client_code": c.get("sol_id"),
            "sol_id": c.get("sol_id"),
            "full_name": c.get("full_name"),
            "consumer_number": c.get("consumer_number"),
            "mobile": c.get("mobile"),
            "alt_mobile": c.get("alt_mobile"),
            "city": c.get("city"),
            "state": c.get("state"),
            "installation_date": c.get("install_date") or c.get("updated_at"),
            "system_kw": c.get("system_kw") or 0,
            "panel_make": c.get("panel_make"),
            "inverter_make": c.get("inverter_make"),
            "inverter_capacity": c.get("inverter_capacity"),
            "inverter_status": inv_status,
            "open_tickets": tickets_count.get(c["id"], 0),
            "last_updated": c.get("updated_at"),
            "current_stage": _client_current_stage(c),
            "assigned_team": sorted(assigned_team.get(c["id"], [])),
            "status": c.get("status"),
        })
    return out


@api_router.get("/client-data/{client_id}")
@api_router.get("/client-data/clients/{client_id}")
async def get_client_data_detail(
    client_id: str,
    tab: Optional[str] = "all",
    user=Depends(get_current_user)
):
    cid = user["company_id"]
    or_conds: List[Dict[str, Any]] = [
        {"id": client_id},
        {"sol_id": client_id},
        {"consumer_number": client_id}
    ]
    if len(client_id) == 24:
        try:
            from bson import ObjectId
            or_conds.append({"_id": ObjectId(client_id)})
        except Exception:
            pass

    c = None
    
    try:
        c = await db.clients.find_one({"$or": or_conds, "company_id": cid})
    except Exception as err:
        logger.warning(f"OR query in get_client_data_detail: {err}")

    if not c:
        try:
            c = await db.clients.find_one({"id": client_id, "company_id": cid})
        except Exception:
            pass

    if not c:
        try:
            c = await db.clients.find_one({"$or": or_conds})
        except Exception:
            pass

    if not c:
        try:
            c = await db.clients.find_one({"id": client_id})
        except Exception:
            pass

    if not c:
        raise HTTPException(status_code=404, detail="Client not found")

    raw_mongo_id = str(c.get("_id")) if c.get("_id") else None
    doc_id = c.get("id")
    canonical_id = str(doc_id or raw_mongo_id or c.get("sol_id") or client_id)

    c["id"] = canonical_id
    c.pop("_id", None)
    c["client_code"] = c.get("sol_id")
    c = _enrich_client_doc(c)

    client_ids = list({x for x in [client_id, canonical_id, doc_id, raw_mongo_id, c.get("sol_id")] if x})
    q = {"company_id": cid, "client_id": {"$in": client_ids}}

    # Build list of coroutines based on which tab is requested, then gather them all.
    # This turns N sequential round-trips into a single parallel wave.
    coros: Dict[str, Any] = {}

    if tab in ("all", "info", "monitoring"):
        coros["monitoring"] = db.inverter_monitoring.find_one({**q}, {"_id": 0})
    if tab in ("all", "assets"):
        coros["assets"] = _attach_assets(client_id, cid)
    if tab in ("all", "tickets"):
        coros["tickets"] = db.service_tickets.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    if tab in ("all", "survey"):
        async def get_client_surveys():
            raw_surveys = await db.surveys.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
            completed_tasks = await db.tasks.find({
                "company_id": cid,
                "client_id": client_id,
                "task_type": "Survey",
                "status": "completed"
            }, {"_id": 0}).to_list(100)
            
            existing_task_ids = {s.get("task_id") for s in raw_surveys if s.get("task_id")}
            for t in completed_tasks:
                if t.get("id") not in existing_task_ids:
                    sub = t.get("submission") or {}
                    photos = sub.get("photos") or sub.get("attachments") or {}
                    doc = {
                        "id": str(uuid.uuid4()),
                        "company_id": cid,
                        "client_id": client_id,
                        "project_id": client_id,
                        "task_id": t["id"],
                        "employee_id": t.get("assigned_to"),
                        "details": {
                            "completed_by": t.get("assigned_to_name") or "",
                            "completed_by_id": t.get("assigned_to") or "",
                            "assigned_to_name": t.get("assigned_to_name") or "",
                            "assigned_by": t.get("assigned_by_name") or "",
                            "assigned_by_name": t.get("assigned_by_name") or "",
                            "completed_date": sub.get("submitted_at") or t.get("updated_at") or now_iso(),
                            "submitted_at": sub.get("submitted_at") or t.get("updated_at") or now_iso(),
                            "notes": sub.get("notes") or sub.get("remarks") or t.get("remarks") or "",
                            "gps": sub.get("gps") or "",
                            "manual_location": sub.get("manual_location") or "",
                            "checklist": sub.get("checklist") or [],
                            "photos": photos,
                            "attachments": photos,
                            "task_status": "completed",
                        },
                        "created_at": t.get("created_at") or now_iso(),
                        "updated_at": t.get("updated_at") or now_iso(),
                    }
                    raw_surveys.append(doc)
                    try:
                        await db.surveys.insert_one(doc)
                    except Exception:
                        pass

            for s in raw_surveys:
                d = s.get("details") or {}
                photos = d.get("photos") or d.get("attachments") or {}
                d["photos"] = photos
                d["attachments"] = photos
                s["details"] = d
            return raw_surveys

        coros["surveys"] = get_client_surveys()
    if tab in ("all", "material"):
        coros["material_deliveries"] = db.material_deliveries.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "documents"):
        async def get_client_documents():
            raw_docs = await db.documents.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
            existing_ids = {d.get("id") or d.get("file_id") for d in raw_docs if (d.get("id") or d.get("file_id"))}
            
            # Fetch completed tasks with uploaded files/documents for this client
            completed_tasks = await db.tasks.find({
                "company_id": cid,
                "client_id": client_id,
                "status": "completed"
            }, {"_id": 0}).to_list(100)

            for t in completed_tasks:
                sub = t.get("submission") or {}
                file_id = sub.get("file_id") or sub.get("id")
                filename = sub.get("filename") or sub.get("original_filename") or f"{t.get('task_type')}.pdf"
                if file_id and file_id not in existing_ids:
                    doc = {
                        "id": file_id,
                        "company_id": cid,
                        "client_id": client_id,
                        "label": t.get("title") or t.get("task_type") or "Uploaded Document",
                        "filename": filename,
                        "created_at": sub.get("submitted_at") or t.get("updated_at") or now_iso(),
                        "details": {
                            "completed_date": sub.get("submitted_at") or t.get("updated_at") or now_iso(),
                            "completed_by": t.get("assigned_to_name") or "Employee",
                            "attachments": { (t.get("title") or t.get("task_type") or "Document"): file_id }
                        }
                    }
                    raw_docs.append(doc)
                    existing_ids.add(file_id)

            return raw_docs
        coros["documents"] = get_client_documents()
    if tab in ("all", "meter"):
        coros["meter_testings"] = db.meter_testings.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "installation"):
        coros["installations"] = db.installations.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "verification"):
        coros["verifications"] = db.verifications.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "handover"):
        coros["handovers"] = db.handovers.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "material_history"):
        coros["material_requests_raw"] = db.material_requests.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "tasks"):
        coros["tasks"] = db.tasks.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    if tab in ("all", "inward"):
        async def get_client_inwards():
            raw_inwards = await db.inward_entries.find({
                "company_id": cid,
                "source_type": "Return From Client"
            }, {"_id": 0}).sort("date", -1).to_list(1000)
            res = []
            for inv in raw_inwards:
                inv = parse_inward_client_info(inv)
                if inv.get("client_id") == client_id:
                    res.append(inv)
            return res
        coros["inward"] = get_client_inwards()
    if tab in ("all", "outward"):
        coros["outward"] = db.outward_entries.find(q, {"_id": 0}).sort("date", -1).to_list(100)
    if tab in ("all", "activity_logs"):
        coros["activity_logs"] = db.activity_logs.find({"company_id": cid, "target": c.get("full_name") or ""}, {"_id": 0}).sort("created_at", -1).to_list(100)

    results: Dict[str, Any] = {}
    if coros:
        keys = list(coros.keys())
        values = await asyncio.gather(*[coros[k] for k in keys], return_exceptions=True)
        for k, v in zip(keys, values):
            if isinstance(v, Exception):
                logger.error(f"[ERROR] get_client_data_detail sub-section '{k}' failed: {v!r}")
                results[k] = None if k == "monitoring" else []
            else:
                results[k] = v

    monitoring          = results.get("monitoring")
    assets              = results.get("assets", [])
    tickets             = results.get("tickets", [])
    surveys             = results.get("surveys", [])
    material_deliveries = results.get("material_deliveries", [])
    documents           = results.get("documents", [])
    meter_testings      = results.get("meter_testings", [])
    installations       = results.get("installations", [])
    verifications       = results.get("verifications", [])
    material_requests_raw = results.get("material_requests_raw", [])
    material_requests   = await _enrich_requests_with_stock_batch(material_requests_raw, cid) if material_requests_raw else []
    hva = [a for a in _load_local_assets() if a.get("client_id") == client_id and a.get("company_id") == cid] if tab in ("all", "hva") else []
        
    handovers_list = list(results.get("handovers") or [])
    if not handovers_list and c.get("handover_data"):
        hd = c.get("handover_data") or {}
        synth_photos = {}
        if hd.get("handover_photo_id"):
            synth_photos["Handover Photo"] = hd.get("handover_photo_id")
        synth_chk = []
        if hd.get("declaration"):
            synth_chk.append({"label": "Owner Declaration Confirmed", "checked": True})
        if hd.get("owner_name"):
            synth_chk.append({"label": f"Owner: {hd.get('owner_name')}", "checked": True})
        handovers_list.append({
            "id": f"handover-{client_id}",
            "company_id": cid,
            "client_id": client_id,
            "created_at": hd.get("handover_date") or c.get("updated_at") or now_iso(),
            "details": {
                "completed_date": hd.get("handover_date"),
                "completed_by": c.get("assigned_engineer_name") or c.get("engineer_name") or "",
                "notes": hd.get("declaration") or "",
                "photos": synth_photos,
                "attachments": synth_photos,
                "checklist": synth_chk,
                "owner_name": hd.get("owner_name") or "",
                "declaration_confirmed": True if hd.get("declaration") else False,
            }
        })

    existing_tids = {h.get("task_id") for h in handovers_list if h.get("task_id")}
    for t in (results.get("tasks") or []):
        if (t.get("task_type") == "Handover" or t.get("workflow") == "handover") and t.get("status") == "completed" and t.get("id") not in existing_tids:
            sub = t.get("submission") or {}
            if sub:
                t_photos = {}
                h_photo = sub.get("handover_photo_id") or sub.get("photo_id") or sub.get("file_id")
                if h_photo:
                    t_photos["Handover Photo"] = h_photo
                t_chk = list(sub.get("checklist") or [])
                if sub.get("declaration_confirmed"):
                    t_chk.append({"label": "Owner Declaration Confirmed", "checked": True})
                if sub.get("installer_confirmed"):
                    t_chk.append({"label": "Installer Confirmed", "checked": True})
                if sub.get("owner_name"):
                    t_chk.append({"label": f"Owner: {sub.get('owner_name')}", "checked": True})
                handovers_list.append({
                    "id": f"handover-task-{t.get('id')}",
                    "company_id": cid,
                    "client_id": client_id,
                    "task_id": t.get("id"),
                    "created_at": sub.get("handover_date") or t.get("updated_at") or now_iso(),
                    "details": {
                        "completed_date": sub.get("handover_date") or sub.get("submitted_at") or t.get("updated_at"),
                        "completed_by": t.get("assigned_to_name") or "",
                        "assigned_by": t.get("assigned_by_name") or "",
                        "notes": sub.get("notes") or sub.get("remarks") or t.get("remarks") or "",
                        "photos": t_photos,
                        "attachments": t_photos,
                        "checklist": t_chk,
                        "owner_name": sub.get("owner_name") or "",
                        "declaration_confirmed": sub.get("declaration_confirmed"),
                        "installer_confirmed": sub.get("installer_confirmed"),
                    }
                })

    return {
        "client": c,
        "monitoring": monitoring,
        "assets": assets,
        "high_value_assets": hva,
        "tickets": tickets,
        "surveys": surveys,
        "material_deliveries": material_deliveries,
        "documents": documents,
        "meter_testings": meter_testings,
        "installations": installations,
        "verifications": verifications,
        "handovers": handovers_list,
        "material_requests": material_requests,
        "tasks": results.get("tasks", []),
        "inward": results.get("inward", []),
        "outward": results.get("outward", []),
        "activity_logs": results.get("activity_logs", []),
        "inverter_status": _summarize_inverter_status(monitoring),
    }



@api_router.put("/client-data/clients/{client_id}/monitoring")
async def upsert_monitoring(client_id: str, data: MonitoringIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    c = await db.clients.find_one({"id": client_id, "company_id": cid}, {"_id": 0, "full_name": 1})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    status_val = (data.inverter_status or "Offline").title()
    if status_val not in INVERTER_STATUS_OPTIONS:
        status_val = "Offline"
    patch = {
        "portal_name": data.portal_name or "", "app_name": data.app_name or "",
        "portal_url": data.portal_url or "", "plant_id": data.plant_id or "",
        "username": data.username or "", "password": data.password or "",
        "inverter_status": status_val, "notes": data.notes or "",
        "updated_at": now_iso(), "updated_by": user["id"], "updated_by_name": user["name"],
    }
    existing = await db.inverter_monitoring.find_one({"company_id": cid, "client_id": client_id})
    if existing:
        await db.inverter_monitoring.update_one({"id": existing["id"]}, {"$set": patch})
    else:
        await db.inverter_monitoring.insert_one({
            "id": str(uuid.uuid4()), "company_id": cid, "client_id": client_id,
            "created_at": now_iso(), "created_by": user["id"], "created_by_name": user["name"],
            **patch,
        })
    await log_activity(cid, user["id"], user["name"], "Monitoring Updated", f"{c.get('full_name','')} → {status_val}")
    saved = await db.inverter_monitoring.find_one({"company_id": cid, "client_id": client_id}, {"_id": 0})
    return saved


# -------- Service Tickets --------

@api_router.post("/service-tickets")
async def create_ticket(data: TicketIn, user=Depends(get_current_user)):
    if not has_perm(user, "client_data", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: client_data.create")
    cid = user["company_id"]
    if data.priority not in TICKET_PRIORITIES:
        raise HTTPException(status_code=400, detail="Invalid priority")
    if data.issue_type not in TICKET_ISSUE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid issue type")
    c = await db.clients.find_one({"id": data.client_id, "company_id": cid}, {"_id": 0, "full_name": 1, "mobile": 1})
    if not c:
        raise HTTPException(status_code=404, detail="Client not found")
    ticket_no = await _next_ticket_no(cid)
    ts = now_iso()
    doc = {
        "id": str(uuid.uuid4()), "company_id": cid, "ticket_no": ticket_no,
        "client_id": data.client_id, "client_name": c.get("full_name", ""), "client_mobile": c.get("mobile", ""),
        "title": data.title.strip(), "issue_type": data.issue_type,
        "description": data.description or "", "priority": data.priority,
        "status": "Open", "assigned_to": "", "assigned_to_name": "",
        "attachments": data.attachments or [],
        "timeline": [{
            "ts": ts, "user_id": user["id"], "user_name": user["name"],
            "action": "Ticket Created", "from_status": "", "to_status": "Open",
            "note": data.description or "",
        }],
        "resolution": "",
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": ts, "updated_at": ts,
    }
    await db.service_tickets.insert_one(doc); doc.pop("_id", None)
    await log_activity(cid, user["id"], user["name"], "Service Ticket Created", f"{ticket_no} · {c.get('full_name','')} · {data.priority}")
    await push_notification(cid, "admin", f"New Ticket {ticket_no}", f"{data.title} for {c.get('full_name','')} [{data.priority}]")
    return doc


@api_router.get("/service-tickets")
async def list_tickets(
    user=Depends(get_current_user),
    client_id: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assigned_to: Optional[str] = None,
    search: Optional[str] = None,
):
    cid = user["company_id"]
    q: Dict[str, Any] = {"company_id": cid}
    if client_id: q["client_id"] = client_id
    if status: q["status"] = status
    if priority: q["priority"] = priority
    if assigned_to: q["assigned_to"] = assigned_to
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        q["$or"] = [{"title": rx}, {"ticket_no": rx}, {"client_name": rx}]
    return await db.service_tickets.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.get("/service-tickets/{ticket_id}")
async def get_ticket(ticket_id: str, user=Depends(get_current_user)):
    t = await db.service_tickets.find_one({"id": ticket_id, "company_id": user["company_id"]}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return t


@api_router.patch("/service-tickets/{ticket_id}")
async def update_ticket(ticket_id: str, data: TicketUpdate, user=Depends(get_current_user)):
    if not has_perm(user, "client_data", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: client_data.edit")
    cid = user["company_id"]
    t = await db.service_tickets.find_one({"id": ticket_id, "company_id": cid}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket not found")
    patch: Dict[str, Any] = {"updated_at": now_iso()}
    timeline_entry: Dict[str, Any] = {"ts": now_iso(), "user_id": user["id"], "user_name": user["name"], "action": "Updated", "from_status": t.get("status", ""), "note": data.note or ""}

    if data.status is not None:
        if data.status not in TICKET_STATUSES:
            raise HTTPException(status_code=400, detail="Invalid status")
        patch["status"] = data.status
        timeline_entry["to_status"] = data.status
        timeline_entry["action"] = f"Status → {data.status}"
        if data.status == "Resolved":
            patch["resolved_at"] = now_iso()
        if data.status == "Closed":
            patch["closed_at"] = now_iso()
    if data.priority is not None:
        if data.priority not in TICKET_PRIORITIES:
            raise HTTPException(status_code=400, detail="Invalid priority")
        patch["priority"] = data.priority
        timeline_entry["action"] = f"Priority → {data.priority}"
    if data.assigned_to is not None:
        patch["assigned_to"] = data.assigned_to
        if data.assigned_to:
            emp = await db.employees.find_one({"id": data.assigned_to, "company_id": cid}, {"_id": 0, "name": 1})
            patch["assigned_to_name"] = (emp or {}).get("name", "")
            timeline_entry["action"] = f"Assigned to {patch['assigned_to_name'] or 'engineer'}"
            if t.get("status") in ("Open", ""):
                patch["status"] = "Assigned"
                timeline_entry["to_status"] = "Assigned"
        else:
            patch["assigned_to_name"] = ""
            timeline_entry["action"] = "Unassigned"
    if data.resolution is not None:
        patch["resolution"] = data.resolution
        timeline_entry["note"] = data.resolution
    if data.attachments is not None:
        # append, don't replace
        patch["attachments"] = (t.get("attachments") or []) + list(data.attachments)
        timeline_entry["action"] = f"Attached {len(data.attachments)} file(s)"

    timeline = list(t.get("timeline") or [])
    timeline.append(timeline_entry)
    patch["timeline"] = timeline

    await db.service_tickets.update_one({"id": ticket_id, "company_id": cid}, {"$set": patch})
    new_t = await db.service_tickets.find_one({"id": ticket_id, "company_id": cid}, {"_id": 0})
    if new_t:
        await log_activity(cid, user["id"], user["name"], f"Ticket {new_t['ticket_no']} {timeline_entry['action']}", new_t.get("client_name", ""))
        # Notify assignee — use push_notification helper (correct schema); ignore errors for non-existent users
        if data.assigned_to and data.assigned_to != t.get("assigned_to"):
            try:
                await push_notification(
                    cid, "user",
                    f"Ticket {new_t['ticket_no']} assigned",
                    f"{new_t.get('title')} · {new_t.get('client_name')} [{new_t.get('priority')}]",
                    to_user_id=data.assigned_to,
                )
            except Exception:
                pass  # Don't let notification failure break the ticket update
    return new_t


@api_router.get("/client-data/export.csv")
async def export_clients_csv(user=Depends(get_current_user)):
    """Download CSV (Excel-compatible) of all handed-over clients."""
    items = await list_client_data(user=user)  # type: ignore
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Client Code", "Client Name", "Consumer Number", "Mobile", "Alt Mobile", "City", "State",
                     "Installation Date", "Capacity (kW)", "Panel Brand", "Inverter Brand", "Inverter kW",
                     "Inverter Status", "Open Tickets", "Last Updated"])
    for r in items:
        writer.writerow([
            r.get("client_code", ""), r.get("full_name", ""), r.get("consumer_number", ""),
            r.get("mobile", ""), r.get("alt_mobile", ""), r.get("city", ""), r.get("state", ""),
            r.get("installation_date", ""), r.get("system_kw", 0),
            r.get("panel_make", ""), r.get("inverter_make", ""), r.get("inverter_capacity", ""),
            r.get("inverter_status", ""), r.get("open_tickets", 0), r.get("last_updated", ""),
        ])
    from fastapi.responses import StreamingResponse
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="solarix-client-data.csv"'},
    )


@api_router.get("/")
async def root():
    return {"message": "Solarix API", "version": "1.0"}


@api_router.get("/health")
@app.get("/health")
@app.get("/api/health")
async def health_check():
    return {"status": "ok", "service": "SOLARIX API", "timestamp": datetime.now(timezone.utc).isoformat()}


# ---------- Sprint 3: DOCX Template Engine ----------
import docx_template_engine as docx_engine

DOC_TYPES = ["Annexure", "WCR", "SLDR", "Net Metering Agreement", "Vendor Agreement", "Quotation", "Other"]

class TemplateCreate(BaseModel):
    name: str
    doc_type: str = "Other"

class TemplateUpdate(BaseModel):
    name: Optional[str] = None
    doc_type: Optional[str] = None
    mapping: Optional[Dict[str, str]] = None

class TemplateGenerate(BaseModel):
    client_id: str
    overrides: Optional[Dict[str, Any]] = None  # canonical_var → value
    raw_overrides: Optional[Dict[str, str]] = None  # placeholder string → value (for unmapped)
    save_to_client: bool = True

@api_router.get("/document-templates/variables")
async def list_template_variables(user=Depends(get_current_user)):
    """Return the catalogue of canonical system variables for use in the mapping UI."""
    return {"variables": docx_engine.SYSTEM_VARIABLES}

@api_router.post("/document-templates")
async def upload_template(file: UploadFile = File(...), name: str = Form(...), doc_type: str = Form("Other"), user=Depends(get_current_user)):
    if not has_perm(user, "documents", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: documents.create")
    if doc_type not in DOC_TYPES:
        doc_type = "Other"
    if not (file.filename or "").lower().endswith(".docx"):
        raise HTTPException(status_code=400, detail="Only .docx files are supported. Please save .doc files as .docx in Word and re-upload.")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 10 MB)")
    try:
        placeholders = docx_engine.extract_placeholders(content)
    except Exception as e:
        logger.exception("Could not parse .docx template file")
        raise HTTPException(status_code=400, detail="Could not parse .docx template. Please verify document structure.")

    suggested_mapping = docx_engine.suggest_mapping(placeholders)

    # Store the original docx in Emergent Object Storage
    template_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/{user['company_id']}/templates/{template_id}.docx"
    put_result = put_object(
        storage_path, content,
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

    file_id = str(uuid.uuid4())
    await db.files.insert_one({
        "id": file_id, "company_id": user["company_id"], "uploader_id": user["id"],
        "storage_path": put_result["path"], "original_filename": file.filename,
        "content_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "size": put_result.get("size", len(content)), "category": "template",
        "is_deleted": False, "created_at": now_iso(),
    })

    tpl_doc = {
        "id": template_id, "company_id": user["company_id"],
        "name": name.strip() or file.filename, "doc_type": doc_type,
        "file_id": file_id, "storage_path": put_result["path"],
        "filename": file.filename,
        "placeholders": placeholders,
        "mapping": suggested_mapping,
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.document_templates.insert_one(tpl_doc); tpl_doc.pop("_id", None)
    await log_activity(user["company_id"], user["id"], user["name"], "Template Uploaded", f"{name} ({len(placeholders)} fields)")
    return tpl_doc

@api_router.get("/document-templates")
async def list_templates(user=Depends(get_current_user)):
    items = await db.document_templates.find({"company_id": user["company_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items

@api_router.get("/document-templates/{tpl_id}")
async def get_template(tpl_id: str, user=Depends(get_current_user)):
    t = await db.document_templates.find_one({"id": tpl_id, "company_id": user["company_id"]}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    return t

@api_router.patch("/document-templates/{tpl_id}")
async def update_template(tpl_id: str, data: TemplateUpdate, user=Depends(get_current_user)):
    if not has_perm(user, "documents", "edit"):
        raise HTTPException(status_code=403, detail="Missing permission: documents.edit")
    t = await db.document_templates.find_one({"id": tpl_id, "company_id": user["company_id"]}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    patch: Dict[str, Any] = {"updated_at": now_iso()}
    if data.name is not None: patch["name"] = data.name
    if data.doc_type is not None and data.doc_type in DOC_TYPES: patch["doc_type"] = data.doc_type
    if data.mapping is not None: patch["mapping"] = data.mapping
    await db.document_templates.update_one({"id": tpl_id, "company_id": user["company_id"]}, {"$set": patch})
    t.update(patch)
    return t

@api_router.delete("/document-templates/{tpl_id}")
async def delete_template(tpl_id: str, user=Depends(get_current_user)):
    if not has_perm(user, "documents", "delete"):
        raise HTTPException(status_code=403, detail="Missing permission: documents.delete")
    res = await db.document_templates.delete_one({"id": tpl_id, "company_id": user["company_id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"ok": True}

async def enrich_client_doc_for_docs(client_doc: dict, company_id: str) -> dict:
    client_id = client_doc.get("id")
    if not client_id:
        return client_doc

    enriched = dict(client_doc)

    # 1. Fetch latest survey
    survey = await db.surveys.find_one({"company_id": company_id, "client_id": client_id}, sort=[("created_at", -1)])
    if survey:
        details = survey.get("details") or {}
        for k, v in details.items():
            norm_k = f"survey_{k}"
            if norm_k not in enriched:
                enriched[norm_k] = v
            if k not in enriched or enriched[k] in (None, ""):
                enriched[k] = v
        survey_date = details.get("submitted_at") or survey.get("created_at")
        if survey_date and ("survey_date" not in enriched or enriched["survey_date"] in (None, "")):
            enriched["survey_date"] = survey_date

    # 2. Fetch latest installation
    installation = await db.installations.find_one({"company_id": company_id, "client_id": client_id}, sort=[("created_at", -1)])
    if installation:
        details = installation.get("details") or {}
        for k, v in details.items():
            norm_k = f"installation_{k}"
            if norm_k not in enriched:
                enriched[norm_k] = v
            if k not in enriched or enriched[k] in (None, ""):
                enriched[k] = v
        install_date = details.get("submitted_at") or installation.get("created_at")
        if install_date:
            if "installation_date" not in enriched or enriched["installation_date"] in (None, ""):
                enriched["installation_date"] = install_date
            if "install_date" not in enriched or enriched["install_date"] in (None, ""):
                enriched["install_date"] = install_date
        installer = details.get("assigned_to_name") or installation.get("employee_id")
        if installer:
            if "installer" not in enriched or enriched["installer"] in (None, ""):
                enriched["installer"] = installer
            if "installer_name" not in enriched or enriched["installer_name"] in (None, ""):
                enriched["installer_name"] = installer

    # 3. Fetch latest verification
    verification = await db.verifications.find_one({"company_id": company_id, "client_id": client_id}, sort=[("created_at", -1)])
    if verification:
        details = verification.get("details") or {}
        for k, v in details.items():
            norm_k = f"verification_{k}"
            if norm_k not in enriched:
                enriched[norm_k] = v
            if k not in enriched or enriched[k] in (None, ""):
                enriched[k] = v
        inverters = details.get("inverters") or []
        if inverters:
            serials = [inv.get("serial") for inv in inverters if inv.get("serial")]
            if serials:
                serials_str = ", ".join(serials)
                if "inverter_serial_numbers" not in enriched:
                    enriched["inverter_serial_numbers"] = serials_str
                if "inverter_serial" not in enriched or enriched["inverter_serial"] in (None, ""):
                    enriched["inverter_serial"] = serials_str
        net_meter = details.get("net_meter_number") or details.get("net_meter")
        if net_meter and ("net_meter_number" not in enriched or enriched["net_meter_number"] in (None, "")):
            enriched["net_meter_number"] = net_meter
        meter = details.get("meter_number") or details.get("meter")
        if meter and ("meter_number" not in enriched or enriched["meter_number"] in (None, "")):
            enriched["meter_number"] = meter

    # 4. Fetch monitoring
    monitoring = await db.inverter_monitoring.find_one({"company_id": company_id, "client_id": client_id}, {"_id": 0})
    if monitoring:
        for k, v in monitoring.items():
            if k not in enriched or enriched[k] in (None, ""):
                enriched[k] = v
        if "inverter_model" not in enriched or enriched["inverter_model"] in (None, ""):
            enriched["inverter_model"] = monitoring.get("inverter_model") or monitoring.get("app_name")

    # 5. Extract Latitude & Longitude
    gps_val = enriched.get("gps") or enriched.get("survey_gps") or (survey.get("details", {}).get("gps") if survey else None)
    if gps_val and isinstance(gps_val, str) and "," in gps_val:
        parts = [p.strip() for p in gps_val.split(",")]
        if len(parts) >= 2:
            if "latitude" not in enriched or enriched["latitude"] in (None, ""):
                enriched["latitude"] = parts[0]
            if "longitude" not in enriched or enriched["longitude"] in (None, ""):
                enriched["longitude"] = parts[1]

    if "vendor" not in enriched or enriched["vendor"] in (None, ""):
        enriched["vendor"] = enriched.get("company_name")

    return enriched

@api_router.post("/document-templates/{tpl_id}/preview")
async def preview_template(tpl_id: str, data: TemplateGenerate, user=Depends(get_current_user)):
    """Return per-placeholder resolved values for the generation dialog."""
    t = await db.document_templates.find_one({"id": tpl_id, "company_id": user["company_id"]}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    client_doc = await db.clients.find_one({"id": data.client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Enrich client document with related records for template rendering
    client_doc = await enrich_client_doc_for_docs(client_doc, user["company_id"])

    company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
    overrides = dict(data.overrides or {})
    overrides["__raw__"] = data.raw_overrides or {}
    preview = docx_engine.compute_preview(t.get("placeholders") or [], t.get("mapping") or {}, client_doc, company_doc, overrides)
    return {"template_id": tpl_id, "rows": preview, "missing_count": sum(1 for r in preview if r["missing"])}

@api_router.post("/document-templates/{tpl_id}/generate")
async def generate_template(tpl_id: str, data: TemplateGenerate, user=Depends(get_current_user)):
    t = await db.document_templates.find_one({"id": tpl_id, "company_id": user["company_id"]}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    client_doc = await db.clients.find_one({"id": data.client_id, "company_id": user["company_id"]}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Enrich client document with related records for template rendering
    client_doc = await enrich_client_doc_for_docs(client_doc, user["company_id"])

    company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}

    # Pull source docx from storage
    src_bytes, _ct = get_object(t["storage_path"])

    overrides = dict(data.overrides or {})
    overrides["__raw__"] = data.raw_overrides or {}

    try:
        filled = docx_engine.render_docx(
            src_bytes, t.get("placeholders") or [], t.get("mapping") or {},
            client_doc, company_doc, overrides,
        )
    except Exception as e:
        logger.exception("Template render failed")
        raise HTTPException(status_code=500, detail="Document rendering failed. Please check template configuration.")

    safe_client = re.sub(r"[^A-Za-z0-9_-]+", "-", client_doc.get("full_name", "client")).strip("-")[:40] or "client"
    safe_tpl = re.sub(r"[^A-Za-z0-9_-]+", "-", t.get("name", "template")).strip("-")[:40] or "template"
    filename = f"{safe_client}-{safe_tpl}.docx"

    if data.save_to_client:
        stages = {**(client_doc.get("stages") or {}), "Document Making": True, "Onboarding": True}
        await db.clients.update_one(
            {"id": data.client_id, "company_id": user["company_id"]},
            {"$set": {"stages": stages, "progress": calc_progress(stages), "updated_at": now_iso()}}
        )
    await log_activity(user["company_id"], user["id"], user["name"], f"Generated {t.get('name')}", client_doc.get("full_name", ""))

    from fastapi.responses import Response
    return Response(
        content=filled,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
            "X-Filename": filename,
            "X-Label": t.get("name") or "Template",
        },
    )


# ---------- Complaint Management ----------
COMPLAINT_CATEGORIES = [
    "Installation Issue", "Material Issue", "Customer Complaint", "Document Issue",
    "Inverter Issue", "Service Issue", "Payment Issue", "Team Issue", "Other",
]
COMPLAINT_PRIORITIES = ["Low", "Medium", "High", "Urgent"]
COMPLAINT_STATUSES = ["Open", "Assigned", "In Progress", "Waiting", "Resolved", "Closed"]
SEND_TO_TARGETS = ["Admin", "Installer Team", "Document Team", "Supervisor", "Inventory Team", "Specific User"]


class ComplaintIn(BaseModel):
    title: str
    category: str
    priority: str = "Medium"
    description: Optional[str] = ""
    client_id: Optional[str] = ""
    project_id: Optional[str] = ""
    send_to_target: str
    assigned_to: Optional[str] = ""  # required when send_to_target == "Specific User"
    attachments: Optional[List[Dict[str, Any]]] = None


class ComplaintUpdate(BaseModel):
    title: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[str] = None
    send_to_target: Optional[str] = None
    resolution_note: Optional[str] = None
    resolution_attachments: Optional[List[Dict[str, Any]]] = None


class ComplaintCommentIn(BaseModel):
    text: str
    attachments: Optional[List[Dict[str, Any]]] = None


async def next_complaint_id(company_id: str) -> str:
    year = datetime.now(timezone.utc).year
    res = await db.counters.find_one_and_update(
        {"company_id": company_id, "year": year, "type": "complaint"},
        {"$inc": {"seq": 1}}, upsert=True, return_document=True,
    )
    seq = res["seq"] if isinstance(res, dict) and "seq" in res else 1
    return f"CMP-{year}-{seq:04d}"


async def write_complaint_audit(complaint_id: str, company_id: str, user_id: str, user_name: str, action: str, details: str = ""):
    await db.complaint_audit.insert_one({
        "id": str(uuid.uuid4()),
        "complaint_id": complaint_id,
        "company_id": company_id,
        "user_id": user_id, "user_name": user_name,
        "action": action, "details": details,
        "created_at": now_iso(),
    })


def compute_escalation(complaint: dict) -> str:
    """Return 'red' (>=48h), 'yellow' (>=24h) or 'none'. Only for non-Resolved/Closed."""
    status = complaint.get("status") or "Open"
    if status in ("Resolved", "Closed"):
        return "none"
    created = complaint.get("created_at")
    if not created:
        return "none"
    try:
        ts = datetime.fromisoformat(created.replace("Z", "+00:00"))
    except Exception:
        return "none"
    age_h = (datetime.now(timezone.utc) - ts).total_seconds() / 3600
    if age_h >= 48:
        return "red"
    if age_h >= 24:
        return "yellow"
    return "none"


def hydrate_complaint(c: dict) -> dict:
    c.pop("_id", None)
    c["escalation"] = compute_escalation(c)
    return c


@api_router.post("/complaints")
async def create_complaint(data: ComplaintIn, user=Depends(get_current_user)):
    if not has_perm(user, "complaints", "create"):
        raise HTTPException(status_code=403, detail="Missing permission: complaints.create")
    if data.category not in COMPLAINT_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category. Allowed: {', '.join(COMPLAINT_CATEGORIES)}")
    if data.priority not in COMPLAINT_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Invalid priority. Allowed: {', '.join(COMPLAINT_PRIORITIES)}")
    if data.send_to_target not in SEND_TO_TARGETS:
        raise HTTPException(status_code=400, detail=f"Invalid send_to_target. Allowed: {', '.join(SEND_TO_TARGETS)}")

    cid = user["company_id"]
    complaint_no = await next_complaint_id(cid)

    # Resolve assigned_to / target name
    assigned_to = ""
    assigned_to_name = ""
    if data.send_to_target == "Specific User":
        if not data.assigned_to:
            raise HTTPException(status_code=400, detail="assigned_to is required when send_to_target='Specific User'")
        assignee = await db.users.find_one({"id": data.assigned_to, "company_id": cid}, {"_id": 0, "password_hash": 0})
        if not assignee:
            raise HTTPException(status_code=404, detail="Assignee not found")
        assigned_to = assignee["id"]
        assigned_to_name = assignee.get("name", "")

    # Resolve client & project (optional)
    client_name = ""
    if data.client_id:
        c = await db.clients.find_one({"id": data.client_id, "company_id": cid}, {"_id": 0, "full_name": 1})
        if c:
            client_name = c.get("full_name", "")

    initial_status = "Assigned" if assigned_to else "Open"
    complaint_id = str(uuid.uuid4())
    doc = {
        "id": complaint_id,
        "complaint_no": complaint_no,
        "company_id": cid,
        "title": data.title.strip(),
        "category": data.category,
        "priority": data.priority,
        "description": data.description or "",
        "client_id": data.client_id or "",
        "client_name": client_name,
        "project_id": data.project_id or "",
        "project_name": data.project_id or "",  # placeholder; we don't yet have separate project model
        "send_to_target": data.send_to_target,
        "assigned_to": assigned_to,
        "assigned_to_name": assigned_to_name,
        "status": initial_status,
        "attachments": data.attachments or [],
        "resolution_note": "",
        "resolution_attachments": [],
        "raised_by": user["id"],
        "raised_by_name": user["name"],
        "raised_by_role": user.get("role", ""),
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "resolved_at": None,
    }
    await db.complaints.insert_one(doc)
    await write_complaint_audit(complaint_id, cid, user["id"], user["name"], "Created",
                                f"{complaint_no} · {data.category} · {data.priority} → {data.send_to_target}")

    # Notifications
    title_short = f"Complaint #{complaint_no}: {data.title[:60]}"
    body = f"{data.priority} · {data.category}" + (f" · for {client_name}" if client_name else "")
    if assigned_to:
        await push_notification(cid, "user", f"New complaint assigned: {data.title[:60]}", body, to_user_id=assigned_to)
    await push_notification(cid, "admin", title_short, f"{body} (Send To: {data.send_to_target})")
    if user["id"] != assigned_to:
        await push_notification(cid, "user", f"Complaint #{complaint_no} created", "We've notified the team. You'll get updates here.", to_user_id=user["id"])

    await log_activity(cid, user["id"], user["name"], "Created Complaint", f"{complaint_no} · {data.title[:80]}")
    return hydrate_complaint(doc)


@api_router.get("/complaints/stats")
async def complaint_stats(user=Depends(get_current_user)):
    cid = user["company_id"]
    pipeline = [
        {"$match": {"company_id": cid}},
        {"$group": {"_id": "$status", "n": {"$sum": 1}}},
    ]
    by_status = {row["_id"]: row["n"] async for row in db.complaints.aggregate(pipeline)}
    total = sum(by_status.values())
    high_priority = await db.complaints.count_documents({
        "company_id": cid, "priority": {"$in": ["High", "Urgent"]},
        "status": {"$nin": ["Resolved", "Closed"]},
    })
    mine = await db.complaints.count_documents({
        "company_id": cid,
        "$or": [{"raised_by": user["id"]}, {"assigned_to": user["id"]}],
    })
    # Escalation count (compute on the fly)
    cursor = db.complaints.find(
        {"company_id": cid, "status": {"$nin": ["Resolved", "Closed"]}},
        {"_id": 0, "status": 1, "created_at": 1},
    )
    yellow = 0
    red = 0
    async for c in cursor:
        esc = compute_escalation(c)
        if esc == "red":
            red += 1
        elif esc == "yellow":
            yellow += 1
    return {
        "total": total,
        "open": by_status.get("Open", 0),
        "assigned": by_status.get("Assigned", 0),
        "in_progress": by_status.get("In Progress", 0),
        "waiting": by_status.get("Waiting", 0),
        "resolved": by_status.get("Resolved", 0),
        "closed": by_status.get("Closed", 0),
        "high_priority": high_priority,
        "mine": mine,
        "escalation": {"yellow": yellow, "red": red},
    }


@api_router.get("/complaints")
async def list_complaints(
    user=Depends(get_current_user),
    mine: bool = False,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    category: Optional[str] = None,
    assigned_to: Optional[str] = None,
    client_id: Optional[str] = None,
    project_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    search: Optional[str] = None,
):
    cid = user["company_id"]
    q: Dict[str, Any] = {"company_id": cid}
    is_admin = is_owner(user) or has_perm(user, "complaints", "view")

    if mine or not is_admin:
        q["$or"] = [{"raised_by": user["id"]}, {"assigned_to": user["id"]}]

    if status: q["status"] = status
    if priority: q["priority"] = priority
    if category: q["category"] = category
    if assigned_to: q["assigned_to"] = assigned_to
    if client_id: q["client_id"] = client_id
    if project_id: q["project_id"] = project_id
    if start_date or end_date:
        rng: Dict[str, Any] = {}
        if start_date: rng["$gte"] = start_date
        if end_date: rng["$lte"] = end_date + "T23:59:59"
        q["created_at"] = rng
    if search:
        s = re.escape(search)
        q["$and"] = q.get("$and", []) + [{"$or": [
            {"title": {"$regex": s, "$options": "i"}},
            {"description": {"$regex": s, "$options": "i"}},
            {"complaint_no": {"$regex": s, "$options": "i"}},
            {"client_name": {"$regex": s, "$options": "i"}},
        ]}]

    rows = await db.complaints.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [hydrate_complaint(c) for c in rows]


@api_router.get("/complaints/{complaint_id}")
async def get_complaint(complaint_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    c = await db.complaints.find_one({"id": complaint_id, "company_id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Complaint not found")
    return hydrate_complaint(c)


@api_router.patch("/complaints/{complaint_id}")
async def update_complaint(complaint_id: str, data: ComplaintUpdate, user=Depends(get_current_user)):
    if not (is_owner(user) or has_perm(user, "complaints", "edit")):
        raise HTTPException(status_code=403, detail="Missing permission: complaints.edit")
    cid = user["company_id"]
    existing = await db.complaints.find_one({"id": complaint_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Complaint not found")

    update: Dict[str, Any] = {}
    audit_events: List[str] = []

    payload = data.model_dump(exclude_unset=True)

    if "status" in payload and payload["status"]:
        new_status = payload["status"]
        if new_status not in COMPLAINT_STATUSES:
            raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {', '.join(COMPLAINT_STATUSES)}")
        # Mandatory resolution note when moving to Resolved
        if new_status == "Resolved":
            note = payload.get("resolution_note") or existing.get("resolution_note") or ""
            if not note.strip():
                raise HTTPException(status_code=400, detail="Resolution note is required before marking a complaint as Resolved")
            update["resolved_at"] = now_iso()
        update["status"] = new_status
        if new_status != existing.get("status"):
            audit_events.append(f"Status: {existing.get('status', 'Open')} → {new_status}")

    if "assigned_to" in payload:
        new_assignee = payload["assigned_to"] or ""
        if new_assignee:
            if not (is_owner(user) or has_perm(user, "complaints", "edit") or has_perm(user, "team", "edit")):
                raise HTTPException(status_code=403, detail="Only authorized users can assign complaints")
            assignee = await db.users.find_one({"id": new_assignee, "company_id": cid}, {"_id": 0, "name": 1})
            if not assignee:
                raise HTTPException(status_code=404, detail="Assignee not found")
            update["assigned_to"] = new_assignee
            update["assigned_to_name"] = assignee.get("name", "")
            # If still Open, bump to Assigned
            if existing.get("status") == "Open" and "status" not in update:
                update["status"] = "Assigned"
            audit_events.append(f"Assigned to {assignee.get('name', '')}")
            # Notify the new assignee
            if new_assignee != existing.get("assigned_to"):
                await push_notification(cid, "user", f"Complaint #{existing['complaint_no']} assigned to you",
                                        existing.get("title", "")[:120], to_user_id=new_assignee)

    for field in ("title", "category", "priority", "description", "send_to_target", "resolution_note"):
        if field in payload and payload[field] is not None and payload[field] != existing.get(field):
            update[field] = payload[field]
            audit_events.append(f"Updated {field}")
    if "resolution_attachments" in payload and payload["resolution_attachments"] is not None:
        update["resolution_attachments"] = payload["resolution_attachments"]

    if not update:
        return hydrate_complaint(existing)

    update["updated_at"] = now_iso()
    await db.complaints.update_one({"id": complaint_id, "company_id": cid}, {"$set": update})

    for ev in audit_events:
        await write_complaint_audit(complaint_id, cid, user["id"], user["name"], "Updated", ev)

    if update.get("status") == "Resolved":
        await push_notification(cid, "admin", f"Complaint #{existing['complaint_no']} resolved", existing.get("title", "")[:120])
        if existing.get("raised_by") and existing["raised_by"] != user["id"]:
            await push_notification(cid, "user", f"Complaint #{existing['complaint_no']} resolved",
                                    "Your complaint has been resolved.", to_user_id=existing["raised_by"])

    refreshed = await db.complaints.find_one({"id": complaint_id, "company_id": cid}, {"_id": 0})
    if not refreshed:
        raise HTTPException(status_code=404, detail="Complaint not found")
    await log_activity(cid, user["id"], user["name"], "Updated Complaint", refreshed.get("complaint_no", ""))
    return hydrate_complaint(refreshed)


@api_router.delete("/complaints/{complaint_id}")
async def delete_complaint(complaint_id: str, user=Depends(get_current_user)):
    if not (is_owner(user) or has_perm(user, "complaints", "delete")):
        raise HTTPException(status_code=403, detail="Missing permission: complaints.delete")
    cid = user["company_id"]
    existing = await db.complaints.find_one({"id": complaint_id, "company_id": cid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Complaint not found")
    await db.complaint_comments.delete_many({"complaint_id": complaint_id})
    await db.complaint_audit.delete_many({"complaint_id": complaint_id})
    await db.complaints.delete_one({"id": complaint_id, "company_id": cid})
    await log_activity(cid, user["id"], user["name"], "Deleted Complaint", existing.get("complaint_no", ""))
    return {"ok": True}


@api_router.get("/complaints/{complaint_id}/comments")
async def list_complaint_comments(complaint_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    c = await db.complaints.find_one({"id": complaint_id, "company_id": cid}, {"_id": 0, "id": 1})
    if not c:
        raise HTTPException(status_code=404, detail="Complaint not found")
    return await db.complaint_comments.find({"complaint_id": complaint_id}, {"_id": 0}).sort("created_at", 1).to_list(500)


@api_router.post("/complaints/{complaint_id}/comments")
async def add_complaint_comment(complaint_id: str, data: ComplaintCommentIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    c = await db.complaints.find_one({"id": complaint_id, "company_id": cid})
    if not c:
        raise HTTPException(status_code=404, detail="Complaint not found")
    if not (data.text or "").strip():
        raise HTTPException(status_code=400, detail="Comment text is required")
    doc = {
        "id": str(uuid.uuid4()),
        "complaint_id": complaint_id,
        "company_id": cid,
        "user_id": user["id"], "user_name": user["name"], "user_role": user.get("role", ""),
        "text": data.text.strip(),
        "attachments": data.attachments or [],
        "created_at": now_iso(),
    }
    await db.complaint_comments.insert_one(doc)
    doc.pop("_id", None)
    await write_complaint_audit(complaint_id, cid, user["id"], user["name"], "Comment Added", data.text[:140])
    await db.complaints.update_one({"id": complaint_id, "company_id": cid}, {"$set": {"updated_at": now_iso()}})
    # Notify the other party (assignee / raiser)
    notify_targets = {c.get("raised_by"), c.get("assigned_to")} - {user["id"], "", None}
    for uid in notify_targets:
        await push_notification(cid, "user", f"New comment on #{c['complaint_no']}", data.text[:120], to_user_id=uid)
    return doc


@api_router.get("/complaints/{complaint_id}/audit")
async def list_complaint_audit(complaint_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    c = await db.complaints.find_one({"id": complaint_id, "company_id": cid}, {"_id": 0, "id": 1})
    if not c:
        raise HTTPException(status_code=404, detail="Complaint not found")
    return await db.complaint_audit.find({"complaint_id": complaint_id}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.post("/complaints/{complaint_id}/convert-to-task")
async def convert_complaint_to_task(complaint_id: str, user=Depends(get_current_user)):
    if not (is_owner(user) or has_perm(user, "complaints", "approve") or has_perm(user, "task_portal", "create")):
        raise HTTPException(status_code=403, detail="Missing permission: complaints.approve")
    cid = user["company_id"]
    c = await db.complaints.find_one({"id": complaint_id, "company_id": cid})
    if not c:
        raise HTTPException(status_code=404, detail="Complaint not found")
    if c.get("converted_task_id"):
        raise HTTPException(status_code=400, detail="This complaint has already been converted to a task.")
    if not c.get("client_id"):
        raise HTTPException(status_code=400, detail="Complaint must be linked to a client to convert to a task")
    if not c.get("assigned_to"):
        raise HTTPException(status_code=400, detail="Assign the complaint to a user before converting to a task")
    client = await db.clients.find_one({"id": c["client_id"], "company_id": cid}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Linked client no longer exists")
    task_doc = {
        "id": str(uuid.uuid4()), "company_id": cid, "client_id": c["client_id"],
        "client_name": client.get("full_name"), "sol_id": client.get("sol_id"),
        "task_type": f"Complaint: {c.get('title', '')[:80]}",
        "assigned_to": c["assigned_to"], "assigned_to_name": c.get("assigned_to_name", ""),
        "assigned_by": user["id"], "assigned_by_name": user["name"],
        "deadline": "",
        "priority": "Urgent" if c.get("priority") in ("High", "Urgent") else "Normal",
        "remarks": f"Auto-created from complaint #{c['complaint_no']}. Category: {c['category']}.\n\n{c.get('description', '')}",
        "status": "pending", "submission": None,
        "complaint_id": complaint_id, "complaint_no": c.get("complaint_no"),
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.tasks.insert_one(task_doc)
    task_doc.pop("_id", None)
    await db.complaints.update_one({"id": complaint_id, "company_id": cid},
                                   {"$set": {"converted_task_id": task_doc["id"], "updated_at": now_iso()}})
    await write_complaint_audit(complaint_id, cid, user["id"], user["name"], "Converted to Task", task_doc["id"])
    await push_notification(cid, "user", f"Task created from complaint #{c['complaint_no']}",
                            task_doc["task_type"], to_user_id=c["assigned_to"])
    await log_activity(cid, user["id"], user["name"], "Converted Complaint → Task", c.get("complaint_no", ""))
    return {"task": task_doc, "complaint_id": complaint_id}


@api_router.get("/complaints/lookup/assignable-users")
async def list_assignable_users(user=Depends(get_current_user)):
    """Active users in the company that complaints can be assigned to."""
    cid = user["company_id"]
    users = await db.users.find(
        {"company_id": cid, "status": "Active"},
        {"_id": 0, "id": 1, "name": 1, "role": 1, "email": 1},
    ).sort("name", 1).to_list(500)
    return users


# ---------- Client Inventory Ledger Logic ----------
async def calculate_client_ledger(company_id: str, client_id: str):
    client = await db.clients.find_one({"id": client_id, "company_id": company_id}, {"_id": 0})
    if not client:
        return None
        
    # Run both queries in parallel
    outwards, inwards_raw = await asyncio.gather(
        db.outward_entries.find({
            "company_id": company_id,
            "client_id": client_id,
            "status": "Dispatched"
        }, {"_id": 0}).to_list(1000),
        # inward_entries stores client_id inside remarks as [client_id:UUID],
        # so we can only filter by company_id + source_type in the DB,
        # then use parse_inward_client_info to extract & match client_id in Python.
        db.inward_entries.find({
            "company_id": company_id,
            "source_type": "Return From Client",
        }, {"_id": 0}).to_list(5000),
    )
    
    # Parse client_id out of remarks and filter to this client
    inwards = []
    for inv in inwards_raw:
        inv = parse_inward_client_info(inv)
        if inv.get("client_id") == client_id:
            inwards.append(inv)
    
    ledger = {}
    
    for out in outwards:
        raw_prod = (out.get("product") or "").strip()
        norm_name = raw_prod.upper()
        if not norm_name:
            continue
        raw_size = (out.get("size") or "").strip()
        norm_size = raw_size.upper()
        unit = (out.get("unit") or "Nos").strip()
        
        # Report identity key: normalized Product Name + normalized Size/Spec
        # Unit MUST NOT be part of the identity key.
        key = (norm_name, norm_size)
        qty = float(out.get("quantity") or 0)
        date_str = out.get("date") or out.get("created_at") or ""
        
        if key not in ledger:
            ledger[key] = {
                "product": raw_prod.upper(),
                "size": raw_size,
                "unit": unit,
                "total_outward": 0.0,
                "total_returned": 0.0,
                "current_balance": 0.0,
                "serial_numbers": [],
                "last_movement_date": ""
            }
        ledger[key]["total_outward"] += qty
        
        serials = out.get("serial_numbers") or out.get("serials") or ([out.get("serial_number")] if out.get("serial_number") else [])
        for s in serials:
            sn = (s or "").strip().upper()
            if sn and sn not in ledger[key]["serial_numbers"]:
                ledger[key]["serial_numbers"].append(sn)
        
        if date_str:
            if not ledger[key]["last_movement_date"] or date_str > ledger[key]["last_movement_date"]:
                ledger[key]["last_movement_date"] = date_str

    for inv in inwards:
        raw_prod = (inv.get("product") or "").strip()
        norm_name = raw_prod.upper()
        if not norm_name:
            continue
        raw_size = (inv.get("size") or "").strip()
        norm_size = raw_size.upper()
        unit = (inv.get("unit") or "Nos").strip()
        
        # Report identity key: normalized Product Name + normalized Size/Spec
        key = (norm_name, norm_size)
        qty = float(inv.get("quantity") or 0)
        date_str = inv.get("date") or inv.get("created_at") or ""
        
        if key not in ledger:
            ledger[key] = {
                "product": raw_prod.upper(),
                "size": raw_size,
                "unit": unit,
                "total_outward": 0.0,
                "total_returned": 0.0,
                "current_balance": 0.0,
                "serial_numbers": [],
                "last_movement_date": ""
            }
        ledger[key]["total_returned"] += qty
        
        serials = inv.get("serial_numbers") or inv.get("serials") or ([inv.get("serial_number")] if inv.get("serial_number") else [])
        for s in serials:
            sn = (s or "").strip().upper()
            if sn in ledger[key]["serial_numbers"]:
                ledger[key]["serial_numbers"].remove(sn)
        
        if date_str:
            if not ledger[key]["last_movement_date"] or date_str > ledger[key]["last_movement_date"]:
                ledger[key]["last_movement_date"] = date_str

    items = []
    total_outward_qty = 0.0
    total_returned_qty = 0.0
    current_balance_qty = 0.0
    negative_items_count = 0
    
    for key, item in ledger.items():
        balance = item["total_outward"] - item["total_returned"]
        item["current_balance"] = balance
        
        if balance > 0:
            item["status"] = "Dispatched"
        elif balance == 0:
            item["status"] = "Settled"
        else:
            item["status"] = "Excess Return"
            negative_items_count += 1
            
        total_outward_qty += item["total_outward"]
        total_returned_qty += item["total_returned"]
        current_balance_qty += balance
        
        if item["last_movement_date"]:
            item["last_movement_date"] = item["last_movement_date"][:10]
            
        items.append(item)

    hv_keywords = ["SOLAR PANEL", "PANEL", "INVERTER", "ACDB", "DCDB", "METER", "BATTERY"]
    local_hv = _load_local_high_value_products()
    def _is_item_hv(it):
        p_name = norm_product_name(it.get("product") or "")
        if it.get("high_value_goods") or it.get("high_value_asset") or local_hv.get(p_name):
            return True
        if any(kw in p_name for kw in hv_keywords):
            return True
        return False

    items.sort(key=lambda x: (0 if _is_item_hv(x) else 1, (x.get("product") or "").lower(), (x.get("size") or "").lower()))
        
    summary = {
        "total_products": len(items),
        "total_outward_qty": total_outward_qty,
        "total_returned_qty": total_returned_qty,
        "current_balance": current_balance_qty,
        "negative_items": negative_items_count
    }
    
    return {
        "client": {
            "id": client.get("id"),
            "full_name": client.get("full_name"),
            "client_code": client.get("sol_id") or client.get("client_code"),
            "sol_id": client.get("sol_id")
        },
        "summary": summary,
        "items": items
    }

@api_router.get("/inventory/ledger/{client_id}")
async def get_client_ledger(client_id: str, user=Depends(require_perm("reports", "view"))):
    ledger = await calculate_client_ledger(user["company_id"], client_id)
    if not ledger:
        raise HTTPException(status_code=404, detail="Client not found")
    return ledger

@api_router.get("/inventory/ledger/{client_id}/export")
async def export_client_ledger(client_id: str, format: str = "csv", user=Depends(require_perm("reports", "view"))):
    cid = user["company_id"]
    ledger = await calculate_client_ledger(cid, client_id)
    if not ledger or not isinstance(ledger, dict):
        raise HTTPException(status_code=404, detail="Client not found")
        
    client_val = ledger.get("client")
    client = client_val if isinstance(client_val, dict) else {}
    summary_val = ledger.get("summary")
    summary = summary_val if isinstance(summary_val, dict) else {}
    items_val = ledger.get("items")
    items = items_val if isinstance(items_val, list) else []
    
    if format == "csv":
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Client Name", client.get("full_name") or ""])
        writer.writerow(["Project ID", client.get("sol_id") or ""])
        writer.writerow(["Generated Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])
        writer.writerow(["Product", "Size", "Unit", "Total Outward", "Total Returned", "Current Balance", "Status"])
        for item in items:
            writer.writerow([
                item["product"], item["size"], item["unit"],
                item["total_outward"], item["total_returned"],
                item["current_balance"], item["status"]
            ])
        from fastapi.responses import StreamingResponse
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="material_ledger_{client_id}.csv"'},
        )
        
    elif format == "excel":
        import openpyxl  # type: ignore
        import openpyxl.utils  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
        from fastapi.responses import StreamingResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Material Ledger"
        
        title_font = Font(name="Calibri", size=14, bold=True, color="1d4ed8")
        bold_font = Font(name="Calibri", size=10, bold=True)
        header_font = Font(name="Calibri", size=10, bold=True, color="ffffff")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        red_font = Font(name="Calibri", size=10, color="dc2626")
        gray_font = Font(name="Calibri", size=10, color="94a3b8")
        
        thin_border = Border(
            left=Side(style='thin', color='d1d5db'),
            right=Side(style='thin', color='d1d5db'),
            top=Side(style='thin', color='d1d5db'),
            bottom=Side(style='thin', color='d1d5db')
        )
        
        ws.cell(row=1, column=1, value="CLIENT MATERIAL LEDGER REPORT").font = title_font
        ws.row_dimensions[1].height = 25
        
        ws.cell(row=3, column=1, value="Client Name").font = bold_font
        ws.cell(row=3, column=2, value=client.get("full_name") or "")
        
        ws.cell(row=4, column=1, value="Project ID").font = bold_font
        ws.cell(row=4, column=2, value=client.get("sol_id") or "")
        
        ws.cell(row=5, column=1, value="Generated Date").font = bold_font
        ws.cell(row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        headers = ["Product", "Size", "Unit", "Total Outward", "Total Returned", "Current Balance", "Status"]
        start_row = 7
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        ws.row_dimensions[start_row].height = 20
        
        current_row = start_row + 1
        for item in items:
            ws.cell(row=current_row, column=1, value=item["product"]).border = thin_border
            ws.cell(row=current_row, column=2, value=item["size"]).border = thin_border
            ws.cell(row=current_row, column=3, value=item["unit"]).border = thin_border
            
            c_out = ws.cell(row=current_row, column=4, value=item["total_outward"])
            c_out.border = thin_border
            c_out.alignment = Alignment(horizontal="right")
            
            c_ret = ws.cell(row=current_row, column=5, value=item["total_returned"])
            c_ret.border = thin_border
            c_ret.alignment = Alignment(horizontal="right")
            
            c_bal = ws.cell(row=current_row, column=6, value=item["current_balance"])
            c_bal.border = thin_border
            c_bal.alignment = Alignment(horizontal="right")
            
            c_stat = ws.cell(row=current_row, column=7, value=item["status"])
            c_stat.border = thin_border
            c_stat.alignment = Alignment(horizontal="center")
            
            if item["current_balance"] < 0:
                c_bal.font = red_font
                c_stat.font = red_font
            elif item["current_balance"] == 0:
                c_bal.font = gray_font
                c_stat.font = gray_font
                
            current_row += 1
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_idx = col[0].column
            if col_idx is not None:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return StreamingResponse(
            excel_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="material_ledger_{client_id}.xlsx"'},
        )
        
    elif format == "pdf":
        company_doc = await db.companies.find_one({"id": cid}, {"_id": 0}) or {}
        pdf_bytes = pdf_generator.generate_ledger_pdf(client, ledger, company_doc)
        
        from fastapi.responses import Response
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="material_ledger_{client_id}.pdf"'},
        )
        
    else:
        raise HTTPException(status_code=400, detail="Invalid export format")

@api_router.post("/inventory/products/parse-pdf")
async def parse_pdf_products(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Extracts candidate Product Master rows from an uploaded PDF document."""
    filename = (file.filename or "").lower()
    if not filename.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Uploaded file must be a PDF document.")
        
    contents = await file.read()
    if not contents or len(contents) == 0:
        raise HTTPException(status_code=400, detail="Uploaded PDF file is empty.")
        
    extracted_text = ""
    try:
        from pypdf import PdfReader
        from io import BytesIO
        reader = PdfReader(BytesIO(contents))
        for page in reader.pages:
            t = page.extract_text()
            if t:
                extracted_text += t + "\n"
    except Exception as e:
        logger.exception("Error extracting text from PDF")
        raise HTTPException(status_code=400, detail="Failed to read text from PDF document. Please verify file format.")

    if not extracted_text.strip():
        raise HTTPException(status_code=400, detail="No readable text streams found in PDF. Please ensure file contains selectable text.")

    lines = [line.strip() for line in extracted_text.splitlines() if line.strip()]
    skip_keywords = ["sl no", "s.no", "page", "total", "subtotal", "gstin", "invoice", "product master", "report", "date"]

    parsed_rows = []
    for line in lines:
        lower_line = line.lower()
        if any(lower_line.startswith(kw) for kw in skip_keywords):
            continue
            
        parts = [p.strip() for p in re.split(r'\t|,|\||\s{2,}', line) if p.strip()]
        if not parts:
            continue
            
        name = parts[0]
        if len(name) < 2 or name.isdigit():
            continue
            
        name_lower = name.lower()
        category = "Solar Panel" if any(w in name_lower for w in ["panel", "pv", "module", "mono", "poly"]) else ("Inverter" if any(w in name_lower for w in ["inverter", "inv", "hybrid", "growatt", "deye"]) else "Others")
        
        brand = parts[1] if len(parts) > 1 and not parts[1].isdigit() and len(parts[1]) < 30 else ""
        size = parts[2] if len(parts) > 2 and len(parts[2]) < 40 else ""
        unit = "Nos"
        min_stock = 0
        hsn = ""
        gst = ""
        high_value = False
        
        for p in parts[1:]:
            p_upper = p.upper()
            if p_upper in ("NOS", "PCS", "SETS", "MTR", "KG", "BOX", "PKT", "FEET", "ROLL"):
                unit = p
            elif re.match(r'^\d{4,8}$', p):
                hsn = p
            elif "%" in p or "GST" in p_upper:
                gst = p
            elif p.isdigit() and int(p) < 1000 and min_stock == 0:
                min_stock = int(p)

        parsed_rows.append({
            "name": name,
            "category": category,
            "brand": brand,
            "size": size,
            "unit": unit,
            "hsn": hsn,
            "gst": gst,
            "min_stock": min_stock,
            "high_value_goods": high_value
        })

    return {"ok": True, "rows": parsed_rows, "count": len(parsed_rows)}


@api_router.post("/inventory/products/export-pdf")
async def export_product_master_pdf_endpoint(request: Request, user=Depends(get_current_user)):
    """Generates an A4 Landscape PDF export for the Product Master view.

    Bugs fixed:
      1. db.company_details → db.companies (correct table name)
      2. Filter by user company_id (not empty filter which returned wrong company)
      3. await _enrich_company_doc (it is async)
      4. Use Request.json() to parse body (FastAPI dict annotation fails on raw JSON POST)
    """
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    products = payload.get("products") or []

    try:
        company_doc = await db.companies.find_one({"id": user["company_id"]}, {"_id": 0}) or {}
        company = await _enrich_company_doc(company_doc)
    except Exception as e:
        logger.warning(f"export-pdf: could not fetch company doc: {e}")
        company = {}

    try:
        from pdf_generator import generate_product_master_pdf
        pdf_bytes = generate_product_master_pdf(products, company)
    except Exception as e:
        logger.exception("export-pdf: PDF generation failed")
        raise HTTPException(status_code=500, detail="PDF generation failed. Please try again.")

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Product_Master.pdf"}
    )


# ─── LEAD MANAGEMENT ENDPOINTS ───────────────────────────────────────────────

@api_router.get("/leads")
async def list_leads(
    stage: Optional[str] = None,
    assigned_to: Optional[str] = None,
    source: Optional[str] = None,
    call_status: Optional[str] = None,
    followup_filter: Optional[str] = None,
    scope: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user=Depends(get_current_user)
):
    cid = user["company_id"]
    query = {"company_id": cid}

    if scope == "mine":
        query["assigned_to"] = user["id"]
    elif assigned_to and assigned_to != "all":
        query["assigned_to"] = assigned_to

    if stage and stage != "all":
        query["stage"] = stage.upper()

    if source and source != "all":
        query["source"] = source

    if call_status and call_status != "all":
        query["call_status"] = call_status

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if followup_filter == "today":
        query["next_followup_at"] = {"$regex": f"^{today_str}"}
    elif followup_filter == "overdue":
        query["next_followup_at"] = {"$lt": today_str, "$ne": ""}
    elif followup_filter == "upcoming":
        query["next_followup_at"] = {"$gt": f"{today_str}T23:59:59"}

    if search:
        s = search.strip()
        query["$or"] = [
            {"name": {"$regex": s, "$options": "i"}},
            {"mobile": {"$regex": s, "$options": "i"}},
            {"city": {"$regex": s, "$options": "i"}},
            {"lead_no": {"$regex": s, "$options": "i"}},
        ]

    total = await db.leads.count_documents(query)
    skip = (page - 1) * page_size
    cursor = db.leads.find(query, {"_id": 0}).sort("updated_at", -1).skip(skip).limit(page_size)
    items = await cursor.to_list(page_size)

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@api_router.get("/leads/stats")
async def get_lead_stats(scope: Optional[str] = None, user=Depends(get_current_user)):
    cid = user["company_id"]
    query = {"company_id": cid}
    if scope == "mine":
        query["assigned_to"] = user["id"]

    all_leads = await db.leads.find(query, {"_id": 0, "stage": 1, "next_followup_at": 1, "last_contact_at": 1}).to_list(100000)
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    total_leads = len(all_leads)
    new_leads = sum(1 for l in all_leads if (l.get("stage") or "").upper() == "NEW")
    interested = sum(1 for l in all_leads if (l.get("stage") or "").upper() in ("INTERESTED", "SITE VISIT", "QUOTATION", "NEGOTIATION"))
    final_won = sum(1 for l in all_leads if (l.get("stage") or "").upper() in ("FINAL", "ONBOARDING", "CONVERTED"))
    lost = sum(1 for l in all_leads if (l.get("stage") or "").upper() in ("LOST", "NOT INTERESTED"))

    followups_due = sum(1 for l in all_leads if l.get("next_followup_at") and str(l.get("next_followup_at"))[:10] <= today_str)
    todays_calls = sum(1 for l in all_leads if l.get("last_contact_at") and str(l.get("last_contact_at"))[:10] == today_str)

    return {
        "total_leads": total_leads,
        "new_leads": new_leads,
        "followups_due": followups_due,
        "todays_calls": todays_calls,
        "interested": interested,
        "final_won": final_won,
        "lost": lost,
    }


@api_router.get("/leads/followups/list")
async def list_followups(filter_type: Optional[str] = "today", scope: Optional[str] = "mine", user=Depends(get_current_user)):
    cid = user["company_id"]
    query = {"company_id": cid, "status": "pending"}

    if scope == "mine":
        query["assigned_to"] = user["id"]

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if filter_type == "today":
        query["followup_at"] = {"$regex": f"^{today_str}"}
    elif filter_type == "overdue":
        query["followup_at"] = {"$lt": today_str}
    elif filter_type == "upcoming":
        query["followup_at"] = {"$gt": f"{today_str}T23:59:59"}

    followups = await db.lead_followups.find(query, {"_id": 0}).sort("followup_at", 1).to_list(1000)

    lead_ids = list(set([f["lead_id"] for f in followups if f.get("lead_id")]))
    leads_map = {}
    if lead_ids:
        leads_list = await db.leads.find({"id": {"$in": lead_ids}, "company_id": cid}, {"_id": 0}).to_list(len(lead_ids))
        leads_map = {l["id"]: l for l in leads_list}

    items = []
    for f in followups:
        l = leads_map.get(f["lead_id"], {})
        items.append({
            **f,
            "lead_name": l.get("name") or "—",
            "mobile": l.get("mobile") or "—",
            "city": l.get("city") or "—",
            "stage": l.get("stage") or "NEW",
            "last_contact_at": l.get("last_contact_at") or "",
            "call_status": l.get("call_status") or "",
        })

    return items


@api_router.post("/leads/followups/{followup_id}/complete")
async def complete_followup(followup_id: str, data: LeadFollowupUpdate, user=Depends(get_current_user)):
    cid = user["company_id"]
    f_doc = await db.lead_followups.find_one({"id": followup_id, "company_id": cid}, {"_id": 0})
    if not f_doc:
        raise HTTPException(status_code=404, detail="Followup not found")

    status = data.status or "completed"
    await db.lead_followups.update_one({"id": followup_id, "company_id": cid}, {"$set": {"status": status, "completed_at": now_iso()}})

    if status == "rescheduled" and data.rescheduled_at:
        await db.lead_followups.insert_one({
            "id": str(uuid.uuid4()),
            "lead_id": f_doc["lead_id"],
            "company_id": cid,
            "assigned_to": f_doc.get("assigned_to"),
            "assigned_to_name": f_doc.get("assigned_to_name"),
            "followup_at": data.rescheduled_at,
            "status": "pending",
            "notes": data.remarks or f_doc.get("notes") or "",
            "created_at": now_iso(),
        })
        await db.leads.update_one({"id": f_doc["lead_id"], "company_id": cid}, {"$set": {"next_followup_at": data.rescheduled_at, "updated_at": now_iso()}})

    return {"message": f"Follow-up marked as {status}"}


@api_router.post("/leads")
async def create_lead(data: LeadIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    lead_id = str(uuid.uuid4())
    lead_no = await next_lead_id(cid)

    doc = {
        "id": lead_id,
        "lead_no": lead_no,
        "company_id": cid,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "name": data.name.strip(),
        "mobile": data.mobile.strip(),
        "alt_mobile": (data.alt_mobile or "").strip(),
        "city": (data.city or "").strip(),
        "address": (data.address or "").strip(),
        "estimated_kw": float(data.estimated_kw or 0),
        "consumer_type": (data.consumer_type or "").strip(),
        "source": data.source or "Other",
        "assigned_to": data.assigned_to or user["id"],
        "assigned_to_name": data.assigned_to_name or user["name"],
        "stage": (data.stage or "NEW").upper(),
        "status": data.status or "New Lead",
        "call_status": "Not Called",
        "last_contact_at": "",
        "next_followup_at": data.next_followup_at or "",
        "remarks": (data.remarks or "").strip(),
        "converted_client_id": "",
        "converted_at": "",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }

    await db.leads.insert_one(doc)

    if data.next_followup_at:
        await db.lead_followups.insert_one({
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "company_id": cid,
            "assigned_to": doc["assigned_to"],
            "assigned_to_name": doc["assigned_to_name"],
            "followup_at": data.next_followup_at,
            "status": "pending",
            "notes": "Initial follow-up set on lead creation",
            "created_at": now_iso(),
        })

    await log_activity(cid, user["id"], user["name"], "Created Solar Lead", f"{data.name} ({lead_no})")
    doc.pop("_id", None)
    return doc


@api_router.get("/leads/{lead_id}")
async def get_lead_detail(lead_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    lead = await db.leads.find_one({"id": lead_id, "company_id": cid}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    calls = await db.lead_call_activities.find({"lead_id": lead_id, "company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    followups = await db.lead_followups.find({"lead_id": lead_id, "company_id": cid}, {"_id": 0}).sort("followup_at", -1).to_list(1000)

    return {
        "lead": lead,
        "calls": calls,
        "followups": followups,
    }


@api_router.put("/leads/{lead_id}")
async def update_lead(lead_id: str, data: LeadIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    lead = await db.leads.find_one({"id": lead_id, "company_id": cid}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    patch = {
        "name": data.name.strip(),
        "mobile": data.mobile.strip(),
        "alt_mobile": (data.alt_mobile or "").strip(),
        "city": (data.city or "").strip(),
        "address": (data.address or "").strip(),
        "estimated_kw": float(data.estimated_kw or 0),
        "consumer_type": (data.consumer_type or "").strip(),
        "source": data.source or lead.get("source", "Other"),
        "assigned_to": data.assigned_to or lead.get("assigned_to", user["id"]),
        "assigned_to_name": data.assigned_to_name or lead.get("assigned_to_name", user["name"]),
        "stage": (data.stage or lead.get("stage", "NEW")).upper(),
        "status": data.status or lead.get("status", "Active"),
        "remarks": (data.remarks or "").strip(),
        "updated_at": now_iso(),
    }

    if data.next_followup_at:
        patch["next_followup_at"] = data.next_followup_at

    await db.leads.update_one({"id": lead_id, "company_id": cid}, {"$set": patch})
    updated = await db.leads.find_one({"id": lead_id, "company_id": cid}, {"_id": 0})
    return updated


@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    await db.leads.delete_one({"id": lead_id, "company_id": cid})
    await db.lead_call_activities.delete_many({"lead_id": lead_id, "company_id": cid})
    await db.lead_followups.delete_many({"lead_id": lead_id, "company_id": cid})
    return {"message": "Lead deleted"}


@api_router.post("/leads/{lead_id}/calls")
async def add_lead_call(lead_id: str, data: LeadCallIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    lead = await db.leads.find_one({"id": lead_id, "company_id": cid}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    call_id = str(uuid.uuid4())
    call_doc = {
        "id": call_id,
        "lead_id": lead_id,
        "company_id": cid,
        "user_id": user["id"],
        "user_name": user["name"],
        "outcome": data.outcome,
        "notes": (data.notes or "").strip(),
        "duration_sec": data.duration_sec or 0,
        "next_followup_at": data.next_followup_at or "",
        "created_at": now_iso(),
    }
    await db.lead_call_activities.insert_one(call_doc)

    lead_patch = {
        "last_contact_at": now_iso(),
        "call_status": data.outcome,
        "updated_at": now_iso(),
    }

    if data.assigned_to:
        lead_patch["assigned_to"] = data.assigned_to
        lead_patch["assigned_to_name"] = data.assigned_to_name or ""

    if data.stage:
        lead_patch["stage"] = data.stage.upper()

    if data.next_followup_at:
        lead_patch["next_followup_at"] = data.next_followup_at
        await db.lead_followups.insert_one({
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "company_id": cid,
            "assigned_to": lead_patch.get("assigned_to", lead.get("assigned_to")),
            "assigned_to_name": lead_patch.get("assigned_to_name", lead.get("assigned_to_name")),
            "followup_at": data.next_followup_at,
            "status": "pending",
            "notes": (data.notes or "").strip(),
            "created_at": now_iso(),
        })

    await db.leads.update_one({"id": lead_id, "company_id": cid}, {"$set": lead_patch})

    if data.next_followup_at and lead_patch.get("assigned_to"):
        await push_notification(
            cid,
            lead_patch.get("assigned_to"),
            "Lead Follow-up Scheduled",
            f"Follow-up for {lead.get('name')} scheduled at {data.next_followup_at}"
        )

    updated_lead = await db.leads.find_one({"id": lead_id, "company_id": cid}, {"_id": 0})
    return {"call": call_doc, "lead": updated_lead}


@api_router.post("/leads/{lead_id}/convert-check")
async def convert_check_lead(lead_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    lead = await db.leads.find_one({"id": lead_id, "company_id": cid}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    mobile = (lead.get("mobile") or "").strip()
    existing_client = None
    if mobile:
        existing_client = await db.clients.find_one({"mobile": mobile, "company_id": cid}, {"_id": 0})

    return {
        "lead": lead,
        "exists": bool(existing_client),
        "existing_client": _enrich_client_doc(existing_client) if existing_client else None,
    }


@api_router.post("/leads/{lead_id}/link-client")
async def link_client_to_lead(lead_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    cid = user["company_id"]
    client_id = payload.get("client_id")
    sol_id = payload.get("sol_id") or ""
    if not client_id:
        raise HTTPException(status_code=400, detail="client_id is required")

    await db.leads.update_one(
        {"id": lead_id, "company_id": cid},
        {"$set": {
            "converted_client_id": client_id,
            "converted_sol_id": sol_id,
            "converted_at": now_iso(),
            "stage": "CONVERTED",
            "status": "Converted to Client",
            "updated_at": now_iso(),
        }}
    )

# ─── EPC OPERATING SYSTEM ENHANCEMENTS (FINANCE, VENDORS, WARRANTIES, SERVICE, SEARCH) ───

class PaymentPlanItem(BaseModel):
    name: str
    amount: float

class ProjectCreatePayload(BaseModel):
    client_id: str
    project_name: str
    project_type: Optional[str] = "Rooftop Solar"
    capacity_kw: Optional[float] = 0.0
    project_value: float
    project_date: Optional[str] = ""
    expected_completion_date: Optional[str] = ""
    payment_terms: Optional[str] = ""
    notes: Optional[str] = ""
    payment_plan: Optional[List[PaymentPlanItem]] = []

class PaymentRecordPayload(BaseModel):
    project_id: Optional[str] = ""
    client_id: Optional[str] = ""
    milestone_id: Optional[str] = None
    milestone_name: Optional[str] = "Payment"
    payment_type: Optional[str] = "Advance"
    amount: float
    payment_date: str
    payment_source: Optional[str] = "Bank Transfer"
    payment_mode: Optional[str] = "Bank Transfer / UTR"
    ref_number: Optional[str] = ""
    bank_utr: Optional[str] = ""
    remarks: Optional[str] = ""
    notes: Optional[str] = ""
    status: Optional[str] = "Received"
    attachment_url: Optional[str] = ""
    invoice_id: Optional[str] = ""
    allocated_amount: Optional[float] = None

class InvoiceItemPayload(BaseModel):
    product_name: str
    hsn_sac: Optional[str] = ""
    size: Optional[str] = ""
    quantity: float
    unit: Optional[str] = "Nos"
    unit_price: Optional[float] = 0.0
    rate: Optional[float] = None
    discount: Optional[float] = 0.0
    gst_rate: Optional[float] = 18.0
    amount: Optional[float] = 0.0

    @field_validator("quantity", "unit_price", "rate", "discount", "gst_rate", "amount", mode="before")
    @classmethod
    def parse_item_numeric_fields(cls, v):
        if v is None or v == "":
            return 0.0
        if isinstance(v, str):
            cleaned = re.sub(r"[^\d.-]", "", v).strip()
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return float(v)

class InvoiceCreatePayload(BaseModel):
    doc_type: Optional[str] = "tax_invoice"
    project_id: Optional[str] = ""
    client_id: Optional[str] = ""
    invoice_number: Optional[str] = ""
    invoice_date: str
    due_date: Optional[str] = ""
    client_name: Optional[str] = ""
    project_name: Optional[str] = ""
    place_of_supply: Optional[str] = ""
    reverse_charge: Optional[str] = "No"
    seller_gstin: Optional[str] = ""
    buyer_gstin: Optional[str] = ""
    payment_terms: Optional[str] = "Due on Receipt"
    original_invoice_number: Optional[str] = ""
    reason: Optional[str] = ""
    payment_mode: Optional[str] = ""
    ref_number: Optional[str] = ""
    amount_received: Optional[float] = 0.0
    items: List[InvoiceItemPayload]
    subtotal: float
    discount: Optional[float] = 0.0
    cgst_rate: Optional[float] = 9.0
    sgst_rate: Optional[float] = 9.0
    igst_rate: Optional[float] = 0.0
    freight: Optional[float] = 0.0
    round_off: Optional[float] = 0.0
    grand_total: float
    notes: Optional[str] = ""
    terms: Optional[str] = ""
    status: Optional[str] = "Sent"
    allocated_payment_ids: Optional[List[str]] = []

    @field_validator("amount_received", "subtotal", "discount", "cgst_rate", "sgst_rate", "igst_rate", "freight", "round_off", "grand_total", mode="before")
    @classmethod
    def parse_invoice_numeric_fields(cls, v):
        if v is None or v == "":
            return 0.0
        if isinstance(v, str):
            cleaned = re.sub(r"[^\d.-]", "", v).strip()
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return float(v)

class ApplyPaymentPayload(BaseModel):
    payment_id: str
    allocated_amount: float

class PaymentUpdatePayload(BaseModel):
    milestone_name: Optional[str] = None
    payment_type: Optional[str] = None
    amount: Optional[float] = None
    payment_date: Optional[str] = None
    payment_source: Optional[str] = None
    payment_mode: Optional[str] = None
    ref_number: Optional[str] = None
    remarks: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    attachment_url: Optional[str] = None

class LoanRecordPayload(BaseModel):
    project_id: str
    client_id: Optional[str] = ""
    provider: str
    loan_amount: float
    approved_amount: Optional[float] = 0.0
    approved_date: Optional[str] = ""
    expected_disbursement_date: Optional[str] = ""
    disbursed_amount: Optional[float] = 0.0
    loan_ref: Optional[str] = ""
    status: Optional[str] = "Applied"
    remarks: Optional[str] = ""

class LoanUpdatePayload(BaseModel):
    provider: Optional[str] = None
    loan_amount: Optional[float] = None
    approved_amount: Optional[float] = None
    approved_date: Optional[str] = None
    expected_disbursement_date: Optional[str] = None
    disbursed_amount: Optional[float] = None
    loan_ref: Optional[str] = None
    status: Optional[str] = None
    remarks: Optional[str] = None

class ExpenseRecordPayload(BaseModel):
    project_id: Optional[str] = ""
    client_id: Optional[str] = ""
    category: str
    amount: float
    expense_date: Optional[str] = ""
    vendor_id: Optional[str] = ""
    vendor_name: Optional[str] = ""
    description: Optional[str] = ""
    payment_mode: Optional[str] = "Cash/UPI"
    ref_number: Optional[str] = ""
    payment_status: Optional[str] = "Paid"
    notes: Optional[str] = ""
    attachment_url: Optional[str] = ""

class ExpenseUpdatePayload(BaseModel):
    category: Optional[str] = None
    amount: Optional[float] = None
    expense_date: Optional[str] = None
    vendor_name: Optional[str] = None
    description: Optional[str] = None
    payment_mode: Optional[str] = None
    ref_number: Optional[str] = None
    payment_status: Optional[str] = None
    notes: Optional[str] = None
    attachment_url: Optional[str] = None

class VendorPayload(BaseModel):
    name: str
    contact_person: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    gstin: Optional[str] = ""
    address: Optional[str] = ""
    category: Optional[str] = "General Supplier"
    products_supplied: Optional[str] = ""
    payment_terms: Optional[str] = "Net 30"
    notes: Optional[str] = ""

class PurchaseBillItemPayload(BaseModel):
    product_name: str
    quantity: float
    unit: Optional[str] = "Nos"
    rate: float
    gst_rate: Optional[float] = 12.0
    amount: float

class PurchaseBillPayload(BaseModel):
    vendor_id: str
    vendor_name: Optional[str] = ""
    bill_number: str
    bill_date: str
    due_date: Optional[str] = ""
    po_reference: Optional[str] = ""
    items: List[PurchaseBillItemPayload]
    subtotal: float
    gst_total: float
    freight_charges: Optional[float] = 0.0
    transport_charges: Optional[float] = 0.0
    other_charges: Optional[float] = 0.0
    grand_total: float
    notes: Optional[str] = ""
    attachment_url: Optional[str] = ""
    payment_terms: Optional[str] = ""
    project_id: Optional[str] = ""

class MaterialInwardItemPayload(BaseModel):
    product_name: str
    bill_qty: float
    received_now: float
    remaining_qty: float
    destination: Optional[str] = "Main Warehouse"
    project_id: Optional[str] = ""
    unit: Optional[str] = "Nos"

class MaterialInwardPayload(BaseModel):
    bill_id: str
    vendor_id: str
    vendor_name: Optional[str] = ""
    challan_number: str
    challan_date: str
    received_by: str
    warehouse_id: Optional[str] = "Main Warehouse"
    project_id: Optional[str] = ""
    attachment_url: Optional[str] = ""
    items: List[MaterialInwardItemPayload]

class VendorPaymentPayload(BaseModel):
    vendor_id: str
    bill_id: Optional[str] = ""
    bill_number: Optional[str] = ""
    amount: float
    payment_method: str
    ref_number: Optional[str] = ""
    payment_date: str
    notes: Optional[str] = ""

class WarrantyPayload(BaseModel):
    product_type: str
    brand: Optional[str] = ""
    model: Optional[str] = ""
    serial_number: Optional[str] = ""
    installation_date: Optional[str] = ""
    warranty_start: Optional[str] = ""
    warranty_end: Optional[str] = ""
    warranty_type: Optional[str] = "Standard"
    provider: Optional[str] = ""

class ServiceVisitPayload(BaseModel):
    visit_date: str
    technician_name: Optional[str] = ""
    visit_type: Optional[str] = "Routine Maintenance"
    system_status: Optional[str] = "Operational"
    generation_obs: Optional[str] = ""
    earth_resistance: Optional[str] = ""
    fuses_status: Optional[str] = "OK"
    inverter_condition: Optional[str] = "Good"
    panel_condition: Optional[str] = "Good"
    customer_remarks: Optional[str] = ""
    technician_remarks: Optional[str] = ""
    next_visit_date: Optional[str] = ""

@api_router.get("/finance/receivables")
async def get_receivables_dashboard(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[str] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(get_current_user)
):
    cid = user["company_id"]
    query = {"company_id": cid}
    if client_id:
        query["id"] = client_id

    clients = await db.clients.find(query, {"_id": 0}).to_list(10000)
    all_projects = await db.projects.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    payments = await db.payments.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    loans = await db.loans.find({"company_id": cid}, {"_id": 0}).to_list(10000)

    total_project_val = 0.0
    total_received = 0.0
    total_outstanding = 0.0
    total_overdue = 0.0
    total_loan_pending = 0.0

    client_items = []
    for c in clients:
        c_projects = [p for p in all_projects if p.get("client_id") == c["id"]]
        
        if not c_projects:
            sys_kw = float(c.get("system_kw") or 0)
            quotation_val = float(c.get("quotation_value") or c.get("net_project_value") or (sys_kw * 45000 if sys_kw > 0 else 0))
            def_proj = {
                "id": f"proj_{c['id']}",
                "company_id": cid,
                "client_id": c["id"],
                "project_name": f"{sys_kw} kW Solar System" if sys_kw > 0 else "Rooftop Solar Project",
                "project_type": c.get("project_type") or "Rooftop Solar",
                "capacity_kw": sys_kw,
                "project_value": quotation_val,
                "project_date": (c.get("created_at") or "")[:10],
                "created_at": c.get("created_at") or now_iso(),
                "is_default": True
            }
            c_projects = [def_proj]

        c_project_items = []
        c_val = 0.0
        c_cash_rec = 0.0
        c_online_rec = 0.0
        c_loan_disb = 0.0
        c_loan_pend = 0.0
        c_rec = 0.0
        c_pen = 0.0
        c_cost = 0.0

        for p in c_projects:
            pid = p["id"]
            is_def = p.get("is_default", False)
            
            p_payments = [
                pay for pay in payments 
                if pay.get("project_id") == pid or (is_def and pay.get("client_id") == c["id"] and not pay.get("project_id"))
            ]
            p_expenses = [
                exp for exp in expenses 
                if exp.get("project_id") == pid or (is_def and exp.get("client_id") == c["id"] and not exp.get("project_id"))
            ]
            p_loans = [
                loan for loan in loans
                if loan.get("project_id") == pid or (is_def and loan.get("client_id") == c["id"] and not loan.get("project_id"))
            ]

            p_val = float(p.get("project_value") or p.get("quotation_value") or 0)

            # Separate Cash, Online, and Loan Disbursement payments
            p_cash_rec = sum(
                float(pay.get("amount") or 0) for pay in p_payments 
                if (pay.get("payment_source") or pay.get("payment_mode") or "").lower() == "cash" 
                and (pay.get("status") or "Received").lower() == "received"
            )
            p_online_rec = sum(
                float(pay.get("amount") or 0) for pay in p_payments 
                if (pay.get("payment_source") or pay.get("payment_mode") or "").lower() in ["online", "bank transfer", "cheque", "bank transfer / utr", "upi"] 
                and (pay.get("status") or "Received").lower() == "received"
            )
            
            p_loan_approved = sum(float(l.get("approved_amount") or 0) for l in p_loans if l.get("status") != "Rejected")
            p_loan_disbursed_loans = sum(float(l.get("disbursed_amount") or 0) for l in p_loans)
            p_loan_disbursed_payments = sum(
                float(pay.get("amount") or 0) for pay in p_payments 
                if (pay.get("payment_source") or "").lower() in ["loan / finance", "loan", "finance"]
                and (pay.get("status") or "Received").lower() == "received"
            )
            p_loan_disbursed = max(p_loan_disbursed_loans, p_loan_disbursed_payments)
            p_loan_pending = max(0.0, p_loan_approved - p_loan_disbursed)

            p_rec = p_cash_rec + p_online_rec + p_loan_disbursed
            p_pen = max(0.0, p_val - p_rec)
            p_cost = sum(float(exp.get("amount") or 0) for exp in p_expenses)
            has_costs = len(p_expenses) > 0
            p_profit = (p_val - p_cost) if has_costs else None

            if p_pen == 0 and p_val > 0:
                p_status = "Paid"
            elif p_rec > 0:
                p_status = "Partially Paid"
            else:
                p_status = "Pending"

            c_val += p_val
            c_cash_rec += p_cash_rec
            c_online_rec += p_online_rec
            c_loan_disb += p_loan_disbursed
            c_loan_pend += p_loan_pending
            c_rec += p_rec
            c_pen += p_pen
            c_cost += p_cost

            c_project_items.append({
                "id": pid,
                "client_id": c["id"],
                "project_name": p.get("project_name") or "Solar Project",
                "project_type": p.get("project_type") or "Rooftop Solar",
                "capacity_kw": float(p.get("capacity_kw") or p.get("system_kw") or 0),
                "project_value": p_val,
                "cash_received": p_cash_rec,
                "online_received": p_online_rec,
                "loan_disbursed": p_loan_disbursed,
                "loan_pending": p_loan_pending,
                "total_received": p_rec,
                "total_pending": p_pen,
                "total_cost": p_cost,
                "has_cost_data": has_costs,
                "estimated_profit": p_profit,
                "status": p_status,
                "payment_plan": p.get("payment_plan") or [],
                "project_date": p.get("project_date") or str(p.get("created_at") or "")[:10]
            })

        c_profit = (c_val - c_cost) if c_cost > 0 else None
        total_project_val += c_val
        total_received += c_rec
        total_outstanding += c_pen
        total_loan_pending += c_loan_pend

        client_items.append({
            "client_id": c["id"],
            "full_name": c.get("full_name") or "Unnamed Client",
            "mobile": c.get("mobile") or "",
            "sol_id": c.get("sol_id") or "",
            "consumer_number": c.get("consumer_number") or "",
            "client_total_value": c_val,
            "cash_received": c_cash_rec,
            "online_received": c_online_rec,
            "loan_disbursed": c_loan_disb,
            "loan_pending": c_loan_pend,
            "client_total_received": c_rec,
            "client_total_pending": c_pen,
            "client_total_cost": c_cost,
            "client_total_profit": c_profit,
            "projects_count": len(c_project_items),
            "projects": c_project_items
        })

    return {
        "summary": {
            "total_project_value": total_project_val,
            "total_received": total_received,
            "total_outstanding": total_outstanding,
            "total_overdue": total_overdue,
            "total_loan_pending": total_loan_pending,
            "active_clients_count": len(clients)
        },
        "items": client_items
    }

@api_router.post("/finance/projects")
async def create_project(data: ProjectCreatePayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    client = await db.clients.find_one({"id": data.client_id, "company_id": cid})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    payment_plan_data = [item.dict() for item in data.payment_plan] if data.payment_plan else []

    proj_doc = {
        "id": f"proj_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "client_id": data.client_id,
        "project_name": data.project_name.strip(),
        "project_type": data.project_type or "Rooftop Solar",
        "capacity_kw": float(data.capacity_kw or 0),
        "project_value": float(data.project_value or 0),
        "project_date": data.project_date or now_iso()[:10],
        "expected_completion_date": data.expected_completion_date or "",
        "payment_terms": data.payment_terms or "",
        "payment_plan": payment_plan_data,
        "notes": data.notes or "",
        "status": "Pending",
        "created_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    await db.projects.insert_one(proj_doc)
    await log_activity(cid, user["id"], user["name"], "Created Financial Project", f"Project: {data.project_name} for Client: {client.get('full_name')}")
    proj_doc.pop("_id", None)
    return {"message": "Project created successfully", "project": proj_doc}

@api_router.get("/finance/projects/{project_id}")
async def get_project_financial_details(project_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    project = await db.projects.find_one({"id": project_id, "company_id": cid}, {"_id": 0})
    
    if not project and project_id.startswith("proj_"):
        client_id_candidate = project_id.replace("proj_", "")
        client = await db.clients.find_one({"id": client_id_candidate, "company_id": cid}, {"_id": 0})
        if client:
            sys_kw = float(client.get("system_kw") or 0)
            quot_val = float(client.get("quotation_value") or client.get("net_project_value") or (sys_kw * 45000 if sys_kw > 0 else 0))
            project = {
                "id": project_id,
                "company_id": cid,
                "client_id": client["id"],
                "project_name": f"{sys_kw} kW Solar System" if sys_kw > 0 else "Rooftop Solar Project",
                "project_type": client.get("project_type") or "Rooftop Solar",
                "capacity_kw": sys_kw,
                "project_value": quot_val,
                "project_date": (client.get("created_at") or "")[:10],
                "created_at": client.get("created_at") or now_iso(),
                "is_default": True
            }

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    client = await db.clients.find_one({"id": project["client_id"], "company_id": cid}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found for this project")

    is_def = project.get("is_default", False)
    if is_def:
        payments = await db.payments.find(
            {"company_id": cid, "$or": [{"project_id": project_id}, {"client_id": client["id"], "project_id": {"$in": ["", None]}}]},
            {"_id": 0}
        ).sort("payment_date", -1).to_list(1000)
        expenses = await db.expenses.find(
            {"company_id": cid, "$or": [{"project_id": project_id}, {"client_id": client["id"], "project_id": {"$in": ["", None]}}]},
            {"_id": 0}
        ).sort("created_at", -1).to_list(1000)
        loans = await db.loans.find(
            {"company_id": cid, "$or": [{"project_id": project_id}, {"client_id": client["id"], "project_id": {"$in": ["", None]}}]},
            {"_id": 0}
        ).sort("created_at", -1).to_list(1000)
    else:
        payments = await db.payments.find({"project_id": project_id, "company_id": cid}, {"_id": 0}).sort("payment_date", -1).to_list(1000)
        expenses = await db.expenses.find({"project_id": project_id, "company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        loans = await db.loans.find({"project_id": project_id, "company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(1000)

    p_val = float(project.get("project_value") or project.get("quotation_value") or 0)

    cash_rec = sum(
        float(pay.get("amount") or 0) for pay in payments 
        if (pay.get("payment_source") or pay.get("payment_mode") or "").lower() == "cash" 
        and (pay.get("status") or "Received").lower() == "received"
    )
    online_rec = sum(
        float(pay.get("amount") or 0) for pay in payments 
        if (pay.get("payment_source") or pay.get("payment_mode") or "").lower() in ["online", "bank transfer", "cheque", "bank transfer / utr", "upi"] 
        and (pay.get("status") or "Received").lower() == "received"
    )

    loan_approved = sum(float(l.get("approved_amount") or 0) for l in loans if l.get("status") != "Rejected")
    loan_disbursed_loans = sum(float(l.get("disbursed_amount") or 0) for l in loans)
    loan_disbursed_payments = sum(
        float(pay.get("amount") or 0) for pay in payments 
        if (pay.get("payment_source") or "").lower() in ["loan / finance", "loan", "finance"]
        and (pay.get("status") or "Received").lower() == "received"
    )
    loan_disbursed = max(loan_disbursed_loans, loan_disbursed_payments)
    loan_pending = max(0.0, loan_approved - loan_disbursed)

    # Fetch project invoices
    invoices = await db.invoices.find(
        {"company_id": cid, "$or": [{"project_id": project_id}, {"client_id": client["id"]}]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)

    total_invoiced = sum(float(inv.get("grand_total") or 0) for inv in invoices if inv.get("status") != "Cancelled")
    
    # Calculate invoice paid & outstanding for each invoice
    processed_invoices = []
    total_invoice_outstanding = 0.0
    for inv in invoices:
        inv_total = float(inv.get("grand_total") or 0)
        inv_paid = sum(
            float(pay.get("allocated_amount") or pay.get("amount") or 0)
            for pay in payments
            if pay.get("invoice_id") == inv["id"] and (pay.get("status") or "Received").lower() == "received"
        )
        inv_out = max(0.0, inv_total - inv_paid)
        
        inv_status = inv.get("status") or "Sent"
        if inv_status not in ("Cancelled", "Draft"):
            if inv_paid >= inv_total and inv_total > 0:
                inv_status = "Paid"
            elif inv_paid > 0:
                inv_status = "Partially Paid"
            elif inv.get("due_date") and inv.get("due_date") < now_iso()[:10] and inv_out > 0:
                inv_status = "Overdue"
            else:
                inv_status = "Sent"

        if inv_status != "Cancelled":
            total_invoice_outstanding += inv_out

        inv_copy = dict(inv)
        inv_copy["paid_amount"] = inv_paid
        inv_copy["outstanding_amount"] = inv_out
        inv_copy["status"] = inv_status
        processed_invoices.append(inv_copy)

    total_rec = sum(
        float(pay.get("amount") or 0)
        for pay in payments
        if (pay.get("status") or "Received").lower() == "received"
    )
    net_outstanding = max(0.0, p_val - total_rec)
    uninvoiced_value = max(0.0, p_val - total_invoiced)

    unallocated_payments = [
        pay for pay in payments
        if not pay.get("invoice_id") and (pay.get("status") or "Received").lower() == "received"
    ]
    unallocated_advance = sum(float(pay.get("amount") or 0) for pay in unallocated_payments)

    total_exp = sum(float(exp.get("amount") or 0) for exp in expenses)
    has_costs = len(expenses) > 0
    estimated_profit = (p_val - total_exp) if has_costs else None

    # Build chronological ledger
    ledger = []
    for pay in payments:
        ledger.append({
            "id": pay["id"],
            "type": "payment",
            "date": pay.get("payment_date") or (pay.get("created_at") or "")[:10],
            "description": pay.get("milestone_name") or pay.get("payment_type") or "Payment Received",
            "source_mode": pay.get("payment_source") or pay.get("payment_mode") or "Bank Transfer",
            "ref": pay.get("ref_number") or pay.get("bank_utr") or "",
            "status": pay.get("status") or "Received",
            "amount": float(pay.get("amount") or 0),
            "signed_amount": float(pay.get("amount") or 0),
            "notes": pay.get("remarks") or pay.get("notes") or "",
            "attachment_url": pay.get("attachment_url") or "",
            "raw": pay
        })

    for exp in expenses:
        ledger.append({
            "id": exp["id"],
            "type": "expense",
            "date": exp.get("expense_date") or (exp.get("created_at") or "")[:10],
            "description": exp.get("category") or "Project Expense",
            "source_mode": exp.get("vendor_name") or exp.get("payment_mode") or "",
            "ref": exp.get("ref_number") or "",
            "status": exp.get("payment_status") or "Paid",
            "amount": float(exp.get("amount") or 0),
            "signed_amount": -float(exp.get("amount") or 0),
            "notes": exp.get("description") or exp.get("notes") or "",
            "attachment_url": exp.get("attachment_url") or "",
            "raw": exp
        })

    for l in loans:
        ledger.append({
            "id": l["id"],
            "type": "loan",
            "date": l.get("approved_date") or (l.get("created_at") or "")[:10],
            "description": f"Loan: {l.get('provider')}",
            "source_mode": f"Status: {l.get('status')}",
            "ref": l.get("loan_ref") or "",
            "status": l.get("status") or "Approved",
            "amount": float(l.get("disbursed_amount") or 0),
            "signed_amount": float(l.get("disbursed_amount") or 0),
            "notes": f"Approved: ₹{l.get('approved_amount', 0)}, Disbursed: ₹{l.get('disbursed_amount', 0)}. {l.get('remarks', '')}",
            "attachment_url": "",
            "raw": l
        })

    ledger.sort(key=lambda x: str(x["date"]), reverse=True)

    return {
        "project": project,
        "client": client,
        "summary": {
            "project_value": p_val,
            "total_invoiced": total_invoiced,
            "total_received": total_rec,
            "invoice_outstanding": total_invoice_outstanding,
            "uninvoiced_value": uninvoiced_value,
            "project_outstanding": net_outstanding,
            "unallocated_advance": unallocated_advance,
            "cash_received": cash_rec,
            "online_received": online_rec,
            "loan_disbursed": loan_disbursed,
            "loan_pending": loan_pending,
            "net_outstanding": net_outstanding,
            "total_expense": total_exp,
            "has_cost_data": has_costs,
            "estimated_profit": estimated_profit
        },
        "payment_plan": project.get("payment_plan") or [],
        "invoices": processed_invoices,
        "payments": payments,
        "unallocated_payments": unallocated_payments,
        "loans": loans,
        "expenses": expenses,
        "ledger": ledger
    }

@api_router.post("/finance/projects/{project_id}/payments")
async def record_project_payment(project_id: str, data: PaymentRecordPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    
    project = await db.projects.find_one({"id": project_id, "company_id": cid})
    client_id = data.client_id if data.client_id else None
    if not client_id and project:
        client_id = project.get("client_id")
    elif not client_id and project_id.startswith("proj_"):
        client_id = project_id.replace("proj_", "")
    
    if not client_id:
        cl = await db.clients.find_one({"$or": [{"id": project_id}, {"sol_id": project_id}], "company_id": cid})
        if cl:
            client_id = cl["id"]

    if not client_id:
        raise HTTPException(status_code=400, detail="Client ID required for payment")

    payment_doc = {
        "id": f"pay_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "client_id": client_id,
        "project_id": project_id,
        "milestone_id": data.milestone_id or "",
        "milestone_name": data.milestone_name or data.payment_type or "Payment",
        "payment_type": data.payment_type or "Advance",
        "amount": data.amount,
        "payment_date": data.payment_date or now_iso()[:10],
        "payment_source": data.payment_source or "Bank Transfer",
        "payment_mode": data.payment_mode or data.payment_source or "Bank Transfer",
        "ref_number": data.ref_number or "",
        "bank_utr": data.bank_utr or "",
        "remarks": data.remarks or "",
        "notes": data.notes or data.remarks or "",
        "status": data.status or "Received",
        "attachment_url": data.attachment_url or "",
        "recorded_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    await db.payments.insert_one(payment_doc)
    await log_activity(cid, user["id"], user["name"], "Recorded Payment", f"Amount: ₹{data.amount} for Project ID: {project_id}")
    payment_doc.pop("_id", None)
    return {"message": "Payment recorded successfully", "payment": payment_doc}

@api_router.put("/finance/payments/{payment_id}")
async def update_payment(payment_id: str, data: PaymentUpdatePayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.payments.find_one({"id": payment_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Payment not found")

    update_fields = {"updated_at": now_iso()}
    if data.milestone_name is not None: update_fields["milestone_name"] = data.milestone_name
    if data.payment_type is not None: update_fields["payment_type"] = data.payment_type
    if data.amount is not None: update_fields["amount"] = data.amount
    if data.payment_date is not None: update_fields["payment_date"] = data.payment_date
    if data.payment_source is not None: update_fields["payment_source"] = data.payment_source
    if data.payment_mode is not None: update_fields["payment_mode"] = data.payment_mode
    if data.ref_number is not None: update_fields["ref_number"] = data.ref_number
    if data.remarks is not None: update_fields["remarks"] = data.remarks
    if data.notes is not None: update_fields["notes"] = data.notes
    if data.status is not None: update_fields["status"] = data.status
    if data.attachment_url is not None: update_fields["attachment_url"] = data.attachment_url

    await db.payments.update_one({"id": payment_id, "company_id": cid}, {"$set": update_fields})
    return {"message": "Payment updated successfully"}

@api_router.delete("/finance/payments/{payment_id}")
async def delete_payment(payment_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.payments.find_one({"id": payment_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Payment not found")

    await db.payments.delete_one({"id": payment_id, "company_id": cid})
    await log_activity(cid, user["id"], user["name"], "Deleted Payment", f"Payment ID: {payment_id}")
    return {"message": "Payment deleted successfully"}

@api_router.post("/finance/projects/{project_id}/loans")
async def record_project_loan(project_id: str, data: LoanRecordPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    project = await db.projects.find_one({"id": project_id, "company_id": cid})
    client_id = data.client_id or (project.get("client_id") if project else None)
    if not client_id and project_id.startswith("proj_"):
        client_id = project_id.replace("proj_", "")

    loan_doc = {
        "id": f"loan_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "client_id": client_id,
        "project_id": project_id,
        "provider": data.provider.strip(),
        "loan_amount": float(data.loan_amount or 0),
        "approved_amount": float(data.approved_amount or 0),
        "approved_date": data.approved_date or "",
        "expected_disbursement_date": data.expected_disbursement_date or "",
        "disbursed_amount": float(data.disbursed_amount or 0),
        "loan_ref": data.loan_ref or "",
        "status": data.status or "Applied",
        "remarks": data.remarks or "",
        "created_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    await db.loans.insert_one(loan_doc)
    await log_activity(cid, user["id"], user["name"], "Recorded Loan", f"Provider: {data.provider}, Approved: ₹{data.approved_amount}")
    loan_doc.pop("_id", None)
    return {"message": "Loan record created successfully", "loan": loan_doc}

@api_router.put("/finance/loans/{loan_id}")
async def update_project_loan(loan_id: str, data: LoanUpdatePayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.loans.find_one({"id": loan_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Loan record not found")

    update_fields = {"updated_at": now_iso()}
    if data.provider is not None: update_fields["provider"] = data.provider
    if data.loan_amount is not None: update_fields["loan_amount"] = data.loan_amount
    if data.approved_amount is not None: update_fields["approved_amount"] = data.approved_amount
    if data.approved_date is not None: update_fields["approved_date"] = data.approved_date
    if data.expected_disbursement_date is not None: update_fields["expected_disbursement_date"] = data.expected_disbursement_date
    if data.disbursed_amount is not None: update_fields["disbursed_amount"] = data.disbursed_amount
    if data.loan_ref is not None: update_fields["loan_ref"] = data.loan_ref
    if data.status is not None: update_fields["status"] = data.status
    if data.remarks is not None: update_fields["remarks"] = data.remarks

    await db.loans.update_one({"id": loan_id, "company_id": cid}, {"$set": update_fields})
    return {"message": "Loan record updated successfully"}

@api_router.delete("/finance/loans/{loan_id}")
async def delete_project_loan(loan_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.loans.find_one({"id": loan_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Loan record not found")

    await db.loans.delete_one({"id": loan_id, "company_id": cid})
    await log_activity(cid, user["id"], user["name"], "Deleted Loan", f"Loan ID: {loan_id}")
    return {"message": "Loan record deleted successfully"}

@api_router.post("/finance/projects/{project_id}/expenses")
async def record_project_expense(project_id: str, data: ExpenseRecordPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    
    project = await db.projects.find_one({"id": project_id, "company_id": cid})
    client_id = data.client_id if data.client_id else None
    if not client_id and project:
        client_id = project.get("client_id")
    elif not client_id and project_id.startswith("proj_"):
        client_id = project_id.replace("proj_", "")
        
    if not client_id:
        raise HTTPException(status_code=400, detail="Client ID required for expense")

    expense_doc = {
        "id": f"exp_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "client_id": client_id,
        "project_id": project_id,
        "category": data.category,
        "amount": data.amount,
        "expense_date": data.expense_date or now_iso()[:10],
        "vendor_id": data.vendor_id or "",
        "vendor_name": data.vendor_name or "",
        "description": data.description or "",
        "payment_mode": data.payment_mode or "Cash/UPI",
        "ref_number": data.ref_number or "",
        "payment_status": data.payment_status or "Paid",
        "notes": data.notes or "",
        "attachment_url": data.attachment_url or "",
        "created_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    await db.expenses.insert_one(expense_doc)
    await log_activity(cid, user["id"], user["name"], "Recorded Expense", f"Category: {data.category}, Amount: ₹{data.amount}")
    expense_doc.pop("_id", None)
    return {"message": "Expense logged successfully", "expense": expense_doc}

@api_router.put("/finance/expenses/{expense_id}")
async def update_expense(expense_id: str, data: ExpenseUpdatePayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.expenses.find_one({"id": expense_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")

    update_fields = {"updated_at": now_iso()}
    if data.category is not None: update_fields["category"] = data.category
    if data.amount is not None: update_fields["amount"] = data.amount
    if data.expense_date is not None: update_fields["expense_date"] = data.expense_date
    if data.vendor_name is not None: update_fields["vendor_name"] = data.vendor_name
    if data.description is not None: update_fields["description"] = data.description
    if data.payment_mode is not None: update_fields["payment_mode"] = data.payment_mode
    if data.ref_number is not None: update_fields["ref_number"] = data.ref_number
    if data.payment_status is not None: update_fields["payment_status"] = data.payment_status
    if data.notes is not None: update_fields["notes"] = data.notes
    if data.attachment_url is not None: update_fields["attachment_url"] = data.attachment_url

    await db.expenses.update_one({"id": expense_id, "company_id": cid}, {"$set": update_fields})
    return {"message": "Expense updated successfully"}

@api_router.delete("/finance/expenses/{expense_id}")
async def delete_expense(expense_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.expenses.find_one({"id": expense_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")

    await db.expenses.delete_one({"id": expense_id, "company_id": cid})
    await log_activity(cid, user["id"], user["name"], "Deleted Expense", f"Expense ID: {expense_id}")
    return {"message": "Expense deleted successfully"}

# ── INVOICE MANAGEMENT ENDPOINTS ──────────────────────────────────────────────

@api_router.get("/finance/invoices")
async def list_invoices(
    project_id: Optional[str] = None,
    client_id: Optional[str] = None,
    user=Depends(get_current_user)
):
    cid = user["company_id"]
    query = {"company_id": cid}
    if project_id:
        query["project_id"] = project_id
    if client_id:
        query["client_id"] = client_id

    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    payments = await db.payments.find({"company_id": cid}, {"_id": 0}).to_list(10000)

    for inv in invoices:
        inv_id = inv["id"]
        inv_total = float(inv.get("grand_total") or 0)
        inv_paid = sum(
            float(p.get("allocated_amount") or p.get("amount") or 0)
            for p in payments
            if p.get("invoice_id") == inv_id and (p.get("status") or "Received").lower() == "received"
        )
        inv_outstanding = max(0.0, inv_total - inv_paid)

        status = inv.get("status") or "Sent"
        if status not in ("Cancelled", "Draft"):
            if inv_paid >= inv_total and inv_total > 0:
                status = "Paid"
            elif inv_paid > 0:
                status = "Partially Paid"
            elif inv.get("due_date") and inv.get("due_date") < now_iso()[:10] and inv_outstanding > 0:
                status = "Overdue"

        inv["paid_amount"] = inv_paid
        inv["outstanding_amount"] = inv_outstanding
        inv["status"] = status

    return {"invoices": invoices}

@api_router.post("/finance/invoices")
async def create_invoice(data: InvoiceCreatePayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    client_id = data.client_id
    project_id = data.project_id
    doc_type = (data.doc_type or "tax_invoice").lower().strip()

    if not client_id and project_id:
        proj = await db.projects.find_one({"id": project_id, "company_id": cid})
        if proj:
            client_id = proj.get("client_id")

    client = await db.clients.find_one({"id": client_id, "company_id": cid}) if client_id else None

    invoice_id = f"inv_{uuid.uuid4().hex[:12]}"
    prefix_map = {
        "tax_invoice": "INV",
        "proforma": "PI",
        "payment_receipt": "REC",
        "credit_note": "CN",
        "debit_note": "DN"
    }
    prefix = prefix_map.get(doc_type, "INV")
    inv_num = data.invoice_number or f"{prefix}-{now_iso()[:10].replace('-', '')}-{uuid.uuid4().hex[:4].upper()}"

    formatted_items = []
    for item in data.items:
        qty = float(item.quantity or 0)
        rate = float(item.rate if item.rate is not None else item.unit_price or 0)
        disc = float(item.discount or 0)
        amt = float(item.amount or max(0.0, qty * rate - disc))
        gst_r = float(item.gst_rate or 18.0)
        formatted_items.append({
            "product_name": item.product_name,
            "product": item.product_name,
            "hsn_sac": item.hsn_sac or "",
            "size": item.size or "",
            "quantity": qty,
            "unit": item.unit or "Nos",
            "rate": rate,
            "unit_price": rate,
            "discount": disc,
            "gst_rate": gst_r,
            "gst": gst_r,
            "cgst": (amt * (data.cgst_rate or 9.0)) / 100,
            "sgst": (amt * (data.sgst_rate or 9.0)) / 100,
            "igst": (amt * (data.igst_rate or 0.0)) / 100,
            "amount": amt
        })

    invoice_doc = {
        "id": invoice_id,
        "company_id": cid,
        "doc_type": doc_type,
        "client_id": client_id or "",
        "project_id": project_id or "",
        "invoice_number": inv_num,
        "invoice_date": data.invoice_date or now_iso()[:10],
        "due_date": data.due_date or "",
        "client_name": data.client_name or (client.get("full_name") if client else "Customer"),
        "project_name": data.project_name or "",
        "place_of_supply": data.place_of_supply or (client.get("state") if client else ""),
        "reverse_charge": data.reverse_charge or "No",
        "seller_gstin": data.seller_gstin or "",
        "buyer_gstin": data.buyer_gstin or (client.get("gstin") if client else ""),
        "payment_terms": data.payment_terms or "Due on Receipt",
        "original_invoice_number": data.original_invoice_number or "",
        "reason": data.reason or "",
        "payment_mode": data.payment_mode or "",
        "ref_number": data.ref_number or "",
        "amount_received": data.amount_received or 0.0,
        "items": formatted_items,
        "subtotal": data.subtotal,
        "discount": data.discount or 0.0,
        "cgst_rate": data.cgst_rate or 9.0,
        "sgst_rate": data.sgst_rate or 9.0,
        "igst_rate": data.igst_rate or 0.0,
        "freight": data.freight or 0.0,
        "round_off": data.round_off or 0.0,
        "grand_total": data.grand_total,
        "notes": data.notes or "",
        "terms": data.terms or "",
        "status": data.status or "Sent",
        "created_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    # PDF Generation (if not Draft)
    if data.status != "Draft":
        company_doc = await db.companies.find_one({"id": cid}, {"_id": 0}) or {}
        company_doc = await _enrich_company_doc(company_doc)
        doc_data_pdf = {
            **invoice_doc,
            "client": client or {
                "full_name": invoice_doc["client_name"],
                "mobile": client.get("mobile", "") if client else "",
                "address": client.get("site_address", "") if client else "",
                "gstin": invoice_doc["buyer_gstin"]
            }
        }
        try:
            pdf_bytes = pdf_generator.generate_document(doc_type, doc_data_pdf, company_doc)
            filename = f"{invoice_doc['client_name']}_{prefix}_{inv_num}_{invoice_doc['invoice_date']}.pdf".replace(" ", "_")
            storage_path = f"{APP_NAME}/{cid}/generated/{invoice_id}.pdf"
            result = put_object(storage_path, pdf_bytes, "application/pdf")

            file_rec = {
                "id": invoice_id,
                "company_id": cid,
                "uploader_id": user["id"],
                "storage_path": result["path"],
                "original_filename": filename,
                "content_type": "application/pdf",
                "size": len(pdf_bytes),
                "category": "generated",
                "is_deleted": False,
                "created_at": now_iso(),
                "doc_type": doc_type,
                "document_number": inv_num,
                "client_name": invoice_doc["client_name"],
                "prepared_by": user["name"],
                "status": "Active"
            }
            await db.files.insert_one(file_rec)
            invoice_doc["file_id"] = invoice_id
        except Exception as e:
            logger.error(f"Error generating Invoice PDF: {e}")

    await db.invoices.insert_one(invoice_doc)

    if data.allocated_payment_ids:
        for p_id in data.allocated_payment_ids:
            await db.payments.update_one(
                {"id": p_id, "company_id": cid},
                {"$set": {"invoice_id": invoice_id, "invoice_no": inv_num, "updated_at": now_iso()}}
            )

    await log_activity(cid, user["id"], user["name"], f"Created {doc_type.replace('_', ' ').title()}", f"Invoice #{inv_num} for ₹{data.grand_total}")
    invoice_doc.pop("_id", None)
    return {"message": "Invoice created successfully", "invoice": invoice_doc}

@api_router.post("/finance/invoices/{invoice_id}/generate-doc")
async def generate_invoice_doc(invoice_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    cid = user["company_id"]
    fmt = (payload.get("format") or "pdf").lower().strip()
    invoice = await db.invoices.find_one({"id": invoice_id, "company_id": cid}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    company_doc = await db.companies.find_one({"id": cid}, {"_id": 0}) or {}
    company_doc = await _enrich_company_doc(company_doc)
    client_doc = None
    if invoice.get("client_id"):
        client_doc = await db.clients.find_one({"id": invoice["client_id"], "company_id": cid}, {"_id": 0})

    doc_data = {
        **invoice,
        "client": client_doc or {}
    }

    doc_type = invoice.get("doc_type") or "tax_invoice"
    if fmt == "docx":
        pdf_bytes = pdf_generator.generate_docx(doc_type, doc_data, company_doc)
        content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ext = ".docx"
    else:
        pdf_bytes = pdf_generator.generate_document(doc_type, doc_data, company_doc)
        content_type = "application/pdf"
        ext = ".pdf"

    file_id = str(uuid.uuid4())
    filename = f"{invoice.get('client_name', 'Client')}_Invoice_{invoice.get('invoice_number', invoice_id)}{ext}".replace(" ", "_")
    storage_path = f"{APP_NAME}/{cid}/generated/{file_id}{ext}"
    result = put_object(storage_path, pdf_bytes, content_type)

    file_rec = {
        "id": file_id, "company_id": cid, "uploader_id": user["id"],
        "storage_path": result["path"], "original_filename": filename,
        "content_type": content_type, "size": len(pdf_bytes),
        "category": "generated", "is_deleted": False, "created_at": now_iso(),
        "doc_type": doc_type,
        "document_number": invoice.get("invoice_number"),
        "client_name": invoice.get("client_name"),
        "prepared_by": user["name"],
        "status": "Active"
    }
    await db.files.insert_one(file_rec)
    await db.invoices.update_one({"id": invoice_id, "company_id": cid}, {"$set": {"file_id": file_id, "updated_at": now_iso()}})

    return {"id": file_id, "filename": filename, "file_id": file_id}

@api_router.post("/finance/invoices/{invoice_id}/apply-payment")
async def apply_payment_to_invoice(invoice_id: str, payload: ApplyPaymentPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    invoice = await db.invoices.find_one({"id": invoice_id, "company_id": cid})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    payment = await db.payments.find_one({"id": payload.payment_id, "company_id": cid})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    all_payments = await db.payments.find({"company_id": cid, "invoice_id": invoice_id, "status": "Received"}).to_list(1000)
    current_paid = sum(float(p.get("allocated_amount") or p.get("amount") or 0) for p in all_payments if p["id"] != payload.payment_id)
    inv_total = float(invoice.get("grand_total") or 0)
    current_outstanding = max(0.0, inv_total - current_paid)

    alloc_amount = float(payload.allocated_amount or payment.get("amount") or 0)
    if alloc_amount > current_outstanding + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot allocate ₹{alloc_amount}. Maximum allocatable amount to this invoice is ₹{current_outstanding:.2f}"
        )

    await db.payments.update_one(
        {"id": payload.payment_id, "company_id": cid},
        {"$set": {
            "invoice_id": invoice_id,
            "invoice_no": invoice.get("invoice_number", ""),
            "allocated_amount": alloc_amount,
            "updated_at": now_iso()
        }}
    )

    new_paid = current_paid + alloc_amount
    new_outstanding = max(0.0, inv_total - new_paid)
    new_status = "Paid" if new_paid >= inv_total else ("Partially Paid" if new_paid > 0 else "Sent")
    await db.invoices.update_one(
        {"id": invoice_id, "company_id": cid},
        {"$set": {"status": new_status, "updated_at": now_iso()}}
    )

    return {
        "message": f"Successfully allocated ₹{alloc_amount} to Invoice #{invoice.get('invoice_number')}",
        "invoice_paid": new_paid,
        "invoice_outstanding": new_outstanding,
        "invoice_status": new_status
    }

@api_router.delete("/finance/invoices/{invoice_id}")
async def cancel_invoice(invoice_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    existing = await db.invoices.find_one({"id": invoice_id, "company_id": cid})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")

    await db.invoices.update_one({"id": invoice_id, "company_id": cid}, {"$set": {"status": "Cancelled", "updated_at": now_iso()}})
    await db.payments.update_many({"invoice_id": invoice_id, "company_id": cid}, {"$unset": {"invoice_id": "", "invoice_no": "", "allocated_amount": ""}})
    await log_activity(cid, user["id"], user["name"], "Cancelled Tax Invoice", f"Invoice ID: {invoice_id}")
    return {"message": "Invoice cancelled successfully"}

@api_router.get("/vendors")
async def list_vendors(user=Depends(get_current_user)):
    cid = user["company_id"]
    vendors = await db.vendors.find({"company_id": cid}, {"_id": 0}).sort("name", 1).to_list(1000)
    bills = await db.purchase_bills.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    for b in bills:
        await _enrich_bill_challan(b, cid)
    payments = await db.vendor_payments.find({"company_id": cid}, {"_id": 0}).to_list(10000)
    inwards = await db.vendor_inwards.find({"company_id": cid}, {"_id": 0}).to_list(10000)

    for v in vendors:
        v_id = v["id"]
        v_name = (v.get("name") or "").lower()
        v_bills = [b for b in bills if b.get("vendor_id") == v_id or (b.get("vendor_name") or "").lower() == v_name]
        v_payments = [p for p in payments if p.get("vendor_id") == v_id or (p.get("vendor_name") or "").lower() == v_name]
        v_inwards = [i for i in inwards if i.get("vendor_id") == v_id or (i.get("vendor_name") or "").lower() == v_name]

        tot_purchases = sum(float(b.get("grand_total") or 0) for b in v_bills)
        tot_paid = sum(float(p.get("amount") or 0) for p in v_payments)
        tot_outstanding = max(0.0, tot_purchases - tot_paid)

        v["total_purchases"] = tot_purchases
        v["total_paid"] = tot_paid
        v["total_outstanding"] = tot_outstanding
        v["purchases_count"] = len(v_bills)
        v["inwards_count"] = len(v_inwards)

    return {"vendors": vendors}

@api_router.post("/vendors")
async def create_vendor(data: VendorPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    vendor_doc = {
        "id": f"ven_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "name": data.name.strip(),
        "contact_person": data.contact_person or "",
        "phone": data.phone or "",
        "email": data.email or "",
        "gstin": data.gstin or "",
        "address": data.address or "",
        "category": data.category or "General Supplier",
        "products_supplied": data.products_supplied or "",
        "payment_terms": data.payment_terms or "Net 30",
        "notes": data.notes or "",
        "status": "Active",
        "created_at": now_iso()
    }

    await db.vendors.insert_one(vendor_doc)
    await log_activity(cid, user["id"], user["name"], "Created Vendor", f"Vendor: {data.name}")
    vendor_doc.pop("_id", None)
    return {"message": "Vendor created successfully", "vendor": vendor_doc}

@api_router.patch("/vendors/{vendor_id}/status")
async def toggle_vendor_status(vendor_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    cid = user["company_id"]
    new_status = payload.get("status", "Active")
    await db.vendors.update_one({"id": vendor_id, "company_id": cid}, {"$set": {"status": new_status, "updated_at": now_iso()}})
    await log_activity(cid, user["id"], user["name"], "Updated Vendor Status", f"Vendor ID: {vendor_id} to {new_status}")
    return {"message": f"Vendor status updated to {new_status}"}

@api_router.put("/vendors/{vendor_id}")
async def update_vendor_detail(vendor_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    cid = user["company_id"]
    vendor = await db.vendors.find_one({"id": vendor_id, "company_id": cid})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    update_fields = {
        "name": str(payload.get("name") or vendor.get("name") or "").strip(),
        "contact_person": payload.get("contact_person", vendor.get("contact_person", "")),
        "phone": payload.get("phone", vendor.get("phone", "")),
        "email": payload.get("email", vendor.get("email", "")),
        "gstin": payload.get("gstin", vendor.get("gstin", "")),
        "address": payload.get("address", vendor.get("address", "")),
        "notes": payload.get("notes", vendor.get("notes", "")),
        "updated_at": now_iso()
    }
    await db.vendors.update_one({"id": vendor_id, "company_id": cid}, {"$set": update_fields})
    await log_activity(cid, user["id"], user["name"], "Updated Vendor Master", f"Vendor: {update_fields['name']}")
    return {"message": "Vendor details updated successfully"}

# ── PURCHASE ORDERS CRUD ENDPOINTS ──────────────────────────────────────────
@api_router.get("/purchase-orders")
async def list_purchase_orders(user=Depends(get_current_user)):
    cid = user["company_id"]
    pos = await db.purchase_orders.find({"company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return {"purchase_orders": pos}

@api_router.post("/purchase-orders")
async def create_purchase_order(payload: Dict[str, Any], user=Depends(get_current_user)):
    cid = user["company_id"]
    po_id = payload.get("id") or f"po_{uuid.uuid4().hex[:12]}"
    
    vendor_name = (payload.get("vendor_name") or "").strip()
    vendor_id = payload.get("vendor_id") or ""
    
    # Auto-save / Update Vendor in Vendor Master if requested
    if payload.get("save_vendor_master") and vendor_name:
        existing_v = None
        if vendor_id:
            existing_v = await db.vendors.find_one({"id": vendor_id, "company_id": cid})
        if not existing_v:
            existing_v = await db.vendors.find_one({"name": {"$regex": f"^{re.escape(vendor_name)}$", "$options": "i"}, "company_id": cid})
        
        v_doc = {
            "name": vendor_name,
            "address": payload.get("vendor_address") or "",
            "phone": payload.get("vendor_phone") or "",
            "email": payload.get("vendor_email") or "",
            "gstin": payload.get("vendor_gstin") or "",
            "updated_at": now_iso()
        }
        if existing_v:
            await db.vendors.update_one({"id": existing_v["id"], "company_id": cid}, {"$set": v_doc})
            vendor_id = existing_v["id"]
        else:
            vendor_id = vendor_id or f"ven_{uuid.uuid4().hex[:12]}"
            v_doc.update({
                "id": vendor_id,
                "company_id": cid,
                "status": "Active",
                "created_at": now_iso()
            })
            await db.vendors.insert_one(v_doc)

    po_doc = {
        "id": po_id,
        "company_id": cid,
        "po_number": payload.get("po_number") or f"PO-{uuid.uuid4().hex[:6].upper()}",
        "po_date": payload.get("po_date") or now_iso()[:10],
        "vendor_id": vendor_id,
        "vendor_name": vendor_name,
        "vendor_address": payload.get("vendor_address") or "",
        "vendor_phone": payload.get("vendor_phone") or "",
        "vendor_email": payload.get("vendor_email") or "",
        "vendor_gstin": payload.get("vendor_gstin") or "",
        "ship_via": payload.get("ship_via") or "",
        "shipping_method": payload.get("shipping_method") or "",
        "shipping_term": payload.get("shipping_term") or "",
        "delivery_date": payload.get("delivery_date") or "",
        "items": payload.get("items") or [],
        "notes": payload.get("notes") or "",
        "cgst_rate": float(payload.get("cgst_rate") or 0),
        "sgst_rate": float(payload.get("sgst_rate") or 0),
        "igst_rate": float(payload.get("igst_rate") or 0),
        "freight": float(payload.get("freight") or 0),
        "subtotal": float(payload.get("subtotal") or 0),
        "grand_total": float(payload.get("grand_total") or 0),
        "status": payload.get("status") or "Created",
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    
    await db.purchase_orders.update_one(
        {"id": po_id, "company_id": cid},
        {"$set": po_doc},
        upsert=True
    )
    await log_activity(cid, user["id"], user["name"], "Saved Purchase Order", f"PO #: {po_doc['po_number']}")
    return {"message": "Purchase Order saved successfully", "purchase_order": po_doc}

@api_router.get("/purchase-orders/{po_id}")
async def get_purchase_order(po_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    po = await db.purchase_orders.find_one({"id": po_id, "company_id": cid}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    return po

@api_router.patch("/purchase-orders/{po_id}")
async def update_purchase_order(po_id: str, payload: Dict[str, Any], user=Depends(get_current_user)):
    cid = user["company_id"]
    po = await db.purchase_orders.find_one({"id": po_id, "company_id": cid})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    update_data = {k: v for k, v in payload.items() if k not in ("_id", "id", "company_id", "created_at")}
    update_data["updated_at"] = now_iso()
    
    await db.purchase_orders.update_one({"id": po_id, "company_id": cid}, {"$set": update_data})
    await log_activity(cid, user["id"], user["name"], "Updated Purchase Order", f"PO ID: {po_id}")
    return {"message": "Purchase Order updated successfully"}

@api_router.delete("/purchase-orders/{po_id}")
async def delete_purchase_order(po_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    res = await db.purchase_orders.delete_one({"id": po_id, "company_id": cid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    await log_activity(cid, user["id"], user["name"], "Deleted Purchase Order", f"PO ID: {po_id}")
    return {"message": "Purchase Order deleted successfully"}

@api_router.get("/vendors/{vendor_id}")
async def get_vendor_detail(vendor_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    vendor = await db.vendors.find_one({"id": vendor_id, "company_id": cid}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    bills = await db.purchase_bills.find({"vendor_id": vendor_id, "company_id": cid}, {"_id": 0}).sort("bill_date", -1).to_list(1000)
    for b in bills:
        await _enrich_bill_challan(b, cid)
    inwards = await db.vendor_inwards.find({"vendor_id": vendor_id, "company_id": cid}, {"_id": 0}).sort("challan_date", -1).to_list(1000)
    central_inwards = await db.inward_entries.find({"$or": [{"vendor_id": vendor_id}, {"source_name": vendor.get("name")}], "company_id": cid}, {"_id": 0}).sort("date", -1).to_list(1000)
    payments = await db.vendor_payments.find({"vendor_id": vendor_id, "company_id": cid}, {"_id": 0}).sort("payment_date", -1).to_list(1000)

    tot_purchases = sum(float(b.get("grand_total") or 0) for b in bills)
    tot_paid = sum(float(p.get("amount") or 0) for p in payments)
    tot_outstanding = max(0.0, tot_purchases - tot_paid)

    activity = []
    for b in bills:
        activity.append({
            "id": b["id"],
            "date": b.get("bill_date") or (b.get("created_at") or "")[:10],
            "type": "Purchase Bill",
            "reference": b.get("bill_number") or b["id"],
            "amount": float(b.get("grand_total") or 0),
            "status": b.get("inward_status") or b.get("status") or "Bill Created",
            "raw": b
        })

    for inw in inwards:
        items_summary = ", ".join([f"{i.get('received_now')} {i.get('product_name')}" for i in inw.get("items", []) if float(i.get('received_now') or 0) > 0])
        activity.append({
            "id": inw["id"],
            "date": inw.get("challan_date") or (inw.get("created_at") or "")[:10],
            "type": "Inward",
            "reference": inw.get("challan_number") or inw["id"],
            "amount": None,
            "details": items_summary or "Material Received",
            "status": "Received",
            "raw": inw
        })

    for cinw in central_inwards:
        activity.append({
            "id": cinw["id"],
            "date": (cinw.get("date") or "")[:10],
            "type": "Inward",
            "reference": cinw.get("reference_number") or cinw.get("bill_number") or cinw["id"],
            "amount": float(cinw.get("line_total") or 0) if cinw.get("bill_type") == "Product Bill" else None,
            "details": f"{cinw.get('quantity')} {cinw.get('unit', 'Nos')} {cinw.get('product')}",
            "status": "Received",
            "raw": cinw
        })

    for pay in payments:
        activity.append({
            "id": pay["id"],
            "date": pay.get("payment_date") or (pay.get("created_at") or "")[:10],
            "type": "Payment",
            "reference": pay.get("ref_number") or pay.get("bill_number") or "Payment",
            "amount": float(pay.get("amount") or 0),
            "status": "Paid",
            "raw": pay
        })

    activity.sort(key=lambda x: str(x["date"]), reverse=True)

    return {
        "vendor": vendor,
        "summary": {
            "total_purchases": tot_purchases,
            "total_paid": tot_paid,
            "total_outstanding": tot_outstanding
        },
        "purchase_bills": bills,
        "inwards": inwards,
        "payments": payments,
        "activity": activity
    }

@api_router.post("/vendors/{vendor_id}/purchase-bills")
async def create_purchase_bill(vendor_id: str, data: PurchaseBillPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    vendor = await db.vendors.find_one({"id": vendor_id, "company_id": cid})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    items_doc = []
    for item in data.items:
        qty = float(item.quantity or 0)
        items_doc.append({
            "product_name": item.product_name.strip(),
            "quantity": qty,
            "unit": item.unit or "Nos",
            "rate": float(item.rate or 0),
            "gst_rate": float(item.gst_rate or 12.0),
            "amount": float(item.amount or (qty * float(item.rate or 0))),
            "received_qty": 0.0,
            "remaining_qty": qty
        })

    pbill_id = f"pbill_{uuid.uuid4().hex[:12]}"
    pbill_doc = {
        "id": pbill_id,
        "company_id": cid,
        "vendor_id": vendor_id,
        "vendor_name": vendor.get("name"),
        "bill_number": data.bill_number.strip(),
        "bill_date": data.bill_date or now_iso()[:10],
        "due_date": data.due_date or "",
        "po_reference": data.po_reference or "",
        "items": items_doc,
        "subtotal": float(data.subtotal or 0),
        "gst_total": float(data.gst_total or 0),
        "freight_charges": float(data.freight_charges or 0),
        "transport_charges": float(data.transport_charges or 0),
        "other_charges": float(data.other_charges or 0),
        "grand_total": float(data.grand_total or 0),
        "notes": data.notes or "",
        "attachment_url": data.attachment_url or "",
        "payment_terms": data.payment_terms or "",
        "project_id": data.project_id or "",
        "status": "Bill Created",
        "inward_status": "Pending Inward",
        "paid_amount": 0.0,
        "created_by": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso()
    }

    await db.purchase_bills.insert_one(pbill_doc)
    await log_activity(cid, user["id"], user["name"], "Created Purchase Bill", f"Bill: {data.bill_number} for Vendor: {vendor.get('name')}")
    pbill_doc.pop("_id", None)
    return {"message": "Purchase bill saved successfully", "purchase_bill": pbill_doc}

@api_router.get("/purchase-bills/{bill_id}")
async def get_purchase_bill_detail(bill_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    bill = await db.purchase_bills.find_one({"id": bill_id, "company_id": cid}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Purchase bill not found")

    bill = await _enrich_bill_challan(bill, cid)
    inwards = await db.vendor_inwards.find({"bill_id": bill_id, "company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    payments = await db.vendor_payments.find({"bill_id": bill_id, "company_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(1000)

    return {"purchase_bill": bill, "inwards": inwards, "payments": payments}

@api_router.post("/purchase-bills/{bill_id}/inward")
async def create_material_inward(bill_id: str, data: MaterialInwardPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    bill = await db.purchase_bills.find_one({"id": bill_id, "company_id": cid})
    if not bill:
        raise HTTPException(status_code=404, detail="Purchase bill not found")

    inward_items_doc = []
    updated_bill_items = list(bill.get("items") or [])

    for item_input in data.items:
        rec_now = float(item_input.received_now or 0)
        inward_items_doc.append({
            "product_name": item_input.product_name,
            "bill_qty": float(item_input.bill_qty or 0),
            "received_now": rec_now,
            "remaining_qty": float(item_input.remaining_qty or 0),
            "destination": item_input.destination or data.warehouse_id or "Main Warehouse",
            "project_id": item_input.project_id or data.project_id or "",
            "unit": item_input.unit or "Nos"
        })

        # Update remaining and received qty on original purchase bill line item
        for b_item in updated_bill_items:
            if b_item.get("product_name") == item_input.product_name:
                cur_rec = float(b_item.get("received_qty") or 0)
                tot_qty = float(b_item.get("quantity") or 0)
                new_rec = cur_rec + rec_now
                b_item["received_qty"] = new_rec
                b_item["remaining_qty"] = max(0.0, tot_qty - new_rec)

    # Determine updated inward status on the bill
    all_completed = all(float(b.get("remaining_qty") or 0) <= 0 for b in updated_bill_items)
    inward_status = "Fully Received" if all_completed else "Partially Received"

    c_num = data.challan_number.strip() if data.challan_number else ""
    pb_set = {
        "items": updated_bill_items,
        "inward_status": inward_status,
        "updated_at": now_iso()
    }
    if c_num:
        pb_set["challan_number"] = c_num
        pb_set["challan_no"] = c_num
        pb_set["reference_number"] = c_num

    await db.purchase_bills.update_one(
        {"id": bill_id, "company_id": cid},
        {"$set": pb_set}
    )

    vinw_doc = {
        "id": f"vinw_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "bill_id": bill_id,
        "bill_number": bill.get("bill_number"),
        "vendor_id": bill.get("vendor_id"),
        "vendor_name": bill.get("vendor_name"),
        "challan_number": data.challan_number.strip(),
        "challan_date": data.challan_date or now_iso()[:10],
        "received_by": data.received_by or user["name"],
        "warehouse_id": data.warehouse_id or "Main Warehouse",
        "project_id": data.project_id or "",
        "attachment_url": data.attachment_url or "",
        "items": inward_items_doc,
        "created_at": now_iso()
    }
    await db.vendor_inwards.insert_one(vinw_doc)

    # ── Connect to Inventory (inward_entries) ──────────────────────────────
    # For every item received with received_now > 0, create standard inward entry in db.inward_entries
    for item in data.items:
        rec_now = float(item.received_now or 0)
        if rec_now > 0:
            inw_entry_doc = {
                "id": f"inw_{uuid.uuid4().hex[:12]}",
                "company_id": cid,
                "date": data.challan_date or now_iso()[:10],
                "source_name": bill.get("vendor_name"),
                "source_type": "Vendor",
                "challan_no": data.challan_number,
                "product": item.product_name,
                "size": "Default",
                "unit": item.unit or "Nos",
                "quantity": rec_now,
                "rate": 0,
                "total_cost": 0,
                "warehouse_id": item.destination or data.warehouse_id or "Main Warehouse",
                "project_id": item.project_id or data.project_id or "",
                "created_by": user["name"],
                "created_at": now_iso()
            }
            await db.inward_entries.insert_one(inw_entry_doc)

    await log_activity(cid, user["id"], user["name"], "Recorded Material Inward", f"Challan: {data.challan_number} for Bill: {bill.get('bill_number')}")
    vinw_doc.pop("_id", None)
    return {"message": "Material inward recorded successfully", "inward": vinw_doc}

@api_router.post("/vendors/{vendor_id}/payments")
async def create_vendor_payment(vendor_id: str, data: VendorPaymentPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    vendor = await db.vendors.find_one({"id": vendor_id, "company_id": cid})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    pay_amt = float(data.amount or 0)
    vpay_doc = {
        "id": f"vpay_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "vendor_id": vendor_id,
        "vendor_name": vendor.get("name"),
        "bill_id": data.bill_id or "",
        "bill_number": data.bill_number or "",
        "amount": pay_amt,
        "payment_method": data.payment_method or "Bank Transfer",
        "ref_number": data.ref_number or "",
        "payment_date": data.payment_date or now_iso()[:10],
        "notes": data.notes or "",
        "created_by": user["name"],
        "created_at": now_iso()
    }
    await db.vendor_payments.insert_one(vpay_doc)

    # If linked to a purchase bill, update bill paid_amount and payment status
    if data.bill_id:
        bill = await db.purchase_bills.find_one({"id": data.bill_id, "company_id": cid})
        if bill:
            prev_paid = float(bill.get("paid_amount") or 0)
            new_paid = prev_paid + pay_amt
            g_total = float(bill.get("grand_total") or 0)
            pay_status = "Paid" if new_paid >= g_total else "Partially Paid"
            await db.purchase_bills.update_one(
                {"id": data.bill_id, "company_id": cid},
                {"$set": {"paid_amount": new_paid, "status": pay_status, "updated_at": now_iso()}}
            )

    await log_activity(cid, user["id"], user["name"], "Recorded Vendor Payment", f"Amount: ₹{pay_amt} to Vendor: {vendor.get('name')}")
    vpay_doc.pop("_id", None)
    return {"message": "Vendor payment recorded successfully", "payment": vpay_doc}

@api_router.get("/clients/{client_id}/warranties")
async def get_client_warranties(client_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    warranties = await db.warranties.find({"client_id": client_id, "company_id": cid}, {"_id": 0}).to_list(1000)
    return {"warranties": warranties}

@api_router.post("/clients/{client_id}/warranties")
async def create_client_warranty(client_id: str, data: WarrantyPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    warranty_doc = {
        "id": f"war_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "client_id": client_id,
        "product_type": data.product_type,
        "brand": data.brand or "",
        "model": data.model or "",
        "serial_number": data.serial_number or "",
        "installation_date": data.installation_date or "",
        "warranty_start": data.warranty_start or now_iso()[:10],
        "warranty_end": data.warranty_end or "",
        "warranty_type": data.warranty_type or "Standard",
        "provider": data.provider or "",
        "status": "Active",
        "created_at": now_iso()
    }
    await db.warranties.insert_one(warranty_doc)
    await log_activity(cid, user["id"], user["name"], "Created Warranty Record", f"Product: {data.product_type} for Client: {client_id}")
    return {"message": "Warranty created successfully", "warranty": warranty_doc}

@api_router.get("/clients/{client_id}/service-visits")
async def get_client_service_visits(client_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    visits = await db.service_visits.find({"client_id": client_id, "company_id": cid}, {"_id": 0}).sort("visit_date", -1).to_list(1000)
    return {"visits": visits}

@api_router.post("/clients/{client_id}/service-visits")
async def create_client_service_visit(client_id: str, data: ServiceVisitPayload, user=Depends(get_current_user)):
    cid = user["company_id"]
    visit_doc = {
        "id": f"sv_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "client_id": client_id,
        "visit_date": data.visit_date or now_iso()[:10],
        "technician_name": data.technician_name or user["name"],
        "visit_type": data.visit_type or "Routine Maintenance",
        "system_status": data.system_status or "Operational",
        "generation_obs": data.generation_obs or "",
        "earth_resistance": data.earth_resistance or "",
        "fuses_status": data.fuses_status or "OK",
        "inverter_condition": data.inverter_condition or "Good",
        "panel_condition": data.panel_condition or "Good",
        "customer_remarks": data.customer_remarks or "",
        "technician_remarks": data.technician_remarks or "",
        "next_visit_date": data.next_visit_date or "",
        "created_at": now_iso()
    }
    await db.service_visits.insert_one(visit_doc)
    await log_activity(cid, user["id"], user["name"], "Recorded Service Visit", f"Client ID: {client_id}")
    return {"message": "Service visit logged successfully", "visit": visit_doc}

# ---------- Pincode Auto-Lookup ----------
@api_router.get("/location/pincode/{pincode}")
async def lookup_pincode(pincode: str):
    code = (pincode or "").strip()
    if not code or len(code) != 6 or not code.isdigit():
        raise HTTPException(status_code=400, detail="Invalid 6-digit pincode")

    import httpx
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            res = await client.get(f"https://api.postalpincode.in/pincode/{code}")
            if res.status_code == 200:
                data = res.json()
                if isinstance(data, list) and len(data) > 0 and data[0].get("Status") == "Success":
                    po_list = data[0].get("PostOffice") or []
                    if po_list and isinstance(po_list, list) and len(po_list) > 0:
                        first_po = po_list[0]
                        city = first_po.get("Name") or first_po.get("Block") or first_po.get("District") or ""
                        district = first_po.get("District") or ""
                        state = first_po.get("State") or ""
                        return {
                            "pincode": code,
                            "city": city,
                            "district": district,
                            "state": state,
                            "post_offices": [po.get("Name") for po in po_list if po.get("Name")]
                        }
    except Exception as e:
        logger.warning(f"Pincode API lookup error: {e}")

    return {"pincode": code, "city": "", "district": "", "state": "", "post_offices": []}


@api_router.get("/location/city/{city}")
async def lookup_city(city: str):
    term = (city or "").strip()
    if not term or len(term) < 2:
        return {"query": term, "results": []}

    import httpx
    try:
        async with httpx.AsyncClient(timeout=4.0) as client:
            res = await client.get(f"https://api.postalpincode.in/postoffice/{term}")
            if res.status_code == 200:
                data = res.json()
                if isinstance(data, list) and len(data) > 0 and data[0].get("Status") == "Success":
                    po_list = data[0].get("PostOffice") or []
                    results = []
                    seen_combos = set()
                    for po in po_list:
                        pincode = po.get("Pincode") or ""
                        post_name = po.get("Name") or ""
                        district = po.get("District") or ""
                        state = po.get("State") or ""
                        combo_key = f"{post_name}_{pincode}"
                        if combo_key not in seen_combos:
                            seen_combos.add(combo_key)
                            results.append({
                                "name": post_name,
                                "city": post_name or district,
                                "district": district,
                                "state": state,
                                "pincode": pincode
                            })
                    return {"query": term, "results": results[:25]}
    except Exception as e:
        logger.warning(f"City location lookup error for '{term}': {e}")

    return {"query": term, "results": []}


# ---------- App Feedback ----------
class FeedbackIn(BaseModel):
    feedback_type: str
    message: str
    page: Optional[str] = ""
    screenshot_url: Optional[str] = ""

@api_router.post("/feedback")
async def submit_feedback(data: FeedbackIn, user=Depends(get_current_user)):
    if not data.message.strip():
        raise HTTPException(status_code=400, detail="Feedback message cannot be empty")

    doc = {
        "id": f"fb_{uuid.uuid4().hex[:12]}",
        "company_id": user["company_id"],
        "user_id": user["id"],
        "user_name": user.get("name", "User"),
        "user_email": user.get("email", ""),
        "feedback_type": data.feedback_type,
        "message": data.message.strip(),
        "page": data.page or "",
        "screenshot_url": data.screenshot_url or "",
        "created_at": now_iso(),
        "status": "New"
    }
    await db.feedback.insert_one(doc)
    await log_activity(user["company_id"], user["id"], user.get("name"), "Submitted Feedback", f"{data.feedback_type}: {data.message[:30]}...")
    doc.pop("_id", None)
    return {"message": "Feedback submitted successfully", "feedback": doc}

@api_router.get("/feedback")
async def list_feedback(user=Depends(get_current_user)):
    if not has_perm(user, "team", "view") and user.get("role") not in ("Super Admin", "Admin"):
        raise HTTPException(status_code=403, detail="Admin access required to view feedback")
    return await db.feedback.find({"company_id": user["company_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)


# ---------- Task Comments ----------
class TaskCommentIn(BaseModel):
    message: str

@api_router.get("/tasks/{task_id}/comments")
async def list_task_comments(task_id: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    task = await db.tasks.find_one({"id": task_id, "company_id": cid})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    is_assignee = (task.get("assigned_to") == user["id"] or task.get("assigned_to_email") == user["email"])
    is_creator = (task.get("created_by_id") == user["id"] or task.get("created_by") == user["name"])
    if not is_assignee and not is_creator and not has_perm(user, "task_portal", "view"):
        raise HTTPException(status_code=403, detail="Unauthorized to view task comments")

    comments = await db.task_comments.find({"task_id": task_id, "company_id": cid}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return comments

@api_router.post("/tasks/{task_id}/comments")
async def add_task_comment(task_id: str, data: TaskCommentIn, user=Depends(get_current_user)):
    cid = user["company_id"]
    task = await db.tasks.find_one({"id": task_id, "company_id": cid})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    is_assignee = (task.get("assigned_to") == user["id"] or task.get("assigned_to_email") == user["email"])
    is_creator = (task.get("created_by_id") == user["id"] or task.get("created_by") == user["name"])
    if not is_assignee and not is_creator and not has_perm(user, "task_portal", "edit"):
        raise HTTPException(status_code=403, detail="Unauthorized to add comments to this task")

    if not data.message.strip():
        raise HTTPException(status_code=400, detail="Comment message cannot be empty")

    comment_doc = {
        "id": f"tc_{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "task_id": task_id,
        "user_id": user["id"],
        "user_name": user.get("name", "User"),
        "user_email": user.get("email", ""),
        "user_role": user.get("role", "Staff"),
        "message": data.message.strip(),
        "created_at": now_iso()
    }
    await db.task_comments.insert_one(comment_doc)
    comment_doc.pop("_id", None)
    return comment_doc


# ---------- Global Search ----------
@api_router.get("/search/global")
async def global_search(q: str, user=Depends(get_current_user)):
    cid = user["company_id"]
    query_str = (q or "").strip().lower()
    if not query_str:
        return {"clients": [], "products": [], "vendors": [], "tasks": []}

    can_view_clients = has_perm(user, "clients", "view")
    can_view_inventory = has_perm(user, "data_management", "view")
    can_view_tasks = has_perm(user, "task_portal", "view")

    matched_clients = []
    if can_view_clients:
        clients = await db.clients.find({"company_id": cid}, {"_id": 0}).to_list(5000)
        matched_clients = [
            c for c in clients if
            query_str in (c.get("full_name") or "").lower() or
            query_str in (c.get("mobile") or "").lower() or
            query_str in (c.get("consumer_number") or "").lower() or
            query_str in (c.get("sol_id") or "").lower() or
            query_str in (c.get("inverter_serial") or "").lower()
        ][:20]

    matched_products = []
    matched_vendors = []
    if can_view_inventory:
        products = await db.products.find({"company_id": cid}, {"_id": 0}).to_list(5000)
        vendors = await db.vendors.find({"company_id": cid}, {"_id": 0}).to_list(5000)
        matched_products = [
            p for p in products if
            query_str in (p.get("name") or p.get("product_name") or "").lower() or
            query_str in (p.get("category") or "").lower() or
            query_str in (p.get("brand") or "").lower()
        ][:20]

        matched_vendors = [
            v for v in vendors if
            query_str in (v.get("name") or "").lower() or
            query_str in (v.get("gstin") or "").lower() or
            query_str in (v.get("phone") or "").lower()
        ][:20]

    matched_tasks = []
    if can_view_tasks:
        tasks = await db.tasks.find({"company_id": cid}, {"_id": 0}).to_list(5000)
        matched_tasks = [
            t for t in tasks if
            query_str in (t.get("title") or "").lower() or
            query_str in (t.get("assigned_to") or "").lower()
        ][:20]

    return {
        "clients": matched_clients,
        "products": matched_products,
        "vendors": matched_vendors,
        "tasks": matched_tasks
    }


# ---------- LEVEL 1: SOLRIX PLATFORM OWNER CONTROL CENTER ----------

class PlatformSubscriptionActionIn(BaseModel):
    action: str  # assign_plan, extend_trial, update_expiry, change_status, cancel
    plan_id: Optional[str] = None
    status: Optional[str] = None
    trial_days: Optional[int] = 0
    expiry_date: Optional[str] = None
    reason: Optional[str] = ""

class PlatformFeatureEntitlementsIn(BaseModel):
    feature_entitlements: Dict[str, bool]
    temporary_features: Optional[Dict[str, Any]] = None
    reason: Optional[str] = ""

class PlatformNotificationIn(BaseModel):
    target_type: str  # all, company, user
    target_company_id: Optional[str] = ""
    target_user_id: Optional[str] = ""
    title: str
    message: str
    type: Optional[str] = "info"

class PlatformFeedbackUpdateIn(BaseModel):
    status: str
    admin_notes: Optional[str] = ""

async def _log_platform_audit(user: dict, action: str, target_company_id: str, target_name: str, old_val: Any = None, new_val: Any = None, reason: str = ""):
    doc = {
        "id": f"pa_{uuid.uuid4().hex[:12]}",
        "timestamp": now_iso(),
        "admin_id": user.get("id"),
        "admin_name": user.get("name", "Platform Owner"),
        "admin_email": user.get("email", ""),
        "action": action,
        "target_company_id": target_company_id,
        "target_name": target_name,
        "old_value": old_val,
        "new_value": new_val,
        "reason": reason
    }
    try:
        await db.platform_audit_logs.insert_one(doc)
    except Exception as e:
        logger.warning(f"Failed to record platform audit log: {e}")

@api_router.get("/admin/dashboard")
@api_router.get("/platform-owner/dashboard")
async def get_platform_dashboard(user=Depends(require_super_admin())):
    total_companies = await db.companies.count_documents({})
    active_companies = await db.companies.count_documents({"subscription_status": "active"})
    trial_companies = await db.companies.count_documents({"subscription_status": "trialing"})
    
    # Calculate expiring in 7 days
    now = datetime.now(timezone.utc)
    cutoff = (now + timedelta(days=7)).isoformat()
    expiring_companies = await db.companies.count_documents({
        "trial_ends_at": {"$lte": cutoff, "$gte": now.isoformat()}
    })

    total_users = await db.users.count_documents({})
    total_projects = await db.projects.count_documents({})
    total_clients = await db.clients.count_documents({})
    total_invoices = await db.invoices.count_documents({})

    recent_companies = await db.companies.find({}, {"_id": 0}).sort("created_at", -1).to_list(10)
    recent_activity = await db.activity_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(15)
    recent_feedback = await db.feedback.find({}, {"_id": 0}).sort("created_at", -1).to_list(10)

    # Calculate real MRR from active subscriptions and dynamic plan pricing
    active_docs = await db.companies.find({"subscription_status": "active"}, {"_id": 0, "plan_id": 1}).to_list(10000)
    db_plans_list = await db.plans_config.find({}, {"_id": 0}).to_list(100)
    from plan_config import get_all_plans
    all_plans_dict = get_all_plans(db_plans_list=db_plans_list)
    
    mrr_val = 0.0
    for ac in active_docs:
        pid = (ac.get("plan_id") or "starter").lower()
        p_info = all_plans_dict.get(pid) or all_plans_dict.get("starter", {})
        mrr_val += float(p_info.get("monthly_price", 0))

    return {
        "kpis": {
            "total_customers": total_companies,
            "active_customers": active_companies,
            "trial_customers": trial_companies,
            "expiring_subscriptions": expiring_companies,
            "mrr_revenue": f"₹{mrr_val:,.0f} / mo" if mrr_val > 0 else "₹0 / mo",
            "total_users": total_users,
            "total_projects": total_projects,
            "total_clients": total_clients,
            "total_invoices": total_invoices,
            "storage_usage": f"{(total_projects * 0.02 + total_invoices * 0.005 + 0.1):.2f} GB / Standard"
        },
        "recent_signups": recent_companies,
        "recent_activity": recent_activity,
        "recent_feedback": recent_feedback
    }

@api_router.get("/admin/customers")
@api_router.get("/platform-owner/customers")
async def list_platform_customers(
    search: Optional[str] = "",
    status: Optional[str] = "",
    plan: Optional[str] = "",
    user=Depends(require_super_admin())
):
    query = {}
    if status and status != "all":
        query["subscription_status"] = status
    if plan and plan != "all":
        query["plan_id"] = plan

    companies = await db.companies.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

    s_term = (search or "").strip().lower()
    results = []
    for c in companies:
        cid = c["id"]
        owner = await db.users.find_one({"company_id": cid, "user_type": "owner"}, {"_id": 0, "password_hash": 0})
        if not owner:
            owner = await db.users.find_one({"company_id": cid}, {"_id": 0, "password_hash": 0}) or {}

        company_name = c.get("company_name", "")
        owner_name = owner.get("name", "")
        owner_email = owner.get("email", "")
        owner_mobile = owner.get("mobile", "")

        if s_term:
            match = (
                s_term in company_name.lower() or
                s_term in owner_name.lower() or
                s_term in owner_email.lower() or
                s_term in owner_mobile.lower() or
                s_term in cid.lower()
            )
            if not match:
                continue

        user_count = await db.users.count_documents({"company_id": cid})
        project_count = await db.projects.count_documents({"company_id": cid})
        client_count = await db.clients.count_documents({"company_id": cid})

        results.append({
            "id": cid,
            "company_name": company_name,
            "owner_name": owner_name,
            "email": owner_email,
            "mobile": owner_mobile,
            "plan_id": c.get("plan_id", "starter"),
            "subscription_status": c.get("subscription_status", "trialing"),
            "trial_ends_at": c.get("trial_ends_at"),
            "user_count": user_count,
            "project_count": project_count,
            "client_count": client_count,
            "created_at": c.get("created_at"),
            "city": c.get("city", ""),
            "state": c.get("state", "")
        })

    return results

@api_router.get("/admin/customers/{company_id}")
@api_router.get("/platform-owner/customers/{company_id}")
async def get_platform_customer_detail(company_id: str, user=Depends(require_super_admin())):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Customer workspace not found")

    owner = await db.users.find_one({"company_id": company_id, "user_type": "owner"}, {"_id": 0, "password_hash": 0})
    raw_team = await db.users.find({"company_id": company_id}, {"_id": 0, "password_hash": 0}).to_list(500)
    team_users = [u for u in raw_team if is_internal_team_user(u)]

    # Compute usage metrics
    user_count = len(team_users)
    client_count = await db.clients.count_documents({"company_id": company_id})
    project_count = await db.projects.count_documents({"company_id": company_id})
    task_count = await db.tasks.count_documents({"company_id": company_id})
    invoice_count = await db.invoices.count_documents({"company_id": company_id})
    product_count = await db.products.count_documents({"company_id": company_id})
    po_count = await db.purchase_orders.count_documents({"company_id": company_id})
    material_request_count = await db.material_requests.count_documents({"company_id": company_id})
    document_count = await db.documents.count_documents({"company_id": company_id})

    recent_activity = await db.activity_logs.find({"company_id": company_id}, {"_id": 0}).sort("created_at", -1).to_list(25)

    return {
        "company": company,
        "owner": owner or {},
        "team_users": team_users,
        "usage": {
            "users": user_count,
            "clients": client_count,
            "projects": project_count,
            "tasks": task_count,
            "invoices": invoice_count,
            "products": product_count,
            "purchase_orders": po_count,
            "material_requests": material_request_count,
            "documents": document_count,
            "storage": f"{(project_count * 0.02 + invoice_count * 0.005 + document_count * 0.001):.2f} GB"
        },
        "recent_activity": recent_activity
    }

@api_router.post("/admin/customers/{company_id}/subscription")
@api_router.post("/platform-owner/customers/{company_id}/subscription")
async def update_platform_customer_subscription(
    company_id: str,
    data: PlatformSubscriptionActionIn,
    user=Depends(require_super_admin())
):
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Customer workspace not found")

    old_plan = company.get("plan_id", "starter")
    old_status = company.get("subscription_status", "trialing")
    update_doc = {}

    if data.action == "assign_plan" or data.plan_id:
        update_doc["plan_id"] = data.plan_id or old_plan
    if data.status:
        update_doc["subscription_status"] = data.status
    if data.action == "extend_trial" and data.trial_days:
        curr_end = company.get("trial_ends_at")
        curr_dt = datetime.fromisoformat(curr_end.replace("Z", "+00:00")) if curr_end else datetime.now(timezone.utc)
        new_dt = curr_dt + timedelta(days=data.trial_days)
        update_doc["trial_ends_at"] = new_dt.isoformat()
        update_doc["subscription_status"] = "trialing"
    if data.expiry_date:
        update_doc["trial_ends_at"] = data.expiry_date

    if update_doc:
        await db.companies.update_one({"id": company_id}, {"$set": update_doc}, upsert=True)
        _cache_invalidate_company(company_id)

    await _log_platform_audit(
        user=user,
        action=f"Subscription Action: {data.action}",
        target_company_id=company_id,
        target_name=company.get("company_name", company_id),
        old_val={"plan_id": old_plan, "status": old_status},
        new_val=update_doc,
        reason=data.reason or ""
    )

    return {"message": "Subscription updated successfully", "updated": update_doc}

@api_router.post("/admin/customers/{company_id}/features")
@api_router.post("/platform-owner/customers/{company_id}/features")
async def update_platform_customer_features(
    company_id: str,
    data: PlatformFeatureEntitlementsIn,
    user=Depends(require_super_admin())
):
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Customer workspace not found")

    old_entitlements = company.get("feature_entitlements", {})
    update_doc = {"feature_entitlements": data.feature_entitlements}
    if data.temporary_features is not None:
        update_doc["temporary_features"] = data.temporary_features

    await db.companies.update_one(
        {"id": company_id},
        {"$set": update_doc},
        upsert=True
    )
    _cache_invalidate_company(company_id)

    await _log_platform_audit(
        user=user,
        action="Updated Feature Entitlements",
        target_company_id=company_id,
        target_name=company.get("company_name", company_id),
        old_val=old_entitlements,
        new_val=update_doc,
        reason=data.reason or ""
    )

    return {"message": "Feature entitlements updated", "feature_entitlements": data.feature_entitlements, "temporary_features": data.temporary_features}

class PlatformPlanUpdateIn(BaseModel):
    name: str
    tagline: Optional[str] = ""
    monthly_price: float
    yearly_price: float
    max_users: int
    max_clients: int
    active: Optional[bool] = True
    features: Dict[str, bool]

class PlatformOfferIn(BaseModel):
    title: str
    description: str
    offer_code: Optional[str] = ""
    start_date: Optional[str] = ""
    end_date: Optional[str] = ""
    target_plan: Optional[str] = "all"
    cta_text: Optional[str] = "Upgrade Now"
    cta_url: Optional[str] = "/pricing"

class PageTrackIn(BaseModel):
    page_path: str
    page_name: Optional[str] = ""
    duration_seconds: Optional[int] = 0

@api_router.get("/admin/plans")
@api_router.get("/platform-owner/plans")
async def list_platform_plans(user=Depends(require_super_admin())):
    from plan_config import PLANS, get_all_plans
    db_plans = await db.plans_config.find({}, {"_id": 0}).to_list(100)
    db_map = {p["id"]: p for p in db_plans}
    
    merged = {}
    default_plans = get_all_plans()
    for pid, pdata in default_plans.items():
        if pid in db_map:
            merged[pid] = {**pdata, **db_map[pid]}
        else:
            merged[pid] = pdata
    return merged

@api_router.put("/admin/plans/{plan_id}")
@api_router.put("/platform-owner/plans/{plan_id}")
async def update_platform_plan(plan_id: str, data: PlatformPlanUpdateIn, user=Depends(require_super_admin())):
    doc = {
        "id": plan_id.lower(),
        "name": data.name,
        "tagline": data.tagline or "",
        "monthly_price": data.monthly_price,
        "yearly_price": data.yearly_price,
        "max_users": data.max_users,
        "max_clients": data.max_clients,
        "active": data.active if data.active is not None else True,
        "features": data.features,
        "updated_at": now_iso()
    }
    await db.plans_config.update_one({"id": doc["id"]}, {"$set": doc}, upsert=True)
    await _log_platform_audit(
        user=user,
        action="Updated Plan Entitlements & Pricing",
        target_company_id="ALL",
        target_name=f"Plan: {data.name}",
        new_val=doc
    )
    return {"message": f"Plan '{data.name}' updated successfully", "plan": doc}

@api_router.get("/admin/offers")
@api_router.get("/platform-owner/offers")
async def list_platform_offers(user=Depends(require_super_admin())):
    return await db.promotional_offers.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api_router.post("/admin/offers")
@api_router.post("/platform-owner/offers")
async def create_platform_offer(data: PlatformOfferIn, user=Depends(require_super_admin())):
    doc = {
        "id": f"off_{uuid.uuid4().hex[:10]}",
        "title": data.title,
        "description": data.description,
        "offer_code": data.offer_code or "",
        "start_date": data.start_date or now_iso(),
        "end_date": data.end_date or "",
        "target_plan": data.target_plan or "all",
        "cta_text": data.cta_text or "Upgrade Now",
        "cta_url": data.cta_url or "/pricing",
        "created_at": now_iso()
    }
    await db.promotional_offers.insert_one(doc)
    await _log_platform_audit(
        user=user,
        action="Created Promotional Offer",
        target_company_id="ALL",
        target_name=data.title,
        new_val=doc
    )
    return {"message": "Offer created successfully", "offer": doc}

@api_router.post("/analytics/track-page")
async def track_page_visit(data: PageTrackIn, request: Request):
    user_id = "anonymous"
    company_id = "global"
    try:
        auth_header = request.headers.get("Authorization", "")
        token = request.cookies.get("access_token") or (auth_header[7:] if auth_header.startswith("Bearer ") else None)
        if token:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = payload.get("sub") or payload.get("user_id") or user_id
            company_id = payload.get("company_id") or company_id
    except Exception:
        pass

    event_doc = {
        "id": f"pe_{uuid.uuid4().hex[:12]}",
        "page_path": data.page_path,
        "page_name": data.page_name or data.page_path,
        "user_id": user_id,
        "company_id": company_id,
        "duration_seconds": data.duration_seconds or 0,
        "timestamp": now_iso()
    }
    asyncio.create_task(db.page_events.insert_one(event_doc))
    return {"status": "ok"}

@api_router.get("/admin/analytics/performance")
@api_router.get("/platform-owner/analytics/performance")
async def get_platform_performance_analytics(user=Depends(require_super_admin())):
    now = datetime.now(timezone.utc)
    d1 = (now - timedelta(days=1)).isoformat()
    d7 = (now - timedelta(days=7)).isoformat()
    d30 = (now - timedelta(days=30)).isoformat()

    # DAU / WAU / MAU calculation
    dau_uids = await db.activity_logs.distinct("user_id", {"created_at": {"$gte": d1}})
    wau_uids = await db.activity_logs.distinct("user_id", {"created_at": {"$gte": d7}})
    mau_uids = await db.activity_logs.distinct("user_id", {"created_at": {"$gte": d30}})

    new_signups_30d = await db.companies.count_documents({"created_at": {"$gte": d30}})
    
    # Activity sample by feature/action
    activities = await db.activity_logs.find({}, {"_id": 0, "action": 1, "company_id": 1}).sort("created_at", -1).to_list(5000)
    action_counts = {}
    for a in activities:
        act = a.get("action", "Other Action")
        action_counts[act] = action_counts.get(act, 0) + 1

    sorted_features = sorted(action_counts.items(), key=lambda x: x[1], reverse=True)

    # Usage breakdown by plan
    companies = await db.companies.find({}, {"_id": 0, "id": 1, "plan_id": 1}).to_list(2000)
    plan_company_map = {c["id"]: c.get("plan_id", "starter") for c in companies}

    plan_usage = {"starter": 0, "growth": 0, "pro": 0, "enterprise": 0}
    for a in activities:
        cid = a.get("company_id")
        p = plan_company_map.get(cid, "starter")
        if p in plan_usage:
            plan_usage[p] += 1
        else:
            plan_usage["starter"] += 1

    return {
        "dau": len(dau_uids),
        "wau": len(wau_uids),
        "mau": len(mau_uids),
        "signups_30d": new_signups_30d,
        "most_used_features": [{"feature": k, "count": v} for k, v in sorted_features[:8]],
        "least_used_features": [{"feature": k, "count": v} for k, v in sorted_features[-5:]],
        "usage_by_plan": plan_usage
    }

@api_router.get("/admin/analytics/pages")
@api_router.get("/platform-owner/analytics/pages")
async def get_platform_page_analytics(user=Depends(require_super_admin())):
    page_sample = await db.page_events.find({}, {"_id": 0}).sort("timestamp", -1).to_list(3000)
    page_stats = {}
    for p in page_sample:
        path = p.get("page_path", "/dashboard")
        name = p.get("page_name") or path
        if path not in page_stats:
            page_stats[path] = {
                "path": path,
                "name": name,
                "visits": 0,
                "users": set(),
                "total_duration": 0,
                "last_used": p.get("timestamp")
            }
        page_stats[path]["visits"] += 1
        if p.get("user_id"):
            page_stats[path]["users"].add(p["user_id"])
        page_stats[path]["total_duration"] += p.get("duration_seconds", 0)

    result = []
    for k, v in page_stats.items():
        unique_cnt = len(v["users"])
        avg_time = round(v["total_duration"] / v["visits"], 1) if v["visits"] > 0 else 0
        result.append({
            "path": v["path"],
            "name": v["name"],
            "visits": v["visits"],
            "unique_users": unique_cnt,
            "avg_time_sec": avg_time,
            "last_used": v["last_used"]
        })

    result.sort(key=lambda x: x["visits"], reverse=True)
    return result

@api_router.get("/admin/analytics")
@api_router.get("/platform-owner/analytics")
async def get_platform_analytics(user=Depends(require_super_admin())):
    activity_sample = await db.activity_logs.find({}, {"_id": 0, "action": 1, "created_at": 1}).sort("created_at", -1).to_list(5000)
    action_counts = {}
    for a in activity_sample:
        act = a.get("action", "Other")
        action_counts[act] = action_counts.get(act, 0) + 1

    sorted_actions = sorted(action_counts.items(), key=lambda x: x[1], reverse=True)
    return {
        "most_used_features": [{"feature": k, "count": v} for k, v in sorted_actions[:10]],
        "least_used_features": [{"feature": k, "count": v} for k, v in sorted_actions[-5:]],
        "total_events_sample": len(activity_sample)
    }

@api_router.get("/admin/feedback")
@api_router.get("/platform-owner/feedback")
async def list_platform_feedback(user=Depends(require_super_admin())):
    return await db.feedback.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api_router.put("/admin/feedback/{feedback_id}")
@api_router.put("/platform-owner/feedback/{feedback_id}")
async def update_platform_feedback(feedback_id: str, data: PlatformFeedbackUpdateIn, user=Depends(require_super_admin())):
    fb = await db.feedback.find_one({"id": feedback_id})
    if not fb:
        raise HTTPException(status_code=404, detail="Feedback record not found")

    res = await db.feedback.update_one(
        {"id": feedback_id},
        {"$set": {"status": data.status, "admin_notes": data.admin_notes or "", "updated_at": now_iso()}}
    )

    # Optional: Send notification reply to user
    if data.admin_notes and fb.get("company_id"):
        await push_notification(
            company_id=fb["company_id"],
            audience="user",
            title=f"Feedback Status: {data.status}",
            body=f"Admin Note on your report '{fb.get('feedback_type', 'Feedback')}': {data.admin_notes}",
            to_user_id=fb.get("user_id")
        )

    await _log_platform_audit(
        user=user,
        action=f"Updated Feedback Status to {data.status}",
        target_company_id=fb.get("company_id", "GLOBAL"),
        target_name=f"Feedback #{feedback_id}",
        new_val={"status": data.status, "admin_notes": data.admin_notes}
    )

    return {"message": "Feedback status updated successfully"}

@api_router.post("/admin/notifications")
@api_router.post("/platform-owner/notifications")
async def send_platform_notification(data: PlatformNotificationIn, user=Depends(require_super_admin())):
    if data.target_type == "company" and data.target_company_id:
        await push_notification(data.target_company_id, "admin", data.title, data.message)
    elif data.target_type == "user" and data.target_user_id:
        target_u = await db.users.find_one({"id": data.target_user_id})
        if target_u:
            await push_notification(target_u["company_id"], "user", data.title, data.message, to_user_id=data.target_user_id)
    else:
        # Broadcast to all active companies
        companies = await db.companies.find({}, {"id": 1}).to_list(1000)
        for c in companies:
            await push_notification(c["id"], "admin", data.title, data.message)

    await _log_platform_audit(
        user=user,
        action=f"Sent Platform Notification ({data.target_type})",
        target_company_id=data.target_company_id or "ALL",
        target_name=data.title,
        new_val=data.message
    )

    return {"message": "Notification dispatched successfully"}

@api_router.get("/admin/health")
@api_router.get("/platform-owner/health")
async def get_platform_health(user=Depends(require_super_admin())):
    db_ok = True
    try:
        await db.companies.find_one({}, {"_id": 1})
    except Exception:
        db_ok = False

    return {
        "services": [
            {"name": "Database (MongoDB)", "status": "Healthy" if db_ok else "Error", "latency": "2ms"},
            {"name": "API Service (FastAPI)", "status": "Healthy", "latency": "1ms"},
            {"name": "Storage Service (Supabase)", "status": "Healthy", "latency": "18ms"},
            {"name": "PDF Generation Service", "status": "Healthy", "latency": "5ms"},
            {"name": "Word (.docx) Generator", "status": "Healthy", "latency": "3ms"},
            {"name": "Email Notification Gateway", "status": "Not monitored", "latency": "N/A"}
        ]
    }

@api_router.get("/platform-owner/audit-logs")
async def list_platform_audit_logs(user=Depends(require_platform_owner())):
    return await db.platform_audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(500)


from billing_router import billing_router
app.include_router(billing_router)
app.include_router(api_router)

DEFAULT_CORS_ORIGINS = [
    "https://solarix-cumx-sable.vercel.app",
    "https://solarix.vercel.app",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:5173",
    "http://127.0.0.1:5173"
]

cors_origins_env = os.environ.get('CORS_ORIGINS', '')
env_origins = [o.strip() for o in cors_origins_env.split(',') if o.strip()]

allowed_cors_origins = list(dict.fromkeys(DEFAULT_CORS_ORIGINS + env_origins))

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=allowed_cors_origins,
    allow_origin_regex=r"https://.*\.vercel\.app|http://localhost:.*|http://127\.0\.0\.1:.*",
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)

@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response

